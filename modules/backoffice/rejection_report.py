"""
modules/backoffice/rejection_report.py
=======================================
Production Rejection Report — Admin / Manager

Reads from job_stage_events (stage_code = 'REJECTED') — single source of truth.
Filters: date range, reason dropdown, lens type, eye, order
Metrics: total, by reason, by lens type, reprocess count, value at risk
"""
from __future__ import annotations
import streamlit as st
from datetime import date, timedelta


def _q(sql, params=None):
    try:
        from modules.sql_adapter import run_query
        return run_query(sql, params or {}) or []
    except Exception as e:
        st.error(f"Query error: {e}")
        return []


_KNOWN_REASONS = [
    "All Reasons",
    "Production Issue",
    "Power Issue — wrong SPH/CYL ground",
    "Lens made very thin",
    "Vibrations during surfacing",
    "Hardcoat / Coating failure",
    "Scratch / surface defect",
    "Wrong blank used",
    "Other",
]


def render_rejection_report():
    from modules.security.roles import require_role
    require_role("admin", "manager")

    st.markdown("## 📋 Production Rejection Report")
    st.caption("All lens rejections recorded during production. Source: job stage events.")

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        f1, f2, f3, f4, f5 = st.columns([2, 2, 2, 2, 2])

        date_from = f1.date_input("From", value=date.today() - timedelta(days=30),
                                   key="rr_from", format="DD/MM/YYYY")
        date_to   = f2.date_input("To",   value=date.today(),
                                   key="rr_to", format="DD/MM/YYYY")
        eye_filter    = f3.selectbox("Eye Side", ["All", "R — Right Eye", "L — Left Eye"],
                                     key="rr_eye")
        reason_filter = f4.selectbox("Rejection Reason", _KNOWN_REASONS, key="rr_reason")
        order_filter  = f5.text_input("Order / Patient", placeholder="Search...",
                                       key="rr_order", label_visibility="collapsed")

    # ── Load data from job_stage_events ──────────────────────────────────────
    params: dict = {"from": str(date_from), "to": str(date_to)}
    where = ["DATE(e.created_at) BETWEEN %(from)s AND %(to)s",
             "e.stage_code = 'REJECTED'",
             "COALESCE(ol.is_deleted, FALSE) = FALSE"]

    if eye_filter != "All":
        params["eye"] = eye_filter[0]
        where.append("UPPER(COALESCE(ol.eye_side,'')) = %(eye)s")

    if reason_filter != "All Reasons":
        params["reason"] = f"%{reason_filter.split(' — ')[0]}%"
        where.append("COALESCE(e.remarks,'') ILIKE %(reason)s")

    if order_filter.strip():
        params["ord"] = f"%{order_filter.strip().lower()}%"
        where.append("(LOWER(o.order_no) LIKE %(ord)s OR LOWER(COALESCE(o.patient_name,'')) LIKE %(ord)s)")

    rows = _q(f"""
        SELECT
            e.id::text                                          AS event_id,
            e.created_at::date                                  AS rej_date,
            e.created_at                                        AS rej_at,
            COALESCE(e.remarks, '—')                            AS reason,
            ol.eye_side,
            ol.quantity,
            jm.reprocess_count,
            o.order_no,
            COALESCE(o.patient_name, o.party_name, '—')        AS patient_name,
            p.product_name,
            p.category,
            COALESCE(p.main_group,'—')                          AS main_group,
            COALESCE(ol.unit_price, 0)                          AS unit_price,
            COALESCE(ol.quantity, 1)                            AS qty,
            jm.current_stage                                    AS current_stage
        FROM job_stage_events e
        JOIN job_master  jm ON jm.id = e.job_id
        JOIN order_lines ol ON ol.id = jm.order_line_id
        JOIN orders      o  ON o.id  = ol.order_id
        JOIN products    p  ON p.id  = ol.product_id
        WHERE {' AND '.join(where)}
        ORDER BY e.created_at DESC
    """, params)

    if not rows:
        st.info("No rejections found for the selected filters.")
        return

    # ── Metrics ───────────────────────────────────────────────────────────────
    total       = len(rows)
    r_eye       = sum(1 for r in rows if str(r.get("eye_side","")).upper() == "R")
    l_eye       = sum(1 for r in rows if str(r.get("eye_side","")).upper() == "L")
    reprocessed = sum(1 for r in rows if int(r.get("reprocess_count") or 0) > 1)
    val_at_risk = sum(float(r.get("unit_price") or 0) * int(r.get("qty") or 1) for r in rows)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Rejections", total)
    m2.metric("Right Eye", r_eye)
    m3.metric("Left Eye",  l_eye)
    m4.metric("Reprocessed > 1×", reprocessed)
    m5.metric("Value at Risk ₹", f"{val_at_risk:,.0f}")

    st.markdown("---")

    # ── By Reason ─────────────────────────────────────────────────────────────
    reason_counts: dict = {}
    for r in rows:
        _raw = str(r.get("reason") or "—")
        # Normalize to known reason bucket
        _bucket = "Other"
        for _k in _KNOWN_REASONS[1:]:
            if _k.split(" — ")[0].lower() in _raw.lower():
                _bucket = _k
                break
        reason_counts[_bucket] = reason_counts.get(_bucket, 0) + 1

    rc1, rc2 = st.columns([2, 3])
    with rc1:
        st.markdown("**By Reason**")
        for rsn, cnt in sorted(reason_counts.items(), key=lambda x: -x[1]):
            pct = round(cnt / total * 100)
            bar = max(4, pct)
            st.markdown(
                f"<div style='margin:3px 0'>"
                f"<div style='display:flex;align-items:center;gap:8px'>"
                f"<div style='width:{bar}%;height:5px;background:#6366f1;"
                f"border-radius:3px;min-width:4px'></div>"
                f"<span style='color:#94a3b8;font-size:0.75rem;flex:1'>{rsn}</span>"
                f"<span style='color:#6366f1;font-size:0.75rem;font-weight:700'>"
                f"{cnt} <span style='color:#475569'>({pct}%)</span></span>"
                f"</div></div>",
                unsafe_allow_html=True
            )

    with rc2:
        st.markdown("**By Lens Type**")
        lens_counts: dict = {}
        for r in rows:
            lt = str(r.get("category") or r.get("main_group") or "Unknown")
            lens_counts[lt] = lens_counts.get(lt, 0) + 1
        for lt, cnt in sorted(lens_counts.items(), key=lambda x: -x[1]):
            pct = round(cnt / total * 100)
            bar = max(4, pct)
            st.markdown(
                f"<div style='margin:3px 0'>"
                f"<div style='display:flex;align-items:center;gap:8px'>"
                f"<div style='width:{bar}%;height:5px;background:#f59e0b;"
                f"border-radius:3px;min-width:4px'></div>"
                f"<span style='color:#94a3b8;font-size:0.75rem;flex:1'>{lt}</span>"
                f"<span style='color:#f59e0b;font-size:0.75rem;font-weight:700'>"
                f"{cnt} <span style='color:#475569'>({pct}%)</span></span>"
                f"</div></div>",
                unsafe_allow_html=True
            )

    st.markdown("---")

    # ── Detailed log ──────────────────────────────────────────────────────────
    st.markdown("**Rejection Log**")

    # Column headers
    _hc = st.columns([1.5, 0.8, 0.6, 2.5, 2.5, 1.2, 1])
    for h, c in zip(["Date", "Order", "Eye", "Product", "Reason", "Stage", "Reprocess×"], _hc):
        c.markdown(
            f"<div style='font-size:0.68rem;color:#6b7280;font-weight:700;"
            f"text-transform:uppercase'>{h}</div>",
            unsafe_allow_html=True
        )

    for r in rows:
        _eye = str(r.get("eye_side") or "").upper()
        _ec  = "#4ade80" if _eye == "R" else "#60a5fa" if _eye == "L" else "#94a3b8"
        _rep = int(r.get("reprocess_count") or 0)
        _stage = str(r.get("current_stage") or "—").replace("_", " ")
        _up  = float(r.get("unit_price") or 0)
        _qty = int(r.get("qty") or 1)

        dc = st.columns([1.5, 0.8, 0.6, 2.5, 2.5, 1.2, 1])
        dc[0].markdown(
            f"<div style='font-size:0.75rem;color:#64748b'>"
            f"{str(r.get('rej_date',''))}</div>",
            unsafe_allow_html=True
        )
        dc[1].markdown(
            f"<div style='font-family:monospace;font-size:0.72rem;color:#94a3b8'>"
            f"{r.get('order_no','—')}</div>"
            f"<div style='font-size:0.68rem;color:#475569'>"
            f"{r.get('patient_name','')}</div>",
            unsafe_allow_html=True
        )
        dc[2].markdown(
            f"<div style='color:{_ec};font-weight:700;font-size:0.8rem'>{_eye}</div>",
            unsafe_allow_html=True
        )
        dc[3].markdown(
            f"<div style='font-size:0.75rem;color:#e2e8f0'>"
            f"{r.get('product_name','—')}</div>"
            + (f"<div style='font-size:0.68rem;color:#475569'>₹{_up:,.0f} × {_qty}</div>"
               if _up > 0 else ""),
            unsafe_allow_html=True
        )
        dc[4].markdown(
            f"<div style='font-size:0.75rem;color:#fbbf24'>{r.get('reason','—')}</div>",
            unsafe_allow_html=True
        )
        dc[5].markdown(
            f"<div style='font-size:0.68rem;color:#64748b'>{_stage}</div>",
            unsafe_allow_html=True
        )
        dc[6].markdown(
            f"<div style='color:{'#ef4444' if _rep > 1 else '#f59e0b' if _rep == 1 else '#475569'};"
            f"font-size:0.78rem;font-weight:700'>"
            f"{'×' + str(_rep) if _rep else '—'}</div>",
            unsafe_allow_html=True
        )

    st.markdown("---")

    # ── Excel export ──────────────────────────────────────────────────────────
    if st.button("⬇️ Export to Excel", key="rr_export"):
        try:
            import io, openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "Rejections"

            headers = ["Date", "Order No", "Patient", "Product", "Category",
                       "Eye", "Reason", "Current Stage", "Reprocess Count",
                       "Unit Price ₹", "Qty", "Value ₹"]
            hfill = PatternFill("solid", fgColor="1E293B")
            hfont = Font(bold=True, color="E2E8F0")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center")

            for ri, r in enumerate(rows, 2):
                _up  = float(r.get("unit_price") or 0)
                _qty = int(r.get("qty") or 1)
                ws.cell(ri, 1,  str(r.get("rej_date") or ""))
                ws.cell(ri, 2,  r.get("order_no") or "")
                ws.cell(ri, 3,  r.get("patient_name") or "")
                ws.cell(ri, 4,  r.get("product_name") or "")
                ws.cell(ri, 5,  r.get("category") or "")
                ws.cell(ri, 6,  r.get("eye_side") or "")
                ws.cell(ri, 7,  r.get("reason") or "")
                ws.cell(ri, 8,  str(r.get("current_stage") or "").replace("_"," "))
                ws.cell(ri, 9,  int(r.get("reprocess_count") or 0))
                ws.cell(ri, 10, round(_up, 2))
                ws.cell(ri, 11, _qty)
                ws.cell(ri, 12, round(_up * _qty, 2))

            # Summary sheet
            ws2 = wb.create_sheet("Summary")
            ws2.append(["By Reason", "Count", "% of Total"])
            for rsn, cnt in sorted(reason_counts.items(), key=lambda x: -x[1]):
                ws2.append([rsn, cnt, round(cnt/total*100, 1)])
            ws2.append([])
            ws2.append(["By Lens Type", "Count"])
            for lt, cnt in sorted(lens_counts.items(), key=lambda x: -x[1]):
                ws2.append([lt, cnt])
            ws2.append([])
            ws2.append(["Metric", "Value"])
            ws2.append(["Total Rejections", total])
            ws2.append(["Right Eye", r_eye])
            ws2.append(["Left Eye",  l_eye])
            ws2.append(["Reprocessed > 1×", reprocessed])
            ws2.append(["Value at Risk ₹", round(val_at_risk, 2)])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                "📥 Download Rejection Report.xlsx",
                data=buf.getvalue(),
                file_name=f"rejection_report_{date_from}_{date_to}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="rr_download"
            )
        except Exception as e:
            st.error(f"Export failed: {e}")

    # ── Rejection Bin (material-side counterpart) ────────────────────────
    st.markdown("---")
    st.markdown("### Rejection Bin - Material Status")
    st.caption(
        "Source: production_rejection_bin. Each row is one rejected lens/item. "
        "Use status later to mark items SCRAPPED or RECLAIMED."
    )

    try:
        bin_rows = _q("""
            SELECT b.id::text AS bin_id,
                   b.rejected_at::date AS rej_date,
                   b.status,
                   b.eye_side,
                   b.qty,
                   COALESCE(p.product_name, '—') AS product_name,
                   COALESCE(p.category, '—') AS category,
                   COALESCE(o.order_no, '—') AS order_no,
                   COALESCE(b.reason, '—') AS reason,
                   COALESCE(b.rejected_by, '—') AS rejected_by,
                   b.blank_id::text AS blank_id
            FROM production_rejection_bin b
            LEFT JOIN products p ON p.id = b.product_id
            LEFT JOIN orders o ON o.id = b.order_id
            WHERE DATE(b.rejected_at) BETWEEN %(from)s AND %(to)s
            ORDER BY b.rejected_at DESC
        """, {"from": str(date_from), "to": str(date_to)})
    except Exception as e:
        st.info(
            "Rejection bin table not yet available. Run migration "
            f"0025_production_rejection_bin.sql. Error: {e}"
        )
        bin_rows = []

    if not bin_rows:
        st.info("No items in the rejection bin for this date range.")
    else:
        from collections import Counter
        status_counts = Counter(r.get("status") or "IN_BIN" for r in bin_rows)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total in Bin", len(bin_rows))
        c2.metric("IN_BIN", status_counts.get("IN_BIN", 0))
        c3.metric("SCRAPPED", status_counts.get("SCRAPPED", 0))
        c4.metric("RECLAIMED", status_counts.get("RECLAIMED", 0))

        prod_counts = Counter(r.get("product_name") or "—" for r in bin_rows)
        st.markdown("**By Product (top 10):**")
        for prod, cnt in sorted(prod_counts.items(), key=lambda x: -x[1])[:10]:
            st.write(f"- {prod}: **{cnt}**")

        with st.expander("Show all rejection-bin rows", expanded=False):
            import pandas as _pd
            st.dataframe(_pd.DataFrame(bin_rows), use_container_width=True, hide_index=True)

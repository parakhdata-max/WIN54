"""Reusable replenishment communication panel.

Used from production/backoffice screens when a stock item, supplier order, or lab
order needs to be communicated through WhatsApp, Excel, email, or phone.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, Iterable, List
import datetime as _dt
import urllib.parse as _urlparse

import streamlit as st


def _clean_mobile(value: str) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    return "91" + digits if len(digits) == 10 else ""


def _line_power(line: Dict) -> str:
    parts = []
    lp = line.get('_lp') or line.get('lens_params') or {}
    if isinstance(lp, str):
        try:
            import json as _pj; lp = _pj.loads(lp)
        except: lp = {}
    for label, key, fmt in (('SPH','sph','+.2f'),('CYL','cyl','+.2f'),('AX','axis','d'),('ADD','add_power','+.2f')):
        val = line.get(key)
        if val in (None,'','None'): val = lp.get(key)
        if val not in (None,'','None',0,'0',0.0):
            try:
                parts.append(f'AX {int(float(val))}' if fmt=='d' else f'{label} {float(val):{fmt}}')
            except Exception:
                parts.append(f'{label} {val}')
    return '  '.join(parts) or '—'
def _positive_int_from_fields(line: Dict, fields: Iterable[str], default: int = 1) -> tuple[int, str]:
    """Return first positive integer field from a line and the source field name.

    Important: do NOT use `a or b or c` for quantities. Database/UI rows can
    contain 0, "0", 0.0 or Decimal("0"). A non-empty string "0" is truthy
    in Python, so an `or` chain can incorrectly stop at zero and make a real
    order line contribute 0 boxes.
    """
    for field in fields:
        raw = line.get(field)
        if raw in (None, "", "None"):
            continue
        try:
            val = int(float(raw))
        except Exception:
            continue
        if val > 0:
            return val, field
    return int(default or 1), "default"


def _normalise_lines(lines: Iterable[Dict]) -> List[Dict]:
    """Normalize selected procurement/replenishment lines.

    For replenishment, the correct source is the *ordered demand* first, then
    fallbacks. So priority is:
        quantity -> qty -> billing_qty -> allocated_qty -> billed_qty -> 1

    This fixes the 117/118 case where a zero/partial allocated/billed value can
    under-count boxes. It also keeps raw PCS and source fields visible for audit.
    """
    out = []
    for idx, line in enumerate(lines or [], 1):
        qty_pcs, qty_source = _positive_int_from_fields(
            line,
            ("quantity", "qty", "billing_qty", "allocated_qty", "billed_qty"),
            default=1,
        )
        box_size, box_source = _positive_int_from_fields(line, ("box_size",), default=1)

        if box_size > 1:
            boxes = max(1, (qty_pcs + box_size - 1) // box_size)
            qty_label = f"{boxes} box{'es' if boxes != 1 else ''}"
        else:
            boxes = max(1, qty_pcs)
            qty_label = f"{boxes} box{'es' if boxes != 1 else ''}"

        line_id = line.get("line_id") or line.get("id") or ""
        out.append({
            "line_no":    idx,
            "line_id":    str(line_id),
            "product_id":  str(line.get("product_id") or ""),
            "order_no":   line.get("order_no") or line.get("customer_order_no") or "",
            "eye":        str(line.get("eye_side") or line.get("eye") or "").upper(),
            "product":    str(line.get("product_name") or line.get("name") or "").split(" | ")[0],
            "brand":      line.get("brand") or line.get("brand_name") or "",
            "supplier_product_name": line.get("supplier_product_name") or "",
            "supplier_brand":        line.get("supplier_brand") or "",
            "supplier_index":        line.get("supplier_index") or "",
            "supplier_coating":      line.get("supplier_coating") or "",
            "supplier_treatment":    line.get("supplier_treatment") or "",
            "sku":        line.get("batch_no") or line.get("sku") or "",
            "power":      _line_power(line),
            "qty":        boxes,       # boxes for WhatsApp/Excel ordering
            "qty_pcs":    qty_pcs,     # raw PCS/demand from DB/order line
            "box_size":   box_size,
            "qty_source": qty_source,
            "box_source": box_source,
            "qty_label":  qty_label,
        })
    return out

def build_replenishment_message(lines: Iterable[Dict], supplier: Dict, order_no: str = "") -> str:
    rows = _normalise_lines(lines)
    supplier_name = supplier.get("name") or supplier.get("party_name") or "Supplier"
    supplier_id = str(supplier.get("id") or supplier.get("supplier_id") or "")
    for row in rows:
        if supplier_id and row.get("product_id"):
            try:
                from modules.backoffice.supplier_product_map_ui import get_supplier_product_name
                _spm = get_supplier_product_name(str(row.get("product_id") or ""), supplier_id)
                if _spm.get("mapped"):
                    row["supplier_product_name"] = row.get("supplier_product_name") or _spm.get("supplier_product_name", "")
                    row["supplier_brand"]        = row.get("supplier_brand") or _spm.get("supplier_brand", "")
                    row["supplier_index"]        = row.get("supplier_index") or _spm.get("supplier_index", "")
                    row["supplier_coating"]      = row.get("supplier_coating") or _spm.get("supplier_coating", "")
                    row["supplier_treatment"]    = row.get("supplier_treatment") or _spm.get("supplier_treatment", "")
            except Exception:
                pass
    import datetime as _mdt
    msg = [
        f"*Stock Replenishment Order*",
        f"*PO No: {order_no or '—'}*",
        f"Date: {_mdt.date.today().strftime('%d/%m/%Y')}",
        f"To: {supplier_name}",
        "",
    ]

    # Group by supplier-facing product/spec → (eye, power) → sum boxes
    from collections import defaultdict as _dd
    _prod_grp: dict = _dd(lambda: _dd(lambda: {"boxes": 0, "pcs": 0, "sku": ""}))
    _prod_order: list = []
    for row in rows:
        _spec_bits = []
        if row.get("supplier_brand"):
            _spec_bits.append(str(row.get("supplier_brand")))
        _spec_bits.append(str(row.get("supplier_product_name") or row["product"]))
        if row.get("supplier_index"):
            _spec_bits.append(f"Index {row.get('supplier_index')}")
        if row.get("supplier_coating"):
            _spec_bits.append(str(row.get("supplier_coating")))
        if row.get("supplier_treatment") and str(row.get("supplier_treatment")) != "Clear":
            _spec_bits.append(str(row.get("supplier_treatment")))
        prod = " · ".join([p for p in _spec_bits if p])
        eye  = row.get("eye", "")
        pwr  = row.get("power", "—")
        key  = (eye, pwr)
        if prod not in _prod_order:
            _prod_order.append(prod)
        _prod_grp[prod][key]["boxes"] += int(row.get("qty") or 0)
        _prod_grp[prod][key]["pcs"]   += int(row.get("qty_pcs") or row.get("qty") or 0)
        _prod_grp[prod][key]["sku"]    = row.get("sku") or ""

    for prod in _prod_order:
        msg.append(f"*{prod}*")
        for (eye, pwr), info in sorted(_prod_grp[prod].items()):
            boxes = info["boxes"]
            pcs   = info["pcs"]
            eye_lbl = f"{eye}:  " if eye else ""
            box_str = f"{boxes} box{'es' if boxes != 1 else ''}"
            if pcs and pcs != boxes:
                box_str += f" ({pcs} PCS)"
            msg.append(f"  {eye_lbl}{pwr}  —  {box_str}")
        msg.append("")

    msg.extend(["Please confirm availability and dispatch date.", "Thank you."])
    return "\n".join(msg)


def build_simple_excel(lines: Iterable[Dict], supplier: Dict, order_no: str = "") -> bytes:
    rows = _normalise_lines(lines)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Replenishment"
        ws.append(["Supplier", supplier.get("name") or supplier.get("party_name") or ""])
        ws.append(["PO / Reference", order_no])
        ws.append(["Date", _dt.date.today().isoformat()])
        ws.append([])
        headers = ["Sr", "Order", "Eye", "Product", "Brand", "SKU / Batch", "Power", "Qty"]
        ws.append(headers)
        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for idx, row in enumerate(rows, 1):
            ws.append([
                idx, row["order_no"], row["eye"], row["product"], row["brand"],
                row["sku"], row["power"], row.get("qty_label") or row["qty"],
            ])
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = [6, 18, 8, 34, 18, 18, 24, 8][col - 1]
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()
    except Exception:
        # Fallback still opens in Excel.
        csv = "Sr,Order,Eye,Product,Brand,SKU / Batch,Power,Qty\n"
        for idx, row in enumerate(rows, 1):
            vals = [idx, row["order_no"], row["eye"], row["product"], row["brand"], row["sku"], row["power"], row.get("qty_label") or row["qty"]]
            csv += ",".join('"' + str(v).replace('"', '""') + '"' for v in vals) + "\n"
        return csv.encode("utf-8")



# ══════════════════════════════════════════════════════════════════════════
# ALCON TORIC TEMPLATE-FILL EXCEL
# ══════════════════════════════════════════════════════════════════════════

_ALCON_DD_CACHE: dict | None = None

def _load_alcon_dd_df():
    """
    Load Data_Drop from the Alcon template file into a pandas DataFrame.
    Cached in module-level variable — loaded once per process.
    Searches for the template in common locations.
    """
    global _ALCON_DD_CACHE
    if _ALCON_DD_CACHE is not None:
        return _ALCON_DD_CACHE

    import os, pandas as pd
    _candidates = [
        os.path.join(os.path.dirname(__file__), "..", "..", "assets", "templates",
                     "Parakh_Alcon_Toric_Order.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "..", "assets", "templates",
                     "Simplify_Toric_order_sheet.xlsx"),
        "/home/claude/alcon_data_drop.pkl",   # dev fallback pickle
    ]
    # Also try any xlsx in current dir that looks like Alcon
    for _root, _dirs, _files in os.walk(os.path.dirname(__file__)):
        for _f in _files:
            if "alcon" in _f.lower() or "toric" in _f.lower() or "simplify" in _f.lower():
                _candidates.insert(0, os.path.join(_root, _f))
        break

    for path in _candidates:
        if not os.path.exists(path):
            continue
        try:
            if path.endswith(".pkl"):
                df = pd.read_pickle(path)
                # Pickle uses Data_Drop column names
                if "Cylinder" in df.columns and "Power" in df.columns:
                    # Normalize to internal format
                    df2 = df[["MD PHL5 Brand","Material Number","Material Number Description",
                               "Power","Cylinder","Axis"]].copy()
                    df2.columns = ["brand","mat_no","mat_desc","power","cyl_raw","axis"]
                    # cyl_raw in pkl is float — convert to semicolon format
                    def _to_semi(v):
                        try: return "{:.2f}".format(float(v)).replace(".", ";")
                        except: return str(v)
                    df2["cyl"] = df2["cyl_raw"].apply(_to_semi)
                    _ALCON_DD_CACHE = df2.dropna(subset=["mat_no"])
                    return _ALCON_DD_CACHE
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                # Try sheet names: Data_Drop
                if "Data_Drop" in wb.sheetnames:
                    ws = wb["Data_Drop"]
                    rows_dd = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[3]:  # brand not null (col D = index 3)
                            rows_dd.append({
                                "brand":    row[3],
                                "mat_no":   row[4],
                                "mat_desc": row[5],
                                "power":    float(row[8]) if row[8] is not None else None,
                                "cyl":      str(row[9]) if row[9] else "",  # already "1;25"
                                "axis":     int(row[10]) if row[10] is not None else None,
                            })
                    df2 = pd.DataFrame(rows_dd).dropna(subset=["mat_no"])
                    wb.close()
                    _ALCON_DD_CACHE = df2
                    return _ALCON_DD_CACHE
        except Exception:
            continue

    return None


def _alcon_lookup(df, brand: str, sph: float, cyl_semi: str, axis: int) -> tuple:
    """
    Look up Material Number and Description from the Alcon Data_Drop.
    cyl_semi: cylinder in semicolon format, e.g. '1;25' (not 1.25)
    Returns (mat_no_str, mat_desc_str) or ("", "Material Not Found")
    """
    if df is None:
        return "", "Data_Drop not available"
    try:
        import numpy as np
        _mask = (
            (df["brand"] == brand) &
            (abs(df["power"].astype(float) - float(sph)) < 0.001) &
            (df["cyl"] == cyl_semi) &
            (df["axis"].astype(int) == int(axis))
        )
        _hits = df[_mask]
        if _hits.empty:
            return "", "Material Not Found"
        _row = _hits.iloc[0]
        return str(_row["mat_no"]), str(_row["mat_desc"])
    except Exception as _le:
        return "", f"Lookup error: {_le}"


def _cyl_to_semi(cyl_float) -> str:
    """Convert 1.25 → '1;25' for Alcon lookup."""
    try:
        v = abs(float(cyl_float))
        return "{:.2f}".format(v).replace(".", ";")
    except Exception:
        return str(cyl_float)


def _is_alcon_toric(line: Dict) -> bool:
    """True if this line is an Alcon toric contact lens."""
    text = (
        str(line.get("product_name","")) + " " +
        str(line.get("brand","")) + " " +
        str(line.get("supplier_name","")) + " " +
        str((line.get("_lp") or line.get("lens_params") or {}).get("brand","") if isinstance(
            line.get("_lp") or line.get("lens_params"), dict) else "")
    ).upper()
    is_alcon = any(x in text for x in [
        "AIR OPTIX", "AIROPTIX", "PRECISION1", "TOTAL30", "ALCON"
    ])
    cyl = line.get("cyl") or (line.get("_lp") or {}).get("cyl") if isinstance(
        line.get("_lp"), dict) else None
    has_cyl = cyl not in (None, "", 0, "0", 0.0)
    return is_alcon and has_cyl


def build_alcon_toric_excel(
    lines: Iterable[Dict],
    supplier: Dict,
    order_no: str = "",
) -> bytes:
    """
    Build an Alcon portal-ready Excel by filling the Alcon input format:
      Col A: OrderType
      Col B: PONumber
      Col C: SoldToCustomer (100423873)
      Col D: ShipToCustomer (100423873)
      Col E: MD_PHL5_Brand
      Col F: Lens_Power (SPH)
      Col G: Lens_Cylinder (1;25 format)
      Col H: Lens_Axis
      Col I: Material_Number  ← looked up from Data_Drop
      Col J: Material_Description ← looked up from Data_Drop
      Col K: ItemQuantity
      Col L: Color_Code (blank)

    Also builds a human-readable summary sheet.
    """
    from io import BytesIO
    import datetime as _dt2

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    _dd = _load_alcon_dd_df()

    # Normalize lines
    _order_lines = list(lines)
    _today = _dt2.date.today().strftime("%d-%b-%Y")
    _sup_name = supplier.get("name") or supplier.get("party_name") or "Alcon India"

    # ── Style helpers ────────────────────────────────────────────────────
    NAVY  = "1E3A5F"; BLUE = "2563EB"; GREEN = "065F46"
    WHITE = "FFFFFF"; YELL = "FEF9C3"; GREY  = "E2E8F0"

    def fill(c): return PatternFill("solid", fgColor=c)
    def bdr():
        s = Side(style="thin", color=GREY)
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Alcon Export (portal format) ─────────────────────────
    ws_exp = wb.create_sheet("Alcon Export")
    ws_exp.sheet_view.showGridLines = True

    exp_headers = [
        "OrderType","PONumber","SoldToCustomer","ShipToCustomer",
        "MD_PHL5_Brand","Lens_Power","Lens_Cylinder","Lens_Axis",
        "Material_Number","Material_Description","ItemQuantity","Color_Code",
    ]
    col_widths = [14,12,16,16,34,12,14,10,16,52,13,12]
    for i,(h,w) in enumerate(zip(exp_headers,col_widths),1):
        ws_exp.column_dimensions[get_column_letter(i)].width = w
        c = ws_exp.cell(1,i); c.value = h
        c.font = Font(name="Arial",bold=True,size=10,color=WHITE)
        c.fill = fill(GREEN); c.alignment = Alignment(horizontal="center",vertical="center")

    _not_found: list = []

    for _li, _line in enumerate(_order_lines, 2):
        _brand = str(_line.get("brand","") or "AirOptix Toric w/ Hydraglyde")
        _sph   = _line.get("sph") or (_line.get("_lp") or {}).get("sph")
        _cyl   = _line.get("cyl") or (_line.get("_lp") or {}).get("cyl")
        _axis  = _line.get("axis") or (_line.get("_lp") or {}).get("axis")
        _qty   = int(_line.get("qty") or _line.get("quantity") or 1)
        _ono   = _line.get("order_no","")

        # Convert CYL to Alcon semicolon format
        _cyl_semi = _cyl_to_semi(_cyl) if _cyl else ""

        # Lookup
        _mat_no, _mat_desc = "", "Material Not Found"
        if _sph is not None and _cyl_semi and _axis is not None:
            try:
                _mat_no, _mat_desc = _alcon_lookup(
                    _dd, _brand, float(_sph), _cyl_semi, int(float(_axis))
                )
            except Exception:
                pass

        if not _mat_no:
            _not_found.append(f"{_ono} {_brand} SPH {_sph} CYL {_cyl_semi} AX {_axis}")

        _row_data = [
            "Sales Order",
            order_no or _ono,
            100423873,   # SoldToCustomer
            100423873,   # ShipToCustomer
            _brand,
            float(_sph) if _sph is not None else "",
            _cyl_semi,
            int(float(_axis)) if _axis is not None else "",
            _mat_no,
            _mat_desc,
            _qty,
            "",          # Color_Code
        ]
        for i, val in enumerate(_row_data, 1):
            c = ws_exp.cell(_li, i); c.value = val
            c.font = Font(name="Arial",size=9)
            c.alignment = Alignment(horizontal="left",vertical="center")
            c.border = bdr()
            alt = _li % 2 == 0
            c.fill = fill("EFF6FF" if alt else WHITE)

    # ── Sheet 2: Order Summary ────────────────────────────────────────
    ws_sum = wb.create_sheet("Order Summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 5
    ws_sum.column_dimensions["B"].width = 30
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 10
    ws_sum.column_dimensions["E"].width = 10
    ws_sum.column_dimensions["F"].width = 8
    ws_sum.column_dimensions["G"].width = 18
    ws_sum.column_dimensions["H"].width = 50

    ws_sum.row_dimensions[1].height = 8
    ws_sum.row_dimensions[2].height = 36
    ws_sum.merge_cells("A2:H2")
    c = ws_sum["A2"]
    c.value = f"PARAKH OPTICAL — ALCON TORIC ORDER  |  {order_no or 'Replenishment'}  |  {_today}"
    c.font = Font(name="Arial",bold=True,size=14,color=WHITE)
    c.fill = fill(NAVY); c.alignment = Alignment(horizontal="center",vertical="center")

    ws_sum.row_dimensions[3].height = 20
    ws_sum.merge_cells("A3:H3")
    c = ws_sum["A3"]
    c.value = f"Customer Code: 100423873  ·  Supplier: {_sup_name}  ·  Date: {_today}"
    c.font = Font(name="Arial",size=10,color=WHITE,italic=True)
    c.fill = fill(BLUE); c.alignment = Alignment(horizontal="center",vertical="center")

    ws_sum.row_dimensions[4].height = 8
    ws_sum.row_dimensions[5].height = 26
    _sum_hdrs = ["#","Product / Brand","SPH","CYL","AXIS","Qty","Material No.","Description"]
    for i,h in enumerate(_sum_hdrs,1):
        c = ws_sum.cell(5,i); c.value=h
        c.font=Font(name="Arial",bold=True,size=10,color=WHITE)
        c.fill=fill(GREEN); c.alignment=Alignment(horizontal="center",vertical="center")

    for _si, _line in enumerate(_order_lines, 6):
        _brand2   = str(_line.get("brand","") or "AirOptix Toric w/ Hydraglyde")
        _sph2     = _line.get("sph") or (_line.get("_lp") or {}).get("sph")
        _cyl2     = _line.get("cyl") or (_line.get("_lp") or {}).get("cyl")
        _axis2    = _line.get("axis") or (_line.get("_lp") or {}).get("axis")
        _qty2     = int(_line.get("qty") or _line.get("quantity") or 1)
        _cyl_s2   = _cyl_to_semi(_cyl2) if _cyl2 else ""
        _mn2, _md2 = _alcon_lookup(_dd, _brand2, float(_sph2) if _sph2 else 0,
                                    _cyl_s2, int(float(_axis2)) if _axis2 else 0) if _sph2 else ("","")
        alt2 = _si % 2 == 0
        rf2  = fill("EFF6FF" if alt2 else WHITE)
        ws_sum.row_dimensions[_si].height = 20
        for _ci,_val in enumerate([_si-5,_brand2,_sph2,_cyl_s2,_axis2,_qty2,_mn2,_md2],1):
            c=ws_sum.cell(_si,_ci); c.value=_val
            c.font=Font(name="Arial",size=9)
            c.fill=rf2; c.alignment=Alignment(horizontal="center" if _ci in (1,3,4,5,6) else "left",vertical="center")
            c.border=bdr()

    if _not_found:
        _nr = len(_order_lines) + 7
        ws_sum.merge_cells(f"A{_nr}:H{_nr}")
        c=ws_sum.cell(_nr,1)
        c.value = "⚠️ Material not found: " + "; ".join(_not_found[:5])
        c.font=Font(name="Arial",size=9,color="DC2626")
        c.fill=fill("FEF2F2")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_replenishment_panel(
    lines: Iterable[Dict],
    supplier: Dict,
    order_no: str = "",
    route: str = "STOCK",
    key_prefix: str = "repl",
    allow_save_ref: bool = True,
):
    """Render WhatsApp, Excel, Mail, Call, and Programed Excel channels."""
    rows = _normalise_lines(lines)
    if not rows:
        st.info("No replenishment lines selected.")
        return

    with st.expander("🔎 Line quantity audit", expanded=False):
        st.caption("Shows the exact individual lines used for replenishment. This is the first check if total boxes look wrong.")
        for r in rows:
            line_tail = f" · {r['line_id'][:8]}…" if r.get("line_id") else ""
            st.markdown(
                f"**{r.get('order_no') or '—'}{line_tail}** · {r.get('eye') or '—'} · "
                f"{r.get('product') or '—'} · {r.get('power') or '—'}  \
"
                f"PCS: `{r.get('qty_pcs')}` from `{r.get('qty_source')}` · "
                f"Box size: `{r.get('box_size')}` · Order qty: **{r.get('qty_label')}**"
            )

    supplier_name = supplier.get("name") or supplier.get("party_name") or "Supplier"
    mobile = _clean_mobile(supplier.get("mobile") or supplier.get("phone") or "")
    email = str(supplier.get("email") or supplier.get("mail") or "")
    msg = build_replenishment_message(lines, supplier, order_no)
    xlsx = build_simple_excel(lines, supplier, order_no)
    fname = f"replenishment_{order_no or _dt.date.today().isoformat()}.xlsx".replace("/", "-")

    tab_wa, tab_xl, tab_mail, tab_call, tab_prog = st.tabs([
        "📲 WhatsApp", "📎 Excel", "📧 Mail", "📞 Call", "🧾 Programed Excel"
    ])

    with tab_wa:
        st.text_area("Message", value=msg, height=170, key=f"{key_prefix}_wa_msg")
        # Missing supplier mobile MUST NOT block sending the order. Staff often
        # need to use a number that isn't on file (personal contact at the
        # supplier, alternate channel, etc). Always show an editable number
        # field — pre-filled with saved mobile if any. Empty number is also
        # valid: opens WhatsApp's chat-picker so user can pick a contact.
        _manual_mobile = st.text_input(
            "WhatsApp number",
            value=str(mobile or ""),
            placeholder="Enter supplier WhatsApp number manually",
            key=f"{key_prefix}_wa_mobile",
            help="Pre-filled from supplier record. Edit freely or leave blank "
                 "to open WhatsApp without a target number.",
        )
        _manual_mobile_clean = _clean_mobile(_manual_mobile)
        if _manual_mobile_clean:
            _wa_url = f"https://wa.me/{_manual_mobile_clean}?text={_urlparse.quote(msg, safe='')}"
            _btn_label = f"📲 Open WhatsApp to {supplier_name}" if supplier_name else "📲 Open WhatsApp"
        else:
            _wa_url = f"https://wa.me/?text={_urlparse.quote(msg, safe='')}"
            _btn_label = "📲 Open WhatsApp (pick contact)"
            if not mobile:
                st.caption("No saved number. Enter manually above or open the WhatsApp picker below.")
        st.link_button(
            _btn_label,
            _wa_url,
            use_container_width=True,
            type="primary",
        )

    with tab_xl:
        st.download_button(
            "Download Simple Excel",
            data=xlsx,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_simple_xlsx",
            use_container_width=True,
        )

    with tab_mail:
        subject = f"Replenishment order {order_no}".strip()
        st.text_area("Email body", value=msg, height=160, key=f"{key_prefix}_email_body")
        if email:
            mailto = f"mailto:{email}?subject={_urlparse.quote(subject)}&body={_urlparse.quote(msg)}"
            st.link_button("Open Mail", mailto, use_container_width=True)
        else:
            st.warning("Supplier email not available. Download Excel and send manually.")

    with tab_call:
        phone = supplier.get("mobile") or supplier.get("phone") or ""
        st.metric("Supplier", supplier_name)
        st.metric("Phone", phone or "-")
        st.text_area("Read on call", value=msg, height=150, key=f"{key_prefix}_call_script")

    # ── Supplier Confirmation Number (common to all routes) ──────────────────
    st.markdown("---")
    if allow_save_ref:
        _conf_col1, _conf_col2 = st.columns([3,1])
        _sup_conf_no = _conf_col1.text_input(
            "📋 Supplier Confirmation No.",
            placeholder="e.g. PO-2024-001 or WhatsApp confirmation ref",
            key=f"{key_prefix}_conf_no",
            help="Record the supplier's own order/reference number for tracking"
        )
        if _conf_col2.button("💾 Save Ref", key=f"{key_prefix}_save_conf",
                              use_container_width=True, type="primary"):
            if _sup_conf_no.strip():
                try:
                    from modules.sql_adapter import run_write as _rw_conf
                    import json as _jc
                    for _ln_conf in list(lines):
                        _lid_conf = str(_ln_conf.get("line_id") or _ln_conf.get("id") or "")
                        if not _lid_conf:
                            continue
                        _lp_conf = dict(_ln_conf.get("_lp") or _ln_conf.get("lens_params") or {})
                        _lp_conf["supplier_confirmation_no"] = _sup_conf_no.strip()
                        if str(_lp_conf.get("replenishment_status") or "").upper() == "PO_SENT":
                            _lp_conf["replenishment_status"] = "ORDERED"
                        _rw_conf("UPDATE order_lines SET lens_params=%(lp)s::jsonb WHERE id=%(id)s::uuid",
                                 {"lp": _jc.dumps(_lp_conf), "id": _lid_conf})
                    import streamlit as _st_conf
                    _st_conf.success(f"✅ Saved: {_sup_conf_no.strip()}")
                except Exception as _ce:
                    import streamlit as _st_conf
                    _st_conf.error(f"Save failed: {_ce}")
            else:
                import streamlit as _st_conf
                _st_conf.warning("Enter a confirmation number first")
    else:
        st.info(
            "Supplier confirmation numbers are not saved for blank inventory replenishment lines."
        )

    with tab_prog:
        _rows_list = list(lines)
        _any_alcon_p  = any(_is_alcon_toric(l) for l in _rows_list)
        _alcon_lines_p = [l for l in _rows_list if _is_alcon_toric(l)]
        _other_lines_p = [l for l in _rows_list if not _is_alcon_toric(l)]

        if _any_alcon_p:
            _dd_ok_p = _load_alcon_dd_df() is not None
            st.info(
                f"Alcon Toric detected ({len(_alcon_lines_p)} line(s)). "
                + ("Data_Drop loaded OK." if _dd_ok_p else
                   "Data_Drop not found — place Alcon template in assets/templates/")
            )
            _alcon_xlsx_b = build_alcon_toric_excel(_alcon_lines_p, supplier, order_no)
            if _alcon_xlsx_b:
                _afn = "Parakh_Alcon_Toric_" + order_no.replace("/","-") + "_" + _dt.date.today().isoformat() + ".xlsx"
                st.download_button(
                    "Download Alcon Toric Order (.xlsx)",
                    data=_alcon_xlsx_b, file_name=_afn,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=key_prefix + "_alcon_xlsx",
                    use_container_width=True, type="primary",
                )
                st.caption("Sheet 1: Alcon Export (paste into portal or email).  Sheet 2: Summary with Material Numbers.")

        if _other_lines_p:
            _gfn = "Order_" + order_no.replace("/","-") + "_" + _dt.date.today().isoformat() + ".xlsx"
            _glbl = "Generic Order Sheet (.xlsx)" + ("  (Non-Alcon lines)" if _any_alcon_p else "")
            st.download_button(
                _glbl, data=build_simple_excel(_other_lines_p, supplier, order_no),
                file_name=_gfn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=key_prefix + "_generic_xlsx", use_container_width=True,
            )

        if not _any_alcon_p and not _other_lines_p:
            st.info("No lines to generate Excel for.")



# ═══════════════════════════════════════════════════════════════════════════
# BLANK REPLENISHMENT — clean rewrite using blank_inventory + blank_allocations
# ═══════════════════════════════════════════════════════════════════════════


def render_blank_replenishment_summary(key_prefix: str = "blank_repl"):
    """
    Blank Replenishment hub — 4 tabs:
      1. Consumption   — blanks consumed by in-house RX in date range
      2. Returns       — blanks restored when orders rolled back
      3. Order Blanks  — raise PO → WA / Excel / Email + record in DB
      4. Upload Invoice — record supplier invoice against open PO
    Tables required:
      blank_replenishment_orders   (id, supplier_id, order_date, status, line_items jsonb, remarks)
      blank_replenishment_invoices (id, po_id, invoice_no, invoice_date, invoice_amount, gst_amount, file_b64, notes)
    """
    import json as _brj
    import datetime as _dt
    import urllib.parse as _up

    def _brq(sql, params=None):
        try:
            from modules.sql_adapter import run_query
            return run_query(sql, params or {}) or []
        except Exception as _e:
            st.error(f"DB read error: {_e}")
            return []

    def _brw(sql, params=None):
        try:
            from modules.sql_adapter import run_write
            run_write(sql, params or {})
            return True
        except Exception as _e:
            st.error(f"DB write error: {_e}")
            return False

    # ── Migration guard — show DDL if tables missing ───────────────────
    # ── Header ────────────────────────────────────────────────────────
    st.markdown(
        "<div style='background:#0f172a;border:1px solid #1e3a5f;"
        "border-left:4px solid #8b5cf6;border-radius:8px;"
        "padding:10px 16px;margin-bottom:12px'>"
        "<span style='color:#c4b5fd;font-size:1rem;font-weight:800'>"
        "🧫 Blank Replenishment</span>"
        "<span style='color:#475569;font-size:0.78rem;margin-left:10px'>"
        "Consumed \xb7 Returned \xb7 Order PO \xb7 Upload Invoice"
        "</span></div>",
        unsafe_allow_html=True,
    )

    _repl_sections = [
        "📊 Consumption",
        "♻️ Returns",
        "📦 Order Blanks",
        "🧾 Upload Invoice",
        "📋 Usage Report",
        "📈 Repl. vs Procured",
        "🏪 Party Audit",
        "🎯 Procurement Dashboard",
    ]
    _repl_section_key = f"{key_prefix}_section"
    if st.session_state.get(_repl_section_key) not in _repl_sections:
        st.session_state[_repl_section_key] = "📊 Consumption"
    _repl_active_section = st.radio(
        "Blank replenishment section",
        _repl_sections,
        key=_repl_section_key,
        horizontal=True,
        label_visibility="collapsed",
    )
    (_tab_consume, _tab_returns, _tab_order, _tab_invoice,
     _tab_report, _tab_repl_report, _tab_party, _tab_dash) = [st.container() for _ in _repl_sections]

    # ══════════════════════════════════════════════════════════════════
    # TAB 1 — CONSUMPTION
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "📊 Consumption":
        with _tab_consume:
            st.caption("Blanks consumed by in-house RX jobs. Tick rows then go to 📦 Order Blanks.")
            with st.container(border=True):
                _c1, _c2, _c3 = st.columns([2, 2, 1])
                _today = _dt.date.today()
                _dfrom = _c1.date_input("Consumed from", value=_today - _dt.timedelta(days=30),
                                         key=f"{key_prefix}_dfrom", format="DD/MM/YYYY")
                _dto   = _c2.date_input("Consumed to",   value=_today,
                                         key=f"{key_prefix}_dto",   format="DD/MM/YYYY")
                _bmin  = _c3.toggle("Below min only", value=False, key=f"{key_prefix}_bmin",
                                     help="Only show blanks where current stock < min_stock.")

                # Brand + Product filters (row 2)
                _fc1, _fc2 = st.columns(2)

                # Load distinct brands + categories for filter dropdowns (from all blanks, not date-filtered)
                _all_brands_cats = _brq("""
                    SELECT DISTINCT COALESCE(brand,'') AS brand, COALESCE(category,'') AS category
                    FROM blank_inventory
                    WHERE COALESCE(is_active, TRUE) = TRUE
                    ORDER BY brand, category
                """)
                _all_brands = sorted(set(r["brand"] for r in _all_brands_cats if r["brand"]))
                _sel_brand  = _fc1.selectbox(
                    "Filter by Brand",
                    ["All Brands"] + _all_brands,
                    key=f"{key_prefix}_f_brand",
                )
                # Category list — filtered by selected brand
                if _sel_brand and _sel_brand != "All Brands":
                    _cat_choices = sorted(set(
                        r["category"] for r in _all_brands_cats
                        if r["brand"] == _sel_brand and r["category"]
                    ))
                else:
                    _cat_choices = sorted(set(r["category"] for r in _all_brands_cats if r["category"]))
                _sel_cat = _fc2.selectbox(
                    "Filter by Product / Category",
                    ["All Products"] + _cat_choices,
                    key=f"{key_prefix}_f_cat",
                )

            if _dfrom > _dto:
                st.error("From date is after To date.")
                # FIX: do NOT return — other tabs must still render
            else:
                _consume_loaded_key = f"{key_prefix}_consumption_loaded"
                _run_consume = st.button(
                    "🔎 Load consumption list",
                    key=f"{key_prefix}_load_consumption",
                    use_container_width=True,
                ) or bool(st.session_state.get(_consume_loaded_key))
                if st.session_state.get(_consume_loaded_key):
                    if st.button("↻ Reset / reload consumption list", key=f"{key_prefix}_reset_consumption"):
                        st.session_state.pop(_consume_loaded_key, None)
                        st.rerun()
                if not _run_consume:
                    st.info("Set filters, then load the consumption list. This keeps the page light.")
                    return
                st.session_state[_consume_loaded_key] = True

                _rows = _brq("""
                    SELECT
                        bi.id::text                         AS blank_id,
                        COALESCE(bi.brand,'')               AS brand,
                        COALESCE(bi.category,'')            AS category,
                        COALESCE(bi.material,'')            AS material,
                        COALESCE(bi.colour,'')              AS colour,
                        COALESCE(bi.add_power,0)::float     AS add_power,
                        COALESCE(bi.base_recommended,0)::float AS base_recommended,
                        CASE
                            WHEN bi.add_power > 0 THEN 'PROGRESSIVE'
                            WHEN LOWER(COALESCE(bi.category,'')) LIKE '%%bifocal%%'
                              OR LOWER(COALESCE(bi.category,'')) LIKE '%%d-bif%%'
                              OR LOWER(COALESCE(bi.category,'')) LIKE '%%d bif%%'
                            THEN 'BIFOCAL'
                            ELSE 'SV'
                        END                                 AS lens_type,
                        COALESCE(bi.qty_right,0)            AS qty_right,
                        COALESCE(bi.qty_left,0)             AS qty_left,
                        COALESCE(bi.qty_independent,0)      AS qty_independent,
                        COALESCE(bi.min_stock,0)            AS min_stock,
                        COALESCE(bi.cost_price,0)::float    AS cost_price,
                        COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='R' THEN 1 ELSE 0 END),0) AS consumed_r,
                        COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='L' THEN 1 ELSE 0 END),0) AS consumed_l,
                        COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,'')) NOT IN ('R','L') THEN 1 ELSE 0 END),0) AS consumed_other,
                        COUNT(ba.id) AS total_consumed
                    FROM blank_inventory bi
                    LEFT JOIN blank_allocations ba
                           ON ba.blank_id = bi.id
                          AND COALESCE(ba.allocated_at,bi.created_at)::date BETWEEN %(df)s AND %(dt)s
                    GROUP BY bi.id,bi.brand,bi.category,bi.material,bi.colour,bi.add_power,
                             bi.base_recommended,bi.qty_right,bi.qty_left,
                             bi.qty_independent,bi.min_stock,bi.cost_price
                    HAVING COUNT(ba.id)>0
                        OR (COALESCE(bi.qty_right,0)+COALESCE(bi.qty_left,0)+COALESCE(bi.qty_independent,0))
                           < COALESCE(bi.min_stock,0)
                    ORDER BY total_consumed DESC, bi.brand, bi.category
                    LIMIT 500
                """, {"df": str(_dfrom), "dt": str(_dto)})

                # Apply in-memory filters
                if _bmin:
                    _rows = [r for r in _rows
                             if int(r["qty_right"]+r["qty_left"]+r["qty_independent"]) < int(r["min_stock"] or 0)]
                if _sel_brand and _sel_brand != "All Brands":
                    _rows = [r for r in _rows if r["brand"] == _sel_brand]
                if _sel_cat and _sel_cat != "All Products":
                    _rows = [r for r in _rows if r["category"] == _sel_cat]

                if not _rows:
                    st.info("✅ No blanks consumed in this period and nothing below min stock.")
                    # FIX: do NOT return — other tabs must still render
                else:
                    _tot = sum(int(r["total_consumed"]) for r in _rows)
                    _blw = sum(1 for r in _rows
                               if int(r["qty_right"]+r["qty_left"]+r["qty_independent"]) < int(r["min_stock"] or 0))
                    _m1, _m2, _m3 = st.columns(3)
                    _m1.metric("Blank SKUs", len(_rows))
                    _m2.metric("Total consumed", _tot)
                    _m3.metric("Below min stock", _blw, delta_color="inverse")

                    _sel_key = f"{key_prefix}_selected"
                    if _sel_key not in st.session_state:
                        st.session_state[_sel_key] = {}

                    _by_brand: dict = {}
                    for r in _rows:
                        _by_brand.setdefault(r["brand"] or "(no brand)", []).append(r)

                    for _bn in sorted(_by_brand.keys()):
                        _brows = _by_brand[_bn]
                        with st.expander(
                            f"🏭 **{_bn}** \xb7 {len(_brows)} blank(s) \xb7 consumed: "
                            f"{sum(int(r['total_consumed']) for r in _brows)}",
                            expanded=(len(_by_brand) == 1),
                        ):
                            st.markdown(
                                "<div style='display:grid;"
                                "grid-template-columns:0.5fr 2.5fr 0.8fr 0.8fr 0.7fr 0.7fr 0.8fr 1.2fr;"
                                "gap:8px;padding:4px 8px;font-size:0.7rem;color:#94a3b8;"
                                "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                                "<span></span><span>Product</span><span>Base</span><span>Add</span>"
                                "<span>R used</span><span>L used</span><span>Stock</span><span>Order (pairs)</span>"
                                "</div>", unsafe_allow_html=True)

                            for r in _brows:
                                _bid   = r["blank_id"]
                                _cat   = r["category"] or "\u2014"
                                _mat   = r["material"] or ""
                                _add   = float(r["add_power"] or 0)
                                _base  = float(r["base_recommended"] or 0)
                                _ltype = str(r.get("lens_type") or "SV").upper()
                                _ru, _lu, _iu = int(r["consumed_r"]), int(r["consumed_l"]), int(r["consumed_other"])
                                _rs, _ls, _is = int(r["qty_right"]), int(r["qty_left"]), int(r["qty_independent"])
                                _mn    = int(r["min_stock"] or 0)
                                _ts    = _rs + _ls + _is
                                _blw_  = _ts < _mn

                                # ── Lens type detection ────────────────────────────────
                                # _isp = needs pair logic (progressive, bifocal, D-bifocal)
                                # _show_add = show Add power column
                                # Base always shows for ALL types (SV, bifocal, progressive)
                                _isp = (
                                    _add > 0.01
                                    or _ltype in ("PROGRESSIVE", "BIFOCAL", "D_BIFOCAL", "DBIFOCAL")
                                    or "bifocal" in _cat.lower()
                                    or "d-bif"   in _cat.lower()
                                    or "d bif"   in _cat.lower()
                                    or "progr"   in _cat.lower()
                                )
                                _show_add = _add > 0.01 or _isp  # show add for bifocal even if stored as 0

                                # ── Pair-aware procurement calculation ─────────────────
                                if _isp:
                                    _net_r = max(0, _ru - _rs)
                                    _net_l = max(0, _lu - _ls)
                                    _pairs         = min(_net_r, _net_l)
                                    _half_r_raw    = _net_r - _pairs
                                    _half_l_raw    = _net_l - _pairs
                                    _extra_r_stock = max(0, _rs - _ru)
                                    _extra_l_stock = max(0, _ls - _lu)
                                    _half_r_order  = max(0, _half_r_raw - _extra_l_stock)
                                    _half_l_order  = max(0, _half_l_raw - _extra_r_stock)
                                    _order_r = _pairs + _half_r_order
                                    _order_l = _pairs + _half_l_order
                                    _pair_parts = []
                                    if _pairs > 0:        _pair_parts.append(f"{_pairs} pairs")
                                    if _half_r_order > 0: _pair_parts.append(f"+{_half_r_order}×R½")
                                    if _half_l_order > 0: _pair_parts.append(f"+{_half_l_order}×L½")
                                    _pair_label   = " ".join(_pair_parts) if _pair_parts else "—"
                                    _order_needed = (_order_r + _order_l) > 0
                                    _sel_r, _sel_l, _sv_need = _order_r, _order_l, 0
                                else:
                                    _sv_need      = max(0, _ru + _lu + _iu - _is)
                                    _pair_label   = str(_sv_need) if _sv_need > 0 else "—"
                                    _order_needed = _sv_need > 0
                                    _sel_r, _sel_l = 0, 0
                                    _pairs = _half_r_order = _half_l_order = 0

                                _cc = st.columns([0.5, 2.5, 0.8, 0.8, 0.7, 0.7, 0.8, 1.2])
                                with _cc[0]:
                                    _chk = st.checkbox("s", value=(_bid in st.session_state[_sel_key]),
                                                       key=f"{key_prefix}_chk_{_bid}",
                                                       label_visibility="collapsed")
                                with _cc[1]:
                                    _pl = _cat + (f" \xb7 {_mat}" if _mat else "") + (f" \xb7 {r['colour']}" if r["colour"] else "")
                                    _bb = "<span style='color:#dc2626;font-size:0.65rem'> ⚠ below min</span>" if _blw_ else ""
                                    st.markdown(
                                        f"<div style='padding-top:6px'>"
                                        f"<span style='color:#e2e8f0;font-weight:600;font-size:0.82rem'>{_pl}</span>{_bb}"
                                        f"</div>", unsafe_allow_html=True)
                                with _cc[2]:
                                    # Base always shown for ALL lens types
                                    _base_str = f"{_base:.2f}" if _base > 0 else "—"
                                    _base_color = "#94a3b8" if _base > 0 else "#475569"
                                    st.markdown(f"<div style='padding-top:8px;color:{_base_color};font-size:0.78rem'>"
                                                f"{_base_str}</div>", unsafe_allow_html=True)
                                with _cc[3]:
                                    # Add shown for progressive/bifocal; — for pure SV
                                    if _show_add:
                                        _add_str = f"+{_add:.2f}" if _add > 0 else "D-Bif"
                                        st.markdown(f"<div style='padding-top:8px;color:#94a3b8'>{_add_str}</div>",
                                                    unsafe_allow_html=True)
                                    else:
                                        st.markdown("<div style='padding-top:8px;color:#475569'>—</div>",
                                                    unsafe_allow_html=True)
                                with _cc[4]:
                                    if _isp:
                                        st.markdown(f"<div style='padding-top:8px;color:#3b82f6;font-weight:700'>{_ru}</div>",
                                                    unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"<div style='padding-top:8px;color:#22c55e;font-weight:700'>"
                                                    f"{_ru+_lu+_iu}</div>", unsafe_allow_html=True)
                                with _cc[5]:
                                    if _isp:
                                        st.markdown(f"<div style='padding-top:8px;color:#a855f7;font-weight:700'>{_lu}</div>",
                                                    unsafe_allow_html=True)
                                    else:
                                        st.markdown("<div style='padding-top:8px;color:#475569'>—</div>",
                                                    unsafe_allow_html=True)
                                with _cc[6]:
                                    _stk = f"R:{_rs} L:{_ls}" if _isp else str(_ts)
                                    _mn_lbl = f" <span style='color:#94a3b8'>/min {_mn}</span>" if _mn else ""
                                    st.markdown(f"<div style='padding-top:6px;color:#e2e8f0;font-size:0.78rem'>"
                                                f"{_stk}{_mn_lbl}</div>", unsafe_allow_html=True)
                                with _cc[7]:
                                    _order_color = "#fbbf24" if _order_needed else "#475569"
                                    st.markdown(f"<div style='padding-top:8px;color:{_order_color};font-weight:700;"
                                                f"font-size:0.78rem'>{_pair_label}</div>",
                                                unsafe_allow_html=True)

                                _is_sel = _bid in st.session_state[_sel_key]
                                if _chk and not _is_sel:
                                    st.session_state[_sel_key][_bid] = {
                                        "blank_id":       _bid,
                                        "brand":          _bn,
                                        "category":       _cat,
                                        "material":       _mat,
                                        "base":           _base,
                                        "add":            _add,
                                        "lens_type":      _ltype,
                                        "is_progressive": _isp,
                                        "show_add":       _show_add,
                                        "r_qty":          _sel_r if _isp else 0,
                                        "l_qty":          _sel_l if _isp else 0,
                                        "i_qty":          (_sv_need if not _isp else 0),
                                        "pairs":          _pairs,
                                        "half_r":         _half_r_order,
                                        "half_l":         _half_l_order,
                                        "pair_label":     _pair_label,
                                    }
                                elif not _chk and _is_sel:
                                    st.session_state[_sel_key].pop(_bid, None)

                    _ns = len(st.session_state.get(_sel_key, {}))
                    if _ns:
                        st.info(f"✅ {_ns} blank(s) selected \u2014 go to **📦 Order Blanks** tab to raise PO.")
                    else:
                        st.caption("Tick blanks above then go to 📦 Order Blanks.")

            # ── Rejections this period ─────────────────────────────────────────
            with st.expander("🗑️ Rejected blanks this period", expanded=False):
                st.caption("Blanks that entered the rejection bin from in-house jobs in the selected date range. "
                           "These also count toward replenishment need — tick the parent blank above to include in PO.")
                _rej_rows = _brq("""
                    SELECT
                        rb.id::text                          AS rb_id,
                        rb.blank_id::text                    AS blank_id,
                        rb.eye_side,
                        rb.qty,
                        rb.reason,
                        rb.status,
                        rb.rejected_by,
                        rb.rejected_at::text                 AS rejected_at,
                        o.order_no,
                        COALESCE(bi.brand,'')                AS brand,
                        COALESCE(bi.category,'')             AS category,
                        COALESCE(bi.material,'')             AS material
                    FROM production_rejection_bin rb
                    LEFT JOIN blank_inventory bi ON bi.id = rb.blank_id
                    LEFT JOIN orders o           ON o.id  = rb.order_id
                    WHERE rb.blank_id IS NOT NULL
                      AND rb.rejected_at::date BETWEEN %(df)s AND %(dt)s
                    ORDER BY rb.rejected_at DESC
                    LIMIT 200
                """, {"df": str(_dfrom), "dt": str(_dto)})

                if not _rej_rows:
                    st.info("No blank rejections in this period.")
                else:
                    _rj_m1, _rj_m2, _rj_m3 = st.columns(3)
                    _rj_m1.metric("Rejected blanks", len(_rej_rows))
                    _rj_m2.metric("Unique SKUs", len(set(r["blank_id"] for r in _rej_rows if r["blank_id"])))
                    _rj_m3.metric("Scrapped", sum(1 for r in _rej_rows if r["status"] == "SCRAPPED"))
                    st.markdown("---")
                    for _rr in _rej_rows:
                        _desc = " · ".join(filter(None, [_rr["brand"], _rr["category"], _rr["material"]]))
                        _eye  = _rr["eye_side"] or "—"
                        _ono  = _rr["order_no"] or "—"
                        _by   = _rr["rejected_by"] or "?"
                        _dt_  = str(_rr["rejected_at"] or "")[:10]
                        _sts  = _rr["status"] or "IN_BIN"
                        _sts_color = {"SCRAPPED": "#ef4444", "REWORKED": "#f59e0b",
                                      "RETURNED_TO_STOCK": "#22c55e"}.get(_sts, "#64748b")
                        st.markdown(
                            f"<div style='background:#0a1628;border-left:3px solid #ef4444;"
                            f"border-radius:4px;padding:6px 12px;margin-bottom:4px;"
                            f"display:flex;justify-content:space-between;align-items:center'>"
                            f"<div><span style='color:#fca5a5;font-weight:700'>🗑 {_desc or _rr['blank_id']}</span>"
                            f"&nbsp;<span style='color:#475569;font-size:0.75rem'>"
                            f"· Eye {_eye} · Qty {_rr['qty']} · Order {_ono}"
                            f"{'· ' + _rr['reason'] if _rr['reason'] else ''}</span></div>"
                            f"<div style='text-align:right'>"
                            f"<span style='color:{_sts_color};font-size:0.7rem;font-weight:700'>{_sts}</span>"
                            f"<br><span style='color:#94a3b8;font-size:0.7rem'>{_dt_} · {_by}</span>"
                            f"</div></div>",
                            unsafe_allow_html=True,
                        )

    # ══════════════════════════════════════════════════════════════════
    # TAB 2 — RETURNS
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "♻️ Returns":
        with _tab_returns:
            st.caption("Blanks returned to stock when an in-house order was rolled back.")
            _rc1, _rc2 = st.columns(2)
            _today_r = _dt.date.today()
            _rfrom = _rc1.date_input("From", value=_today_r - _dt.timedelta(days=30),
                                      key=f"{key_prefix}_rfrom", format="DD/MM/YYYY")
            _rto   = _rc2.date_input("To",   value=_today_r,
                                      key=f"{key_prefix}_rto",   format="DD/MM/YYYY")
            _ret = _brq("""
                SELECT bl.created_at::date AS return_date,
                       o.order_no,
                       bi.brand, bi.category, bi.material,
                       bl.eye_side, bl.qty_change, bl.created_by
                FROM blank_stock_ledger bl
                JOIN blank_inventory bi ON bi.id = bl.blank_id
                LEFT JOIN orders o       ON o.id  = bl.ref_id
                WHERE bl.ref_type = 'ROLLBACK'
                  AND bl.qty_change > 0
                  AND bl.created_at::date BETWEEN %(df)s AND %(dt)s
                ORDER BY bl.created_at DESC
                LIMIT 200
            """, {"df": str(_rfrom), "dt": str(_rto)})
            if not _ret:
                st.info("No blanks returned in this period.")
            else:
                _m1, _m2 = st.columns(2)
                _m1.metric("Returns in period", len(_ret))
                _m2.metric("Unique orders", len(set(r["order_no"] or "" for r in _ret)))
                st.markdown("---")
                for r in _ret:
                    _bname = f"{r['brand']} {r['category']}" + (f" \xb7 {r['material']}" if r["material"] else "")
                    _ono   = r["order_no"] or "\u2014"
                    _eye   = r["eye_side"] or ""
                    _by    = r["created_by"] or "?"
                    _rd    = str(r["return_date"])
                    st.markdown(
                        f"<div style='background:#0a1628;border-left:3px solid #22c55e;"
                        f"border-radius:4px;padding:6px 12px;margin-bottom:4px;"
                        f"display:flex;justify-content:space-between;align-items:center'>"
                        f"<div><span style='color:#86efac;font-weight:700'>♻️ {_bname}</span>"
                        f"&nbsp;<span style='color:#475569;font-size:0.75rem'>\xb7 Eye {_eye} \xb7 {_ono}</span></div>"
                        f"<span style='color:#94a3b8;font-size:0.75rem'>{_rd} by {_by}</span>"
                        f"</div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    # TAB 3 — ORDER BLANKS  (supplier-wise with shift capability)
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "📦 Order Blanks":
        with _tab_order:
            _sel_key3 = f"{key_prefix}_selected"
            _sel3 = st.session_state.get(_sel_key3, {})
            if not _sel3:
                st.info("No blanks selected. Go to 📊 Consumption tab, tick blanks, then return here.")
            else:
                _clear_col, _keep_col = st.columns([1, 5])
                if _clear_col.button("Clear selected", key=f"{key_prefix}_clear_selected_blanks", use_container_width=True):
                    st.session_state.pop(_sel_key3, None)
                    st.session_state.pop(f"{key_prefix}_sup_assign", None)
                    st.rerun()
                # ── Load supplier list once ────────────────────────────────
                _sup_list = _brq("""
                    SELECT id::text AS id, party_name,
                           COALESCE(mobile,'')  AS mobile,
                           COALESCE(email,'')   AS email
                    FROM parties
                    WHERE UPPER(COALESCE(party_type,'')) IN ('SUPPLIER','VENDOR','LAB','EXTERNAL_LAB')
                      AND COALESCE(is_active,TRUE) = TRUE
                    ORDER BY party_name
                """)
                _sup_by_id = {s["id"]: s for s in _sup_list}
                _sup_name_opts = [""] + [s["id"] for s in _sup_list]

                # ── Load last-used supplier per blank from PO history ─────
                _last_sup_by_blank: dict = {}
                if _sel3:
                    _bid_list = list(_sel3.keys())
                    _ph = ",".join([f"%(b{i})s::uuid" for i in range(len(_bid_list))])
                    _ph_params = {f"b{i}": _bid_list[i] for i in range(len(_bid_list))}
                    _lsup_rows = _brq(f"""
                        SELECT DISTINCT ON (li_item->>'blank_id')
                            li_item->>'blank_id'  AS blank_id,
                            bro.supplier_id::text AS supplier_id,
                            p.party_name          AS supplier_name,
                            bro.order_date        AS order_date
                        FROM blank_replenishment_orders bro
                        CROSS JOIN LATERAL jsonb_array_elements(bro.line_items) AS li_item
                        LEFT JOIN parties p ON p.id = bro.supplier_id
                        WHERE (li_item->>'blank_id')::uuid IN ({_ph})
                          AND bro.status IN ('SENT','RECEIVED','PARTIALLY_RECEIVED')
                        ORDER BY li_item->>'blank_id', bro.order_date DESC
                    """, _ph_params) if _ph_params else []
                    for _lr in _lsup_rows:
                        if _lr["blank_id"]:
                            _last_sup_by_blank[_lr["blank_id"]] = {
                                "supplier_id":   _lr["supplier_id"],
                                "supplier_name": _lr["supplier_name"] or "?",
                                "order_date":    str(_lr["order_date"] or "")[:10],
                            }

                st.markdown(f"**{len(_sel3)} blank(s) — assign supplier per blank, then raise PO(s)**")
                st.caption("Each blank can go to a different supplier. Partial splits allowed — adjust quantities per supplier.")

                # ── Per-blank supplier assignment with split support ───────
                # session key: _sup_assign_{key_prefix}_{blank_id} → list of {sup_id, r_qty, l_qty, i_qty}
                _assign_key = f"{key_prefix}_sup_assign"
                if _assign_key not in st.session_state:
                    st.session_state[_assign_key] = {}

                _assign: dict = st.session_state[_assign_key]

                # Initialise assignments for newly-added blanks
                for _bid, _meta in _sel3.items():
                    if _bid not in _assign:
                        _def_sup = _last_sup_by_blank.get(_bid, {}).get("supplier_id", "")
                        _assign[_bid] = [{
                            "sup_id": _def_sup,
                            "r_qty":  int(_meta.get("r_qty", 0)),
                            "l_qty":  int(_meta.get("l_qty", 0)),
                            "i_qty":  int(_meta.get("i_qty", 0)),
                        }]
                # Remove stale blank assignments
                for _bid in list(_assign.keys()):
                    if _bid not in _sel3:
                        del _assign[_bid]

                # ── Render per-blank assignment UI ────────────────────────
                for _bid, _meta in _sel3.items():
                    _lbl = " · ".join(filter(None, [
                        _meta.get("brand"), _meta.get("category"), _meta.get("material"),
                    ]))
                    # Base always shown for all lens types
                    _base_v = float(_meta.get("base") or 0)
                    if _base_v > 0: _lbl += f" · Base {_base_v:.2f}"
                    # Add shown for progressive/bifocal
                    _add_v  = float(_meta.get("add") or 0)
                    if _meta.get("show_add") or _add_v > 0.01:
                        _lbl += f" · Add +{_add_v:.2f}" if _add_v > 0.01 else " · D-Bif"
                    _isp = _meta.get("is_progressive", False)
                    _total_r = int(_meta.get("r_qty", 0))
                    _total_l = int(_meta.get("l_qty", 0))
                    _total_i = int(_meta.get("i_qty", 0))
                    _plbl    = _meta.get("pair_label", "")

                    _last_sup_info = _last_sup_by_blank.get(_bid)
                    _last_hint = (f" · Last ordered from **{_last_sup_info['supplier_name']}** "
                                  f"({_last_sup_info['order_date']})"
                                  if _last_sup_info else " · No order history")

                    with st.expander(
                        f"**{_lbl}** · {_plbl or str(_total_r+_total_l+_total_i)+' pcs'}{_last_hint}",
                        expanded=True,
                    ):
                        _splits = _assign.get(_bid, [])

                        # ── Render each split row ──────────────────────────
                        _new_splits = []
                        _used_r, _used_l, _used_i = 0, 0, 0

                        for _si, _split in enumerate(_splits):
                            _sc1, _sc2, _sc3, _sc4 = st.columns([3, 1, 1, 0.5])
                            _s_sup = _sc1.selectbox(
                                "Supplier",
                                _sup_name_opts,
                                index=(_sup_name_opts.index(_split["sup_id"])
                                       if _split["sup_id"] in _sup_name_opts else 0),
                                format_func=lambda x: _sup_by_id.get(x, {}).get("party_name", "— Select —") if x else "— Select supplier —",
                                key=f"{key_prefix}_sas_{_bid}_{_si}",
                                label_visibility="collapsed",
                            )
                            if _isp:
                                _s_r = _sc2.number_input("R", min_value=0,
                                                          value=int(_split.get("r_qty", 0)),
                                                          step=1,
                                                          key=f"{key_prefix}_sar_{_bid}_{_si}",
                                                          label_visibility="visible")
                                _s_l = _sc3.number_input("L", min_value=0,
                                                          value=int(_split.get("l_qty", 0)),
                                                          step=1,
                                                          key=f"{key_prefix}_sal_{_bid}_{_si}",
                                                          label_visibility="visible")
                                _s_i = 0
                            else:
                                _s_r, _s_l = 0, 0
                                _s_i = _sc2.number_input("Qty", min_value=0,
                                                          value=int(_split.get("i_qty", 0)),
                                                          step=1,
                                                          key=f"{key_prefix}_sai_{_bid}_{_si}",
                                                          label_visibility="visible")
                            # Remove split button
                            _remove = False
                            if len(_splits) > 1:
                                with _sc4:
                                    st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
                                    if st.button("✕", key=f"{key_prefix}_rm_{_bid}_{_si}",
                                                 use_container_width=True):
                                        _remove = True
                            if not _remove:
                                _new_splits.append({"sup_id": _s_sup,
                                                    "r_qty": _s_r, "l_qty": _s_l, "i_qty": _s_i})
                                _used_r += _s_r; _used_l += _s_l; _used_i += _s_i

                        # ── Remainder / validation ─────────────────────────
                        _rem_r = _total_r - _used_r
                        _rem_l = _total_l - _used_l
                        _rem_i = _total_i - _used_i
                        _has_rem = (_rem_r + _rem_l + _rem_i) > 0
                        _over    = _used_r > _total_r or _used_l > _total_l or _used_i > _total_i

                        if _over:
                            st.warning(f"⚠ Quantities exceed requirement — "
                                       f"R: need {_total_r} assigned {_used_r}  "
                                       f"L: need {_total_l} assigned {_used_l}")
                        elif _has_rem:
                            if _isp:
                                st.caption(f"Unassigned: R {_rem_r}  L {_rem_l}")
                            else:
                                st.caption(f"Unassigned: {_rem_i} pcs")

                        # ── Add split button ───────────────────────────────
                        if _has_rem and st.button(
                            f"➕ Add another supplier for remaining "
                            + (f"R:{_rem_r} L:{_rem_l}" if _isp else f"{_rem_i} pcs"),
                            key=f"{key_prefix}_addsplit_{_bid}",
                        ):
                            _new_splits.append({
                                "sup_id": "",
                                "r_qty": _rem_r if _isp else 0,
                                "l_qty": _rem_l if _isp else 0,
                                "i_qty": _rem_i if not _isp else 0,
                            })

                        _assign[_bid] = _new_splits

                st.session_state[_assign_key] = _assign

                # ── Consolidate: group splits by supplier ──────────────────
                # supplier_id → list of (blank_id, meta, r_qty, l_qty, i_qty)
                _by_sup: dict = {}
                _unassigned_items = []
                for _bid, _splits in _assign.items():
                    _meta = _sel3.get(_bid, {})
                    for _sp in _splits:
                        _sid = _sp.get("sup_id", "")
                        _rq  = int(_sp.get("r_qty", 0))
                        _lq  = int(_sp.get("l_qty", 0))
                        _iq  = int(_sp.get("i_qty", 0))
                        if (_rq + _lq + _iq) == 0:
                            continue
                        if not _sid:
                            _unassigned_items.append((_bid, _meta, _rq, _lq, _iq))
                            continue
                        _by_sup.setdefault(_sid, []).append((_bid, _meta, _rq, _lq, _iq))

                if _unassigned_items:
                    st.warning(f"⚠ {len(_unassigned_items)} blank(s) have no supplier assigned — select a supplier above.")

                # ── Preview: one card per supplier ────────────────────────
                st.markdown("---")
                st.markdown("**📋 PO Preview — one PO per supplier**")

                if not _by_sup:
                    st.info("Assign at least one supplier above to preview POs.")
                else:
                    for _sid, _sitems in _by_sup.items():
                        _sinfo = _sup_by_id.get(_sid, {})
                        _sname = _sinfo.get("party_name", "Unknown")
                        _ol_prev = []
                        for (_bid, _meta, _rq, _lq, _iq) in _sitems:
                            _isp = _meta.get("is_progressive", False)
                            if _isp:
                                _fp = min(_rq, _lq)
                                _hr = _rq - _fp; _hl = _lq - _fp
                                _qp = []
                                if _fp: _qp.append(f"{_fp} pairs")
                                if _hr: _qp.append(f"{_hr} R only")
                                if _hl: _qp.append(f"{_hl} L only")
                                _qs = " + ".join(_qp) or "—"
                            else:
                                _qs = f"{_iq} pcs"
                            _bl = " · ".join(filter(None, [_meta.get("brand"), _meta.get("category"), _meta.get("material")]))
                            _bv = float(_meta.get("base") or 0)
                            _av = float(_meta.get("add") or 0)
                            if _bv > 0: _bl += f" · Base {_bv:.2f}"
                            if _meta.get("show_add") or _av > 0.01:
                                _bl += f" · Add +{_av:.2f}" if _av > 0.01 else " · D-Bif"
                            _ol_prev.append(f"• {_bl} — {_qs}")

                        with st.container(border=True):
                            _ph1, _ph2 = st.columns([4, 1])
                            _ph1.markdown(f"**{_sname}** · {len(_sitems)} blank SKU(s)")
                            # WA / Email send buttons inline
                            _mob_p = "".join(d for d in (_sinfo.get("mobile","")) if d.isdigit())
                            if _mob_p and not _mob_p.startswith("91"):
                                _mob_p = "91" + _mob_p
                            _wm_p = (
                                f"Dear {_sname},\n\nPlease supply the following lens blanks:\n\n"
                                + "\n".join(_ol_prev)
                                + "\n\nKindly confirm availability and delivery date.\nThank you."
                            )
                            _wa_p = (f"https://wa.me/{_mob_p}?text={_up.quote(_wm_p)}"
                                     if _mob_p else f"https://wa.me/?text={_up.quote(_wm_p)}")
                            _ph2.link_button("📲 WA", _wa_p, use_container_width=True)

                            for _line in _ol_prev:
                                st.markdown(
                                    f"<div style='color:#e2e8f0;font-size:0.82rem;padding:2px 0'>{_line}</div>",
                                    unsafe_allow_html=True)

                            _em_p = _sinfo.get("email","")
                            if _em_p:
                                _subj_p = _up.quote(f"Blank Lens Order — {_dt.date.today()}")
                                st.link_button("📧 Email", f"mailto:{_em_p}?subject={_subj_p}&body={_up.quote(_wm_p)}",
                                               use_container_width=True)

                # ── Excel download (consolidated, supplier column) ─────────
                st.markdown("---")
                try:
                    import io, openpyxl as _oxl
                    _wb2 = _oxl.Workbook(); _ws2 = _wb2.active; _ws2.title = "Blank Order"
                    _ws2.append(["Supplier","Brand","Category","Material","Base","Add","R Qty","L Qty","Qty (SV)","Pairs Summary"])
                    for _sid, _sitems in _by_sup.items():
                        _sname2 = _sup_by_id.get(_sid, {}).get("party_name","")
                        for (_bid, _meta, _rq, _lq, _iq) in _sitems:
                            _isp2 = _meta.get("is_progressive", False)
                            _bv2  = float(_meta.get("base") or 0)
                            _av2  = float(_meta.get("add") or 0)
                            _add_str2 = f"+{_av2:.2f}" if _av2 > 0.01 else ("D-Bif" if _meta.get("show_add") else "")
                            _fp2 = min(_rq,_lq) if _isp2 else 0
                            _hr2 = (_rq-_fp2) if _isp2 else 0
                            _hl2 = (_lq-_fp2) if _isp2 else 0
                            _ps2 = []
                            if _fp2: _ps2.append(f"{_fp2} pairs")
                            if _hr2: _ps2.append(f"{_hr2} R only")
                            if _hl2: _ps2.append(f"{_hl2} L only")
                            _ws2.append([
                                _sname2,
                                _meta.get("brand",""), _meta.get("category",""), _meta.get("material",""),
                                f"{_bv2:.2f}" if _bv2 > 0 else "",   # Base — always
                                _add_str2,                             # Add — bifocal/progressive only
                                _rq if _isp2 else "", _lq if _isp2 else "", _iq if not _isp2 else "",
                                " + ".join(_ps2) if _ps2 else str(_iq),
                            ])
                    _xb2 = io.BytesIO(); _wb2.save(_xb2); _xb2.seek(0)
                    st.download_button("📊 Download full Excel", data=_xb2.getvalue(),
                                       file_name=f"blank_order_{_dt.date.today()}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
                except ImportError:
                    st.caption("openpyxl not installed — Excel download unavailable.")

                # ── Record all POs at once ─────────────────────────────────
                st.markdown("---")
                _por = st.text_input("PO Reference / Note (applied to all POs)",
                                      key=f"{key_prefix}_po_ref",
                                      placeholder="e.g. Urgent reorder June 2026")
                _po_btn_disabled = (not _by_sup) or bool(_unassigned_items)
                if st.button(
                    f"✅ Record {len(_by_sup)} PO(s) Sent",
                    key=f"{key_prefix}_po_record",
                    type="primary",
                    use_container_width=True,
                    disabled=_po_btn_disabled,
                ):
                    import uuid as _u_po
                    _po_saved = 0
                    for _sid, _sitems in _by_sup.items():
                        _items_with_id2 = []
                        for (_bid, _meta, _rq, _lq, _iq) in _sitems:
                            _items_with_id2.append({
                                **_meta,
                                "blank_id": _bid,
                                "r_qty": _rq,
                                "l_qty": _lq,
                                "i_qty": _iq,
                            })
                        _pok = _brw("""
                            INSERT INTO blank_replenishment_orders
                                (id, supplier_id, order_date, status, line_items, remarks, created_at)
                            VALUES
                                (%(id)s::uuid, %(sid)s::uuid, NOW(), 'SENT',
                                 %(items)s::jsonb, %(rmk)s, NOW())
                        """, {
                            "id":    str(_u_po.uuid4()),
                            "sid":   _sid,
                            "items": _brj.dumps(_items_with_id2),
                            "rmk":   _por or "",
                        })
                        if _pok:
                            _po_saved += 1
                    if _po_saved:
                        st.success(f"✅ {_po_saved} PO(s) recorded. Switch to 🧾 Upload Invoice once goods arrive.")
                        st.session_state.pop(_sel_key3, None)
                        st.session_state.pop(_assign_key, None)
                        st.rerun()
                    else:
                        st.warning("Could not record PO to DB. Check blank_replenishment_orders table exists.")

    # ══════════════════════════════════════════════════════════════════
    # TAB 4 — UPLOAD INVOICE
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "🧾 Upload Invoice":
        with _tab_invoice:
            st.caption("Record a supplier invoice against an open blank PO.")

            # ── Open PO tracker ───────────────────────────────────────────
            _all_pos = _brq("""
                SELECT bro.id::text         AS po_id,
                       bro.supplier_id::text AS supplier_id,
                       p.party_name         AS supplier,
                       bro.order_date::date AS order_date,
                       bro.status,
                       COALESCE(bro.remarks,'') AS remarks,
                       (
                           SELECT COUNT(*)
                           FROM blank_replenishment_invoices bri
                           WHERE bri.po_id = bro.id
                       ) AS invoice_count,
                       (
                           SELECT COALESCE(SUM(bri.invoice_amount),0)
                           FROM blank_replenishment_invoices bri
                           WHERE bri.po_id = bro.id
                       ) AS total_invoiced
                FROM blank_replenishment_orders bro
                LEFT JOIN parties p ON p.id = bro.supplier_id
                ORDER BY bro.order_date DESC
                LIMIT 100
            """)
            _supplier_rows_inv = _brq("""
                SELECT id::text AS supplier_id, party_name
                FROM parties
                WHERE UPPER(COALESCE(party_type,'')) IN ('SUPPLIER','VENDOR','LAB','EXTERNAL_LAB')
                  AND COALESCE(is_active, TRUE) = TRUE
                ORDER BY party_name
            """)
            _supplier_ids_inv = [r["supplier_id"] for r in _supplier_rows_inv]
            _supplier_by_id_inv = {r["supplier_id"]: r["party_name"] for r in _supplier_rows_inv}

            _open_pos  = [p for p in _all_pos if p["status"] in ("SENT","PARTIALLY_RECEIVED")]
            _closed_pos = [p for p in _all_pos if p["status"] not in ("SENT","PARTIALLY_RECEIVED")]

            _po_m1, _po_m2, _po_m3 = st.columns(3)
            _po_m1.metric("Open POs", len(_open_pos))
            _po_m2.metric("Received POs", len(_closed_pos))
            _po_m3.metric("Total POs", len(_all_pos))

            if _open_pos:
                with st.expander(f"📋 Open POs awaiting delivery ({len(_open_pos)})", expanded=True):
                    for _p in _open_pos:
                        _sts_color = {"SENT": "#f59e0b", "PARTIALLY_RECEIVED": "#3b82f6"}.get(_p["status"], "#64748b")
                        _inv_note_po = f" · {_p['invoice_count']} invoice(s) · ₹{float(_p['total_invoiced'] or 0):,.0f}" if _p["invoice_count"] else ""
                        st.markdown(
                            f"<div style='background:#0a1628;border-left:3px solid {_sts_color};"
                            f"border-radius:4px;padding:6px 12px;margin-bottom:4px;"
                            f"display:flex;justify-content:space-between;align-items:center'>"
                            f"<div><span style='color:#e2e8f0;font-weight:700'>{_p['supplier'] or '—'}</span>"
                            f"<span style='color:#475569;font-size:0.75rem'> · {_p['order_date']}"
                            f"{(' · ' + _p['remarks'][:30]) if _p['remarks'] else ''}{_inv_note_po}</span></div>"
                            f"<span style='color:{_sts_color};font-size:0.7rem;font-weight:700'>{_p['status']}</span>"
                            f"</div>", unsafe_allow_html=True)

            # ── Received PO history ───────────────────────────────────────
            if _closed_pos:
                with st.expander(f"✅ Received POs history ({len(_closed_pos)})", expanded=False):
                    for _p in _closed_pos:
                        _inv_note_po = f" · {_p['invoice_count']} invoice(s) · ₹{float(_p['total_invoiced'] or 0):,.0f}" if _p["invoice_count"] else ""
                        st.markdown(
                            f"<div style='background:#0a1628;border-left:3px solid #22c55e;"
                            f"border-radius:4px;padding:5px 12px;margin-bottom:3px;"
                            f"display:flex;justify-content:space-between;align-items:center'>"
                            f"<div><span style='color:#86efac;font-weight:600'>{_p['supplier'] or '—'}</span>"
                            f"<span style='color:#475569;font-size:0.75rem'> · {_p['order_date']}"
                            f"{(' · ' + _p['remarks'][:30]) if _p['remarks'] else ''}{_inv_note_po}</span></div>"
                            f"<span style='color:#22c55e;font-size:0.7rem'>{_p['status']}</span>"
                            f"</div>", unsafe_allow_html=True)

            if _all_pos and _supplier_ids_inv:
                with st.expander("🛠 Correct PO / invoice supplier", expanded=False):
                    st.caption("Use only when the PO was raised to one supplier but the actual invoice/goods came from another supplier.")
                    _corr_po_by_id = {p["po_id"]: p for p in _all_pos}

                    def _corr_po_fmt(x):
                        p = _corr_po_by_id[x]
                        inv_note = f" · {p['invoice_count']} inv" if p.get("invoice_count") else ""
                        return f"{p['supplier'] or '—'} — {p['order_date']} — {p['status']}{inv_note}"

                    _corr_po = st.selectbox(
                        "PO / invoice to correct",
                        [p["po_id"] for p in _all_pos],
                        format_func=_corr_po_fmt,
                        key=f"{key_prefix}_corr_po",
                    )
                    _corr_current = str(_corr_po_by_id[_corr_po].get("supplier_id") or "")
                    _corr_index = _supplier_ids_inv.index(_corr_current) if _corr_current in _supplier_ids_inv else 0
                    _corr_supplier = st.selectbox(
                        "Correct supplier",
                        _supplier_ids_inv,
                        index=_corr_index,
                        format_func=lambda x: _supplier_by_id_inv.get(x, "—"),
                        key=f"{key_prefix}_corr_supplier",
                    )
                    _corr_reason = st.text_input(
                        "Correction reason",
                        placeholder="Example: invoice received from actual supplier",
                        key=f"{key_prefix}_corr_reason",
                    )
                    _old_supplier_name = _corr_po_by_id[_corr_po].get("supplier") or "—"
                    _new_supplier_name = _supplier_by_id_inv.get(_corr_supplier, "—")
                    if _corr_supplier == _corr_current:
                        st.info("Selected supplier is already set on this PO.")
                    elif st.button("Apply supplier correction", key=f"{key_prefix}_apply_supplier_correction", type="primary"):
                        try:
                            from modules.sql_adapter import run_transaction
                            _corr_note = (
                                f"Supplier corrected {_old_supplier_name} -> {_new_supplier_name}"
                                + (f" | {_corr_reason.strip()}" if _corr_reason.strip() else "")
                            )
                            run_transaction([
                                ("""
                                    UPDATE blank_replenishment_orders
                                    SET supplier_id = %(sid)s::uuid,
                                        remarks = TRIM(BOTH ' |' FROM COALESCE(remarks,'') || ' | ' || %(note)s),
                                        updated_at = NOW()
                                    WHERE id = %(po)s::uuid
                                """, {"sid": _corr_supplier, "po": _corr_po, "note": _corr_note}),
                            ])
                            st.success(f"Supplier corrected to {_new_supplier_name}. Reports will now follow the corrected supplier.")
                            st.rerun()
                        except Exception as _corr_e:
                            st.error(f"Supplier correction failed and was rolled back: {_corr_e}")

            st.markdown("---")
            st.markdown("**Record invoice against open PO**")

            _pos = [p for p in _all_pos if p["status"] in ("SENT","PARTIALLY_RECEIVED")]
            if not _pos:
                st.info("No open blank POs. Raise a PO from the 📦 Order Blanks tab first.")
            else:
                _po_by_id = {p["po_id"]: p for p in _pos}
                def _po_fmt(x):
                    p = _po_by_id[x]
                    rmk = f" ({p['remarks'][:25]})" if p["remarks"] else ""
                    return f"{p['supplier']} — {p['order_date']} — {p['status']}{rmk}"
                _sel_po = st.selectbox("Select PO to invoice", [p["po_id"] for p in _pos],
                                       format_func=_po_fmt, key=f"{key_prefix}_inv_po")

                # ── Load PO line items to show per-line received qty ─────
                _po_lines_data = _brq("""
                    SELECT bro.line_items, bro.status,
                           bro.supplier_id::text AS supplier_id,
                           COALESCE(p.party_name,'') AS supplier,
                           bro.order_date::date AS order_date
                    FROM blank_replenishment_orders bro
                    LEFT JOIN parties p ON p.id = bro.supplier_id
                    WHERE bro.id = %(po)s::uuid LIMIT 1
                """, {"po": _sel_po})

                _po_line_items = []
                if _po_lines_data and _po_lines_data[0].get("line_items"):
                    import json as _inv_json
                    _raw = _po_lines_data[0]["line_items"]
                    if isinstance(_raw, str):
                        try: _raw = _inv_json.loads(_raw)
                        except: _raw = []
                    _po_line_items = [li for li in _raw if li.get("blank_id")]

                _po_supplier_id = str((_po_lines_data[0].get("supplier_id") if _po_lines_data else "") or "")
                _actual_supplier_id = _po_supplier_id
                if _supplier_ids_inv:
                    _sup_index = _supplier_ids_inv.index(_po_supplier_id) if _po_supplier_id in _supplier_ids_inv else 0
                    _actual_supplier_id = st.selectbox(
                        "Actual invoice supplier",
                        _supplier_ids_inv,
                        index=_sup_index,
                        format_func=lambda x: _supplier_by_id_inv.get(x, "—"),
                        key=f"{key_prefix}_actual_supplier",
                        help="Change this if the PO was raised to one supplier but the invoice/goods came from another.",
                    )
                    if _actual_supplier_id != _po_supplier_id:
                        st.warning(
                            "Supplier correction will be saved with this invoice. "
                            f"PO supplier will change to {_supplier_by_id_inv.get(_actual_supplier_id, 'selected supplier')}."
                        )

                # ── Auto-calculate remaining qty from previous invoices ───
                # Query blank_stock_ledger for all INVOICE_RECEIPT entries
                # against this PO to know what has already been received.
                _already_recd: dict = {}   # blank_id → {r, l, i}
                if _po_line_items:
                    _blank_ids_po = [str(li["blank_id"]) for li in _po_line_items]
                    _ph = ", ".join([f"%(b{i})s::uuid" for i in range(len(_blank_ids_po))])
                    _ph_p = {f"b{i}": _blank_ids_po[i] for i in range(len(_blank_ids_po))}
                    _prev_rcvd = _brq(f"""
                        SELECT
                            blank_id::text,
                            eye_side,
                            COALESCE(SUM(qty_change), 0) AS total_recd
                        FROM blank_stock_ledger
                        WHERE ref_type = 'INVOICE_RECEIPT'
                          AND ref_id   = %(po)s::uuid
                          AND blank_id IN ({_ph})
                        GROUP BY blank_id, eye_side
                    """, {"po": _sel_po, **_ph_p}) if _blank_ids_po else []

                    for _pr in _prev_rcvd:
                        _bid_pr = str(_pr["blank_id"])
                        if _bid_pr not in _already_recd:
                            _already_recd[_bid_pr] = {"r": 0, "l": 0, "i": 0}
                        _eye_pr = str(_pr["eye_side"] or "").upper()
                        _qty_pr = int(_pr["total_recd"] or 0)
                        if   _eye_pr == "R": _already_recd[_bid_pr]["r"] += _qty_pr
                        elif _eye_pr == "L": _already_recd[_bid_pr]["l"] += _qty_pr
                        else:                _already_recd[_bid_pr]["i"] += _qty_pr

                # Show previous receipt summary if this is a second/third invoice
                _total_prev = sum(
                    sum(v.values()) for v in _already_recd.values()
                )
                if _total_prev > 0:
                    st.info(
                        f"📦 This PO has **{_total_prev} unit(s)** already received via previous invoice(s). "
                        f"Remaining quantities are pre-filled below — adjust as needed."
                    )

                st.markdown("---")

                # ── Invoice details ───────────────────────────────────────
                _iv1, _iv2, _iv3 = st.columns(3)
                _inv_no   = _iv1.text_input("Invoice Number", key=f"{key_prefix}_inv_no",
                                             placeholder="INV-2024-001")
                _inv_date = _iv2.date_input("Invoice Date", key=f"{key_prefix}_inv_date",
                                             value=_dt.date.today(), format="DD/MM/YYYY")
                _inv_amt  = _iv3.number_input("Invoice Amount (₹)", min_value=0.0, step=100.0,
                                               key=f"{key_prefix}_inv_amt")
                _inv_gst  = st.number_input("GST Amount (₹)", min_value=0.0, step=10.0,
                                             key=f"{key_prefix}_inv_gst")
                _inv_file = st.file_uploader("Invoice PDF / Image (optional)",
                                              type=["pdf","jpg","jpeg","png"],
                                              key=f"{key_prefix}_inv_file")
                _inv_note = st.text_input("Notes", key=f"{key_prefix}_inv_note",
                                           placeholder="Partial delivery, qty difference, etc.")

                # ── Per-line received qty editor ───────────────────────────
                st.markdown("**Received quantities — edit if supplier sent partial or different qty**")
                st.caption("Defaults to ordered qty. Change any line if actual received differs. "
                           "Partial → PO marked PARTIALLY_RECEIVED. Full → RECEIVED.")

                _recd_qtys: dict = {}   # blank_id → {r, l, i}
                _is_partial = False

                if _po_line_items:
                    # Column headers
                    st.markdown(
                        "<div style='display:grid;"
                        "grid-template-columns:3fr 1fr 1fr 1fr 1fr 1fr 1fr;"
                        "gap:6px;padding:4px 8px;font-size:0.7rem;color:#94a3b8;"
                        "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                        "<span>Blank</span>"
                        "<span>Ord R</span><span>Ord L</span><span>Ord Qty</span>"
                        "<span>Recd R</span><span>Recd L</span><span>Recd Qty</span>"
                        "</div>", unsafe_allow_html=True)

                    for _li_idx, _li in enumerate(_po_line_items):
                        _bid_li  = str(_li.get("blank_id",""))
                        _ord_r   = int(_li.get("r_qty") or 0)
                        _ord_l   = int(_li.get("l_qty") or 0)
                        _ord_i   = int(_li.get("i_qty") or 0)
                        _isp_li  = (_ord_r + _ord_l) > 0
                        _bname_li = " · ".join(filter(None,[
                            _li.get("brand"), _li.get("category"), _li.get("material")
                        ]))
                        _bv_li = float(_li.get("base") or 0)
                        _av_li = float(_li.get("add")  or 0)
                        if _bv_li > 0: _bname_li += f" Base {_bv_li:.0f}"
                        if _av_li > 0.01: _bname_li += f" +{_av_li:.2f}"

                        # Auto-calculate remaining = ordered − already received
                        _prev = _already_recd.get(_bid_li, {"r": 0, "l": 0, "i": 0})
                        _rem_r = max(0, _ord_r - _prev["r"])
                        _rem_l = max(0, _ord_l - _prev["l"])
                        _rem_i = max(0, _ord_i - _prev["i"])

                        # Show already-received hint on the blank label if any prior receipt
                        _prev_hint = ""
                        if (_prev["r"] + _prev["l"] + _prev["i"]) > 0:
                            if _isp_li:
                                _prev_hint = f" <span style='color:#f59e0b;font-size:0.7rem'>(prev: R:{_prev['r']} L:{_prev['l']})</span>"
                            else:
                                _prev_hint = f" <span style='color:#f59e0b;font-size:0.7rem'>(prev: {_prev['i']})</span>"

                        _rc1,_rc2,_rc3,_rc4,_rc5,_rc6,_rc7 = st.columns([3,1,1,1,1,1,1])
                        _rc1.markdown(
                            f"<div style='padding-top:8px;color:#e2e8f0;font-size:0.8rem'>"
                            f"{_bname_li}{_prev_hint}</div>",
                            unsafe_allow_html=True)
                        # Ordered (read-only)
                        _rc2.markdown(f"<div style='padding-top:8px;color:#94a3b8;text-align:center'>{_ord_r or '—'}</div>", unsafe_allow_html=True)
                        _rc3.markdown(f"<div style='padding-top:8px;color:#94a3b8;text-align:center'>{_ord_l or '—'}</div>", unsafe_allow_html=True)
                        _rc4.markdown(f"<div style='padding-top:8px;color:#94a3b8;text-align:center'>{_ord_i or '—'}</div>", unsafe_allow_html=True)
                        # Received — default = remaining balance
                        if _isp_li:
                            _rec_r = _rc5.number_input("R", min_value=0, value=_rem_r, max_value=_rem_r, step=1,
                                                        key=f"{key_prefix}_recr_{_bid_li}_{_li_idx}",
                                                        label_visibility="collapsed")
                            _rec_l = _rc6.number_input("L", min_value=0, value=_rem_l, max_value=_rem_l, step=1,
                                                        key=f"{key_prefix}_recl_{_bid_li}_{_li_idx}",
                                                        label_visibility="collapsed")
                            _rec_i = 0
                        else:
                            _rec_r, _rec_l = 0, 0
                            _rec_i = _rc7.number_input("Qty", min_value=0, value=_rem_i, max_value=_rem_i, step=1,
                                                        key=f"{key_prefix}_reci_{_bid_li}_{_li_idx}",
                                                        label_visibility="collapsed")

                        _recd_qtys[_bid_li] = {"r": _rec_r, "l": _rec_l, "i": _rec_i,
                                               "ord_r": _ord_r, "ord_l": _ord_l, "ord_i": _ord_i}

                        # Check if partial — compare total received so far + this invoice vs ordered
                        _total_r_after = _prev["r"] + _rec_r
                        _total_l_after = _prev["l"] + _rec_l
                        _total_i_after = _prev["i"] + _rec_i
                        if _total_r_after < _ord_r or _total_l_after < _ord_l or _total_i_after < _ord_i:
                            _is_partial = True

                    # Show partial warning
                    if _is_partial:
                        st.warning("⚠ Received quantities are less than ordered — PO will be marked **PARTIALLY_RECEIVED**. "
                                   "You can record another invoice for the remaining balance later.")
                    else:
                        st.success("✅ All quantities match ordered — PO will be marked **RECEIVED**.")
                else:
                    st.info("PO has no line items with blank_id — stock update will be skipped. "
                            "Invoice record will still be saved.")

                # ── Save button ───────────────────────────────────────────
                if st.button("💾 Record Invoice & Update Stock", key=f"{key_prefix}_inv_save",
                             type="primary", use_container_width=True,
                             disabled=not _inv_no.strip()):
                    import uuid as _u_inv, base64 as _b64_inv
                    _fb64 = ""
                    if _inv_file:
                        _fb64 = _b64_inv.b64encode(_inv_file.read()).decode()

                    _tx_steps = []
                    _invoice_id = str(_u_inv.uuid4())
                    if _actual_supplier_id and _actual_supplier_id != _po_supplier_id:
                        _old_sup_name = _supplier_by_id_inv.get(_po_supplier_id, _po_lines_data[0].get("supplier") if _po_lines_data else "—")
                        _new_sup_name = _supplier_by_id_inv.get(_actual_supplier_id, "—")
                        _tx_steps.append(("""
                            UPDATE blank_replenishment_orders
                            SET supplier_id = %(sid)s::uuid,
                                remarks = TRIM(BOTH ' |' FROM COALESCE(remarks,'') || ' | ' || %(note)s),
                                updated_at = NOW()
                            WHERE id = %(po)s::uuid
                        """, {
                            "sid": _actual_supplier_id,
                            "po": _sel_po,
                            "note": f"Supplier corrected at invoice upload {_old_sup_name} -> {_new_sup_name}",
                        }))

                    # 1. Save invoice record
                    _tx_steps.append(("""
                        INSERT INTO blank_replenishment_invoices
                            (id, po_id, invoice_no, invoice_date,
                             invoice_amount, gst_amount, file_b64, notes, created_at)
                        VALUES
                            (%(id)s::uuid, %(po)s::uuid, %(ino)s, %(idate)s,
                             %(iamt)s, %(igst)s, %(fb64)s, %(note)s, NOW())
                    """, {
                        "id":    _invoice_id,
                        "po":    _sel_po,
                        "ino":   _inv_no.strip(),
                        "idate": str(_inv_date),
                        "iamt":  _inv_amt,
                        "igst":  _inv_gst,
                        "fb64":  _fb64,
                        "note":  _inv_note or "",
                    }))

                    # 2. Update PO status — PARTIALLY_RECEIVED or RECEIVED
                    _new_po_status = "PARTIALLY_RECEIVED" if _is_partial else "RECEIVED"
                    _tx_steps.append(("""
                        UPDATE blank_replenishment_orders
                        SET status = %(sts)s, updated_at = NOW()
                        WHERE id = %(po)s::uuid
                    """, {"po": _sel_po, "sts": _new_po_status}))

                    # 3. Update blank_inventory and 4. write stock ledger.
                    _stk_updated = 0
                    for _bid_inv, _qdata in _recd_qtys.items():
                        _r_add = int(_qdata["r"] or 0)
                        _l_add = int(_qdata["l"] or 0)
                        _i_add = int(_qdata["i"] or 0)
                        _upd = []
                        _upd_params = {"bid": _bid_inv}
                        if _r_add > 0:
                            _upd.append("qty_right = COALESCE(qty_right,0) + %(r_add)s")
                            _upd_params["r_add"] = _r_add
                        if _l_add > 0:
                            _upd.append("qty_left = COALESCE(qty_left,0) + %(l_add)s")
                            _upd_params["l_add"] = _l_add
                        if _i_add > 0:
                            _upd.append("qty_independent = COALESCE(qty_independent,0) + %(i_add)s")
                            _upd_params["i_add"] = _i_add
                        if _upd:
                            _stk_updated += 1
                            _tx_steps.append((
                                f"UPDATE blank_inventory SET {', '.join(_upd)}, "
                                f"updated_at = NOW() WHERE id = %(bid)s::uuid",
                                _upd_params,
                            ))

                        for _eye, _qty in (("R", _r_add), ("L", _l_add), ("I", _i_add)):
                            if _qty <= 0:
                                continue
                            _tx_steps.append(("""
                                INSERT INTO blank_stock_ledger
                                    (blank_id, eye_side, qty_change, ref_type,
                                     ref_id, remarks, created_at, created_by)
                                VALUES
                                    (%(bid)s::uuid, %(eye)s, %(qty)s, 'INVOICE_RECEIPT',
                                     %(po)s::uuid, %(rmk)s, NOW(), %(by)s)
                            """, {
                                "bid": _bid_inv,
                                "eye": _eye,
                                "qty": _qty,
                                "po":  _sel_po,
                                "rmk": f"Received via invoice {_inv_no.strip()}",
                                "by":  st.session_state.get("user_name", "system"),
                            }))

                    try:
                        from modules.sql_adapter import run_transaction
                        _iok = bool(run_transaction(_tx_steps))
                    except Exception as _txe:
                        st.error(f"Invoice/stock update was rolled back: {_txe}")
                        _iok = False

                    if _iok:

                        # 5. Success message
                        _partial_note = " (Partial — remaining balance still open)" if _is_partial else ""
                        st.success(
                            f"✅ Invoice **{_inv_no}** recorded. "
                            f"PO → **{_new_po_status}**{_partial_note}. "
                            f"Stock updated for **{_stk_updated}** blank SKU(s)."
                        )
                        st.rerun()
                    else:
                        st.error("Invoice could not be saved. Check DB logs.")

    # ══════════════════════════════════════════════════════════════════
    # TAB 5 — USAGE REPORT  (product × blank × operator)
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "📋 Usage Report":
        with _tab_report:
            st.caption(
                "Which blanks were issued for which product, and by which operator. "
                "Grouped by product — shows blank SKU, R/L counts, and the operator who processed."
            )
            with st.container(border=True):
                _ur1, _ur2, _ur3 = st.columns([2, 2, 1])
                _today_ur = _dt.date.today()
                _ur_from  = _ur1.date_input("From", value=_today_ur - _dt.timedelta(days=30),
                                             key=f"{key_prefix}_ur_from", format="DD/MM/YYYY")
                _ur_to    = _ur2.date_input("To",   value=_today_ur,
                                             key=f"{key_prefix}_ur_to",   format="DD/MM/YYYY")
                _ur_group = _ur3.selectbox("Group by", ["Product", "Operator", "Blank SKU"],
                                            key=f"{key_prefix}_ur_group")

            if _ur_from > _ur_to:
                st.error("From date is after To date.")
            else:
                # ── Main query: blank_allocations → blank_inventory → order_lines → orders
                _ur_rows = _brq("""
                    SELECT
                        COALESCE(bi.brand,'')                       AS brand,
                        COALESCE(bi.category,'')                    AS category,
                        COALESCE(bi.material,'')                    AS material,
                        UPPER(COALESCE(ba.eye_side,''))             AS eye_side,
                        COALESCE(ba.allocated_by, '')              AS operator,
                        ol.lens_params->>'product_name'             AS product_name,
                        ol.lens_params->>'product_code'             AS product_code,
                        o.order_no,
                        ba.allocated_at::date                       AS alloc_date,
                        COUNT(ba.id)                                AS qty
                    FROM blank_allocations ba
                    JOIN blank_inventory bi  ON bi.id  = ba.blank_id
                    JOIN order_lines ol      ON ol.id  = ba.order_line_id
                    JOIN orders o            ON o.id   = ol.order_id
                    WHERE COALESCE(ba.allocated_at, bi.created_at)::date
                          BETWEEN %(df)s AND %(dt)s
                      AND COALESCE(ol.is_deleted, FALSE) = FALSE
                    GROUP BY bi.brand, bi.category, bi.material,
                             ba.eye_side, ba.allocated_by,
                             ol.lens_params->>'product_name',
                             ol.lens_params->>'product_code',
                             o.order_no,
                             ba.allocated_at::date
                    ORDER BY alloc_date DESC, brand, category
                    LIMIT 2000
                """, {"df": str(_ur_from), "dt": str(_ur_to)})

                # ── Rejection rows in same period (for combined view)
                _ur_rej = _brq("""
                    SELECT
                        COALESCE(bi.brand,'')       AS brand,
                        COALESCE(bi.category,'')    AS category,
                        COALESCE(bi.material,'')    AS material,
                        UPPER(COALESCE(rb.eye_side,'')) AS eye_side,
                        COALESCE(rb.rejected_by,'') AS operator,
                        ol.lens_params->>'product_name' AS product_name,
                        ol.lens_params->>'product_code' AS product_code,
                        o.order_no,
                        rb.rejected_at::date        AS alloc_date,
                        rb.qty,
                        rb.reason
                    FROM production_rejection_bin rb
                    LEFT JOIN blank_inventory bi ON bi.id  = rb.blank_id
                    LEFT JOIN order_lines ol     ON ol.id  = rb.order_line_id
                    LEFT JOIN orders o           ON o.id   = rb.order_id
                    WHERE rb.blank_id IS NOT NULL
                      AND rb.rejected_at::date BETWEEN %(df)s AND %(dt)s
                    ORDER BY rb.rejected_at DESC
                    LIMIT 500
                """, {"df": str(_ur_from), "dt": str(_ur_to)})

                if not _ur_rows and not _ur_rej:
                    st.info("No blank usage in this period.")
                else:
                    # ── Summary metrics
                    _tot_alloc = sum(int(r["qty"]) for r in _ur_rows)
                    _tot_rej   = sum(int(r["qty"]) for r in _ur_rej)
                    _ops       = set(r["operator"] for r in _ur_rows if r["operator"])
                    _sm1, _sm2, _sm3, _sm4 = st.columns(4)
                    _sm1.metric("Total blanks issued", _tot_alloc)
                    _sm2.metric("Rejected blanks",     _tot_rej)
                    _sm3.metric("Unique operators",    len(_ops))
                    _sm4.metric("Unique products",
                                len(set((r["product_name"] or r["product_code"] or "—")
                                        for r in _ur_rows)))

                    st.markdown("---")

                    if _ur_group == "Product":
                        # ── Group by product, then blank SKU, then operator
                        _by_prod: dict = {}
                        for r in _ur_rows:
                            _pname = r["product_name"] or r["product_code"] or "Unknown Product"
                            _by_prod.setdefault(_pname, []).append(r)

                        for _pname in sorted(_by_prod.keys()):
                            _prows = _by_prod[_pname]
                            _ptot  = sum(int(r["qty"]) for r in _prows)

                            with st.expander(
                                f"**{_pname}** · {_ptot} blanks issued",
                                expanded=(len(_by_prod) == 1),
                            ):
                                # Aggregate by blank SKU + eye_side
                                _agg: dict = {}
                                for r in _prows:
                                    _bkey = " · ".join(filter(None, [r["brand"], r["category"], r["material"]]))
                                    _eye  = r["eye_side"] or "—"
                                    _op   = r["operator"] or "?"
                                    _k    = (_bkey, _eye)
                                    if _k not in _agg:
                                        _agg[_k] = {"qty": 0, "operators": set()}
                                    _agg[_k]["qty"]       += int(r["qty"])
                                    _agg[_k]["operators"].add(_op)

                                # Header row
                                st.markdown(
                                    "<div style='display:grid;"
                                    "grid-template-columns:3fr 0.6fr 0.8fr 2fr;"
                                    "gap:8px;padding:4px 8px;"
                                    "font-size:0.7rem;color:#94a3b8;font-weight:700;"
                                    "text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                                    "<span>Blank SKU</span><span>Eye</span>"
                                    "<span>Qty</span><span>Operator(s)</span>"
                                    "</div>", unsafe_allow_html=True)

                                for (_bkey, _eye), _adat in sorted(_agg.items()):
                                    _ops_str = ", ".join(sorted(_adat["operators"]))
                                    _eye_color = {"R": "#3b82f6", "L": "#a855f7"}.get(_eye, "#64748b")
                                    st.markdown(
                                        f"<div style='display:grid;"
                                        f"grid-template-columns:3fr 0.6fr 0.8fr 2fr;"
                                        f"gap:8px;padding:5px 8px;"
                                        f"font-size:0.8rem;border-bottom:1px solid #0f172a'>"
                                        f"<span style='color:#e2e8f0'>{_bkey or '—'}</span>"
                                        f"<span style='color:{_eye_color};font-weight:700'>{_eye}</span>"
                                        f"<span style='color:#fbbf24;font-weight:700'>{_adat['qty']}</span>"
                                        f"<span style='color:#94a3b8;font-size:0.75rem'>{_ops_str}</span>"
                                        f"</div>", unsafe_allow_html=True)

                                # Rejections for this product
                                _pr = [r for r in _ur_rej
                                       if (r["product_name"] or r["product_code"] or "Unknown Product") == _pname]
                                if _pr:
                                    st.markdown(
                                        f"<div style='margin-top:6px;padding:4px 8px;"
                                        f"background:#1a0a0a;border-left:3px solid #ef4444;"
                                        f"border-radius:4px;font-size:0.75rem;color:#fca5a5'>"
                                        f"🗑 {sum(int(r['qty']) for r in _pr)} blank(s) rejected for this product"
                                        f"</div>", unsafe_allow_html=True)

                    elif _ur_group == "Operator":
                        _by_op: dict = {}
                        for r in _ur_rows:
                            _op = r["operator"] or "Unknown"
                            _by_op.setdefault(_op, []).append(r)

                        for _op in sorted(_by_op.keys()):
                            _orows = _by_op[_op]
                            _otot  = sum(int(r["qty"]) for r in _orows)
                            with st.expander(f"👤 **{_op}** · {_otot} blanks processed", expanded=False):
                                # Aggregate: product → blank → R/L count
                                _pagg: dict = {}
                                for r in _orows:
                                    _pname = r["product_name"] or r["product_code"] or "Unknown"
                                    _bkey  = " · ".join(filter(None, [r["brand"], r["category"], r["material"]]))
                                    _eye   = r["eye_side"] or "—"
                                    _pk    = (_pname, _bkey, _eye)
                                    _pagg[_pk] = _pagg.get(_pk, 0) + int(r["qty"])

                                st.markdown(
                                    "<div style='display:grid;"
                                    "grid-template-columns:2.5fr 2fr 0.6fr 0.7fr;"
                                    "gap:8px;padding:4px 8px;"
                                    "font-size:0.7rem;color:#94a3b8;font-weight:700;"
                                    "text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                                    "<span>Product</span><span>Blank SKU</span>"
                                    "<span>Eye</span><span>Qty</span>"
                                    "</div>", unsafe_allow_html=True)

                                for (_pname, _bkey, _eye), _qty in sorted(_pagg.items()):
                                    _eye_color = {"R": "#3b82f6", "L": "#a855f7"}.get(_eye, "#64748b")
                                    st.markdown(
                                        f"<div style='display:grid;"
                                        f"grid-template-columns:2.5fr 2fr 0.6fr 0.7fr;"
                                        f"gap:8px;padding:5px 8px;"
                                        f"font-size:0.8rem;border-bottom:1px solid #0f172a'>"
                                        f"<span style='color:#e2e8f0'>{_pname}</span>"
                                        f"<span style='color:#94a3b8'>{_bkey or '—'}</span>"
                                        f"<span style='color:{_eye_color};font-weight:700'>{_eye}</span>"
                                        f"<span style='color:#fbbf24;font-weight:700'>{_qty}</span>"
                                        f"</div>", unsafe_allow_html=True)

                    else:  # Group by Blank SKU
                        _by_blank: dict = {}
                        for r in _ur_rows:
                            _bkey = " · ".join(filter(None, [r["brand"], r["category"], r["material"]]))
                            _by_blank.setdefault(_bkey or "Unknown", []).append(r)

                        for _bkey in sorted(_by_blank.keys()):
                            _brows = _by_blank[_bkey]
                            _r_tot = sum(int(r["qty"]) for r in _brows if r["eye_side"] == "R")
                            _l_tot = sum(int(r["qty"]) for r in _brows if r["eye_side"] == "L")
                            _i_tot = sum(int(r["qty"]) for r in _brows
                                         if r["eye_side"] not in ("R", "L"))
                            _btot  = sum(int(r["qty"]) for r in _brows)

                            _eye_summary = (
                                (f"R: {_r_tot}  " if _r_tot else "") +
                                (f"L: {_l_tot}  " if _l_tot else "") +
                                (f"Ind: {_i_tot}" if _i_tot else "")
                            ).strip()

                            with st.expander(
                                f"**{_bkey}** · {_btot} issued · {_eye_summary}",
                                expanded=False,
                            ):
                                _pagg2: dict = {}
                                for r in _brows:
                                    _pn  = r["product_name"] or r["product_code"] or "Unknown"
                                    _eye = r["eye_side"] or "—"
                                    _op  = r["operator"] or "?"
                                    _k2  = (_pn, _eye, _op)
                                    _pagg2[_k2] = _pagg2.get(_k2, 0) + int(r["qty"])

                                st.markdown(
                                    "<div style='display:grid;"
                                    "grid-template-columns:3fr 0.6fr 0.7fr 2fr;"
                                    "gap:8px;padding:4px 8px;"
                                    "font-size:0.7rem;color:#94a3b8;font-weight:700;"
                                    "text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                                    "<span>Product</span><span>Eye</span>"
                                    "<span>Qty</span><span>Operator</span>"
                                    "</div>", unsafe_allow_html=True)

                                for (_pn, _eye, _op), _qty in sorted(_pagg2.items()):
                                    _eye_color = {"R": "#3b82f6", "L": "#a855f7"}.get(_eye, "#64748b")
                                    st.markdown(
                                        f"<div style='display:grid;"
                                        f"grid-template-columns:3fr 0.6fr 0.7fr 2fr;"
                                        f"gap:8px;padding:5px 8px;"
                                        f"font-size:0.8rem;border-bottom:1px solid #0f172a'>"
                                        f"<span style='color:#e2e8f0'>{_pn}</span>"
                                        f"<span style='color:{_eye_color};font-weight:700'>{_eye}</span>"
                                        f"<span style='color:#fbbf24;font-weight:700'>{_qty}</span>"
                                        f"<span style='color:#94a3b8;font-size:0.75rem'>{_op}</span>"
                                        f"</div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    # TAB 6 — REPLENISHMENT vs PROCUREMENT REPORT
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "📈 Repl. vs Procured":
        with _tab_repl_report:
            st.caption(
                "Consumed vs Ordered vs Received vs Balance — per blank SKU. "
                "Shows the full replenishment loop: what was used, what was ordered to cover it, "
                "what arrived, and what is still outstanding."
            )
            with st.container(border=True):
                _rr1, _rr2, _rr3, _rr4 = st.columns([2, 2, 2, 1])
                _today_rr = _dt.date.today()
                _rr_from  = _rr1.date_input("Consumed from", value=_today_rr - _dt.timedelta(days=90),
                                             key=f"{key_prefix}_rr_from", format="DD/MM/YYYY")
                _rr_to    = _rr2.date_input("Consumed to",   value=_today_rr,
                                             key=f"{key_prefix}_rr_to",   format="DD/MM/YYYY")
                _rr_sup   = _rr3.text_input("Supplier", placeholder="All suppliers",
                                             key=f"{key_prefix}_rr_sup",
                                             label_visibility="collapsed")
                _rr_brand = _rr4.text_input("Brand",    placeholder="All",
                                             key=f"{key_prefix}_rr_brand",
                                             label_visibility="collapsed")

            import json as _rrj

            # ── Step 1: Consumption per blank in period ────────────────────
            _cons_rows = _brq("""
                SELECT
                    bi.id::text                         AS blank_id,
                    COALESCE(bi.brand,'')               AS brand,
                    COALESCE(bi.category,'')            AS category,
                    COALESCE(bi.material,'')            AS material,
                    COALESCE(bi.add_power,0)::float     AS add_power,
                    COALESCE(bi.base_recommended,0)::float AS base_recommended,
                    COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='R' THEN 1 ELSE 0 END),0) AS consumed_r,
                    COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='L' THEN 1 ELSE 0 END),0) AS consumed_l,
                    COALESCE(SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,'')) NOT IN ('R','L') THEN 1 ELSE 0 END),0) AS consumed_sv,
                    COUNT(ba.id) AS total_consumed
                FROM blank_inventory bi
                LEFT JOIN blank_allocations ba
                       ON ba.blank_id = bi.id
                      AND COALESCE(ba.allocated_at,bi.created_at)::date BETWEEN %(df)s AND %(dt)s
                WHERE COALESCE(bi.is_active,TRUE) = TRUE
                  AND (%(brand)s = '' OR LOWER(COALESCE(bi.brand,'')) LIKE %(brand_like)s)
                GROUP BY bi.id,bi.brand,bi.category,bi.material,bi.add_power,bi.base_recommended
                HAVING COUNT(ba.id) > 0
                ORDER BY total_consumed DESC, bi.brand, bi.category
                LIMIT 200
            """, {
                "df":         str(_rr_from),
                "dt":         str(_rr_to),
                "brand":      (_rr_brand or "").strip(),
                "brand_like": f"%{(_rr_brand or '').lower().strip()}%",
            })

            # ── Step 2: All POs with line_items (no date restriction — covers all POs) ──
            _po_rows_rr = _brq("""
                SELECT
                    bro.id::text                        AS po_id,
                    COALESCE(p.party_name,'Unknown')    AS supplier,
                    bro.order_date::date                AS order_date,
                    bro.status,
                    bro.line_items,
                    COALESCE(inv_agg.total_invoiced,0)  AS total_invoiced,
                    COALESCE(inv_agg.inv_count,0)       AS inv_count,
                    COALESCE(inv_agg.total_gst,0)       AS total_gst,
                    inv_agg.last_invoice_date
                FROM blank_replenishment_orders bro
                LEFT JOIN parties p ON p.id = bro.supplier_id
                LEFT JOIN (
                    SELECT po_id,
                           COUNT(*)             AS inv_count,
                           SUM(invoice_amount)  AS total_invoiced,
                           SUM(gst_amount)      AS total_gst,
                           MAX(invoice_date)    AS last_invoice_date
                    FROM blank_replenishment_invoices
                    GROUP BY po_id
                ) inv_agg ON inv_agg.po_id = bro.id
                WHERE (%(sup)s = '' OR LOWER(COALESCE(p.party_name,'')) LIKE %(sup_like)s)
                ORDER BY bro.order_date DESC
                LIMIT 500
            """, {
                "sup":      (_rr_sup or "").strip(),
                "sup_like": f"%{(_rr_sup or '').lower().strip()}%",
            })

            _receipt_rows_rr = _brq("""
                SELECT
                    bsl.ref_id::text                 AS po_id,
                    bsl.blank_id::text               AS blank_id,
                    UPPER(COALESCE(bsl.eye_side,'')) AS eye_side,
                    COALESCE(SUM(bsl.qty_change),0)  AS received_qty
                FROM blank_stock_ledger bsl
                JOIN blank_replenishment_orders bro ON bro.id = bsl.ref_id
                LEFT JOIN parties p ON p.id = bro.supplier_id
                WHERE bsl.ref_type = 'INVOICE_RECEIPT'
                  AND (%(sup)s = '' OR LOWER(COALESCE(p.party_name,'')) LIKE %(sup_like)s)
                GROUP BY bsl.ref_id, bsl.blank_id, UPPER(COALESCE(bsl.eye_side,''))
            """, {
                "sup":      (_rr_sup or "").strip(),
                "sup_like": f"%{(_rr_sup or '').lower().strip()}%",
            })

            # ── Step 3: Build blank_id → {ordered_r, ordered_l, ordered_sv,
            #                               received_r, received_l, received_sv,
            #                               supplier, po_date, status}
            _blank_po: dict = {}   # blank_id → aggregated PO data
            for _rr in _po_rows_rr:
                _li = _rr.get("line_items") or []
                if isinstance(_li, str):
                    try: _li = _rrj.loads(_li)
                    except: _li = []
                for _item in _li:
                    _bid_po = str(_item.get("blank_id") or "")
                    if not _bid_po:
                        continue
                    if _bid_po not in _blank_po:
                        _blank_po[_bid_po] = {
                            "ordered_r": 0, "ordered_l": 0, "ordered_sv": 0,
                            "received_r": 0, "received_l": 0, "received_sv": 0,
                            "suppliers": set(), "open_pos": 0, "last_po": "",
                            "invoiced": 0.0,
                        }
                    _bp = _blank_po[_bid_po]
                    _rq = int(_item.get("r_qty") or 0)
                    _lq = int(_item.get("l_qty") or 0)
                    _iq = int(_item.get("i_qty") or 0)
                    _bp["ordered_r"]  += _rq
                    _bp["ordered_l"]  += _lq
                    _bp["ordered_sv"] += _iq
                    if _rr["status"] in ("SENT","PARTIALLY_RECEIVED"):
                        _bp["open_pos"] += 1
                    _bp["suppliers"].add(_rr["supplier"])
                    _bp["invoiced"] += float(_rr.get("total_invoiced") or 0) / max(1, len(_li))
                    if str(_rr["order_date"] or "") > _bp["last_po"]:
                        _bp["last_po"] = str(_rr["order_date"] or "")

            _po_meta_by_id = {str(_rr.get("po_id") or ""): _rr for _rr in _po_rows_rr}
            for _rec in _receipt_rows_rr:
                _bid_rec = str(_rec.get("blank_id") or "")
                if not _bid_rec:
                    continue
                if _bid_rec not in _blank_po:
                    _blank_po[_bid_rec] = {
                        "ordered_r": 0, "ordered_l": 0, "ordered_sv": 0,
                        "received_r": 0, "received_l": 0, "received_sv": 0,
                        "suppliers": set(), "open_pos": 0, "last_po": "",
                        "invoiced": 0.0,
                    }
                _bp = _blank_po[_bid_rec]
                _qty_rec = int(_rec.get("received_qty") or 0)
                _eye_rec = str(_rec.get("eye_side") or "").upper()
                if _eye_rec == "R":
                    _bp["received_r"] += _qty_rec
                elif _eye_rec == "L":
                    _bp["received_l"] += _qty_rec
                else:
                    _bp["received_sv"] += _qty_rec
                _po_meta = _po_meta_by_id.get(str(_rec.get("po_id") or ""), {})
                if _po_meta.get("supplier"):
                    _bp["suppliers"].add(_po_meta["supplier"])

            if not _cons_rows:
                st.info("No blank consumption in this period.")
            else:
                # ── Summary metrics ───────────────────────────────────────
                _tot_consumed  = sum(int(r["total_consumed"]) for r in _cons_rows)
                _tot_ord_pairs = sum(
                    min(_blank_po.get(r["blank_id"],{}).get("ordered_r",0),
                        _blank_po.get(r["blank_id"],{}).get("ordered_l",0))
                    if (_blank_po.get(r["blank_id"],{}).get("ordered_r",0) +
                        _blank_po.get(r["blank_id"],{}).get("ordered_l",0)) > 0
                    else _blank_po.get(r["blank_id"],{}).get("ordered_sv",0)
                    for r in _cons_rows
                )
                _tot_recv_pairs = sum(
                    min(_blank_po.get(r["blank_id"],{}).get("received_r",0),
                        _blank_po.get(r["blank_id"],{}).get("received_l",0))
                    if (_blank_po.get(r["blank_id"],{}).get("received_r",0) +
                        _blank_po.get(r["blank_id"],{}).get("received_l",0)) > 0
                    else _blank_po.get(r["blank_id"],{}).get("received_sv",0)
                    for r in _cons_rows
                )
                _unordered = sum(
                    1 for r in _cons_rows
                    if r["blank_id"] not in _blank_po
                )

                _sm1,_sm2,_sm3,_sm4,_sm5,_sm6 = st.columns(6)
                _sm1.metric("SKUs consumed",   len(_cons_rows))
                _sm2.metric("Total consumed",  _tot_consumed)
                _sm3.metric("Pairs ordered",   _tot_ord_pairs)
                _sm4.metric("Pairs received",  _tot_recv_pairs)
                _sm5.metric("Balance (open)",  _tot_ord_pairs - _tot_recv_pairs,
                            delta_color="inverse" if (_tot_ord_pairs - _tot_recv_pairs) > 0 else "off")
                _sm6.metric("SKUs not ordered yet", _unordered,
                            delta_color="inverse" if _unordered > 0 else "off")

                st.markdown("---")

                # ── Per-blank comparison table ────────────────────────────
                st.markdown("**Per blank: Consumed → Ordered → Received → Balance**")
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:2.5fr 0.8fr 0.8fr 0.9fr 0.9fr 0.9fr 0.9fr 1fr 1.5fr;"
                    "gap:6px;padding:4px 8px;font-size:0.68rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Blank</span>"
                    "<span>Base</span><span>Add</span>"
                    "<span>Used R</span><span>Used L</span>"
                    "<span>Ordered</span><span>Received</span>"
                    "<span>Balance</span><span>Supplier / Status</span>"
                    "</div>", unsafe_allow_html=True)

                for r in _cons_rows:
                    _bid   = r["blank_id"]
                    _bname = " · ".join(filter(None,[r["brand"],r["category"],r["material"]]))
                    _bv    = float(r.get("base_recommended") or 0)
                    _av    = float(r.get("add_power") or 0)
                    _cr    = int(r["consumed_r"])
                    _cl    = int(r["consumed_l"])
                    _csv   = int(r["consumed_sv"])
                    _isp   = _av > 0.01 or "prog" in r.get("category","").lower() or "bifocal" in r.get("category","").lower()

                    _bp    = _blank_po.get(_bid, {})
                    _ord_r = _bp.get("ordered_r",  0)
                    _ord_l = _bp.get("ordered_l",  0)
                    _ord_sv= _bp.get("ordered_sv", 0)
                    _rec_r = _bp.get("received_r", 0)
                    _rec_l = _bp.get("received_l", 0)
                    _rec_sv= _bp.get("received_sv",0)

                    if _isp:
                        _ord_pairs = min(_ord_r, _ord_l) + max(0, _ord_r-min(_ord_r,_ord_l)) + max(0, _ord_l-min(_ord_r,_ord_l))
                        _rec_pairs = min(_rec_r, _rec_l)
                        _ord_disp  = f"{min(_ord_r,_ord_l)}pr" + (f"+{_ord_r-min(_ord_r,_ord_l)}R" if _ord_r > _ord_l else "") + (f"+{_ord_l-min(_ord_r,_ord_l)}L" if _ord_l > _ord_r else "")
                        _rec_disp  = f"{_rec_pairs}pr"
                        _bal       = _ord_r + _ord_l - _rec_r - _rec_l
                    else:
                        _ord_disp  = str(_ord_sv)
                        _rec_disp  = str(_rec_sv)
                        _bal       = _ord_sv - _rec_sv

                    _bal_color = "#ef4444" if _bal > 0 else ("#22c55e" if _bal == 0 and _bp else "#475569")
                    _not_ordered = not _bp
                    _sup_info   = ", ".join(sorted(_bp.get("suppliers", set()))) if _bp else "—"
                    _open_flag  = f" ({_bp.get('open_pos',0)} open)" if _bp.get("open_pos",0) > 0 else ""
                    if _not_ordered:
                        _status_str = "⚠ Not ordered"
                        _bal_color  = "#dc2626"
                        _ord_disp   = "—"
                        _rec_disp   = "—"
                        _bal        = 0

                    _row_bg = "background:#1a0a0a;" if _not_ordered else ""
                    _bv_str = f"{_bv:.0f}D" if _bv > 0 else "—"
                    _av_str = f"+{_av:.2f}" if _av > 0.01 else "—"
                    _bal_str = "⚠ Not ordered" if _not_ordered else (f"{_bal} pending" if _bal > 0 else "✅ Full")
                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:2.5fr 0.8fr 0.8fr 0.9fr 0.9fr 0.9fr 0.9fr 1fr 1.5fr;"
                        f"gap:6px;padding:5px 8px;font-size:0.78rem;"
                        f"border-bottom:1px solid #0f172a;{_row_bg}'>"
                        f"<span style='color:#e2e8f0'>{_bname}</span>"
                        f"<span style='color:#94a3b8'>{_bv_str}</span>"
                        f"<span style='color:#94a3b8'>{_av_str}</span>"
                        f"<span style='color:#3b82f6;font-weight:700'>{_cr if _isp else '—'}</span>"
                        f"<span style='color:#a855f7;font-weight:700'>{_cl if _isp else _csv}</span>"
                        f"<span style='color:#fbbf24'>{_ord_disp}</span>"
                        f"<span style='color:#22c55e'>{_rec_disp}</span>"
                        f"<span style='color:{_bal_color};font-weight:700'>{_bal_str}</span>"
                        f"<span style='color:#475569;font-size:0.72rem'>{_sup_info}{_open_flag}</span>"
                        f"</div>", unsafe_allow_html=True)

                st.markdown("---")

                # ── PO detail expanders ───────────────────────────────────
                with st.expander("📋 All POs in period (detail)", expanded=False):
                    for _rr in _po_rows_rr[:50]:
                        _sts   = _rr["status"]
                        _sts_c = {"SENT":"#f59e0b","PARTIALLY_RECEIVED":"#3b82f6",
                                  "RECEIVED":"#22c55e"}.get(_sts,"#64748b")
                        _li3   = _rr.get("line_items") or []
                        if isinstance(_li3, str):
                            try: _li3 = _rrj.loads(_li3)
                            except: _li3 = []
                        _po_pairs = 0
                        _po_lines = []
                        for _item in _li3:
                            _rq3=int(_item.get("r_qty") or 0)
                            _lq3=int(_item.get("l_qty") or 0)
                            _iq3=int(_item.get("i_qty") or 0)
                            _fp3=min(_rq3,_lq3) if (_rq3+_lq3)>0 else _iq3
                            _po_pairs += _fp3
                            _bl3=" · ".join(filter(None,[_item.get("brand"),_item.get("category"),_item.get("material")]))
                            _bv3=float(_item.get("base") or 0)
                            _av3=float(_item.get("add") or 0)
                            if _bv3>0: _bl3+=f" Base {_bv3:.0f}"
                            if _av3>0.01: _bl3+=f" Add +{_av3:.2f}"
                            _qty_s=f"R:{_rq3} L:{_lq3}" if (_rq3+_lq3)>0 else f"Qty {_iq3}"
                            _po_lines.append(f"{_bl3} — {_qty_s} ({_fp3} pairs)")
                        _inv_s=(f"  ·  ₹{float(_rr['total_invoiced'] or 0):,.0f} ({_rr['inv_count']} inv)"
                                if _rr["inv_count"] else "")
                        with st.expander(
                            f"**{_rr['supplier']}** · {_rr['order_date']} · "
                            f"{_po_pairs} pairs · {_sts}{_inv_s}",
                            expanded=False):
                            for _pl in _po_lines:
                                st.markdown(f"<div style='color:#e2e8f0;font-size:0.8rem;padding:2px 0'>• {_pl}</div>",
                                            unsafe_allow_html=True)
                            if _rr.get("remarks"):
                                st.caption(f"Note: {_rr['remarks']}")
                            if _rr.get("last_invoice_date"):
                                st.caption(f"Last invoice: {_rr['last_invoice_date']}")

    # ══════════════════════════════════════════════════════════════════
    # TAB 7 — PARTY-WISE BLANK CONSUMPTION AUDIT
    # Shows which party's orders used which blank, by which operator,
    # and what stock was available at the time of assignment.
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "🏪 Party Audit":
        with _tab_party:
            st.caption(
                "Party-wise audit: which blank was assigned to which party's orders, "
                "by which operator, and what R/L stock was available at that moment. "
                "Helps identify mismatches — e.g. GKB blank given when Aroline had 20 pcs in stock."
            )
            with st.container(border=True):
                _pa1, _pa2, _pa3, _pa4 = st.columns([2, 2, 2, 1])
                _today_pa = _dt.date.today()
                _pa_from  = _pa1.date_input("From", value=_today_pa - _dt.timedelta(days=30),
                                             key=f"{key_prefix}_pa_from", format="DD/MM/YYYY")
                _pa_to    = _pa2.date_input("To",   value=_today_pa,
                                             key=f"{key_prefix}_pa_to",   format="DD/MM/YYYY")
                _pa_party = _pa3.text_input("Party / Optician", placeholder="Filter by party name",
                                             key=f"{key_prefix}_pa_party",
                                             label_visibility="collapsed")
                _pa_blank = _pa4.text_input("Blank brand", placeholder="e.g. GKB",
                                             key=f"{key_prefix}_pa_blank",
                                             label_visibility="collapsed")

            # ── Main audit query ──────────────────────────────────────────
            # For each allocation: get the blank used, the order/party, operator,
            # and a snapshot of R/L stock on that blank AT TIME of allocation.
            # We reconstruct stock-at-time using blank_stock_ledger — sum all
            # ledger entries BEFORE the allocation timestamp gives stock in hand.
            _pa_rows = _brq("""
                SELECT
                    ba.id::text                             AS alloc_id,
                    ba.allocated_at::timestamptz            AS allocated_at,
                    COALESCE(ba.allocated_by, 'Unknown') AS operator,
                    UPPER(COALESCE(ba.eye_side,''))          AS eye_side,
                    COALESCE(bi.brand,'')                   AS blank_brand,
                    COALESCE(bi.category,'')                AS blank_category,
                    COALESCE(bi.material,'')                AS blank_material,
                    COALESCE(bi.add_power,0)::float         AS add_power,
                    COALESCE(bi.base_recommended,0)::float  AS base_recommended,
                    o.order_no,
                    COALESCE(o.party_name,'')               AS party_name,
                    COALESCE(o.patient_name,'')             AS patient_name,
                    ol.lens_params->>'product_name'         AS product_name,
                    -- Current stock (for reference)
                    COALESCE(bi.qty_right,0)    AS cur_r,
                    COALESCE(bi.qty_left,0)     AS cur_l,
                    -- Stock AT TIME of allocation:
                    -- current_stock + all outflows after allocation - all inflows after allocation
                    COALESCE(bi.qty_right,0) + COALESCE((
                        SELECT SUM(CASE
                            WHEN bsl.eye_side='R' AND bsl.qty_change < 0 THEN ABS(bsl.qty_change)
                            WHEN bsl.eye_side='R' AND bsl.qty_change > 0 THEN -bsl.qty_change
                            ELSE 0 END)
                        FROM blank_stock_ledger bsl
                        WHERE bsl.blank_id = bi.id
                          AND bsl.created_at > ba.allocated_at
                    ), 0)                                   AS stock_r_at_time,
                    COALESCE(bi.qty_left,0) + COALESCE((
                        SELECT SUM(CASE
                            WHEN bsl.eye_side='L' AND bsl.qty_change < 0 THEN ABS(bsl.qty_change)
                            WHEN bsl.eye_side='L' AND bsl.qty_change > 0 THEN -bsl.qty_change
                            ELSE 0 END)
                        FROM blank_stock_ledger bsl
                        WHERE bsl.blank_id = bi.id
                          AND bsl.created_at > ba.allocated_at
                    ), 0)                                   AS stock_l_at_time
                FROM blank_allocations ba
                JOIN blank_inventory bi ON bi.id  = ba.blank_id
                JOIN order_lines ol     ON ol.id  = ba.order_line_id
                JOIN orders o           ON o.id   = ol.order_id
                WHERE ba.allocated_at::date BETWEEN %(df)s AND %(dt)s
                  AND (%(party)s = '' OR LOWER(COALESCE(o.party_name,'')) LIKE %(party_like)s)
                  AND (%(blank)s = '' OR LOWER(COALESCE(bi.brand,'')) LIKE %(blank_like)s)
                ORDER BY o.party_name, ba.allocated_at DESC
                LIMIT 1000
            """, {
                "df":         str(_pa_from),
                "dt":         str(_pa_to),
                "party":      (_pa_party or "").strip(),
                "party_like": f"%{(_pa_party or '').lower().strip()}%",
                "blank":      (_pa_blank or "").strip(),
                "blank_like": f"%{(_pa_blank or '').lower().strip()}%",
            })

            if not _pa_rows:
                st.info("No assignments found for the selected filters and period.")
            else:
                # ── Summary ───────────────────────────────────────────────
                _pa_parties  = set(r["party_name"] for r in _pa_rows if r["party_name"])
                _pa_blanks   = set(r["blank_brand"] for r in _pa_rows if r["blank_brand"])
                _pa_ops      = set(r["operator"]    for r in _pa_rows if r["operator"])
                _pa_m1, _pa_m2, _pa_m3, _pa_m4 = st.columns(4)
                _pa_m1.metric("Assignments",     len(_pa_rows))
                _pa_m2.metric("Parties",         len(_pa_parties))
                _pa_m3.metric("Blank brands",    len(_pa_blanks))
                _pa_m4.metric("Operators",       len(_pa_ops))

                st.markdown("---")

                # ── Group by party ────────────────────────────────────────
                _by_party: dict = {}
                for r in _pa_rows:
                    _pn = r["party_name"] or "Unknown Party"
                    _by_party.setdefault(_pn, []).append(r)

                for _pname in sorted(_by_party.keys()):
                    _prows = _by_party[_pname]
                    _pblnks = set(r["blank_brand"] for r in _prows)
                    _pops   = set(r["operator"]    for r in _prows)
                    with st.expander(
                        f"**{_pname}** · {len(_prows)} assignment(s) · "
                        f"Blanks: {', '.join(sorted(_pblnks))} · "
                        f"Ops: {', '.join(sorted(_pops))}",
                        expanded=False,
                    ):
                        # Column headers
                        st.markdown(
                            "<div style='display:grid;"
                            "grid-template-columns:1.2fr 2fr 1.2fr 0.6fr 1fr 1fr 1fr 1.5fr;"
                            "gap:6px;padding:4px 8px;font-size:0.68rem;color:#94a3b8;"
                            "font-weight:700;text-transform:uppercase;"
                            "border-bottom:1px solid #1e293b'>"
                            "<span>Date/Time</span><span>Product</span>"
                            "<span>Blank Used</span><span>Eye</span>"
                            "<span>Base·Add</span>"
                            "<span>Stock at Time</span><span>Current Stock</span>"
                            "<span>Operator</span>"
                            "</div>", unsafe_allow_html=True)

                        for r in _prows:
                            _alloc_dt   = str(r["allocated_at"] or "")[:16]
                            _prod       = (r["product_name"] or r["patient_name"] or "—")[:28]
                            _blank_desc = f"{r['blank_brand']} {r['blank_category']}"
                            _eye        = r["eye_side"] or "—"
                            _eye_color  = {"R": "#3b82f6", "L": "#a855f7"}.get(_eye, "#64748b")
                            _base_v     = float(r["base_recommended"] or 0)
                            _add_v      = float(r["add_power"] or 0)
                            _base_add   = (f"{_base_v:.0f}" if _base_v > 0 else "—") + \
                                          (f" +{_add_v:.2f}" if _add_v > 0.01 else "")
                            _sr = int(r["stock_r_at_time"] or 0)
                            _sl = int(r["stock_l_at_time"] or 0)
                            _cr = int(r["cur_r"] or 0)
                            _cl = int(r["cur_l"] or 0)
                            # Flag if stock was available for OPPOSITE blank at time
                            # (detection needs a separate query — show raw stock for now)
                            _stk_color = "#22c55e" if (_sr > 0 and _sl > 0) \
                                         else ("#f59e0b" if (_sr + _sl) > 0 else "#ef4444")

                            st.markdown(
                                f"<div style='display:grid;"
                                f"grid-template-columns:1.2fr 2fr 1.2fr 0.6fr 1fr 1fr 1fr 1.5fr;"
                                f"gap:6px;padding:5px 8px;font-size:0.78rem;"
                                f"border-bottom:1px solid #0f172a'>"
                                f"<span style='color:#94a3b8;font-size:0.72rem'>{_alloc_dt}</span>"
                                f"<span style='color:#e2e8f0'>{_prod}</span>"
                                f"<span style='color:#fbbf24;font-weight:600'>{_blank_desc}</span>"
                                f"<span style='color:{_eye_color};font-weight:700'>{_eye}</span>"
                                f"<span style='color:#94a3b8;font-size:0.75rem'>{_base_add}</span>"
                                f"<span style='color:{_stk_color};font-size:0.75rem'>"
                                f"R:{_sr} L:{_sl}</span>"
                                f"<span style='color:#475569;font-size:0.72rem'>"
                                f"R:{_cr} L:{_cl}</span>"
                                f"<span style='color:#94a3b8;font-size:0.72rem'>{r['operator']}</span>"
                                f"</div>", unsafe_allow_html=True)

                        # ── Operator summary for this party ───────────────
                        _op_agg: dict = {}
                        for r in _prows:
                            _op = r["operator"] or "?"
                            _op_agg.setdefault(_op, {"count": 0, "blanks": set()})
                            _op_agg[_op]["count"] += 1
                            _op_agg[_op]["blanks"].add(r["blank_brand"])
                        if len(_op_agg) > 1 or (len(_op_agg) == 1 and len(list(_op_agg.values())[0]["blanks"]) > 1):
                            st.markdown(
                                "<div style='margin-top:8px;padding:6px 10px;"
                                "background:#0a1628;border-radius:4px;font-size:0.75rem;color:#94a3b8'>"
                                "Operator breakdown: " +
                                "  ·  ".join(
                                    f"<span style='color:#e2e8f0'>{op}</span>: "
                                    f"{d['count']} assign · "
                                    f"{', '.join(sorted(d['blanks']))}"
                                    for op, d in sorted(_op_agg.items())
                                ) + "</div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    # TAB 8 — PROCUREMENT DASHBOARD
    # Open POs · Top consumed · Supplier performance · Consolidated reports
    # ══════════════════════════════════════════════════════════════════
    if _repl_active_section == "🎯 Procurement Dashboard":
        with _tab_dash:
            import json as _dj
            import datetime as _ddt

            _today_d = _ddt.date.today()
            _dash_from = _today_d - _ddt.timedelta(days=90)

            # ── Refresh button ────────────────────────────────────────────
            _dc0, _dc1 = st.columns([8, 1])
            with _dc1:
                if st.button("🔄 Refresh", key=f"{key_prefix}_dash_refresh",
                             use_container_width=True):
                    st.rerun()

            # ── Load all PO data ─────────────────────────────────────────
            _dash_pos = _brq("""
                SELECT
                    bro.id::text                            AS po_id,
                    COALESCE(p.party_name,'Unknown')        AS supplier,
                    bro.order_date::date                    AS order_date,
                    bro.status,
                    bro.line_items,
                    COALESCE(inv_agg.inv_count,0)           AS inv_count,
                    COALESCE(inv_agg.total_invoiced,0)      AS total_invoiced,
                    inv_agg.last_invoice_date,
                    inv_agg.first_invoice_date
                FROM blank_replenishment_orders bro
                LEFT JOIN parties p ON p.id = bro.supplier_id
                LEFT JOIN (
                    SELECT
                        po_id,
                        COUNT(*)               AS inv_count,
                        SUM(invoice_amount)    AS total_invoiced,
                        MIN(invoice_date)      AS first_invoice_date,
                        MAX(invoice_date)      AS last_invoice_date
                    FROM blank_replenishment_invoices
                    GROUP BY po_id
                ) inv_agg ON inv_agg.po_id = bro.id
                ORDER BY bro.order_date DESC
                LIMIT 500
            """)

            _dash_receipts = _brq("""
                SELECT
                    bsl.ref_id::text                 AS po_id,
                    UPPER(COALESCE(bsl.eye_side,'')) AS eye_side,
                    COALESCE(SUM(bsl.qty_change),0)  AS received_qty
                FROM blank_stock_ledger bsl
                JOIN blank_replenishment_orders bro ON bro.id = bsl.ref_id
                WHERE bsl.ref_type = 'INVOICE_RECEIPT'
                  AND bro.order_date::date >= %(df)s
                GROUP BY bsl.ref_id, UPPER(COALESCE(bsl.eye_side,''))
            """, {"df": str(_dash_from)})
            _dash_rec_by_po: dict = {}
            for _dr in _dash_receipts:
                _pid = str(_dr.get("po_id") or "")
                _dash_rec_by_po.setdefault(_pid, {"r": 0, "l": 0, "i": 0})
                _eye = str(_dr.get("eye_side") or "").upper()
                _qty = int(_dr.get("received_qty") or 0)
                if _eye == "R":
                    _dash_rec_by_po[_pid]["r"] += _qty
                elif _eye == "L":
                    _dash_rec_by_po[_pid]["l"] += _qty
                else:
                    _dash_rec_by_po[_pid]["i"] += _qty

            # ── Load blank consumption (last 90 days) ─────────────────────
            _dash_consume = _brq("""
                SELECT
                    COALESCE(bi.brand,'')        AS brand,
                    COALESCE(bi.category,'')     AS category,
                    COALESCE(bi.material,'')     AS material,
                    COALESCE(bi.add_power,0)::float AS add_power,
                    COALESCE(bi.base_recommended,0)::float AS base_recommended,
                    COUNT(ba.id)                 AS total_used,
                    SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='R' THEN 1 ELSE 0 END) AS used_r,
                    SUM(CASE WHEN UPPER(COALESCE(ba.eye_side,''))='L' THEN 1 ELSE 0 END) AS used_l
                FROM blank_allocations ba
                JOIN blank_inventory bi ON bi.id = ba.blank_id
                WHERE COALESCE(ba.allocated_at, bi.created_at)::date
                      >= %(df)s
                GROUP BY bi.brand, bi.category, bi.material, bi.add_power, bi.base_recommended
                ORDER BY total_used DESC
                LIMIT 20
            """, {"df": str(_dash_from)})

            # ── Pre-compute supplier stats ─────────────────────────────────
            # sup → {ordered_pairs, received_pairs, open_pos, recv_pos,
            #         total_invoiced, delivery_days_list, fill_pct}
            _sup_stats: dict = {}
            for _dp in _dash_pos:
                _sn = _dp["supplier"]
                if _sn not in _sup_stats:
                    _sup_stats[_sn] = {
                        "ordered": 0, "received": 0,
                        "open_pos": 0, "recv_pos": 0,
                        "invoiced": 0.0, "delivery_days": [],
                    }
                _ss = _sup_stats[_sn]
                _li = _dp.get("line_items") or []
                if isinstance(_li, str):
                    try: _li = _dj.loads(_li)
                    except: _li = []
                for _item in _li:
                    _rq = int(_item.get("r_qty") or 0)
                    _lq = int(_item.get("l_qty") or 0)
                    _iq = int(_item.get("i_qty") or 0)
                    _fp = min(_rq, _lq) if (_rq + _lq) > 0 else _iq
                    _ss["ordered"] += _fp
                _rec_pack = _dash_rec_by_po.get(str(_dp.get("po_id") or ""), {"r": 0, "l": 0, "i": 0})
                _ss["received"] += min(_rec_pack["r"], _rec_pack["l"]) if (_rec_pack["r"] + _rec_pack["l"]) > 0 else _rec_pack["i"]
                if _dp["status"] in ("SENT", "PARTIALLY_RECEIVED"):
                    _ss["open_pos"] += 1
                elif _dp["status"] == "RECEIVED":
                    _ss["recv_pos"] += 1
                    # Delivery days = invoice date − order date
                    if _dp.get("first_invoice_date") and _dp.get("order_date"):
                        try:
                            _inv_d = _dp["first_invoice_date"]
                            _ord_d = _dp["order_date"]
                            if hasattr(_inv_d, 'days'):
                                pass
                            else:
                                import datetime as _ddt2
                                if isinstance(_inv_d, str):
                                    _inv_d = _ddt2.date.fromisoformat(str(_inv_d)[:10])
                                if isinstance(_ord_d, str):
                                    _ord_d = _ddt2.date.fromisoformat(str(_ord_d)[:10])
                                _days = (_inv_d - _ord_d).days
                                if 0 <= _days <= 365:
                                    _ss["delivery_days"].append(_days)
                        except Exception:
                            pass
                _ss["invoiced"] += float(_dp.get("total_invoiced") or 0)

            # ════════════════════════════════════════════════════
            # SECTION 1: OPEN PO SUMMARY
            # ════════════════════════════════════════════════════
            _open_all = [p for p in _dash_pos if p["status"] in ("SENT","PARTIALLY_RECEIVED")]
            _recv_all = [p for p in _dash_pos if p["status"] == "RECEIVED"]
            _total_open_pairs   = 0
            _total_recv_pairs   = 0
            _total_ordered_pairs = 0
            for _dp in _dash_pos:
                _li = _dp.get("line_items") or []
                if isinstance(_li, str):
                    try: _li = _dj.loads(_li)
                    except: _li = []
                _fp = sum(min(int(i.get("r_qty",0)), int(i.get("l_qty",0)))
                          if (int(i.get("r_qty",0))+int(i.get("l_qty",0))) > 0
                          else int(i.get("i_qty",0)) for i in _li)
                _total_ordered_pairs += _fp
                _rec_pack = _dash_rec_by_po.get(str(_dp.get("po_id") or ""), {"r": 0, "l": 0, "i": 0})
                _rec_fp = min(_rec_pack["r"], _rec_pack["l"]) if (_rec_pack["r"] + _rec_pack["l"]) > 0 else _rec_pack["i"]
                _total_recv_pairs += _rec_fp
                if _dp["status"] in ("SENT","PARTIALLY_RECEIVED"):
                    _total_open_pairs += max(0, _fp - _rec_fp)

            st.markdown(
                "<div style='background:#0f172a;border:1px solid #1e3a5f;"
                "border-left:4px solid #3b82f6;border-radius:8px;"
                "padding:10px 16px;margin-bottom:14px'>"
                "<span style='color:#93c5fd;font-size:1rem;font-weight:800'>"
                "📋 Open Procurement</span></div>",
                unsafe_allow_html=True)

            _d1,_d2,_d3,_d4,_d5 = st.columns(5)
            _d1.metric("Open POs",         len(_open_all))
            _d2.metric("Open pairs",        _total_open_pairs)
            _d3.metric("Received POs",      len(_recv_all))
            _d4.metric("Received pairs",    _total_recv_pairs)
            _d5.metric("Total value",       f"₹{sum(float(p.get('total_invoiced',0) or 0) for p in _recv_all):,.0f}")

            # Open PO table per supplier
            if _open_all:
                st.markdown("**Open POs by supplier**")
                # Header
                st.markdown(
                    "<div style='display:grid;grid-template-columns:2.5fr 1fr 1fr 1fr 1fr;"
                    "gap:8px;padding:4px 8px;font-size:0.7rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Supplier</span><span>Open POs</span>"
                    "<span>Pairs pending</span><span>Oldest PO</span><span>Status</span>"
                    "</div>", unsafe_allow_html=True)

                # Aggregate open POs per supplier
                _open_by_sup: dict = {}
                for _dp in _open_all:
                    _sn = _dp["supplier"]
                    _open_by_sup.setdefault(_sn, {"pos": 0, "pairs": 0, "oldest": str(_dp["order_date"] or "")})
                    _open_by_sup[_sn]["pos"] += 1
                    _ods = str(_dp["order_date"] or "")
                    if _ods < _open_by_sup[_sn]["oldest"]:
                        _open_by_sup[_sn]["oldest"] = _ods
                    _li = _dp.get("line_items") or []
                    if isinstance(_li, str):
                        try: _li = _dj.loads(_li)
                        except: _li = []
                    for _item in _li:
                        _rq = int(_item.get("r_qty") or 0)
                        _lq = int(_item.get("l_qty") or 0)
                        _iq = int(_item.get("i_qty") or 0)
                        _open_by_sup[_sn]["pairs"] += (
                            min(_rq, _lq) if (_rq+_lq) > 0 else _iq)

                for _sn, _sv in sorted(_open_by_sup.items(),
                                       key=lambda x: -x[1]["pairs"]):
                    # Age in days
                    try:
                        import datetime as _ddt3
                        _age = (_today_d - _ddt3.date.fromisoformat(_sv["oldest"])).days
                        _age_str = f"{_age}d ago"
                        _age_color = "#ef4444" if _age > 14 else ("#f59e0b" if _age > 7 else "#22c55e")
                    except Exception:
                        _age_str = _sv["oldest"]
                        _age_color = "#94a3b8"
                    st.markdown(
                        f"<div style='display:grid;grid-template-columns:2.5fr 1fr 1fr 1fr 1fr;"
                        f"gap:8px;padding:6px 8px;font-size:0.82rem;"
                        f"border-bottom:1px solid #0f172a'>"
                        f"<span style='color:#e2e8f0;font-weight:600'>{_sn}</span>"
                        f"<span style='color:#f59e0b'>{_sv['pos']}</span>"
                        f"<span style='color:#fbbf24;font-weight:700'>{_sv['pairs']} pr</span>"
                        f"<span style='color:{_age_color};font-size:0.75rem'>{_age_str}</span>"
                        f"<span style='color:#f59e0b;font-size:0.75rem'>OPEN</span>"
                        f"</div>", unsafe_allow_html=True)
            else:
                st.success("✅ No open POs — all blanks received.")

            st.markdown("---")

            # ════════════════════════════════════════════════════
            # SECTION 2: TOP CONSUMED BLANKS (last 90 days)
            # ════════════════════════════════════════════════════
            st.markdown(
                "<div style='background:#0f172a;border:1px solid #1e3a5f;"
                "border-left:4px solid #8b5cf6;border-radius:8px;"
                "padding:10px 16px;margin-bottom:14px'>"
                "<span style='color:#c4b5fd;font-size:1rem;font-weight:800'>"
                f"🔥 Top Consumed Blanks — last 90 days</span></div>",
                unsafe_allow_html=True)

            if _dash_consume:
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:3fr 0.8fr 0.8fr 0.7fr 0.7fr 0.7fr;"
                    "gap:8px;padding:4px 8px;font-size:0.7rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Blank</span><span>Base</span><span>Add</span>"
                    "<span>Total</span><span>R</span><span>L</span>"
                    "</div>", unsafe_allow_html=True)
                for _dc in _dash_consume:
                    _bname = " · ".join(filter(None, [_dc["brand"], _dc["category"], _dc["material"]]))
                    _base_v = float(_dc.get("base_recommended") or 0)
                    _add_v  = float(_dc.get("add_power") or 0)
                    _tot    = int(_dc["total_used"])
                    _bar_w  = min(100, int(_tot / max(1, int(_dash_consume[0]["total_used"])) * 100))
                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:3fr 0.8fr 0.8fr 0.7fr 0.7fr 0.7fr;"
                        f"gap:8px;padding:5px 8px;font-size:0.82rem;"
                        f"border-bottom:1px solid #0f172a'>"
                        f"<span style='color:#e2e8f0'>{_bname}"
                        f"<div style='height:3px;background:#7c3aed;width:{_bar_w}%;margin-top:3px;border-radius:2px'></div>"
                        f"</span>"
                        f"<span style='color:#94a3b8'>{_base_v:.0f}D' if _base_v > 0 else '—'</span>"
                        f"<span style='color:#94a3b8'>{'+'+ f'{_add_v:.2f}' if _add_v > 0.01 else '—'}</span>"
                        f"<span style='color:#fbbf24;font-weight:700'>{_tot}</span>"
                        f"<span style='color:#3b82f6'>{int(_dc['used_r'])}</span>"
                        f"<span style='color:#a855f7'>{int(_dc['used_l'])}</span>"
                        f"</div>", unsafe_allow_html=True)
            else:
                st.info("No blank consumption data in last 90 days.")

            st.markdown("---")

            # ════════════════════════════════════════════════════
            # SECTION 3: SUPPLIER PERFORMANCE
            # ════════════════════════════════════════════════════
            st.markdown(
                "<div style='background:#0f172a;border:1px solid #1e3a5f;"
                "border-left:4px solid #10b981;border-radius:8px;"
                "padding:10px 16px;margin-bottom:14px'>"
                "<span style='color:#6ee7b7;font-size:1rem;font-weight:800'>"
                "📊 Supplier Performance</span></div>",
                unsafe_allow_html=True)

            if _sup_stats:
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:2.5fr 1fr 1fr 1fr 1fr 1fr 1fr;"
                    "gap:8px;padding:4px 8px;font-size:0.7rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Supplier</span><span>Ordered</span><span>Received</span>"
                    "<span>Pending</span><span>Fill Rate</span>"
                    "<span>Avg Delivery</span><span>Invoiced ₹</span>"
                    "</div>", unsafe_allow_html=True)

                for _sn in sorted(_sup_stats.keys()):
                    _ss = _sup_stats[_sn]
                    _fill = (int(_ss["received"] / max(1, _ss["ordered"]) * 100))
                    _fill_color = "#22c55e" if _fill >= 95 else ("#f59e0b" if _fill >= 80 else "#ef4444")
                    _pending = _ss["ordered"] - _ss["received"]
                    _pend_color = "#ef4444" if _pending > 10 else ("#f59e0b" if _pending > 0 else "#22c55e")
                    _avg_days = (int(sum(_ss["delivery_days"]) / len(_ss["delivery_days"]))
                                 if _ss["delivery_days"] else None)
                    _days_str = f"{_avg_days}d" if _avg_days is not None else "—"
                    _days_color = ("#ef4444" if _avg_days and _avg_days > 10
                                   else "#f59e0b" if _avg_days and _avg_days > 5 else "#22c55e")

                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:2.5fr 1fr 1fr 1fr 1fr 1fr 1fr;"
                        f"gap:8px;padding:6px 8px;font-size:0.82rem;"
                        f"border-bottom:1px solid #0f172a'>"
                        f"<span style='color:#e2e8f0;font-weight:600'>{_sn}</span>"
                        f"<span style='color:#94a3b8'>{_ss['ordered']} pr</span>"
                        f"<span style='color:#22c55e'>{_ss['received']} pr</span>"
                        f"<span style='color:{_pend_color};font-weight:700'>{_pending} pr</span>"
                        f"<span style='color:{_fill_color};font-weight:700'>{_fill}%</span>"
                        f"<span style='color:{_days_color}'>{_days_str}</span>"
                        f"<span style='color:#94a3b8'>₹{_ss['invoiced']:,.0f}</span>"
                        f"</div>", unsafe_allow_html=True)
            else:
                st.info("No supplier data available.")

            st.markdown("---")

            # ════════════════════════════════════════════════════
            # SECTION 4: CONSOLIDATED BLANK PROCUREMENT REPORT
            # All PO line items flattened — filterable, downloadable
            # ════════════════════════════════════════════════════
            st.markdown(
                "<div style='background:#0f172a;border:1px solid #1e3a5f;"
                "border-left:4px solid #f59e0b;border-radius:8px;"
                "padding:10px 16px;margin-bottom:14px'>"
                "<span style='color:#fcd34d;font-size:1rem;font-weight:800'>"
                "📋 Consolidated Blank Procurement Report</span>"
                "<span style='color:#475569;font-size:0.78rem;margin-left:10px'>"
                "All PO lines · filterable · Excel download</span></div>",
                unsafe_allow_html=True)

            # Flatten all PO line_items
            _flat_lines = []
            for _dp in _dash_pos:
                _li = _dp.get("line_items") or []
                if isinstance(_li, str):
                    try: _li = _dj.loads(_li)
                    except: _li = []
                for _item in _li:
                    _rq = int(_item.get("r_qty") or 0)
                    _lq = int(_item.get("l_qty") or 0)
                    _iq = int(_item.get("i_qty") or 0)
                    _fp = min(_rq, _lq) if (_rq+_lq) > 0 else _iq
                    _bv = float(_item.get("base") or 0)
                    _av = float(_item.get("add") or 0)
                    _flat_lines.append({
                        "order_date":  str(_dp["order_date"] or ""),
                        "supplier":    _dp["supplier"],
                        "status":      _dp["status"],
                        "brand":       _item.get("brand",""),
                        "category":    _item.get("category",""),
                        "material":    _item.get("material",""),
                        "base":        f"{_bv:.0f}D" if _bv > 0 else "",
                        "add":         f"+{_av:.2f}" if _av > 0.01 else "",
                        "r_qty":       _rq,
                        "l_qty":       _lq,
                        "i_qty":       _iq,
                        "pairs":       _fp,
                        "invoiced":    float(_dp.get("total_invoiced") or 0),
                    })

            if _flat_lines:
                # Filter controls
                _cf1, _cf2 = st.columns(2)
                _cf_sup   = _cf1.text_input("Filter supplier",   key=f"{key_prefix}_cf_sup",
                                             label_visibility="collapsed",
                                             placeholder="All suppliers")
                _cf_blank = _cf2.text_input("Filter blank brand", key=f"{key_prefix}_cf_blank",
                                             label_visibility="collapsed",
                                             placeholder="All brands")
                _cf_status = st.multiselect("Status", ["SENT","PARTIALLY_RECEIVED","RECEIVED"],
                                             default=["SENT","PARTIALLY_RECEIVED","RECEIVED"],
                                             key=f"{key_prefix}_cf_status")
                _filtered = [r for r in _flat_lines
                             if (_cf_sup   == "" or _cf_sup.lower()   in r["supplier"].lower())
                             and (_cf_blank == "" or _cf_blank.lower() in r["brand"].lower())
                             and r["status"] in (_cf_status or ["SENT","PARTIALLY_RECEIVED","RECEIVED"])]

                # Summary of filtered
                _ft_m1, _ft_m2, _ft_m3 = st.columns(3)
                _ft_m1.metric("Lines",        len(_filtered))
                _ft_m2.metric("Total pairs",  sum(r["pairs"] for r in _filtered))
                _ft_m3.metric("Total invoiced", f"₹{sum(r['invoiced'] for r in _filtered):,.0f}")

                # Table header
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:1fr 2fr 2fr 1fr 0.8fr 0.8fr 0.6fr 0.6fr 0.7fr 0.8fr;"
                    "gap:6px;padding:4px 8px;font-size:0.68rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Date</span><span>Supplier</span><span>Blank</span>"
                    "<span>Status</span><span>Base</span><span>Add</span>"
                    "<span>R</span><span>L</span><span>Pairs</span><span>Invoiced</span>"
                    "</div>", unsafe_allow_html=True)

                for _fr in _filtered[:200]:  # cap at 200 rows for display
                    _sts_c2 = {"SENT":"#f59e0b","PARTIALLY_RECEIVED":"#3b82f6",
                               "RECEIVED":"#22c55e"}.get(_fr["status"],"#64748b")
                    _bname2 = " ".join(filter(None, [_fr["brand"], _fr["category"], _fr["material"]]))
                    _inv_str2 = f"₹{_fr['invoiced']:,.0f}" if _fr["invoiced"] > 0 else "—"
                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:1fr 2fr 2fr 1fr 0.8fr 0.8fr 0.6fr 0.6fr 0.7fr 0.8fr;"
                        f"gap:6px;padding:4px 8px;font-size:0.76rem;"
                        f"border-bottom:1px solid #0f172a'>"
                        f"<span style='color:#94a3b8;font-size:0.7rem'>{_fr['order_date']}</span>"
                        f"<span style='color:#e2e8f0'>{_fr['supplier']}</span>"
                        f"<span style='color:#e2e8f0'>{_bname2}</span>"
                        f"<span style='color:{_sts_c2};font-size:0.7rem'>{_fr['status'][:8]}</span>"
                        f"<span style='color:#94a3b8'>{_fr['base']}</span>"
                        f"<span style='color:#94a3b8'>{_fr['add']}</span>"
                        f"<span style='color:#3b82f6'>{_fr['r_qty'] or '—'}</span>"
                        f"<span style='color:#a855f7'>{_fr['l_qty'] or '—'}</span>"
                        f"<span style='color:#fbbf24;font-weight:700'>{_fr['pairs']}</span>"
                        f"<span style='color:#94a3b8;font-size:0.7rem'>{_inv_str2}</span>"
                        f"</div>", unsafe_allow_html=True)

                # Excel download of full consolidated report
                try:
                    import io as _io2, openpyxl as _oxl2
                    _wb3 = _oxl2.Workbook(); _ws3 = _wb3.active
                    _ws3.title = "Procurement"
                    _ws3.append(["Date","Supplier","Brand","Category","Material",
                                 "Status","Base","Add","R Qty","L Qty","Pairs","Invoiced ₹"])
                    for _fr in _flat_lines:
                        _ws3.append([
                            _fr["order_date"], _fr["supplier"],
                            _fr["brand"], _fr["category"], _fr["material"],
                            _fr["status"], _fr["base"], _fr["add"],
                            _fr["r_qty"] or "", _fr["l_qty"] or "", _fr["pairs"],
                            _fr["invoiced"] or "",
                        ])
                    _xb3 = _io2.BytesIO(); _wb3.save(_xb3); _xb3.seek(0)
                    st.download_button(
                        "📊 Download Full Procurement Report (Excel)",
                        data=_xb3.getvalue(),
                        file_name=f"blank_procurement_{_today_d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except ImportError:
                    st.caption("openpyxl not installed — Excel download unavailable.")
            else:
                st.info("No procurement data yet.")

            st.markdown("---")

            # ════════════════════════════════════════════════════
            # SECTION 5: CONSOLIDATED INVOICE REPORT
            # ════════════════════════════════════════════════════
            st.markdown(
                "<div style='background:#0f172a;border:1px solid #1e3a5f;"
                "border-left:4px solid #ec4899;border-radius:8px;"
                "padding:10px 16px;margin-bottom:14px'>"
                "<span style='color:#f9a8d4;font-size:1rem;font-weight:800'>"
                "🧾 Consolidated Invoice Report</span></div>",
                unsafe_allow_html=True)

            _all_invoices = _brq("""
                SELECT
                    bri.invoice_no,
                    bri.invoice_date,
                    bri.invoice_amount,
                    bri.gst_amount,
                    COALESCE(bri.invoice_amount,0) + COALESCE(bri.gst_amount,0) AS total_with_gst,
                    bri.notes,
                    bri.created_at::date AS recorded_date,
                    COALESCE(p.party_name,'Unknown') AS supplier,
                    bro.order_date::date             AS po_date,
                    bro.status                       AS po_status
                FROM blank_replenishment_invoices bri
                JOIN blank_replenishment_orders bro ON bro.id = bri.po_id
                LEFT JOIN parties p ON p.id = bro.supplier_id
                ORDER BY bri.invoice_date DESC
                LIMIT 200
            """)

            if _all_invoices:
                _tot_inv_amt  = sum(float(r["invoice_amount"] or 0) for r in _all_invoices)
                _tot_gst_amt  = sum(float(r["gst_amount"] or 0)     for r in _all_invoices)
                _tot_with_gst = _tot_inv_amt + _tot_gst_amt

                _im1, _im2, _im3, _im4 = st.columns(4)
                _im1.metric("Total invoices",  len(_all_invoices))
                _im2.metric("Net amount",      f"₹{_tot_inv_amt:,.0f}")
                _im3.metric("GST",             f"₹{_tot_gst_amt:,.0f}")
                _im4.metric("Total with GST",  f"₹{_tot_with_gst:,.0f}")

                # Header
                st.markdown(
                    "<div style='display:grid;"
                    "grid-template-columns:1.2fr 1.5fr 1.5fr 1fr 1fr 1fr 2fr;"
                    "gap:6px;padding:4px 8px;font-size:0.68rem;color:#94a3b8;"
                    "font-weight:700;text-transform:uppercase;border-bottom:1px solid #1e293b'>"
                    "<span>Date</span><span>Invoice No</span><span>Supplier</span>"
                    "<span>Amount</span><span>GST</span><span>Total</span><span>Notes</span>"
                    "</div>", unsafe_allow_html=True)

                for _inv in _all_invoices:
                    _inv_a = float(_inv["invoice_amount"] or 0)
                    _gst_a = float(_inv["gst_amount"] or 0)
                    _tot_a = _inv_a + _gst_a
                    st.markdown(
                        f"<div style='display:grid;"
                        f"grid-template-columns:1.2fr 1.5fr 1.5fr 1fr 1fr 1fr 2fr;"
                        f"gap:6px;padding:5px 8px;font-size:0.78rem;"
                        f"border-bottom:1px solid #0f172a'>"
                        f"<span style='color:#94a3b8'>{str(_inv['invoice_date'] or '')[:10]}</span>"
                        f"<span style='color:#e2e8f0;font-weight:600'>{_inv['invoice_no'] or '—'}</span>"
                        f"<span style='color:#e2e8f0'>{_inv['supplier']}</span>"
                        f"<span style='color:#94a3b8'>₹{_inv_a:,.0f}</span>"
                        f"<span style='color:#94a3b8'>₹{_gst_a:,.0f}</span>"
                        f"<span style='color:#fbbf24;font-weight:700'>₹{_tot_a:,.0f}</span>"
                        f"<span style='color:#475569;font-size:0.72rem'>{(_inv['notes'] or '')[:40]}</span>"
                        f"</div>", unsafe_allow_html=True)

                # Excel download
                try:
                    import io as _io3, openpyxl as _oxl3
                    _wb4 = _oxl3.Workbook(); _ws4 = _wb4.active
                    _ws4.title = "Invoices"
                    _ws4.append(["Invoice Date","Invoice No","Supplier","PO Date",
                                 "Net Amount","GST","Total","Notes","Recorded Date"])
                    for _inv in _all_invoices:
                        _ws4.append([
                            str(_inv["invoice_date"] or "")[:10],
                            _inv["invoice_no"] or "",
                            _inv["supplier"],
                            str(_inv["po_date"] or "")[:10],
                            float(_inv["invoice_amount"] or 0),
                            float(_inv["gst_amount"] or 0),
                            float(_inv["invoice_amount"] or 0) + float(_inv["gst_amount"] or 0),
                            _inv["notes"] or "",
                            str(_inv["recorded_date"] or "")[:10],
                        ])
                    _xb4 = _io3.BytesIO(); _wb4.save(_xb4); _xb4.seek(0)
                    st.download_button(
                        "📊 Download Invoice Report (Excel)",
                        data=_xb4.getvalue(),
                        file_name=f"blank_invoices_{_today_d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except ImportError:
                    st.caption("openpyxl not installed.")
            else:
                st.info("No invoices recorded yet.")

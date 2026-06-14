"""
modules/backoffice/stock_pipeline.py
==========================================
Stock Pipeline (📦) — reservation display, replenishment send, Stock Procurement save.

Extracted from production_page.py.
Entry points called from production_page.py:
  render_stock_pipeline
  render_stock_procurement
"""
from __future__ import annotations
from modules.backoffice.production_shared import (
    _sync_supplier_orders_id_sequence,
    _q,
    _render_pipeline_cards,
    _go_to_billing
)

import streamlit as st

def _scan_norm(value: str) -> str:
    s = "".join(ch for ch in str(value or "") if ch.isalnum()).lower()
    if s.startswith("o") and len(s) > 1:
        s = s[1:]
    return s

def _scan_match(needle: str, *hay_values) -> bool:
    raw = str(needle or "").strip().lower()
    norm = _scan_norm(needle)
    if not raw and not norm:
        return True
    for val in hay_values:
        text = str(val or "").lower()
        if raw and raw in text:
            return True
        if norm and norm in _scan_norm(text):
            return True
    return False

# --- Cancelled status normalisation -----------------------------------------
_CANCELLED_STATUSES = {
    "CANCELLED", "CANCELLED_RELEASED", "SUPPLIER_CANCELLED",
    "DISCARDED", "NOT_NEEDED", "CANCELLED / NOT NEEDED",
}


def _normalize_repl_status(status_str: str) -> str:
    """Normalise various cancelled / not-needed spellings to CANCELLED_RELEASED."""
    s = str(status_str or "").upper().strip()
    if s in _CANCELLED_STATUSES:
        return "CANCELLED_RELEASED"
    return s or "PENDING"


def _is_cancelled_status(status_str: str) -> bool:
    """Return True for any cancelled-like status."""
    return str(status_str or "").upper().strip() in _CANCELLED_STATUSES


def _render_stock_not_needed_view():
    """
    Cancelled / Not Needed view — minimal, recovery-only.

    Shows ONLY lines where the user (or supplier) decided they no longer want
    this procurement. No supplier picker, no qty editor, no replenishment
    expander. Just a flat list with one Restore button per line.

    Restore puts the line back to PENDING and recovers the previously-released
    allocation onto order_lines.allocated_qty, so the line resumes its normal
    pipeline flow.
    """
    import json as _jss
    import datetime as _dts

    st.markdown(
        "<div style='background:#1f1610;border:1px solid #a16207;"
        "border-left:4px solid #f59e0b;border-radius:8px;"
        "padding:10px 16px;margin-bottom:12px'>"
        "<span style='color:#fbbf24;font-size:1rem;font-weight:800'>🚫 Not Needed</span>"
        "<span style='color:#94a3b8;font-size:0.78rem;margin-left:10px'>"
        "Stock lines you cancelled or discarded. Click <b>Restore</b> on any "
        "line to bring it back into the active pipeline — its allocation is "
        "reinstated and it returns to PENDING for normal replenishment."
        "</span></div>",
        unsafe_allow_html=True,
    )

    # ── Search filter (single field — keep it simple) ────────────────────
    with st.container(border=True):
        _f1, _f2 = st.columns([5, 2])
        _q_text = _f1.text_input(
            "Search",
            placeholder="Order no / patient / product",
            key="stk_nn_q",
            label_visibility="collapsed",
        )
        _flt_kind = _f2.selectbox(
            "Type",
            ["All Cancelled", "🚫 Cancelled / Released", "🗑️ Discarded", "❌ Supplier Cancelled"],
            key="stk_nn_flt_kind",
            label_visibility="collapsed",
        )

    _kind_map = {
        "🚫 Cancelled / Released": "CANCELLED",
        "🗑️ Discarded":           "DISCARDED",
        "❌ Supplier Cancelled":   "SUPPLIER_CANCELLED",
    }
    _status_in_clause = (
        f"= '{_kind_map[_flt_kind]}'"
        if _flt_kind in _kind_map
        else "IN ('CANCELLED','CANCELLED_RELEASED','DISCARDED','SUPPLIER_CANCELLED','NOT_NEEDED')"
    )

    # Note: cancelled lines have allocated_qty=0 (released), so the main
    # stock query's allocated_qty>0 filter would hide them. We query
    # directly here, status-only, route-agnostic.
    try:
        _rows = _q(f"""
            SELECT
                o.order_no,
                COALESCE(o.patient_name, o.party_name, '—') AS patient_name,
                ol.id::text                    AS line_id,
                ol.eye_side,
                ol.quantity,
                COALESCE(p.product_name,'')    AS product_name,
                COALESCE(p.brand,'')           AS brand,
                ol.sph, ol.cyl, ol.axis, ol.add_power,
                ol.lens_params,
                UPPER(COALESCE(ol.lens_params->>'replenishment_status','')) AS status,
                COALESCE(ol.lens_params->>'replenishment_po_no','')          AS po_no,
                COALESCE(ol.lens_params->>'replenishment_supplier_name','')  AS sup_name,
                COALESCE(ol.lens_params->>'cancel_reason','')                AS cancel_reason,
                o.created_at
            FROM order_lines ol
            JOIN orders o      ON o.id = ol.order_id
            LEFT JOIN products p ON p.id = ol.product_id
            WHERE COALESCE(ol.is_deleted, FALSE) = FALSE
              AND COALESCE(ol.is_service_line, FALSE) = FALSE
              AND UPPER(COALESCE(ol.eye_side,'')) NOT IN ('S','SERVICE')
              AND o.status NOT IN ('CANCELLED','CLOSED')
              AND UPPER(COALESCE(ol.lens_params->>'replenishment_status','')) {_status_in_clause}
            ORDER BY o.created_at DESC
            LIMIT 500
        """, {}) or []
    except Exception as _se:
        st.error(f"Could not load cancelled lines: {_se}")
        return

    # Apply text search filter (client-side — simple ILIKE on three fields)
    if _q_text.strip():
        _rows = [
            r for r in _rows
            if _scan_match(_q_text, r.get("order_no",""), r.get("patient_name",""), r.get("product_name",""))
        ]

    if not _rows:
        st.info("✅ No cancelled / not-needed lines in this view.")
        return

    st.caption(f"**{len(_rows)} cancelled line(s)** — click Restore to recover.")

    # ── List ─────────────────────────────────────────────────────────────
    for _row in _rows:
        _lid     = str(_row["line_id"])
        _lp      = _row.get("lens_params") or {}
        if isinstance(_lp, str):
            try: _lp = _jss.loads(_lp)
            except Exception: _lp = {}
        _status_raw = str(_row.get("status","")).upper()
        _status = _normalize_repl_status(_status_raw)
        _badge   = {
            "CANCELLED":          ("🚫 Cancelled / Released", "#f59e0b"),
            "DISCARDED":          ("🗑️ Discarded",            "#64748b"),
            "SUPPLIER_CANCELLED": ("❌ Supplier Cancelled",    "#dc2626"),
        }.get(_status, ("Cancelled", "#94a3b8"))
        _bl, _bc = _badge

        # Power summary inline
        _pw = []
        try:
            if _row.get("sph") is not None: _pw.append(f"SPH {float(_row['sph']):+.2f}")
            if _row.get("cyl") and abs(float(_row["cyl"])) > 0.01: _pw.append(f"CYL {float(_row['cyl']):+.2f}")
            if _row.get("axis"): _pw.append(f"AX {int(float(_row['axis']))}")
            if _row.get("add_power") and float(_row["add_power"]) > 0: _pw.append(f"ADD +{float(_row['add_power']):.2f}")
        except Exception: pass
        _pwr_str = "  ".join(_pw)

        _c1, _c2 = st.columns([9, 2])
        with _c1:
            st.markdown(
                f"<div style='border:1px solid #1e293b;"
                f"border-left:4px solid {_bc};"
                f"border-radius:6px;padding:8px 12px;margin:4px 0'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center'>"
                f"<span style='color:#e2e8f0;font-weight:700'>"
                f"👁 {_row.get('eye_side','')} {_row.get('product_name','')}"
                f"</span>"
                f"<span style='color:{_bc};font-size:0.72rem;font-weight:700'>{_bl}</span>"
                f"</div>"
                f"<div style='color:#94a3b8;font-size:0.72rem;margin-top:2px'>"
                f"{_row.get('order_no','')} · {_row.get('patient_name','')}"
                + (f" · Qty {_row.get('quantity',1)}" if _row.get('quantity') else "")
                + "</div>"
                + (f"<div style='color:#38bdf8;font-size:0.7rem;font-family:monospace'>{_pwr_str}</div>"
                   if _pwr_str else "")
                + (f"<div style='color:#64748b;font-size:0.7rem'>"
                   f"Supplier was: {_row.get('sup_name','')} · PO: {_row.get('po_no','—')}</div>"
                   if _row.get('sup_name') else "")
                + (f"<div style='color:#f59e0b;font-size:0.7rem'>"
                   f"Reason: {_row.get('cancel_reason','')}</div>"
                   if _row.get('cancel_reason') else "")
                + "</div>",
                unsafe_allow_html=True,
            )
        with _c2:
            if st.button(
                "↩️ Restore",
                key=f"stk_nn_restore_{_lid}",
                type="primary",
                use_container_width=True,
                help="Recover this line: reinstate allocation and return to PENDING",
            ):
                try:
                    from modules.sql_adapter import run_write as _rw_rest
                    # Recover quantity: prefer the saved prev_allocated_qty,
                    # fall back to billed_qty, then quantity, then 1.
                    _restored_qty = int(
                        _lp.get("replenishment_prev_allocated_qty") or
                        _row.get("quantity") or
                        1
                    )
                    _reopen_lp = dict(_lp or {})
                    _reopen_lp.pop("replenishment_prev_allocated_qty", None)
                    for _old_key in (
                        "replenishment_po_no",
                        "replenishment_supplier_id",
                        "replenishment_supplier_name",
                        "supplier_confirmation_no",
                        "replenishment_ref_note",
                    ):
                        _reopen_lp.pop(_old_key, None)
                    _reopen_lp["replenishment_status"] = "PENDING"
                    # Stamp a recovery event in the timeline if present.
                    _tl = _reopen_lp.get("replenishment_timeline")
                    if not isinstance(_tl, list): _tl = []
                    _tl.append({
                        "stage": "RESTORED",
                        "at": _dts.datetime.now().astimezone().isoformat(),
                        "source": "Not Needed view",
                    })
                    _reopen_lp["replenishment_timeline"] = _tl
                    _rw_rest("""
                        UPDATE order_lines
                        SET allocated_qty = %(alloc)s,
                            batch_status  = 'ALLOCATED',
                            lens_params   = %(lp)s::jsonb
                        WHERE id = %(lid)s::uuid
                    """, {"lid": _lid, "alloc": _restored_qty, "lp": _jss.dumps(_reopen_lp)})
                    _rw_rest("""
                        UPDATE supplier_order_items
                        SET item_status = 'CANCELLED'
                        WHERE customer_line_id::text = %(lid)s
                          AND UPPER(COALESCE(item_status,'')) NOT IN
                              ('RECEIVED','PROCURED','PURCHASE_ACKED','LOCKED','CANCELLED','VOID')
                    """, {"lid": _lid})
                    st.success(
                        f"↩️ Restored {_row.get('order_no','')} ({_row.get('eye_side','')}) — "
                        f"allocation reinstated, line back to PENDING."
                    )
                    st.rerun()
                except Exception as _re:
                    st.error(f"Restore failed: {_re}")


def _render_stock_po_view():
    """Stock PO view — all stock replenishment POs created from this tab.

    Shows: PO no, Supplier, Status, Refs, Line summary.
    Actions: Resend (WhatsApp / Mail / Excel) for SENT POs; Cancel for unconfirmed.
    """
    import urllib.parse as _up_po
    import datetime as _dt_po

    st.markdown("### 📄 Stock Purchase Orders")
    st.caption("All POs generated from Stock pipeline. Resend communication or cancel POs that haven't been confirmed yet.")

    # ── Filters ──────────────────────────────────────────────────────────
    with st.container(border=True):
        _pf1, _pf2, _pf3 = st.columns([3, 2, 2])
        _po_search = _pf1.text_input(
            "Search",
            placeholder="🔍 PO no / supplier name",
            key="stk_po_search",
            label_visibility="collapsed",
        )
        _po_status_flt = _pf2.selectbox(
            "Status",
            ["All", "SENT (awaiting)", "ACKNOWLEDGED", "RECEIVED", "CANCELLED"],
            key="stk_po_status_flt",
            label_visibility="collapsed",
        )
        _po_days = _pf3.selectbox(
            "Window",
            ["Last 30 days", "Last 90 days", "Last 1 year", "All time"],
            key="stk_po_days",
            label_visibility="collapsed",
        )

    _date_clause = ""
    if _po_days == "Last 30 days":
        _date_clause = "AND so.order_date >= CURRENT_DATE - INTERVAL '30 days'"
    elif _po_days == "Last 90 days":
        _date_clause = "AND so.order_date >= CURRENT_DATE - INTERVAL '90 days'"
    elif _po_days == "Last 1 year":
        _date_clause = "AND so.order_date >= CURRENT_DATE - INTERVAL '365 days'"

    _status_clause = ""
    _qp_po = {}
    if _po_status_flt == "SENT (awaiting)":
        _status_clause = "AND UPPER(COALESCE(so.status,'')) = 'SENT'"
    elif _po_status_flt == "ACKNOWLEDGED":
        _status_clause = "AND UPPER(COALESCE(so.status,'')) IN ('ACKNOWLEDGED','PARTIAL')"
    elif _po_status_flt == "RECEIVED":
        _status_clause = "AND UPPER(COALESCE(so.status,'')) = 'RECEIVED'"
    elif _po_status_flt == "CANCELLED":
        _status_clause = "AND UPPER(COALESCE(so.status,'')) = 'CANCELLED'"

    _search_clause = ""
    if _po_search.strip():
        _search_clause = "AND (LOWER(so.supplier_order_id) LIKE %(s)s OR LOWER(COALESCE(so.supplier_name,'')) LIKE %(s)s)"
        _qp_po["s"] = f"%{_po_search.strip().lower()}%"

    rows = _q(f"""
        SELECT so.id, so.supplier_order_id, so.supplier_name, so.supplier_id::text AS supplier_id,
               so.order_date, so.status, so.total_items, so.total_qty, so.total_value,
               so.created_at,
               COUNT(soi.id) AS line_count,
               COUNT(soi.id) FILTER (WHERE soi.received_qty >= soi.ordered_qty) AS received_lines,
               STRING_AGG(DISTINCT ol.lens_params->>'supplier_confirmation_no', ', ')
                  FILTER (WHERE ol.lens_params->>'supplier_confirmation_no' IS NOT NULL
                          AND ol.lens_params->>'supplier_confirmation_no' <> '') AS supplier_refs
        FROM supplier_orders so
        LEFT JOIN supplier_order_items soi ON soi.supplier_order_id = so.id
        LEFT JOIN order_lines ol ON ol.id::text = NULLIF(soi.customer_line_id::text, '')
        WHERE COALESCE(so.created_by,'') IN ('stock_replenishment', 'stock_replenishment_bulk')
        {_date_clause}
        {_status_clause}
        {_search_clause}
        GROUP BY so.id, so.supplier_order_id, so.supplier_name, so.supplier_id,
                 so.order_date, so.status, so.total_items, so.total_qty, so.total_value, so.created_at
        ORDER BY so.created_at DESC NULLS LAST, so.id DESC
        LIMIT 200
    """, _qp_po) or []

    if not rows:
        st.info("No stock replenishment POs match these filters.")
        return

    # ── Summary metrics ──────────────────────────────────────────────────
    _total_pos = len(rows)
    _total_value = sum(float(r.get("total_value") or 0) for r in rows)
    _open_pos = sum(1 for r in rows if str(r.get("status","")).upper() in ("SENT", "DRAFT"))
    _received_pos = sum(1 for r in rows if str(r.get("status","")).upper() == "RECEIVED")
    _m1, _m2, _m3, _m4 = st.columns(4)
    _m1.metric("POs", _total_pos)
    _m2.metric("Open / Awaiting", _open_pos, delta_color="inverse")
    _m3.metric("Received", _received_pos)
    _m4.metric("Total Value", f"₹{_total_value:,.0f}")

    st.divider()

    # ── PO list ─────────────────────────────────────────────────────────
    for po in rows:
        poid = po.get("id")
        _po_no = po.get("supplier_order_id") or f"PO-{poid}"
        _sup_name = po.get("supplier_name") or "—"
        _status = str(po.get("status","") or "").upper()
        _refs = po.get("supplier_refs") or ""

        # Status colour
        _stat_clr = {
            "SENT":         "#f59e0b",
            "DRAFT":        "#94a3b8",
            "ACKNOWLEDGED": "#3b82f6",
            "PARTIAL":      "#3b82f6",
            "RECEIVED":     "#22c55e",
            "CANCELLED":    "#64748b",
        }.get(_status, "#94a3b8")

        _is_open = _status in ("SENT", "DRAFT")

        _hdr = (
            f"**{_po_no}** · {_sup_name} · "
            f"{int(po.get('line_count') or 0)} line(s) · "
            f"₹{float(po.get('total_value') or 0):,.0f}"
        )

        with st.expander(_hdr, expanded=False):
            # Status row
            _sc1, _sc2, _sc3, _sc4 = st.columns([1.5, 1.5, 1.5, 2])
            _sc1.markdown(
                f"<div style='color:{_stat_clr};font-weight:700;font-size:0.9rem'>"
                f"● {_status}</div>",
                unsafe_allow_html=True,
            )
            _sc2.caption(f"Date: {str(po.get('order_date') or '')[:10]}")
            _sc3.caption(f"Qty: {po.get('total_qty') or 0}")
            if _refs:
                _sc4.markdown(
                    f"<div style='color:#22c55e;font-size:0.78rem;font-weight:600'>"
                    f"📋 Refs: {_refs[:60]}{'...' if len(_refs)>60 else ''}</div>",
                    unsafe_allow_html=True,
                )
            else:
                _sc4.caption("📋 No supplier ref yet")

            # Line items
            items = _q("""
                SELECT soi.item_no, soi.product_name, soi.eye_side, soi.sph, soi.cyl,
                       soi.axis, soi.add_power, soi.ordered_qty, soi.received_qty,
                       soi.pending_qty, soi.item_status, o.order_no, o.patient_name,
                       ol.lens_params->>'supplier_confirmation_no' AS line_ref
                FROM supplier_order_items soi
                LEFT JOIN order_lines ol ON ol.id::text = NULLIF(soi.customer_line_id::text, '')
                LEFT JOIN orders o ON o.id = ol.order_id
                WHERE soi.supplier_order_id = %(poid)s
                ORDER BY soi.item_no
            """, {"poid": poid}) or []

            if items:
                try:
                    import pandas as _pd
                    _df = _pd.DataFrame([{
                        "#": it.get("item_no"),
                        "Order": it.get("order_no") or "—",
                        "Patient": it.get("patient_name") or "—",
                        "Product": (it.get("product_name") or "")[:40],
                        "Eye": it.get("eye_side") or "",
                        "Qty": it.get("ordered_qty") or 0,
                        "Recv": it.get("received_qty") or 0,
                        "Ref": it.get("line_ref") or "—",
                        "Status": it.get("item_status") or "—",
                    } for it in items])
                    st.dataframe(_df, use_container_width=True, hide_index=True)
                except Exception:
                    for it in items:
                        st.markdown(
                            f"{it.get('item_no')}. {it.get('product_name','')} "
                            f"({it.get('eye_side','')}) qty {it.get('ordered_qty',0)} · "
                            f"{it.get('order_no','')} {it.get('patient_name','')}"
                        )

            # ── Actions for open POs ────────────────────────────────────
            if _is_open:
                st.divider()
                st.caption("**Actions**")

                # Build comms message
                _msg_lines = [
                    f"*Stock PO {_po_no}*",
                    f"Date: {str(po.get('order_date') or '')[:10]}",
                    f"To: {_sup_name}",
                    "",
                    f"Total: {int(po.get('total_qty') or 0)} pcs across {int(po.get('line_count') or 0)} line(s)",
                    "",
                ]
                for it in items[:30]:
                    _eye = str(it.get("eye_side","")).upper()
                    _msg_lines.append(
                        f"  {it.get('product_name','')[:40]} · {_eye} · {it.get('ordered_qty',0)} pcs"
                    )
                if len(items) > 30:
                    _msg_lines.append(f"  ...and {len(items)-30} more")
                _msg_lines += ["", "Please confirm dispatch.", "Thank you."]
                _msg = "\n".join(_msg_lines)
                _plain_msg = _msg.replace("*", "")

                # Lookup supplier contact
                _sup_party = []
                if po.get("supplier_id"):
                    _sup_party = _q("""
                        SELECT COALESCE(mobile,'') AS mobile, COALESCE(email,'') AS email
                        FROM parties WHERE id = %(sid)s::uuid LIMIT 1
                    """, {"sid": po["supplier_id"]}) or []
                _sup_mob = _sup_party[0]["mobile"] if _sup_party else ""
                _sup_em  = _sup_party[0]["email"] if _sup_party else ""
                _wa_mob = "".join(c for c in _sup_mob if c.isdigit())
                if _wa_mob and not _wa_mob.startswith("91"):
                    _wa_mob = "91" + _wa_mob
                _wa_url = (
                    f"https://wa.me/{_wa_mob}?text={_up_po.quote(_msg, safe='')}"
                    if _wa_mob else
                    f"https://wa.me/?text={_up_po.quote(_msg, safe='')}"
                )
                _mail_sub = f"Stock PO {_po_no}"
                _mailto = (
                    f"mailto:{_sup_em}?subject={_up_po.quote(_mail_sub)}&body={_up_po.quote(_plain_msg)}"
                    if _sup_em else ""
                )

                _ac1, _ac2, _ac3, _ac4 = st.columns(4)
                _ac1.link_button(
                    "📲 Resend WhatsApp",
                    url=_wa_url,
                    use_container_width=True,
                )
                if _sup_em:
                    _ac2.link_button(
                        "📧 Resend Mail",
                        url=_mailto,
                        use_container_width=True,
                    )
                else:
                    _ac2.button(
                        "📧 No email",
                        key=f"stk_po_noemail_{poid}",
                        disabled=True,
                        use_container_width=True,
                    )

                # Excel download
                try:
                    import io as _io_po
                    import openpyxl as _xl_po
                    from openpyxl.styles import Font as _F2, PatternFill as _PF2
                    _wb = _xl_po.Workbook()
                    _ws = _wb.active
                    _ws.title = "Stock PO"
                    _ws.append(["PO No", _po_no])
                    _ws.append(["Supplier", _sup_name])
                    _ws.append(["Date", str(po.get("order_date") or "")[:10]])
                    _ws.append([])
                    _ws.append(["#", "Product", "Eye", "SPH", "CYL", "AXIS", "ADD", "Qty", "Order", "Patient"])
                    for _c in _ws[5]:
                        _c.font = _F2(bold=True, color="FFFFFF")
                        _c.fill = _PF2(start_color="6366F1", end_color="6366F1", fill_type="solid")
                    for it in items:
                        _ws.append([
                            it.get("item_no"),
                            it.get("product_name") or "",
                            it.get("eye_side") or "",
                            it.get("sph"),
                            it.get("cyl"),
                            it.get("axis"),
                            it.get("add_power"),
                            it.get("ordered_qty") or 0,
                            it.get("order_no") or "",
                            it.get("patient_name") or "",
                        ])
                    _buf = _io_po.BytesIO()
                    _wb.save(_buf)
                    _buf.seek(0)
                    _ac3.download_button(
                        "📊 Excel",
                        data=_buf,
                        file_name=f"{_po_no.replace('/', '-')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"stk_po_xl_{poid}",
                        use_container_width=True,
                    )
                except ImportError:
                    _ac3.caption("Excel: openpyxl not installed")
                except Exception as _xl_e:
                    _ac3.caption(f"Excel error: {_xl_e}")

                # Cancel PO action
                with _ac4:
                    _cancel_key = f"stk_po_cancel_confirm_{poid}"
                    if st.session_state.get(_cancel_key):
                        _ck1, _ck2 = st.columns(2)
                        if _ck1.button("✓ Yes", key=f"stk_po_cancel_yes_{poid}", use_container_width=True):
                            try:
                                from modules.sql_adapter import run_write as _rw_cp
                                _rw_cp(
                                    "UPDATE supplier_orders SET status='CANCELLED', updated_at=NOW() WHERE id=%(id)s",
                                    {"id": poid},
                                )
                                _rw_cp(
                                    "UPDATE supplier_order_items SET item_status='CANCELLED' "
                                    "WHERE supplier_order_id=%(id)s AND COALESCE(received_qty,0) < COALESCE(ordered_qty,0)",
                                    {"id": poid},
                                )
                                # Also reset replenishment_status on linked order_lines so they
                                # can be re-ordered. Only flip PENDING-ish / PO_SENT lines, not
                                # ORDERED / PROCURED — those have moved past PO recall.
                                _rw_cp("""
                                    UPDATE order_lines
                                    SET lens_params = COALESCE(lens_params,'{}'::jsonb)
                                        || '{"replenishment_status":"PENDING"}'::jsonb
                                    WHERE id::text IN (
                                        SELECT NULLIF(customer_line_id::text,'')
                                        FROM supplier_order_items
                                        WHERE supplier_order_id=%(id)s
                                          AND customer_line_id IS NOT NULL
                                    )
                                    AND UPPER(COALESCE(lens_params->>'replenishment_status','')) IN ('PO_SENT','')
                                """, {"id": poid})
                                st.session_state.pop(_cancel_key, None)
                                st.success(f"❌ PO {_po_no} cancelled. Linked lines reset to PENDING.")
                                st.rerun()
                            except Exception as _ce:
                                st.error(f"Cancel failed: {_ce}")
                        if _ck2.button("✕ No", key=f"stk_po_cancel_no_{poid}", use_container_width=True):
                            st.session_state.pop(_cancel_key, None)
                            st.rerun()
                    else:
                        if st.button(
                            "❌ Cancel PO",
                            key=f"stk_po_cancel_{poid}",
                            use_container_width=True,
                            help="Cancel this PO. Linked PO_SENT lines will reset to PENDING.",
                        ):
                            st.session_state[_cancel_key] = True
                            st.rerun()

            elif _status == "RECEIVED":
                st.success("✅ PO received in full.")
            elif _status == "CANCELLED":
                st.caption("This PO was cancelled.")


def _render_stock_pipeline():
    """
    Stock pipeline — shows reserved stock lines for fulfilment and replenishment.

    Flow:
      Backoffice stock assignment saved → appears here immediately
      → Reserve stock blocks it from punching/backoffice availability
      → Replenishment can be ordered/discarded
      → Purchase acknowledgement can be recorded when supplier bill arrives
    """
    import json as _jss
    import urllib.parse as _uparse
    import datetime as _dts

    st.markdown("### 📦 Stock Replenishment Pipeline")
    st.caption(
        "Customer-side STOCK lines are ready for billing directly. "
        "This tab is only for replenishing shelf stock consumed/reserved by those orders."
    )

    # ── Sub-tab selector ─────────────────────────────────────────────────────
    # Active Pipeline → normal flow (PENDING/PO_SENT/ORDERED) with replenishment UI.
    # Not Needed     → only CANCELLED/DISCARDED lines, with a single Restore button.
    #                  No replenishment panel, no supplier picker. Pure recovery view.
    _stk_mode = st.radio(
        "View",
        ["📦 Active Pipeline", "🚫 Not Needed (Cancelled)", "📄 Stock POs"],
        key="stk_mode_radio",
        horizontal=True,
        label_visibility="collapsed",
    )
    if _stk_mode == "🚫 Not Needed (Cancelled)":
        _render_stock_not_needed_view()
        return
    if _stk_mode == "📄 Stock POs":
        _render_stock_po_view()
        return

    # ── Filters ──────────────────────────────────────────────────────────────
    with st.container(border=True):
        sf1, sf2, sf3, sf4 = st.columns([3, 2, 2, 2])
        with sf1:
            _flt_ord = sf1.text_input("Order/Patient", placeholder="🔍 Order no / patient",
                                      key="stk_flt_ord", label_visibility="collapsed")
        with sf2:
            _flt_from = sf2.date_input("From", value=None, key="stk_flt_from",
                                       label_visibility="collapsed", format="DD/MM/YYYY",
                                       help="Billed from date")
        with sf3:
            _show_acked = sf3.toggle("Show acknowledged", value=False, key="stk_show_acked")
        # Cancelled lines now have their own sub-tab ("🚫 Not Needed").
        # No need for a toggle here — Active Pipeline always excludes them.
        _show_cancelled = False
        with sf4:
            _flt_status = sf4.selectbox("Status", [
                "All",
                "🔴 Pending",
                "📤 PO Sent (awaiting confirmation)",
                "⏳ In Procurement",
            ], key="stk_flt_status", label_visibility="collapsed")
        sf5, sf6, sf7 = st.columns([2, 2, 2])
        with sf5:
            _sup_rows = _q("""
                SELECT DISTINCT COALESCE(pt.party_name, ol.lens_params->>'supplier_name', '') AS sname
                FROM order_lines ol
                LEFT JOIN purchase_acknowledgements pa ON pa.order_line_id = ol.id
                LEFT JOIN parties pt ON pt.id = pa.supplier_id
                WHERE ol.lens_params->>'manufacturing_route' = 'STOCK'
                  AND COALESCE(ol.allocated_qty,0) > 0
                  AND COALESCE(pt.party_name, ol.lens_params->>'supplier_name', '') != ''
            """, {}) or []
            _sup_opts = ["All Suppliers"] + [r["sname"] for r in _sup_rows if r.get("sname")]
            _flt_sup = sf5.selectbox("Supplier", _sup_opts,
                                      key="stk_flt_sup", label_visibility="collapsed")
        with sf6:
            _brand_rows = _q("""
                SELECT DISTINCT p.brand FROM order_lines ol
                JOIN products p ON p.id = ol.product_id
                WHERE ol.lens_params->>'manufacturing_route' = 'STOCK'
                  AND COALESCE(ol.allocated_qty,0) > 0
                  AND p.brand IS NOT NULL
            """, {}) or []
            _brand_opts = ["All Brands"] + [r["brand"] for r in _brand_rows if r.get("brand")]
            _flt_brand = sf6.selectbox("Brand", _brand_opts,
                                        key="stk_flt_brand", label_visibility="collapsed")
        with sf7:
            _cat_rows = _q("""
                SELECT DISTINCT p.main_group FROM order_lines ol
                JOIN products p ON p.id = ol.product_id
                WHERE ol.lens_params->>'manufacturing_route' = 'STOCK'
                  AND COALESCE(ol.allocated_qty,0) > 0
            """, {}) or []
            _cat_opts = ["All Types"] + [r["main_group"] for r in _cat_rows if r.get("main_group")]
            _flt_cat = sf7.selectbox("Type", _cat_opts,
                                      key="stk_flt_cat", label_visibility="collapsed")

    # ── Query: billed stock lines ─────────────────────────────────────────────
    _where_extra = ""
    _params = {}
    if not _show_acked:
        _where_extra += """
            AND NOT EXISTS (
                SELECT 1 FROM purchase_acknowledgements pa2
                WHERE pa2.order_line_id = ol.id
                  AND COALESCE(pa2.purchase_price, 0) > 0
                  AND COALESCE(pa2.is_price_locked, FALSE) = TRUE
            )"""
    if _flt_from:
        _where_extra += " AND DATE(o.created_at) >= %(df)s"
        _params["df"] = str(_flt_from)
    if _flt_ord.strip():
        _where_extra += (
            " AND (LOWER(o.order_no) LIKE %(ord)s "
            " OR regexp_replace(LOWER(COALESCE(o.order_no,'')), '[^a-z0-9]', '', 'g') LIKE %(ord_norm)s "
            " OR LOWER(COALESCE(o.patient_name,o.party_name,'')) LIKE %(ord)s)"
        )
        _params["ord"] = f"%{_flt_ord.strip().lower()}%"
        _params["ord_norm"] = f"%{_scan_norm(_flt_ord)}%"
    if _flt_sup != "All Suppliers":
        _where_extra += " AND (LOWER(COALESCE(pt.party_name, ol.lens_params->>'supplier_name','')) = %(sup)s)"
        _params["sup"] = _flt_sup.lower()
    if _flt_brand != "All Brands":
        _where_extra += " AND p.brand = %(brand)s"
        _params["brand"] = _flt_brand
    if _flt_cat != "All Types":
        _where_extra += " AND p.main_group = %(cat)s"
        _params["cat"] = _flt_cat
    _reserved_filter = (
        """
              AND (
                    COALESCE(ol.allocated_qty,0) > 0
                 OR UPPER(COALESCE(ol.batch_status,'')) = 'CANCELLED'
                 OR UPPER(COALESCE(ol.lens_params->>'replenishment_status','')) IN ('CANCELLED','DISCARDED')
              )
        """
        if _show_cancelled else
        "              AND COALESCE(ol.allocated_qty,0) > 0"
    )

    try:
        rows = _q(f"""
            SELECT
                o.id::text                                     AS order_id,
                o.order_no,
                COALESCE(o.patient_name, o.party_name, '—')   AS patient_name,
                o.status,
                o.created_at,
                ol.id::text                                    AS line_id,
                ol.eye_side,
                ol.quantity,
                COALESCE(NULLIF(ol.billed_qty, 0), NULLIF(ol.allocated_qty, 0), ol.quantity, 1) AS billed_qty,
                ol.product_id::text                            AS product_id,
                ol.lens_params,
                ol.sph, ol.cyl, ol.axis, ol.add_power,
                -- Phase-3 (2026-05-25): surface lens spec for display + PO outputs.
                -- Source priority: lens_params JSON (canonical for in-flight orders).
                NULLIF(TRIM(COALESCE(ol.lens_params->>'coating',
                                     ol.lens_params->>'COATING', '')), '')   AS coating,
                NULLIF(TRIM(COALESCE(ol.lens_params->>'lens_index',
                                     ol.lens_params->>'index',
                                     ol.lens_params->>'INDEX', '')), '')     AS lens_index,
                NULLIF(TRIM(COALESCE(ol.lens_params->>'treatment',
                                     ol.lens_params->>'TREATMENT', '')), '') AS treatment,
                NULLIF(TRIM(COALESCE(ol.lens_params->>'diameter',
                                     ol.lens_params->>'dia',
                                     ol.lens_params->>'DIA', '')), '') AS diameter,
                NULLIF(TRIM(COALESCE(ol.lens_params->>'corridor',
                                     ol.lens_params->>'CORRIDOR', '')), '') AS corridor,
                NULLIF(TRIM(COALESCE(ol.lens_params->>'fitting_height',
                                     ol.lens_params->>'fitting',
                                     ol.lens_params->>'FITTING_HEIGHT', '')), '') AS fitting_height,
                p.product_name,
                p.brand,
                p.category,
                COALESCE(p.unit, 'PCS')    AS unit,
                COALESCE(p.box_size, 1)    AS box_size,
                -- Supplier from purchase_acknowledgements (DB truth) or lens_params
                pa.supplier_id::text                           AS pa_supplier_id,
                COALESCE(pt.party_name,
                         ol.lens_params->>'supplier_name',
                         '')                                   AS supplier_name,
                pa.challan_no                                  AS pa_challan,
                pa.purchase_price                              AS pa_price,
                pa.is_price_locked                             AS pa_locked,
                pa.billing_status                              AS pa_status,
                pa.received_qty                                AS pa_recv_qty,
                COALESCE(ol.allocated_qty,0)                    AS allocated_qty,
                COALESCE(ol.billed_qty,0)                       AS billed_qty_raw,
                COALESCE(ol.dispatched_qty,0)                   AS dispatched_qty,
                CASE
                    WHEN UPPER(COALESCE(ol.batch_status,'')) = 'CANCELLED'
                         THEN 'CANCELLED'
                    ELSE COALESCE(ol.lens_params->>'replenishment_status','PENDING')
                END AS replenishment_status,
                -- Preferred supplier from product_supplier_map
                pref_sup.pref_sup_id,
                pref_sup.pref_sup_name,
                pref_sup.pref_sup_mobile,
                pref_sup.pref_sup_email,
                -- Inventory purchase price for auto-fill
                -- Try purchase_price first, fallback to purchase_rate
                COALESCE((
                    SELECT COALESCE(NULLIF(s.purchase_price,0), s.purchase_rate, 0)
                    FROM inventory_stock s
                    WHERE s.product_id = ol.product_id
                      AND COALESCE(s.is_active,TRUE) = TRUE
                    ORDER BY s.created_at DESC LIMIT 1
                ), 0) AS inv_purchase_price
            FROM order_lines ol
            JOIN orders  o  ON o.id  = ol.order_id
            JOIN products p ON p.id  = ol.product_id
            -- Only STOCK routed lines
            LEFT JOIN purchase_acknowledgements pa ON pa.order_line_id = ol.id
            LEFT JOIN parties pt ON pt.id = pa.supplier_id
            -- Preferred supplier from product_supplier_map (rank=1)
            LEFT JOIN LATERAL (
                SELECT psm.supplier_id::text AS pref_sup_id,
                       pt2.party_name        AS pref_sup_name,
                       pt2.mobile            AS pref_sup_mobile,
                       pt2.email             AS pref_sup_email
                FROM product_supplier_map psm
                JOIN parties pt2 ON pt2.id = psm.supplier_id
                WHERE psm.product_id = ol.product_id
                  AND COALESCE(psm.is_active, TRUE) = TRUE
                ORDER BY psm.rank ASC
                LIMIT 1
            ) pref_sup ON TRUE
            WHERE (
                      ol.lens_params->>'manufacturing_route' = 'STOCK'
                   OR ol.lens_params->>'batch_status' = 'ALLOCATED'
                   OR UPPER(COALESCE(ol.batch_status,'')) = 'CANCELLED'
                   OR (ol.lens_params->>'batch_no' IS NOT NULL
                       AND COALESCE(ol.lens_params->>'manufacturing_route','') != 'INHOUSE')
                  )
              AND UPPER(COALESCE(ol.eye_side,'')) NOT IN ('S','SERVICE')
              AND COALESCE(ol.is_service_line, FALSE) = FALSE
              AND COALESCE(ol.is_deleted, FALSE) = FALSE
              -- Show from reservation onwards. Cancelled/not-needed lines may
              -- have allocated_qty=0 after release, so the top toggle includes
              -- them for recovery/reopen.
{_reserved_filter}
            {_where_extra}
            ORDER BY o.created_at DESC, o.order_no, ol.eye_side
        """, _params) or []
    except Exception as _se:
        st.error(f"Could not load stock lines: {_se}")
        return

    # ── Apply status filter ──────────────────────────────────────────────────
    _STATUS_FILTER_MAP = {
        "🔴 Pending":              lambda l: str(l.get("replenishment_status","")).upper() in ("PENDING",""),
        "📤 PO Sent (awaiting confirmation)": lambda l: str(l.get("replenishment_status","")).upper() == "PO_SENT",
        "⏳ In Procurement":       lambda l: str(l.get("replenishment_status","")).upper() == "ORDERED",
        "🚫 Discarded":            lambda l: str(l.get("replenishment_status","")).upper() == "DISCARDED",
        "❌ Supplier Cancelled":   lambda l: str(l.get("replenishment_status","")).upper() == "SUPPLIER_CANCELLED",
        "✅ Cancelled / Released": lambda l: str(l.get("replenishment_status","")).upper() == "CANCELLED",
    }
    if not _show_cancelled:
        rows = [r for r in rows if str(r.get("replenishment_status","")).upper() != "CANCELLED"]

    if _flt_status != "All" and _flt_status in _STATUS_FILTER_MAP:
        rows = [r for r in rows if _STATUS_FILTER_MAP[_flt_status](r)]

    if not rows:
        st.info("✅ No stock lines matching current filters.")
        return

    # ── Selection Action Bar ─────────────────────────────────────────────────
    # Read which lines were selected in previous render cycle
    _sel_lids = [
        r["line_id"] for r in rows
        if st.session_state.get(f"stk_chk_{r['line_id']}", False)
        and str(r.get("replenishment_status") or "").upper() in ("", "PENDING")
    ]
    _sel_rows = [r for r in rows if r["line_id"] in _sel_lids]

    if _sel_rows:
        with st.container(border=True):
            _tb1, _tb2 = st.columns([6, 1])
            _tb1.markdown(f"**📦 {len(_sel_rows)} line(s) selected** · scroll down to review and send ↓")
            with _tb2:
                if st.button("✕ Clear", key="stk_sel_clear", use_container_width=True):
                    for _sr in rows:
                        st.session_state.pop(f"stk_chk_{_sr['line_id']}", None)
                    st.rerun()

    # ── Select All toggle ─────────────────────────────────────────────────────
    _sa1, _sa2 = st.columns([1, 8])
    with _sa1:
        if st.checkbox("All", key="stk_sel_all", label_visibility="visible"):
            for _r in rows:
                st.session_state[f"stk_chk_{_r['line_id']}"] = (
                    str(_r.get("replenishment_status") or "").upper() in ("", "PENDING")
                )

    st.divider()

    # ── Group by order ────────────────────────────────────────────────────────
    from collections import defaultdict as _dds
    _orders = _dds(lambda: {"info": {}, "lines": []})
    for row in rows:
        _lp = row.get("lens_params") or {}
        if isinstance(_lp, str):
            try: _lp = _jss.loads(_lp)
            except: _lp = {}
        row["_lp"] = _lp
        _orders[row["order_id"]]["info"] = {
            "order_no":     row["order_no"],
            "patient_name": row["patient_name"],
            "status":       row["status"],
            "order_id":     row["order_id"],
        }
        _orders[row["order_id"]]["lines"].append(row)

    _stk_tbl_col, _stk_cap_col = st.columns([1, 8])
    with _stk_tbl_col:
        _stk_table_view = st.toggle("⊞", value=st.session_state.get("stk_tbl_view", True),
                                     key="stk_tbl_view", help="Compact card view")
    with _stk_cap_col:
        st.caption(f"{len(rows)} line(s) · {len(_orders)} order(s)")

    if _stk_table_view:
        # Build groups for shared card renderer
        _stk_groups = {}
        for _sk, _sd2 in _orders.items():
            _sl2 = _sd2["lines"]
            _stk_groups[_sk] = {
                "order_no":   _sd2["info"]["order_no"],
                "patient":    _sd2["info"]["patient_name"],
                "order_id":   _sk,
                "created_at": str(_sl2[0].get("created_at",""))[:10] if _sl2 else "",
                "lines":      _sl2,
            }
            # Attach power from lens_params if available
            for _l2 in _sl2:
                _lp2 = _l2.get("_lp") or {}
                if isinstance(_lp2, dict):
                    for _pk in ("sph","cyl","axis","add_power"):
                        if _lp2.get(_pk) is not None and not _l2.get(_pk):
                            _l2[_pk] = _lp2.get(_pk)
        _render_pipeline_cards(
            groups=_stk_groups,
            route_key="STOCK",
            stage_label_fn=lambda l: (
                "✅ Cancelled / Released" if str(l.get("replenishment_status","")).upper() == "CANCELLED"
                else "🚫 Discarded"       if str(l.get("replenishment_status","")).upper() == "DISCARDED"
                else "📤 PO Sent"         if str(l.get("replenishment_status","")).upper() == "PO_SENT"
                else "❌ Supplier Cancelled" if str(l.get("replenishment_status","")).upper() == "SUPPLIER_CANCELLED"
                else "⏳ Sent to Procurement" if str(l.get("replenishment_status","")).upper() == "ORDERED"
                else "🔴 Pending"
            ),
            stage_code_fn=lambda l: (
                "CANCELLED_RELEASED" if _normalize_repl_status(l.get("replenishment_status","")) == "CANCELLED_RELEASED"
                else "DISCARDED" if str(l.get("replenishment_status","")).upper() == "DISCARDED"
                else "PO_SENT"   if str(l.get("replenishment_status","")).upper() == "PO_SENT"
                else "SUPPLIER_CANCELLED" if str(l.get("replenishment_status","")).upper() == "SUPPLIER_CANCELLED"
                else "ORDERED"   if str(l.get("replenishment_status","")).upper() == "ORDERED"
                else "PENDING"
            ),
            open_billing_fn=_go_to_billing,
        )
        st.divider()

    def _stock_repl_write_status(line_id: str, lp: dict, status: str, supplier_id: str = "", supplier_name: str = "", po_no: str = "", supplier_ref: str = ""):
        from modules.sql_adapter import run_query as _rq_repl2, run_write as _rw_repl
        _db_lp_rows = _rq_repl2(
            "SELECT COALESCE(lens_params,'{}')::text AS lp FROM order_lines WHERE id=%(lid)s::uuid LIMIT 1",
            {"lid": line_id},
        ) or []
        if _db_lp_rows:
            try:
                _new_lp = _jss.loads(_db_lp_rows[0].get("lp") or "{}")
            except Exception:
                _new_lp = dict(lp or {})
        else:
            _new_lp = dict(lp or {})
        _new_lp["replenishment_status"] = status
        if supplier_id:
            _new_lp["replenishment_supplier_id"] = supplier_id
        if supplier_name:
            _new_lp["replenishment_supplier_name"] = supplier_name
        if po_no:
            _new_lp["replenishment_po_no"] = po_no
        if supplier_ref:
            _new_lp["supplier_confirmation_no"] = supplier_ref
        _tl = _new_lp.get("replenishment_timeline")
        if not isinstance(_tl, list):
            _tl = []
        _tl.append({
            "stage": status,
            "at": _dts.datetime.now().astimezone().isoformat(),
            "source": "Stock replenishment",
            "po_no": po_no or _new_lp.get("replenishment_po_no") or "",
            "supplier": supplier_name or _new_lp.get("replenishment_supplier_name") or "",
            "supplier_ref": supplier_ref or _new_lp.get("supplier_confirmation_no") or "",
        })
        _new_lp["replenishment_timeline"] = _tl
        _rw_repl(
            "UPDATE order_lines SET lens_params=%(lp)s::jsonb WHERE id=%(lid)s::uuid",
            {"lp": _jss.dumps(_new_lp), "lid": line_id},
        )
        try:
            _rw_repl(
                """
                UPDATE supplier_order_items
                SET item_status = %(st)s
                WHERE customer_line_id::text = %(lid)s
                """,
                {"st": status, "lid": line_id},
            )
        except Exception:
            pass

        try:
            from modules.procurement.procurement_ledger import ensure_queue_item
            if status == "ORDERED":
                ensure_queue_item(line_id, source="STOCK_PROCUREMENT", status="ORDERED")
        except Exception:
            pass

    def _stock_repl_create_po(line: dict, supplier_id: str, supplier_name: str) -> str:
        if not supplier_id:
            return ""
        from modules.sql_adapter import run_query as _rq_repl, run_write as _rw_repl
        _lid = str(line.get("line_id") or "")
        _existing = _rq_repl("""
            SELECT so.supplier_order_id AS po_number
            FROM supplier_order_items soi
            JOIN supplier_orders so ON so.id = soi.supplier_order_id
            WHERE soi.customer_line_id::text = %(lid)s
              AND so.status NOT IN ('CANCELLED','VOID')
              AND UPPER(COALESCE(soi.item_status,'')) NOT IN
                  ('CANCELLED','VOID','DISCARDED','NOT_NEEDED','SUPPLIER_CANCELLED')
            LIMIT 1
        """, {"lid": _lid})
        if _existing:
            return _existing[0].get("po_number") or ""
        try:
            from modules.db.order_number_registry import alloc_doc_number
            _po_no = alloc_doc_number("PURCHASE_ORDER")
        except Exception as _e:
            _po_no = f"PO/{_dts.datetime.now().strftime('%y%m%d%H%M%S')}"
        _qty = int(line.get("allocated_qty") or line.get("billed_qty") or line.get("quantity") or 1)
        _unit = float(line.get("unit_price") or line.get("purchase_rate") or 0)
        _sync_supplier_orders_id_sequence()
        _hdr = _rq_repl("""
            INSERT INTO supplier_orders (
                supplier_order_id, supplier_id, supplier_name, order_date, status,
                total_items, total_qty, total_value, created_by, created_at
            ) VALUES (
                %(po)s, %(sid)s::uuid, %(sname)s, CURRENT_DATE, 'SENT',
                1, %(qty)s, %(val)s, 'stock_replenishment', NOW()
            )
            RETURNING id
        """, {
            "po": _po_no, "sid": supplier_id, "sname": supplier_name,
            "qty": _qty, "val": round(_unit * _qty, 2),
        })
        if not _hdr:
            return ""
        _po_id = int(_hdr[0]["id"])
        _lp = line.get("_lp") or {}
        _rw_repl("""
            INSERT INTO supplier_order_items (
                supplier_order_id, item_no, product_id, product_name,
                eye_side, sph, cyl, axis, add_power,
                ordered_qty, received_qty, pending_qty, unit_price, total_price,
                customer_line_id, item_status
            ) VALUES (
                %(poid)s, 1, %(pid)s::uuid, %(pname)s,
                %(eye)s, %(sph)s, %(cyl)s, %(axis)s, %(add)s,
                %(qty)s, 0, %(qty)s, %(unit)s, %(total)s,
                %(lid)s, 'PENDING'
            )
            ON CONFLICT DO NOTHING
        """, {
            "poid": _po_id,
            "pid": line.get("product_id") or "00000000-0000-0000-0000-000000000000",
            "pname": str(line.get("product_name") or "")[:120],
            "eye": str(line.get("eye_side") or ""),
            "sph": line.get("sph") or _lp.get("sph"),
            "cyl": line.get("cyl") or _lp.get("cyl"),
            "axis": line.get("axis") or _lp.get("axis"),
            "add": line.get("add_power") or _lp.get("add_power"),
            "qty": _qty,
            "unit": _unit,
            "total": round(_unit * _qty, 2),
            "lid": _lid,
        })
        return _po_no

    def _stock_repl_create_bulk_po(lines: list[dict], supplier_id: str, supplier_name: str,
                                    qty_overrides: dict | None = None) -> tuple[str, list[str]]:
        """
        Create one supplier_order header for all selected stock replenishment
        lines, with one supplier_order_items row per customer order line.

        qty_overrides (Phase-3 2026-05-26): optional dict of {line_id: boxes}
        from the bulk-send panel where staff explicitly picked the box count.
        Falls back to box_size-aware derivation if not supplied.

        Existing PO-linked lines are skipped and returned as already handled.
        """
        if not supplier_id or not lines:
            return "", []
        from modules.sql_adapter import run_query as _rq_repl, run_write as _rw_repl

        _new_lines = []
        for _ln in lines:
            _lid = str(_ln.get("line_id") or "")
            if not _lid:
                continue
            _existing = _rq_repl("""
                SELECT so.supplier_order_id AS po_number
                FROM supplier_order_items soi
                JOIN supplier_orders so ON so.id = soi.supplier_order_id
                WHERE soi.customer_line_id::text = %(lid)s
                  AND so.status NOT IN ('CANCELLED','VOID')
                  AND UPPER(COALESCE(soi.item_status,'')) NOT IN
                      ('CANCELLED','VOID','DISCARDED','NOT_NEEDED','SUPPLIER_CANCELLED')
                LIMIT 1
            """, {"lid": _lid}) or []
            if not _existing:
                _new_lines.append(_ln)

        if not _new_lines:
            return "", []

        try:
            from modules.db.order_number_registry import alloc_doc_number
            _po_no = alloc_doc_number("PURCHASE_ORDER")
        except Exception:
            _po_no = f"PO/{_dts.datetime.now().strftime('%y%m%d%H%M%S')}"

        def _line_qty(_ln: dict) -> int:
            # Phase-3 (2026-05-26): qty must be in the SAME unit as unit_price.
            # unit_price for CL is per-BOX; for ophthalmic it is per-PIECE.
            # Earlier writers stored allocated_qty (PCS) regardless, which made
            # CL POs show ordered_qty=6 × unit_price=1603 = ₹9620 per eye —
            # 6× the correct value. Now we prefer the explicit override from
            # the bulk-send panel (which already knows box vs pc per row),
            # then fall back to box_size-aware derivation.
            _lid_q = str(_ln.get("line_id") or "")
            if qty_overrides and _lid_q in qty_overrides:
                try:
                    _v = int(float(qty_overrides[_lid_q]))
                    if _v > 0:
                        return _v
                except Exception:
                    pass
            _lp_q = _ln.get("_lp") or {}
            _bxs = _lp_q.get("replenishment_ordered_boxes")
            if _bxs not in (None, "", 0, "0"):
                try:
                    return max(1, int(float(_bxs)))
                except Exception:
                    pass
            # Fallback: derive from pcs and box_size (CEIL-style match to UI math).
            _pcs = int(_ln.get("allocated_qty")
                        or _ln.get("billed_qty")
                        or _ln.get("quantity") or 1)
            _bs_q = int(_ln.get("box_size") or 1)
            if _bs_q > 1:
                _q = _pcs // _bs_q + (1 if _pcs % _bs_q else 0)
                return max(1, _q)
            return max(1, _pcs)

        def _line_unit(_ln: dict) -> float:
            return float(_ln.get("inv_purchase_price") or _ln.get("purchase_rate") or _ln.get("unit_price") or 0)

        _total_qty = sum(_line_qty(_ln) for _ln in _new_lines)
        _total_val = round(sum(_line_qty(_ln) * _line_unit(_ln) for _ln in _new_lines), 2)
        _sync_supplier_orders_id_sequence()
        _hdr = _rq_repl("""
            INSERT INTO supplier_orders (
                supplier_order_id, supplier_id, supplier_name, order_date, status,
                total_items, total_qty, total_value, created_by, created_at
            ) VALUES (
                %(po)s, %(sid)s::uuid, %(sname)s, CURRENT_DATE, 'SENT',
                %(items)s, %(qty)s, %(val)s, 'stock_replenishment_bulk', NOW()
            )
            RETURNING id
        """, {
            "po": _po_no,
            "sid": supplier_id,
            "sname": supplier_name,
            "items": len(_new_lines),
            "qty": _total_qty,
            "val": _total_val,
        }) or []
        if not _hdr:
            return "", []
        _po_id = int(_hdr[0]["id"])
        _saved_lids = []
        for _item_no, _ln in enumerate(_new_lines, 1):
            _lid = str(_ln.get("line_id") or "")
            _lp = _ln.get("_lp") or {}
            _qty = _line_qty(_ln)
            _unit = _line_unit(_ln)
            _rw_repl("""
                INSERT INTO supplier_order_items (
                    supplier_order_id, item_no, product_id, product_name,
                    eye_side, sph, cyl, axis, add_power,
                    ordered_qty, received_qty, pending_qty, unit_price, total_price,
                    customer_line_id, item_status
                ) VALUES (
                    %(poid)s, %(item_no)s, %(pid)s::uuid, %(pname)s,
                    %(eye)s, %(sph)s, %(cyl)s, %(axis)s, %(add)s,
                    %(qty)s, 0, %(qty)s, %(unit)s, %(total)s,
                    %(lid)s, 'PENDING'
                )
                ON CONFLICT DO NOTHING
            """, {
                "poid": _po_id,
                "item_no": _item_no,
                "pid": _ln.get("product_id") or "00000000-0000-0000-0000-000000000000",
                "pname": str(_ln.get("product_name") or "")[:120],
                "eye": str(_ln.get("eye_side") or ""),
                "sph": _ln.get("sph") or _lp.get("sph"),
                "cyl": _ln.get("cyl") or _lp.get("cyl"),
                "axis": _ln.get("axis") or _lp.get("axis"),
                "add": _ln.get("add_power") or _lp.get("add_power"),
                "qty": _qty,
                "unit": _unit,
                "total": round(_unit * _qty, 2),
                "lid": _lid,
            })
            _saved_lids.append(_lid)
        return _po_no, _saved_lids

    for oid, odata in _orders.items():
        info  = odata["info"]
        lines = odata["lines"]
        _acked = sum(1 for l in lines if l.get("pa_price") and float(l.get("pa_price") or 0) > 0)
        _total = len(lines)
        _all_acked = _acked == _total

        with st.container(border=True):
            # ── Order header ──────────────────────────────────────────────
            _h1, _h2 = st.columns([4, 2])
            with _h1:
                st.markdown(
                    f"<div style='padding:2px 0'>"
                    f"<span style='font-weight:800;color:#e2e8f0'>"
                    f"{'✅' if _all_acked else '📦'} {info['order_no']}</span>"
                    f" <span style='color:#64748b;font-size:0.82rem'>"
                    f"— {info['patient_name']}</span><br>"
                    f"<span style='color:#475569;font-size:0.72rem'>"
                    f"{_acked}/{_total} acknowledged</span></div>",
                    unsafe_allow_html=True
                )
            with _h2:
                _ord_date = str(odata["lines"][0].get("created_at",""))[:10]
                st.caption(f"📅 {_ord_date}")

            # ── Per-line purchase acknowledgement ─────────────────────────
            for line in lines:
                _lid      = str(line["line_id"])
                _eye      = str(line.get("eye_side") or "").upper()
                _pname    = str(line.get("product_name") or "").split(" | ")[0]
                _qty      = int(line.get("allocated_qty") or line.get("billed_qty") or line.get("quantity") or 1)
                _lp        = line.get("_lp") or {}
                # Power details from order_line columns or lens_params
                _sph_v = line.get('sph'); _cyl_v = line.get('cyl')
                _ax_v  = line.get('axis'); _add_v = line.get('add_power')
                if _sph_v is None: _sph_v = _lp.get('sph')
                if _cyl_v is None: _cyl_v = _lp.get('cyl')
                _pwr_p = []
                if _sph_v not in (None,''): _pwr_p.append(f'SPH {float(_sph_v):+.2f}')
                if _cyl_v not in (None,'',0,'0'): _pwr_p.append(f'CYL {float(_cyl_v):+.2f}')
                if _ax_v not in (None,'',0): _pwr_p.append(f'AX {int(float(_ax_v))}')
                if _add_v not in (None,'',0): _pwr_p.append(f'ADD {float(_add_v):+.2f}')
                _pwr_str = '  '.join(_pwr_p)
                # Preferred supplier from product_supplier_map
                _pref_sid    = str(line.get('pref_sup_id')    or '')
                _pref_sname  = str(line.get('pref_sup_name')  or '')
                _pref_smob   = str(line.get('pref_sup_mobile') or '')
                _pref_semail = str(line.get('pref_sup_email')  or '')
                # Use DB-extracted field first (more reliable), fallback to lens_params
                _repl_status = _normalize_repl_status(
                    line.get("replenishment_status") or
                    _lp.get("replenishment_status")
                )

                _eye_lbl = f"👁 {_eye}" if _eye in ("R","L") else "🖼"
                _status_color = (
                    "#22c55e" if _is_cancelled_status(_repl_status)
                    else "#64748b" if _repl_status == "DISCARDED"
                    else "#38bdf8" if _repl_status == "PO_SENT"
                    else "#f97316" if _repl_status == "SUPPLIER_CANCELLED"
                    else "#3b82f6" if _repl_status == "ORDERED"
                    else "#ef4444"   # PENDING
                )
                _status_label = (
                    "🚫 Not Needed"        if _is_cancelled_status(_repl_status)
                    else "🗑️ Discarded"    if _repl_status == "DISCARDED"
                    else "📤 PO Sent"       if _repl_status == "PO_SENT"
                    else "❌ Supplier Cancelled" if _repl_status == "SUPPLIER_CANCELLED"
                    else "⏳ In Procurement" if _repl_status == "ORDERED"
                    else "⚠️ Pending"
                )
                _sup_name  = str(line.get("supplier_name") or _lp.get("replenishment_supplier_name") or "")

                # Phase-3 (2026-05-25): build coating/index/treatment summary
                # for in-card display. Read DB-extracted columns first, fall back
                # to lens_params keys so partially-populated rows still render.
                _coating_v = str(line.get("coating")    or _lp.get("coating")
                                 or _lp.get("COATING")  or "").strip()
                _index_v   = str(line.get("lens_index") or _lp.get("lens_index")
                                 or _lp.get("index")    or _lp.get("INDEX") or "").strip()
                _treat_v   = str(line.get("treatment")  or _lp.get("treatment")
                                 or _lp.get("TREATMENT") or "").strip()
                _spec_bits = []
                if _index_v:   _spec_bits.append(f"Index {_index_v}")
                if _coating_v: _spec_bits.append(_coating_v)
                if _treat_v and _treat_v.lower() not in ("clear", "none", "-", "—"):
                    _spec_bits.append(_treat_v)
                _spec_str = " · ".join(_spec_bits)

                # Checkbox for multi-select
                _chk_key = f"stk_chk_{_lid}"
                _can_select_for_po = _repl_status in ("", "PENDING")
                if not _can_select_for_po:
                    st.session_state.pop(_chk_key, None)
                st.checkbox(
                    "Select procurement line",
                    key=_chk_key,
                    value=(st.session_state.get(_chk_key, False) and _can_select_for_po),
                    disabled=not _can_select_for_po,
                    label_visibility="collapsed",
                    help=None if _can_select_for_po else "Already sent/ordered. Enter supplier ref or use cancelled/recover flow.",
                )
                _ref_saved = str(_lp.get("supplier_confirmation_no") or "")
                _c_reason  = str(_lp.get("cancel_reason") or "")
                st.markdown(
                    f"<div style='border:1px solid #1e293b;"
                    f"border-left:4px solid {_status_color};"
                    f"border-radius:6px;padding:8px 12px;margin:4px 0'>"
                    f"<div style='display:flex;justify-content:space-between'>"
                    f"<span style='color:#111827;font-weight:900;font-size:0.92rem'>{_eye_lbl} {_pname}</span>"
                    f"<span style='color:{_status_color};font-size:0.72rem;"
                    f"font-weight:700'>{_status_label}</span></div>"
                    + (f"<div style='color:#38bdf8;font-size:0.7rem;font-family:monospace'>{_pwr_str}</div>" if _pwr_str else "")
                    + (f"<div style='color:#a78bfa;font-size:0.7rem'>{_spec_str}</div>" if _spec_str else "")
                    + (f"<div style='color:#374151;font-size:0.74rem;font-weight:700'>Supplier: {_sup_name}"
                       + (f"  ·  PO: {_lp.get('replenishment_po_no')}" if _lp.get("replenishment_po_no") else "")
                       + "</div>" if _sup_name else "")
                    + (f"<div style='color:#22c55e;font-size:0.7rem;font-weight:600'>📋 Ref: {_ref_saved}</div>" if _ref_saved else "")
                    + "</div>",
                    unsafe_allow_html=True
                )

                # Replenishment actions moved to bottom bulk-selection panel.
                # Individual lines remain compact/read-only in Active Pipeline view.
                _po_note = f" · {_lp.get('replenishment_po_no')}" if _lp.get("replenishment_po_no") else ""
                if _repl_status in ("PO_SENT", "ORDERED", "DISCARDED", "SUPPLIER_CANCELLED") or _is_cancelled_status(_repl_status):
                    st.caption(f"🔁 Replenishment: {_repl_status}{_po_note}")
                if _repl_status == "PO_SENT":
                    # Auto-expand the Save Ref UI for PO_SENT lines so staff
                    # don't miss the next step. Previously this was inside a
                    # collapsed expander and staff didn't realize the line was
                    # waiting for them. The line cannot reach Procurement Queue
                    # until supplier ref is saved (which promotes to ORDERED).
                    st.markdown(
                        "<div style='background:#1f1610;border:1px solid #a16207;"
                        "border-left:4px solid #f59e0b;border-radius:6px;"
                        "padding:6px 10px;margin:6px 0;font-size:0.78rem'>"
                        "<span style='color:#fbbf24;font-weight:700'>⏳ Awaiting supplier confirmation</span>"
                        "<span style='color:#94a3b8;margin-left:6px'>"
                        "Enter supplier ref below → line moves to Procurement Queue."
                        "</span></div>",
                        unsafe_allow_html=True,
                    )
                    with st.expander("📋 Save supplier reference → Procurement Queue", expanded=True):
                        st.caption("Enter the supplier order / confirmation number. Saving moves this line to Procurement Queue.")
                        _ref_key = f"stk_sup_ref_{_lid}"
                        _ref_val = st.text_input(
                            "Supplier order / confirmation no.",
                            value=str(_lp.get("supplier_confirmation_no") or ""),
                            key=_ref_key,
                        ).strip()
                        if st.button(
                            "✅ Save Ref → Procurement Queue",
                            key=f"stk_sup_ref_save_{_lid}",
                            type="primary",
                            use_container_width=True,
                            disabled=not bool(_ref_val),
                        ):
                            # Phase-3 (2026-05-26): cascade the supplier confirmation
                            # to all order_lines that share the SAME PO number.
                            # Previously, R and L eyes — each having their own
                            # "Save Ref" button — required two clicks; missing
                            # the second one left that eye at PO_SENT and
                            # invisible to the procurement queue (queue filters
                            # by replenishment_status='ORDERED').
                            _this_po = str(_lp.get("replenishment_po_no") or "")
                            _sup_id  = str(_lp.get("replenishment_supplier_id")
                                            or line.get("pref_sup_id") or "")
                            _sup_nm  = str(_lp.get("replenishment_supplier_name")
                                            or line.get("pref_sup_name")
                                            or line.get("supplier_name") or "")

                            # Collect sibling line_ids: current + others on same PO
                            # that are still PO_SENT (don't disturb already-ORDERED).
                            _siblings: list[tuple[str, dict]] = [(_lid, _lp)]
                            if _this_po:
                                try:
                                    from modules.sql_adapter import run_query as _rq_cas
                                    _sib_rows = _rq_cas("""
                                        SELECT ol.id::text AS lid, ol.lens_params AS lp
                                          FROM order_lines ol
                                         WHERE COALESCE(ol.is_deleted, FALSE) = FALSE
                                           AND ol.lens_params->>'replenishment_po_no' = %(po)s
                                           AND ol.id::text <> %(self)s
                                           AND UPPER(COALESCE(ol.lens_params->>'replenishment_status','')) = 'PO_SENT'
                                    """, {"po": _this_po, "self": _lid}) or []
                                    import json as _js_cas
                                    for _sr_cas in _sib_rows:
                                        _slp = _sr_cas.get("lp")
                                        if isinstance(_slp, str):
                                            try: _slp = _js_cas.loads(_slp)
                                            except Exception: _slp = {}
                                        _siblings.append((str(_sr_cas["lid"]), _slp or {}))
                                except Exception as _cas_e:
                                    import logging as _casl
                                    _casl.getLogger(__name__).warning(
                                        "[save_ref] sibling lookup failed: %s", _cas_e
                                    )

                            _cas_n = 0
                            for _cas_lid, _cas_lp in _siblings:
                                try:
                                    _stock_repl_write_status(
                                        _cas_lid,
                                        _cas_lp,
                                        "ORDERED",
                                        _sup_id,
                                        _sup_nm,
                                        _this_po,
                                        _ref_val,
                                    )
                                    _cas_n += 1
                                except Exception as _ws_e:
                                    import logging as _wsl
                                    _wsl.getLogger(__name__).warning(
                                        "[save_ref] write failed for %s: %s",
                                        _cas_lid, _ws_e
                                    )

                            # Preserve current tab on rerun (Issue #2 hot path).
                            try:
                                _cur_panel = st.session_state.get("prod_lazy_panel", "")
                                if _cur_panel:
                                    st.session_state["_prod_lazy_panel_next"] = _cur_panel
                            except Exception:
                                pass
                            _extra_eyes = _cas_n - 1
                            _msg = "✅ Supplier ref saved. Line moved to Procurement Queue."
                            if _extra_eyes > 0:
                                _msg = (
                                    f"✅ Supplier ref saved on {_cas_n} eyes "
                                    f"(this + {_extra_eyes} sibling on same PO). "
                                    "All moved to Procurement Queue."
                                )
                            st.success(_msg)
                            st.rerun()


    # ═══════════════════════════════════════════════════════════════════════════
    # ── ORDER PANEL (bottom) — appears only when lines are selected ────────────
    # Flow: select ↑ → scroll down → club → adjust qty → pick send mode → confirm
    # ═══════════════════════════════════════════════════════════════════════════
    if _sel_rows:
        st.divider()
        # Replenishment Order panel — collapsed by default so the line list
        # stays the focus. User scrolls down, expands once they've finished
        # picking lines, configures supplier + qty + send mode, fires.
        with st.expander(
            f"📦 Replenishment Order  ·  {len(_sel_rows)} line(s) selected  "
            f"·  click to review & send ↓",
            expanded=False,
        ):
            st.caption(f"{len(_sel_rows)} line(s) selected · review, adjust and send below")

            # ── 1. Supplier ───────────────────────────────────────────────────
            st.markdown("**1 · Supplier**")
            _all_sup_list = _q("""
                SELECT id::text AS id, party_name,
                       COALESCE(mobile,'')  AS mobile,
                       COALESCE(email,'')   AS email
                FROM parties
                WHERE UPPER(COALESCE(party_type,'')) IN ('SUPPLIER','VENDOR','LAB','EXTERNAL_LAB')
                  AND COALESCE(is_active,TRUE) = TRUE
                ORDER BY party_name
            """, {}) or []
            _all_sup_ids   = [""] + [s["id"] for s in _all_sup_list]
            _all_sup_by_id = {"": {"party_name": "— Select Supplier —", "mobile": "", "email": ""}}
            _all_sup_by_id.update({s["id"]: s for s in _all_sup_list})
            _auto_sup_id  = str(_sel_rows[0].get("pref_sup_id") or
                                (_sel_rows[0].get("_lp") or {}).get("replenishment_supplier_id") or "")
            _auto_sup_idx = _all_sup_ids.index(_auto_sup_id) if _auto_sup_id in _all_sup_ids else 0
            _chosen_sup_id = st.selectbox(
                "Supplier", _all_sup_ids, index=_auto_sup_idx,
                format_func=lambda x: _all_sup_by_id.get(x, {}).get("party_name", x),
                key="stk_multi_sup",
            )
            _chosen_sup       = _all_sup_by_id.get(_chosen_sup_id, {})
            _chosen_sup_name  = str(_chosen_sup.get("party_name") or "")
            _chosen_sup_mob   = str(_chosen_sup.get("mobile") or "")
            _chosen_sup_email = str(_chosen_sup.get("email") or "")

            st.divider()

            # ── 2. Club & edit qty ────────────────────────────────────────────
            st.markdown("**2 · Review & adjust quantities** *(boxes for CL packs · pcs for ophthalmic / loose)*")
            st.caption("Same product + eye + power are merged. Edit qty before sending.")

            # pcs → boxes: ceiling so we never order short
            def _to_boxes_op(qty_pcs: int, box_size: int) -> int:
                bs = max(1, int(box_size or 1))
                return max(1, (int(qty_pcs) + bs - 1) // bs)

            def _pwr_label(sr: dict) -> str:
                _lp2 = sr.get("_lp") or {}
                parts = []
                def _f(v, fmt):
                    try:
                        fv = float(v)
                        if fmt == "sph": return f"SPH {fv:+.2f}"
                        if fmt == "cyl" and abs(fv) > 0.01: return f"CYL {fv:+.2f}"
                        if fmt == "ax"  and int(fv) != 0:   return f"AX {int(fv)}°"
                        if fmt == "add" and fv > 0:          return f"ADD {fv:+.2f}"
                    except: pass
                    return ""
                _s = sr.get("sph")  or _lp2.get("sph");  r = _f(_s,"sph");  r and parts.append(r)
                _c = sr.get("cyl")  or _lp2.get("cyl");  r = _f(_c,"cyl");  r and parts.append(r)
                _a = sr.get("axis") or _lp2.get("axis"); r = _f(_a,"ax");   r and parts.append(r)
                _d = sr.get("add_power") or _lp2.get("add_power"); r = _f(_d,"add"); r and parts.append(r)
                def _named(label, *vals):
                    for _v in vals:
                        _s = str(_v or "").strip()
                        if _s:
                            return f"{label} {_s}"
                    return f"{label} —"
                parts.append(_named("DIA", sr.get("diameter"), _lp2.get("diameter"), _lp2.get("dia"), _lp2.get("DIA")))
                parts.append(_named("Corridor", sr.get("corridor"), _lp2.get("corridor"), _lp2.get("CORRIDOR")))
                parts.append(_named("Fit Ht", sr.get("fitting_height"), _lp2.get("fitting_height"), _lp2.get("fitting"), _lp2.get("FITTING_HEIGHT")))
                return "  ".join(parts) or "Plano"

            def _supplier_spec_label(sr: dict) -> str:
                """Supplier-facing product/spec for stock replenishment messages."""
                _pid = str(sr.get("product_id") or "")
                _bits = []
                if _chosen_sup_id and _pid:
                    try:
                        from modules.backoffice.supplier_product_map_ui import get_supplier_product_name
                        _spm = get_supplier_product_name(_pid, _chosen_sup_id)
                        if _spm.get("mapped"):
                            if _spm.get("supplier_brand"):
                                _bits.append(str(_spm.get("supplier_brand")))
                            if _spm.get("supplier_product_name"):
                                _bits.append(str(_spm.get("supplier_product_name")))
                            if _spm.get("supplier_index"):
                                _bits.append(f"Index {_spm.get('supplier_index')}")
                            if _spm.get("supplier_coating"):
                                _bits.append(str(_spm.get("supplier_coating")))
                            if _spm.get("supplier_treatment") and str(_spm.get("supplier_treatment")) != "Clear":
                                _bits.append(str(_spm.get("supplier_treatment")))
                    except Exception:
                        _bits = []
                if not _bits:
                    # Phase-3 (2026-05-25): when no supplier-product mapping exists,
                    # fall back to product_name plus the actual lens spec — coating,
                    # index, treatment — so WhatsApp / Excel / Print / Call PO outputs
                    # include them. Previously the bare product name leaked through
                    # and suppliers had no idea which index/coating to ship.
                    _bits = [str(sr.get("product_name") or "").split(" | ")[0]]
                    _lp_fb = sr.get("_lp") or sr.get("lens_params") or {}
                    if isinstance(_lp_fb, str):
                        try:
                            import json as _json_fb
                            _lp_fb = _json_fb.loads(_lp_fb)
                        except Exception:
                            _lp_fb = {}
                    _idx_fb = str(sr.get("lens_index") or _lp_fb.get("lens_index")
                                  or _lp_fb.get("index") or _lp_fb.get("INDEX") or "").strip()
                    _coat_fb = str(sr.get("coating") or _lp_fb.get("coating")
                                   or _lp_fb.get("COATING") or "").strip()
                    _tr_fb = str(sr.get("treatment") or _lp_fb.get("treatment")
                                 or _lp_fb.get("TREATMENT") or "").strip()
                    if _idx_fb:  _bits.append(f"Index {_idx_fb}")
                    if _coat_fb: _bits.append(_coat_fb)
                    if _tr_fb and _tr_fb.lower() not in ("clear", "none", "-", "—"):
                        _bits.append(_tr_fb)
                return " · ".join([b for b in _bits if b])

            # Robust positive-int reader: Python `or`-chain treats string "0" as
            # truthy so a line with quantity="0" silently contributed 0 boxes.
            # This explicit loop skips any value that is zero/empty/None.
            def _qty_pos_int(*vals, default: int = 1) -> int:
                for _v in vals:
                    if _v in (None, "", "None"):
                        continue
                    try:
                        _n = int(float(_v))
                    except Exception as _e:
                        continue
                    if _n > 0:
                        return _n
                return int(default)

            # Phase-3 (2026-05-26): Model-2 compound unit labels.
            # Ophthalmic lenses (unit=PCS, box_size=1) → "2 pcs (2 boxes)" is silly,
            # so for box_size==1 we just show "2 pcs". For real boxed items
            # (CL 6PK, 30PK etc.) we show "1 box (6 pcs)".
            # Reads each clubbed line's actual `unit` and `box_size` so spectacle
            # and CL render correctly side-by-side in the same panel.
            def _unit_compound(boxes: int, pcs: int, box_size: int,
                               unit: str = "") -> str:
                """Compound 'N box (M pcs)' or 'N pcs' depending on pack size."""
                try:
                    _b  = int(boxes or 0)
                    _p  = int(pcs   or 0)
                    _bs = int(box_size or 1)
                except Exception:
                    return f"{boxes} {unit or 'unit'}"
                _u = (unit or "").strip().lower()
                _is_box = _bs > 1
                if _is_box:
                    _box_word = "box" if _b == 1 else "boxes"
                    _pcs_word = "pc"  if _p == 1 else "pcs"
                    return f"{_b} {_box_word} ({_p} {_pcs_word})"
                _pcs_word = "pc" if _p == 1 else "pcs"
                return f"{_p} {_pcs_word}"

            # Deduplicate: (product_name, eye, power) → {boxes, pcs, line_ids, details}
            _club_keys  = []
            _club: dict = {}
            _audit_rows = []
            for _sr in _sel_rows:
                _pn   = str(_sr.get("product_name") or "").split(" | ")[0]
                _spn  = _supplier_spec_label(_sr)
                _eye  = str(_sr.get("eye_side") or "").upper()
                _pwr  = _pwr_label(_sr)
                _bs   = _qty_pos_int(_sr.get("box_size"), default=1)
                _unit_raw = str(_sr.get("unit") or "").strip()
                # Priority: quantity → qty → billing_qty → allocated_qty → billed_qty
                # (allocated_qty is a partial reservation; quantity is the full demand)
                _qpcs = _qty_pos_int(
                    _sr.get("quantity"),
                    _sr.get("qty"),
                    _sr.get("billing_qty"),
                    _sr.get("allocated_qty"),
                    _sr.get("billed_qty"),
                    default=1,
                )
                _boxes = _to_boxes_op(_qpcs, _bs)
                _lid_s = str(_sr.get("line_id") or _sr.get("id") or "")
                _ord_s = str(_sr.get("order_no") or "")
                _ck    = (_spn, _eye, _pwr)
                _detail = {
                    "order_no": _ord_s, "line_id": _lid_s,
                    "eye": _eye, "pwr": _pwr, "product": _pn,
                    "supplier_product": _spn,
                    "pcs": _qpcs, "box_size": _bs, "boxes": _boxes,
                    "unit": _unit_raw,
                }
                _audit_rows.append(_detail)
                if _ck not in _club:
                    _club_keys.append(_ck)
                    _club[_ck] = {"pname": _pn, "eye": _eye, "pwr": _pwr,
                                  "supplier_product": _spn,
                                  "box_size": _bs,
                                  "boxes": _boxes,
                                  "pcs": _qpcs,
                                  "unit": _unit_raw,
                                  "line_ids": [_lid_s],
                                  "lines": [_sr],
                                  "details": [_detail]}
                else:
                    _club[_ck]["boxes"] += _boxes
                    _club[_ck]["pcs"]   += _qpcs
                    _club[_ck]["line_ids"].append(_lid_s)
                    _club[_ck]["lines"].append(_sr)
                    _club[_ck]["details"].append(_detail)

            # Per-line audit — always visible so staff can verify DB quantities
            with st.expander("🔎 Individual selected lines / DB quantity audit", expanded=True):
                for _ar in _audit_rows:
                    _ar_label = _unit_compound(
                        _ar.get('boxes', 0), _ar.get('pcs', 0),
                        _ar.get('box_size', 1), _ar.get('unit', ''),
                    )
                    st.caption(
                        f"{_ar['order_no'] or '—'} · {_ar['eye']} · {_ar['product']} · "
                        f"{_ar['pwr']} · {_ar['pcs']} PCS ÷ {_ar['box_size']} = "
                        f"{_ar_label} · "
                        f"line {_ar['line_id'][:8]}…"
                    )
                # Total: aggregate per-unit type. Mixed audits (CL + spectacle)
                # would have heterogeneous units; show each unit's total.
                _box_total = sum(int(x.get('boxes', 0)) for x in _audit_rows
                                  if int(x.get('box_size', 1) or 1) > 1)
                _pcs_total = sum(int(x.get('pcs', 0))   for x in _audit_rows
                                  if int(x.get('box_size', 1) or 1) <= 1)
                _tot_parts = []
                if _box_total:
                    _tot_parts.append(f"{_box_total} box{'es' if _box_total != 1 else ''}")
                if _pcs_total:
                    _tot_parts.append(f"{_pcs_total} pc{'s' if _pcs_total != 1 else ''}")
                st.markdown(
                    f"**Total selected:** {' + '.join(_tot_parts) if _tot_parts else '—'}"
                )

            _edit_boxes: dict = {}
            for _ck in _club_keys:
                _cl   = _club[_ck]
                _cl_bs   = int(_cl.get("box_size", 1) or 1)
                _cl_unit = str(_cl.get("unit", "") or "")
                _cl_is_box = _cl_bs > 1
                _ql, _qr = st.columns([6, 2])
                with _ql:
                    _dupe = f" <span style='color:#64748b;font-size:10px'>×{len(_cl['line_ids'])} orders</span>" if len(_cl["line_ids"]) > 1 else ""
                    _eye_lbl = f"👁 {_cl['eye']}" if _cl["eye"] in ("R","L") else _cl["eye"]
                    _send_name = _cl.get("supplier_product") or _cl["pname"]
                    _detail_txt = " · ".join(
                        f"{d['order_no'] or '—'}: "
                        + _unit_compound(d.get('boxes', 0), d.get('pcs', 0),
                                          d.get('box_size', 1), d.get('unit', ''))
                        for d in _cl.get("details", [])
                    )
                    st.markdown(
                        f"<div style='padding:5px 2px'>"
                        f"<span style='color:#e2e8f0;font-weight:600'>{_eye_lbl}  {_send_name}</span>{_dupe}"
                        f"<br><span style='color:#38bdf8;font-family:monospace;font-size:11px'>{_cl['pwr']}</span>"
                        f"<br><span style='color:#94a3b8;font-size:11px'>{_detail_txt}</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                with _qr:
                    _safe_key = "_".join(str(x)[:12] for x in _ck).replace(" ","_").replace("+","p").replace("-","m").replace(".","d")
                    # Hash of line_ids + boxes forces Streamlit to discard its
                    # cached value whenever the selection or computed qty changes
                    # (this was the root cause of R -2.00 staying at 1 after two
                    # matching orders were selected).
                    _qty_sig = str(abs(hash(tuple(
                        (d['line_id'], d['pcs'], d['boxes'])
                        for d in _cl.get('details', [])
                    ))))[:10]
                    # Phase-3 (2026-05-26): switch input label + default value to
                    # PCS for ophthalmic / loose stock so staff edit the number
                    # they actually think in. CL packs keep editing in boxes.
                    if _cl_is_box:
                        _input_label = "Boxes"
                        _input_default = int(_cl["boxes"])
                        _input_help    = f"1 box = {_cl_bs} pcs · computed from {_cl.get('pcs', 0)} PCS"
                    else:
                        _input_label = "Pcs"
                        _input_default = int(_cl.get("pcs", _cl.get("boxes", 1)))
                        _input_help    = f"Loose / per-piece item (box_size=1) · {_cl.get('pcs', 0)} PCS required"
                    _edit_val = st.number_input(
                        _input_label, value=_input_default, min_value=0, step=1,
                        key=f"stk_ob_{_safe_key}_{_qty_sig}",
                        label_visibility="collapsed",
                        help=_input_help,
                    )
                    # Internally _edit_boxes still keys by clubbing-key and stores
                    # whatever the user typed. For non-box items the number IS
                    # pcs — downstream conversions multiply by box_size (=1) so
                    # the math stays correct without a second mapping table.
                    _edit_boxes[_ck] = _edit_val

            st.divider()

            # ── 3. Reference number ───────────────────────────────────────────
            # Phase-3 (2026-05-26): PO numbers are now auto-allocated from the
            # PURCHASE_ORDER series in order_number_registry — same gap-free
            # transactional mechanism that issues R/CH/INV/PAY numbers.
            #
            # The real allocation happens at Confirm time inside
            # _stock_repl_create_bulk_po → alloc_doc_number("PURCHASE_ORDER")
            # so the counter is only consumed on a successful commit.
            #
            # Here we just PEEK the next number for display in the outgoing
            # message. Staff can no longer type — the field is read-only.
            # If two terminals confirm POs simultaneously the second gets
            # the actual-next number (the previewed one was an estimate).
            st.markdown("**3 · PO Number** *(auto-allocated on Confirm)*")
            try:
                from modules.sql_adapter import run_query as _rq_po_peek
                _peek = _rq_po_peek("""
                    SELECT prefix, fiscal_year, COALESCE(last_number, 0) + 1 AS next_n
                    FROM order_number_registry
                    WHERE series = 'PURCHASE_ORDER'
                    LIMIT 1
                """, {}) or []
                if _peek:
                    _pf = _peek[0]
                    _ref_no = f"{_pf.get('prefix','PO')}/{_pf.get('fiscal_year','')}/" \
                              f"{int(_pf.get('next_n', 1)):04d}"
                else:
                    _ref_no = ""
            except Exception as _pe_po:
                _ref_no = ""
            if _ref_no:
                st.text_input(
                    "PO No (auto)",
                    value=_ref_no,
                    key="stk_sel_ref_no_preview",
                    disabled=True,
                    label_visibility="collapsed",
                    help="Auto-allocated from PURCHASE_ORDER series. "
                         "Final number is assigned on Confirm — gap-free, transactional.",
                )
                st.caption(f"➜ Will be saved as **{_ref_no}** when you Confirm. "
                           "If another terminal confirms first, your PO gets the next number after that.")
            else:
                st.warning(
                    "⚠️ PO numbering registry not initialised. PO will use a "
                    "timestamp fallback. Run migration 0xx to seed PURCHASE_ORDER series."
                )

            st.divider()

            # ── 4. Send via ───────────────────────────────────────────────────
            st.markdown("**4 · Send order via**")

            # Build clubbed message text once (shared across all tabs)
            import urllib.parse as _up2
            _today_str2 = _dts.datetime.now().strftime("%d/%m/%Y")
            _msg_parts = [
                "*Stock Replenishment Order*",
                f"Date: {_today_str2}",
                f"Ref: {_ref_no.strip() or '—'}",
                f"To: {_chosen_sup_name or '—'}",
                "",
            ]
            _by_prod2: dict = {}
            _prod_order2: list = []
            for _ck in _club_keys:
                _cl  = _club[_ck]
                _pn  = _cl.get("supplier_product") or _cl["pname"]
                _msg_bs   = int(_cl.get("box_size", 1) or 1)
                _msg_unit = str(_cl.get("unit", "") or "")
                _msg_is_box = _msg_bs > 1
                _qty_now = int(_edit_boxes.get(_ck, _cl["boxes"]) or 0)
                # Phase-3 (2026-05-26): qty_now == "boxes" for CL, "pcs" for ophthalmic.
                # Convert to pcs for the compound label calc.
                if _msg_is_box:
                    _msg_pcs   = _qty_now * _msg_bs
                    _msg_boxes = _qty_now
                else:
                    _msg_pcs   = _qty_now
                    _msg_boxes = _qty_now   # synonymous when box_size=1
                _qty_label = _unit_compound(_msg_boxes, _msg_pcs, _msg_bs, _msg_unit)
                if _pn not in _by_prod2:
                    _prod_order2.append(_pn)
                    _by_prod2[_pn] = []
                _by_prod2[_pn].append(
                    f"  {'R' if _cl['eye']=='R' else 'L' if _cl['eye']=='L' else _cl['eye']}:"
                    f"  {_cl['pwr']}  —  {_qty_label}"
                )
            for _pn2 in _prod_order2:
                _msg_parts.append(f"*{_pn2}*")
                _msg_parts.extend(_by_prod2[_pn2])
                _msg_parts.append("")
            _msg_parts += ["Please confirm availability and dispatch date.", "Thank you."]
            _full_msg = "\n".join(_msg_parts)
            _plain_msg = _full_msg.replace("*", "")
            _msg_key_sig = str(abs(hash(_full_msg)))[:12]

            _tw, _tm, _ts, _te, _tc, _tp = st.tabs([
                "📲 WhatsApp", "📧 Mail", "📄 Prescribed Sheet", "📊 Excel", "📞 Telephonic", "🧾 Programed Excel"
            ])

            with _tw:
                _wa_msg_key = f"stk_wa_msg_{_msg_key_sig}"
                if not str(st.session_state.get(_wa_msg_key) or "").strip():
                    st.session_state[_wa_msg_key] = _full_msg
                _wa_text = st.text_area("Message", value=_full_msg, height=220,
                                        key=_wa_msg_key, label_visibility="collapsed")
                _wa_mob2 = "".join(c for c in _chosen_sup_mob if c.isdigit())
                if _wa_mob2 and not _wa_mob2.startswith("91"):
                    _wa_mob2 = "91" + _wa_mob2
                _wa_url2 = (f"https://wa.me/{_wa_mob2}?text={_up2.quote(_wa_text, safe='')}"
                            if _wa_mob2 else f"https://wa.me/?text={_up2.quote(_wa_text, safe='')}")
                st.caption(f"📱 {_chosen_sup_mob}" if _chosen_sup_mob else "⚠️ No number — link opens chat picker")
                st.link_button("📲 Open WhatsApp", url=_wa_url2, use_container_width=True)

            with _tm:
                _mail_sub = f"Stock Replenishment Order — {_ref_no.strip() or _today_str2}"
                _mail_body_key = f"stk_mail_body_{_msg_key_sig}"
                if not str(st.session_state.get(_mail_body_key) or "").strip():
                    st.session_state[_mail_body_key] = _plain_msg
                _mail_body = st.text_area("Body", value=_plain_msg, height=220,
                                          key=_mail_body_key, label_visibility="collapsed")
                _mail_sub_e = st.text_input("Subject", value=_mail_sub, key="stk_mail_sub")
                _mailto = (f"mailto:{_chosen_sup_email}?subject={_up2.quote(_mail_sub_e)}&body={_up2.quote(_mail_body)}"
                           if _chosen_sup_email else
                           f"mailto:?subject={_up2.quote(_mail_sub_e)}&body={_up2.quote(_mail_body)}")
                st.caption(f"✉️ {_chosen_sup_email}" if _chosen_sup_email else "⚠️ No email for supplier")
                st.link_button("📧 Open Mail Client", url=_mailto, use_container_width=True)

            with _ts:
                st.caption("Review then print / save as PDF (Ctrl+P / ⌘+P)")
                st.markdown(
                    f"**REPLENISHMENT ORDER**\n\n"
                    f"Date: {_today_str2}  |  Ref: {_ref_no.strip() or '—'}  |  "
                    f"Supplier: {_chosen_sup_name or '—'}\n\n"
                    + "\n\n".join(
                        f"**{_pn3}**\n" + "\n".join(_by_prod2[_pn3])
                        for _pn3 in _prod_order2
                    )
                    + "\n\n---\n*Please confirm by return message or call.*"
                )

            with _te:
                st.caption("Download Excel to attach to email")
                try:
                    import io as _io2, importlib.util as _iu2
                    if _iu2.find_spec("openpyxl"):
                        import openpyxl as _oxl2
                        _wb2 = _oxl2.Workbook()
                        _ws2 = _wb2.active
                        _ws2.title = "Replenishment Order"
                        _ws2.append(["REPLENISHMENT ORDER"])
                        _ws2.append([f"Date: {_today_str2}", f"Ref: {_ref_no.strip() or '—'}", f"Supplier: {_chosen_sup_name or '—'}"])
                        _ws2.append([])
                        _ws2.append(["Product", "Eye", "Power", "Qty", "Unit", "Pcs (each)", "Total Pcs"])
                        for _ck2 in _club_keys:
                            _cl2 = _club[_ck2]
                            _xl_bs   = int(_cl2.get("box_size", 1) or 1)
                            _xl_unit = str(_cl2.get("unit", "") or "")
                            _xl_is_box = _xl_bs > 1
                            _xl_qty  = int(_edit_boxes.get(_ck2, _cl2["boxes"]) or 0)
                            _xl_unit_lbl = "Box" if _xl_is_box else "Pcs"
                            _xl_total_pcs = _xl_qty * _xl_bs if _xl_is_box else _xl_qty
                            _ws2.append([
                                _cl2.get("supplier_product") or _cl2["pname"],
                                _cl2["eye"], _cl2["pwr"],
                                _xl_qty, _xl_unit_lbl, _xl_bs, _xl_total_pcs,
                            ])
                        _buf2 = _io2.BytesIO(); _wb2.save(_buf2); _buf2.seek(0)
                        st.download_button("⬇️ Download Excel",
                            data=_buf2.getvalue(),
                            file_name=f"replenishment_{_ref_no.strip() or _today_str2}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
                    else:
                        st.info("openpyxl not available. Ask admin to install it.")
                except Exception as _xe: st.error(f"Excel error: {_xe}")

            with _tc:
                st.markdown("**Call script — read while on call**")
                _call_txt = [
                    f"Calling: **{_chosen_sup_name or '—'}**  ·  {_chosen_sup_mob or 'no number'}",
                    f"Date: {_today_str2}  |  Ref: {_ref_no.strip() or 'TBD'}",
                    "\nItems:",
                ]
                for _ck2 in _club_keys:
                    _cl2 = _club[_ck2]
                    _ca_bs   = int(_cl2.get("box_size", 1) or 1)
                    _ca_unit = str(_cl2.get("unit", "") or "")
                    _ca_is_box = _ca_bs > 1
                    _ca_qty  = int(_edit_boxes.get(_ck2, _cl2["boxes"]) or 0)
                    if _ca_is_box:
                        _ca_lbl = _unit_compound(_ca_qty, _ca_qty * _ca_bs, _ca_bs, _ca_unit)
                    else:
                        _ca_lbl = _unit_compound(_ca_qty, _ca_qty, _ca_bs, _ca_unit)
                    _call_txt.append(
                        f"  • {_cl2.get('supplier_product') or _cl2['pname']}  Eye {_cl2['eye']}  {_cl2['pwr']}  —  "
                        f"{_ca_lbl}"
                    )
                st.markdown("\n".join(_call_txt))
                if _chosen_sup_mob:
                    _tel = "".join(c for c in _chosen_sup_mob if c.isdigit() or c=="+")
                    if not _tel.startswith("+"): _tel = "+91" + _tel.lstrip("91") if _tel.startswith("91") else "+91"+_tel
                    st.link_button(f"📞 Call {_chosen_sup_mob}", url=f"tel:{_tel}", use_container_width=True)
                else:
                    st.caption("⚠️ No mobile number — add in Parties first")

            with _tp:
                st.markdown("**Programed Excel**")
                try:
                    from modules.backoffice.replenishment_panel import (
                        _build_excel_alcon,
                        _build_excel_generic,
                        _is_alcon_line,
                    )
                    _program_lines = []
                    for _ck2 in _club_keys:
                        _cl2 = _club[_ck2]
                        # Phase-3 (2026-05-26): edit field is in BOXES for CL,
                        # in PCS for ophthalmic. The downstream Alcon / generic
                        # builders treat the value as "how many supplier-units"
                        # to put on the order line, which matches what staff
                        # see in the edit panel — so we pass through as-is.
                        _qty_user = int(_edit_boxes.get(_ck2, _cl2["boxes"]) or 0)
                        if _qty_user <= 0:
                            continue
                        _src = (_cl2.get("lines") or [{}])[0]
                        _program_lines.append({
                            **_src,
                            "qty": _qty_user,
                            "quantity": _qty_user,
                            "billed_qty": _qty_user,
                            "eye_side": _cl2["eye"],
                            "product_name": _cl2.get("supplier_product") or _cl2["pname"],
                        })
                    _is_alcon = any(_is_alcon_line(_ln) for _ln in _program_lines)
                    _xlsx = (
                        _build_excel_alcon(_program_lines)
                        if _is_alcon else
                        _build_excel_generic(_program_lines, {"name": _chosen_sup_name})
                    )
                    st.download_button(
                        "⬇️ Download Programed Excel",
                        data=_xlsx,
                        file_name=f"programed_replenishment_{_ref_no.strip() or _today_str2}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    st.caption("Alcon lines use the programmed Alcon format; others use generic order Excel.")
                except Exception as _pe:
                    st.info(f"Programed Excel not available: {_pe}")

            st.divider()

            # ── 5. Confirm ────────────────────────────────────────────────────
            st.markdown("**5 · Confirm — creates PO and marks all selected lines as PO_SENT**")
            _conf_disabled = not bool(_chosen_sup_id)
            if _conf_disabled:
                st.caption("⚠️ Select a supplier first")
            if st.button("✅ Confirm Order — Save All",
                         type="primary", use_container_width=True,
                         key="stk_multi_confirm", disabled=_conf_disabled):
                # Phase-3 (2026-05-26): build line_id → boxes map from the
                # clubbed view so the bulk-PO writer stores ordered_qty in
                # the SAME unit as unit_price. For CL (box_size>1) this is
                # the number of boxes; for ophthalmic (box_size=1) it equals pcs.
                _qty_overrides_send: dict = {}
                try:
                    for _ck_o in _club_keys:
                        _cl_o = _club.get(_ck_o) or {}
                        _bs_o = int(_cl_o.get("box_size", 1) or 1)
                        _u_o  = str(_cl_o.get("unit", "") or "").strip().lower()
                        _isbox_o = _bs_o > 1
                        # The number the user actually picked in the box/pc input.
                        # For non-box items this is already pcs; for box items
                        # this is the box count — perfect for ordered_qty.
                        _qty_o = int(_edit_boxes.get(_ck_o, _cl_o.get("boxes", 1) or 1) or 1)
                        for _lid_o in _cl_o.get("line_ids", []):
                            if _lid_o:
                                _qty_overrides_send[str(_lid_o)] = _qty_o
                except Exception as _bld_e:
                    import logging as _ovl
                    _ovl.getLogger(__name__).warning(
                        "[bulk_po] qty_overrides build failed: %s", _bld_e
                    )

                _po_no, _saved_lids = _stock_repl_create_bulk_po(
                    _sel_rows,
                    _chosen_sup_id,
                    _chosen_sup_name,
                    qty_overrides=_qty_overrides_send,
                )
                if not _po_no:
                    st.error("Could not create PO. Lines may already have active PO numbers.")
                    st.stop()
                _saved = 0
                for _sr2 in _sel_rows:
                    _sr_lid2 = str(_sr2["line_id"])
                    if _sr_lid2 not in _saved_lids:
                        continue
                    _sr_lp4  = dict(_sr2.get("_lp") or {})
                    _ck4 = (_supplier_spec_label(_sr2),
                             str(_sr2.get("eye_side") or "").upper(),
                             _pwr_label(_sr2))
                    # Phase-3 (2026-05-26): persist BOTH the supplier-unit count
                    # (kept as replenishment_ordered_boxes for backward compat
                    # with existing readers) AND the per-piece total, plus the
                    # unit string. Downstream readers can pick either.
                    _ckc       = _club.get(_ck4) or {}
                    _ord_bs    = int(_ckc.get("box_size", 1) or 1)
                    _ord_unit  = str(_ckc.get("unit", "") or "")
                    _ord_isbox = _ord_bs > 1
                    _ord_qty   = int(_edit_boxes.get(_ck4, 1))
                    _ord_pcs   = _ord_qty * _ord_bs if _ord_isbox else _ord_qty
                    _sr_lp4["replenishment_ordered_boxes"] = _ord_qty
                    _sr_lp4["replenishment_ordered_pcs"]   = _ord_pcs
                    _sr_lp4["replenishment_unit"]          = _ord_unit or ("Box" if _ord_isbox else "PCS")
                    # Phase-3 (2026-05-26): ref_note is now historical-only.
                    # With auto-allocated PO numbers it would duplicate
                    # replenishment_po_no. Only stamp it when staff added a
                    # genuine free-form note (different from the previewed PO).
                    if _ref_no.strip() and _ref_no.strip() != (_po_no or "").strip():
                        _sr_lp4["replenishment_ref_note"] = _ref_no.strip()
                    _stock_repl_write_status(
                        _sr_lid2,
                        _sr_lp4,
                        "PO_SENT",
                        _chosen_sup_id,
                        _chosen_sup_name,
                        _po_no,
                    )
                    _saved += 1
                st.success(f"✅ PO {_po_no} sent · {_saved} line(s) PO_SENT · supplier confirmation pending")
                for _sr2 in rows:
                    st.session_state.pop(f"stk_chk_{_sr2['line_id']}", None)
                st.rerun()


def _render_stock_procurement(procured_view: bool = False):
    """
    Inventory Movement tab.

    All stock items that need purchasing (Contact Lenses, Stock Lenses,
    Ophthalmic lenses, Frames — any route) appear here.

    Workflow:
      1. Tick items → Add supplier → Add invoice/challan no + price → Save
      2. Creates Purchase Acknowledgement record
      3. Option: generate Challan PDF or Invoice PDF inline
      4. Upload invoice PDF/Image → stored in DB + file system for audit

    Purchase manager's view: tick, supplier, save. Done.
    """
    import json as _pj
    import datetime as _pd
    import base64 as _b64
    import os as _pos

    def _q2(sql, params=None):
        try:
            from modules.sql_adapter import run_query
            return run_query(sql, params or {}) or []
        except Exception as _e:
            st.error(f"DB: {_e}")
            return []

    def _w2(sql, params=None):
        try:
            from modules.sql_adapter import run_write
            run_write(sql, params or {})
            return True
        except Exception as _e:
            st.error(f"Write: {_e}")
            return False

    def _invoice_path_from_notes(raw: str) -> str:
        """Extract a file path from legacy notes values such as 'note | path'."""
        raw = str(raw or "").strip()
        if not raw or raw.startswith("b64:"):
            return ""
        for part in reversed(raw.split(" | ")):
            part = part.strip()
            if (
                "\\" in part or "/" in part or
                part.lower().endswith((".pdf", ".png", ".jpg", ".jpeg", ".webp"))
            ):
                return part
        return raw if raw.lower().endswith((".pdf", ".png", ".jpg", ".jpeg", ".webp")) else ""

    def _save_invoice_upload(uploaded, ref_no: str = ""):
        """Save original invoice and a compressed image preview when possible."""
        if uploaded is None:
            return "", ""
        import pathlib as _plb
        import re as _re
        original_dir = _plb.Path("uploads/purchase_invoices")
        preview_dir = _plb.Path("uploads/purchase_invoice_previews")
        original_dir.mkdir(parents=True, exist_ok=True)
        preview_dir.mkdir(parents=True, exist_ok=True)
        safe_ref = _re.sub(r"[^A-Za-z0-9_.-]+", "-", str(ref_no or "draft")).strip("-") or "draft"
        safe_name = _re.sub(r"[^A-Za-z0-9_.-]+", "-", str(uploaded.name or "invoice")).strip("-") or "invoice"
        original_path = original_dir / f"{_pd.date.today().isoformat()}_{safe_ref}_{safe_name}"
        data = uploaded.getvalue()
        with open(original_path, "wb") as fh:
            fh.write(data)

        preview_path = ""
        if original_path.suffix.lower() in (".png", ".jpg", ".jpeg", ".webp"):
            try:
                from PIL import Image
                img = Image.open(original_path)
                img.thumbnail((1600, 1600))
                if img.mode not in ("RGB", "L"):
                    img = img.convert("RGB")
                preview_path_obj = preview_dir / f"{original_path.stem}_preview.jpg"
                img.save(preview_path_obj, format="JPEG", quality=82, optimize=True)
                preview_path = str(preview_path_obj)
            except Exception:
                preview_path = ""
        return str(original_path), preview_path

    def _ensure_invoice_preview(path: str) -> str:
        """Create/reuse compressed preview for an existing image attachment."""
        import pathlib as _plb2
        import os as _os_prev
        path = str(path or "")
        if not path.lower().endswith((".png", ".jpg", ".jpeg", ".webp")) or not _os_prev.path.exists(path):
            return ""
        preview_dir = _plb2.Path("uploads/purchase_invoice_previews")
        preview_dir.mkdir(parents=True, exist_ok=True)
        src = _plb2.Path(path)
        preview = preview_dir / f"{src.stem}_preview.jpg"
        if preview.exists():
            return str(preview)
        try:
            from PIL import Image
            img = Image.open(src)
            img.thumbnail((1600, 1600))
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            img.save(preview, format="JPEG", quality=82, optimize=True)
            return str(preview)
        except Exception:
            return ""

    def _render_attachment_popover(label: str, raw_path: str, key: str):
        path = _invoice_path_from_notes(raw_path)
        if not path:
            return
        import os as _os_attach
        if not _os_attach.path.exists(path):
            st.caption(f"⚠️ Attached file missing: {_os_attach.path.basename(path)}")
            return
        with st.popover(label, use_container_width=False):
            lower = path.lower()
            if lower.endswith((".png", ".jpg", ".jpeg", ".webp")):
                preview_path = _ensure_invoice_preview(path)
                st.image(preview_path if preview_path and _os_attach.path.exists(preview_path) else path, use_container_width=True)
            with open(path, "rb") as fh:
                st.download_button(
                    "⬇️ Download attached invoice",
                    fh.read(),
                    file_name=_os_attach.path.basename(path),
                    mime="application/pdf" if lower.endswith(".pdf") else "image/jpeg",
                    key=key,
                    use_container_width=True,
                )

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown(
        "<div style='background:#0f172a;border:1px solid #1e3a5f;"
        "border-left:4px solid #3b82f6;border-radius:8px;"
        "padding:10px 16px;margin-bottom:12px'>"
        f"<span style='color:#93c5fd;font-size:1rem;font-weight:800'>{'📊 Inventory Movement' if procured_view else '📦 Stock Replenishment Receipts'}</span>"
        "<span style='color:#475569;font-size:0.78rem;margin-left:10px'>"
        + (
            "Purchased / received supplier documents. Search supplier, order, product, RX or stock in one place."
            if procured_view else
            "Record supplier invoice/challan for stock replenishment POs. Customer STOCK billing does not wait here."
        )
        + "</span></div>",
        unsafe_allow_html=True,
    )

    # ── Filters ───────────────────────────────────────────────────────────
    with st.container(border=True):
        _fc1, _fc2, _fc3, _fc4, _fc5 = st.columns([3, 2, 2, 2, 1])
        _sp_search = _fc1.text_input("Search", placeholder="🔍 Sales/Purchase invoice, order, product, power",
                                      key="stkpr_search", label_visibility="collapsed")
        _sp_from   = _fc2.date_input("From", value=_pd.date.today() - _pd.timedelta(days=180 if procured_view else 60),
                                      key="stkpr_from", label_visibility="collapsed",
                                      format="DD/MM/YYYY")
        _sp_to     = _fc3.date_input("To",   value=_pd.date.today(),
                                      key="stkpr_to",   label_visibility="collapsed",
                                      format="DD/MM/YYYY")
        _sp_status_options = ["Purchase", "Sales", "Both"] if procured_view else ["Pending Purchase", "All", "Purchased"]
        _sp_status = _fc4.selectbox(
            "View",
            _sp_status_options,
            index=(0 if procured_view else 0),
            key="stkpr_status",
            label_visibility="collapsed",
        )
        if _fc5.button("🔄", key="stkpr_refresh", use_container_width=True):
            st.rerun()

        _party_filter_pick = "All"
        if procured_view:
            _party_rows = _q2("""
                SELECT DISTINCT COALESCE(supplier_name, '') AS name
                FROM purchase_acknowledgements
                WHERE COALESCE(supplier_name, '') <> ''
                UNION
                SELECT DISTINCT COALESCE(po.supplier_name, '') AS name
                FROM procurement_orders po
                WHERE COALESCE(po.supplier_name, '') <> ''
                UNION
                SELECT DISTINCT COALESCE(pt.party_name, o.party_name, o.patient_name, '') AS name
                FROM invoices i
                LEFT JOIN parties pt ON pt.id = i.party_id
                LEFT JOIN LATERAL (
                    SELECT o2.party_name, o2.patient_name
                    FROM orders o2
                    WHERE o2.id::text = ANY(i.order_ids)
                    LIMIT 1
                ) o ON TRUE
                WHERE COALESCE(pt.party_name, o.party_name, o.patient_name, '') <> ''
                ORDER BY name
            """) or []
            _party_opts = ["All"] + [str(r.get("name") or "").strip() for r in _party_rows if str(r.get("name") or "").strip()]
            _party_filter_pick = st.selectbox(
                "Patient / Party / Supplier",
                _party_opts,
                key="stkpr_party_filter",
            )

        _pwr_filters = {"sph": None, "cyl": None, "axis": None, "add_power": None}
        if procured_view:
            st.markdown(
                "<div style='font-size:0.74rem;color:#374151;font-weight:700;margin-top:6px'>"
                "Power filter for Inventory Movement</div>",
                unsafe_allow_html=True,
            )
            _pf1, _pf2, _pf3, _pf4, _pf5 = st.columns([1, 1, 1, 1, 1])
            _pwr_filters["sph"] = _pf1.text_input("SPH", placeholder="-1.00", key="stkpr_pf_sph")
            _pwr_filters["cyl"] = _pf2.text_input("CYL", placeholder="-0.50", key="stkpr_pf_cyl")
            _pwr_filters["axis"] = _pf3.text_input("AXIS", placeholder="90", key="stkpr_pf_axis")
            _pwr_filters["add_power"] = _pf4.text_input("ADD", placeholder="+2.00", key="stkpr_pf_add")
            if _pf5.button("Clear Power", key="stkpr_pf_clear", use_container_width=True):
                for _pf_key in ("stkpr_pf_sph", "stkpr_pf_cyl", "stkpr_pf_axis", "stkpr_pf_add"):
                    st.session_state.pop(_pf_key, None)
                st.rerun()

    # ── Query: all non-inhouse stock/procurement lines from reservation/order onwards ───────────────
    _where_parts = [
        (
            """(
                UPPER(COALESCE(ol.lens_params->>'manufacturing_route','')) IN ('STOCK','VENDOR','EXTERNAL_LAB')
                OR EXISTS (
                    SELECT 1 FROM purchase_acknowledgements pa_route
                    WHERE pa_route.order_line_id = ol.id
                      AND UPPER(COALESCE(pa_route.billing_status,'')) NOT IN ('VOID','CANCELLED')
                )
            )"""
            if procured_view else
            "UPPER(COALESCE(ol.lens_params->>'manufacturing_route','')) IN ('STOCK','VENDOR','EXTERNAL_LAB')"
        ),
        "COALESCE(ol.is_deleted, FALSE) = FALSE",
        "COALESCE(ol.is_service_line, FALSE) = FALSE",
        "UPPER(COALESCE(ol.eye_side,'')) NOT IN ('S','SERVICE','O')",
        "o.status NOT IN ('CANCELLED','CLOSED')",
        "UPPER(COALESCE(ol.lens_params->>'manufacturing_route','')) != 'INHOUSE'",
        (
            "(o.created_at::date BETWEEN %(dfrom)s AND %(dto)s"
            " OR EXISTS ("
            "   SELECT 1 FROM purchase_acknowledgements pav_d"
            "   WHERE pav_d.order_line_id = ol.id"
            "   AND COALESCE(pav_d.document_date, pav_d.acknowledged_at::date)"
            "       BETWEEN %(dfrom)s AND %(dto)s"
            " ))"
        ) if procured_view else
        "o.created_at::date BETWEEN %(dfrom)s AND %(dto)s",
        (
            "(COALESCE(ol.allocated_qty,0) > 0"
            " OR EXISTS ("
            "   SELECT 1 FROM purchase_acknowledgements pav_a"
            "   WHERE pav_a.order_line_id = ol.id"
            "   AND UPPER(COALESCE(pav_a.billing_status,'')) NOT IN ('VOID','CANCELLED')"
            " ))"
        ) if procured_view else
        "COALESCE(ol.allocated_qty,0) > 0",
    ]
    _qparams = {"dfrom": str(_sp_from), "dto": str(_sp_to)}
    _movement_mode = _sp_status if procured_view else "Purchase"

    if _sp_search.strip():
        _where_parts.append(
            "(LOWER(o.order_no) LIKE %(srch)s "
            " OR regexp_replace(LOWER(COALESCE(o.order_no,'')), '[^a-z0-9]', '', 'g') LIKE %(srch_norm)s"
            " OR LOWER(COALESCE(o.patient_name,o.party_name,'')) LIKE %(srch)s"
            " OR LOWER(p.product_name) LIKE %(srch)s"
            " OR LOWER(COALESCE(pa.invoice_no,'')) LIKE %(srch)s"
            " OR LOWER(COALESCE(pa.challan_no,'')) LIKE %(srch)s"
            " OR LOWER(COALESCE(pa.batch_no,'')) LIKE %(srch)s"
            " OR ol.sph::text = %(srch_exact)s"
            " OR COALESCE(ol.lens_params->>'sph','') = %(srch_exact)s)"
        )
        _qparams["srch"] = f"%{_sp_search.strip().lower()}%"
        _qparams["srch_norm"] = f"%{_scan_norm(_sp_search)}%"
        _qparams["srch_exact"] = _sp_search.strip()

    if procured_view and _party_filter_pick != "All":
        _where_parts.append(
            "(COALESCE(pt.party_name, ol.lens_params->>'supplier_name', '') = %(party_pick)s "
            "OR COALESCE(o.patient_name, o.party_name, '') = %(party_pick)s)"
        )
        _qparams["party_pick"] = _party_filter_pick

    if not procured_view and _sp_status == "Pending Purchase":
        _where_parts.append(
            "NOT EXISTS ("
            "  SELECT 1 FROM purchase_acknowledgements pa2"
            "  WHERE pa2.order_line_id = ol.id"
            "  AND COALESCE(pa2.purchase_price,0) > 0"
            ")"
        )
    elif not procured_view and _sp_status == "Purchased":
        _where_parts.append(
            "EXISTS ("
            "  SELECT 1 FROM purchase_acknowledgements pa2"
            "  WHERE pa2.order_line_id = ol.id"
            "  AND UPPER(COALESCE(pa2.billing_status,'')) NOT IN ('VOID','CANCELLED')"
            ")"
        )

    _where_sql = " AND ".join(_where_parts)

    try:
        _rows = [] if (procured_view and _movement_mode == "Sales") else None
    except Exception:
        _rows = None

    try:
        if _rows is None:
            _rows = _q2(f"""
            SELECT
                o.id::text                                    AS order_id,
                o.order_no,
                COALESCE(o.patient_name, o.party_name, '—')  AS patient_name,
                o.created_at,
                ol.id::text                                   AS line_id,
                ol.eye_side,
                COALESCE(ol.quantity, 1)                      AS quantity,
                ol.lens_params,
                ol.product_id::text                           AS product_id,
                p.product_name,
                COALESCE(p.category,'')                       AS category,
                COALESCE(p.main_group,'')                     AS main_group,
                COALESCE(p.box_size, 1)                       AS box_size,
                COALESCE(p.is_batch_applicable, FALSE)        AS is_batch_applicable,
                ol.sph, ol.cyl, ol.axis, ol.add_power,
                -- Current purchase acknowledgement
                pa.id::text                                   AS pa_id,
                pa.supplier_id::text                          AS pa_sup_id,
                COALESCE(pt.party_name, ol.lens_params->>'supplier_name', '')  AS pa_sup_name,
                COALESCE(pa.challan_no, '')                   AS pa_challan,
                COALESCE(
                    NULLIF(pa.invoice_no, ''),
                    NULLIF(substring(COALESCE(pa.notes, '') FROM 'invoice:([^| ]+)'), ''),
                    ''
                )                                             AS pa_invoice_no,
                COALESCE(pa.purchase_price, 0)::numeric       AS pa_price,
                COALESCE(pa.received_qty, 0)                  AS pa_recv_qty,
                COALESCE(pa.batch_no, '')                     AS pa_batch_no,
                COALESCE(pa.expiry_date::text, '')            AS pa_expiry_date,
                COALESCE(pa.document_date, pa.inventory_posted_at::date, pa.acknowledged_at::date)::text
                                                               AS pa_document_date,
                pa.is_price_locked                            AS pa_locked,
                COALESCE(pa.billing_status, '')               AS pa_status,
                -- Invoice file: check notes (may be "note | /path/to/file"), then procurement_receipts
                COALESCE(
                    NULLIF(pa.notes, ''),
                    (SELECT pr.invoice_file_path FROM procurement_receipts pr
                     JOIN procurement_order_items poi ON poi.procurement_order_id = pr.procurement_order_id
                     WHERE poi.order_line_id = ol.id
                       AND NULLIF(pr.invoice_file_path,'') IS NOT NULL
                     ORDER BY pr.created_at DESC LIMIT 1),
                    ''
                )                                             AS pa_invoice_file,
                -- Preferred supplier from product master
                COALESCE((
                    SELECT pt2.party_name FROM product_supplier_map psm
                    JOIN parties pt2 ON pt2.id = psm.supplier_id
                    WHERE psm.product_id = ol.product_id
                    ORDER BY psm.supplier_id
                    LIMIT 1
                ), ol.lens_params->>'supplier_name', '')       AS preferred_supplier_name,
                COALESCE((
                    SELECT psm.supplier_id::text FROM product_supplier_map psm
                    WHERE psm.product_id = ol.product_id
                    ORDER BY psm.supplier_id
                    LIMIT 1
                ), ol.lens_params->>'supplier_id', '')         AS preferred_supplier_id,
                -- Last purchase price from inventory_stock
                COALESCE((
                    SELECT COALESCE(NULLIF(s.purchase_price,0), s.purchase_rate, 0)
                    FROM inventory_stock s
                    WHERE s.product_id = ol.product_id
                      AND COALESCE(s.is_active, TRUE) = TRUE
                    ORDER BY s.created_at DESC LIMIT 1
                ), 0)::numeric                                 AS last_purchase_price
            FROM order_lines ol
            JOIN orders   o  ON o.id  = ol.order_id
            JOIN products p  ON p.id  = ol.product_id
            LEFT JOIN purchase_acknowledgements pa  ON pa.order_line_id = ol.id
            LEFT JOIN parties pt ON pt.id = pa.supplier_id
            WHERE {_where_sql}
            ORDER BY
                CASE WHEN pa.id IS NULL THEN 0 ELSE 1 END,
                o.created_at DESC,
                o.order_no,
                CASE WHEN ol.eye_side='R' THEN 0 WHEN ol.eye_side='L' THEN 1 ELSE 2 END
            """, _qparams) or []
    except Exception as _qe:
        st.error(f"Query error: {_qe}")
        return

    if procured_view and _movement_mode in ("Purchase", "Both"):
        _direct_where = [
            "pa.order_line_id IS NULL",
            "UPPER(COALESCE(pa.billing_status,'')) NOT IN ('VOID','CANCELLED')",
            (
                "("
                "pa.document_date BETWEEN %(dfrom)s AND %(dto)s"
                " OR pa.inventory_posted_at::date BETWEEN %(dfrom)s AND %(dto)s"
                " OR pa.acknowledged_at::date BETWEEN %(dfrom)s AND %(dto)s"
                ")"
            ),
        ]
        _direct_params = {"dfrom": str(_sp_from), "dto": str(_sp_to)}
        if _sp_search.strip():
            _direct_where[-1] = (
                "("
                "pa.document_date BETWEEN %(dfrom)s AND %(dto)s"
                " OR pa.inventory_posted_at::date BETWEEN %(dfrom)s AND %(dto)s"
                " OR pa.acknowledged_at::date BETWEEN %(dfrom)s AND %(dto)s"
                " OR LOWER(COALESCE(pa.invoice_no,'')) LIKE %(srch)s"
                " OR LOWER(COALESCE(pa.challan_no,'')) LIKE %(srch)s"
                " OR LOWER(COALESCE(pa.supplier_name,'')) LIKE %(srch)s"
                " OR LOWER(COALESCE(pa.our_product_name, pa.product_name, pa.supplier_product_name,'')) LIKE %(srch)s"
                " OR LOWER(COALESCE(pa.supplier_product_name,'')) LIKE %(srch)s"
                " OR LOWER(COALESCE(pa.supplier_product_description,'')) LIKE %(srch)s"
                ")"
            )
            _direct_params["srch"] = f"%{_sp_search.strip().lower()}%"
        if False:
            _direct_where.append("FALSE")

        _direct_rows = _q2(f"""
            SELECT
                ''::text                                      AS order_id,
                COALESCE(NULLIF(pa.invoice_no,''), NULLIF(pa.challan_no,''), 'DIRECT-UPLOAD') AS order_no,
                COALESCE(pa.supplier_name, 'Direct invoice upload') AS patient_name,
                COALESCE(pa.inventory_posted_at, pa.acknowledged_at, NOW()) AS created_at,
                ('direct_' || pa.id::text)                    AS line_id,
                COALESCE(pa.eye_side, '')                     AS eye_side,
                COALESCE(pa.received_qty, pa.qty, 1)          AS quantity,
                jsonb_build_object(
                    'manufacturing_route', 'DIRECT_INVOICE',
                    'batch_no', COALESCE(pa.batch_no,''),
                    'expiry_date', COALESCE(pa.expiry_date::text,'')
                )                                             AS lens_params,
                COALESCE(pa.our_product_id, pa.product_id)::text AS product_id,
                COALESCE(pa.our_product_name, p.product_name, pa.product_name, pa.supplier_product_name, '—') AS product_name,
                COALESCE(p.category,'')                       AS category,
                COALESCE(p.main_group,'')                     AS main_group,
                COALESCE(p.box_size, 1)                       AS box_size,
                COALESCE(p.is_batch_applicable, FALSE)        AS is_batch_applicable,
                NULL::numeric                                AS sph,
                NULL::numeric                                AS cyl,
                NULL::integer                                AS axis,
                NULL::numeric                                AS add_power,
                pa.id::text                                   AS pa_id,
                pa.supplier_id::text                          AS pa_sup_id,
                COALESCE(pa.supplier_name, pt.party_name, '') AS pa_sup_name,
                COALESCE(pa.challan_no, '')                   AS pa_challan,
                COALESCE(pa.invoice_no, '')                   AS pa_invoice_no,
                COALESCE(pa.purchase_price, 0)::numeric       AS pa_price,
                COALESCE(pa.received_qty, pa.qty, 0)          AS pa_recv_qty,
                COALESCE(pa.batch_no, '')                     AS pa_batch_no,
                COALESCE(pa.expiry_date::text, '')            AS pa_expiry_date,
                COALESCE(pa.document_date, pa.inventory_posted_at::date, pa.acknowledged_at::date)::text
                                                               AS pa_document_date,
                pa.is_price_locked                            AS pa_locked,
                COALESCE(pa.billing_status, '')               AS pa_status,
                COALESCE(pa.notes, '')                        AS pa_invoice_file,
                COALESCE(pa.supplier_name, '')                AS preferred_supplier_name,
                COALESCE(pa.supplier_id::text, '')            AS preferred_supplier_id,
                COALESCE((
                    SELECT COALESCE(NULLIF(s.purchase_price,0), s.purchase_rate, 0)
                    FROM inventory_stock s
                    WHERE s.product_id = COALESCE(pa.our_product_id, pa.product_id)
                      AND COALESCE(s.is_active, TRUE) = TRUE
                    ORDER BY s.created_at DESC LIMIT 1
                ), 0)::numeric                                AS last_purchase_price,
                TRUE                                          AS direct_invoice_upload,
                COALESCE(pa.audit_status, '')                 AS direct_audit_status,
                COALESCE(pa.supplier_product_name, '')        AS supplier_product_name,
                COALESCE(pa.supplier_product_description, '') AS supplier_product_description,
                COALESCE(pa.batch_no, '')                     AS direct_batch_no,
                COALESCE(pa.expiry_date::text, '')            AS direct_expiry_date
            FROM purchase_acknowledgements pa
            LEFT JOIN products p ON p.id = COALESCE(pa.our_product_id, pa.product_id)
            LEFT JOIN parties pt ON pt.id = pa.supplier_id
            WHERE {" AND ".join(_direct_where)}
            ORDER BY COALESCE(pa.inventory_posted_at, pa.acknowledged_at, NOW()) DESC,
                     COALESCE(pa.invoice_no, pa.challan_no, '')
        """, _direct_params) or []
        _rows.extend(_direct_rows)

        _ledger_where = [
            "poi.order_line_id IS NULL",
            "UPPER(COALESCE(poi.status,'')) NOT IN ('VOID','CANCELLED')",
            "UPPER(COALESCE(po.source, poi.source, '')) IN ('DIRECT_PURCHASE','SCANNED_INVOICE')",
            "(po.document_date BETWEEN %(dfrom)s AND %(dto)s OR poi.received_at::date BETWEEN %(dfrom)s AND %(dto)s)",
        ]
        _ledger_params = {"dfrom": str(_sp_from), "dto": str(_sp_to)}
        if _sp_search.strip():
            _ledger_where[-1] = (
                "(po.document_date BETWEEN %(dfrom)s AND %(dto)s "
                "OR poi.received_at::date BETWEEN %(dfrom)s AND %(dto)s "
                "OR LOWER(COALESCE(po.document_no,'')) LIKE %(srch)s "
                "OR LOWER(COALESCE(po.supplier_name, poi.supplier_name,'')) LIKE %(srch)s "
                "OR LOWER(COALESCE(poi.product_name,'')) LIKE %(srch)s)"
            )
            _ledger_params["srch"] = f"%{_sp_search.strip().lower()}%"
        if procured_view and _party_filter_pick != "All":
            _ledger_where.append("COALESCE(po.supplier_name, poi.supplier_name, '') = %(party_pick)s")
            _ledger_params["party_pick"] = _party_filter_pick

        _ledger_rows = _q2(f"""
            SELECT
                ''::text                                      AS order_id,
                COALESCE(NULLIF(po.document_no,''), po.procurement_no, 'DIRECT-PURCHASE') AS order_no,
                COALESCE(po.supplier_name, poi.supplier_name, 'Direct purchase') AS patient_name,
                COALESCE(poi.received_at, po.received_at, po.created_at, NOW()) AS created_at,
                ('directpo_' || poi.id::text)                 AS line_id,
                COALESCE(poi.eye_side, '')                    AS eye_side,
                COALESCE(poi.qty_received, poi.qty_ordered, 1) AS quantity,
                COALESCE(poi.power_json, '{{}}'::jsonb) ||
                    jsonb_build_object(
                        'manufacturing_route', 'DIRECT_PURCHASE',
                        'direct_source', COALESCE(po.source, poi.source, 'DIRECT_PURCHASE')
                    )                                         AS lens_params,
                poi.product_id::text                          AS product_id,
                COALESCE(poi.product_name, '—')               AS product_name,
                ''                                            AS category,
                COALESCE(poi.route, '')                       AS main_group,
                1                                             AS box_size,
                FALSE                                         AS is_batch_applicable,
                NULL::numeric                                 AS sph,
                NULL::numeric                                 AS cyl,
                NULL::integer                                 AS axis,
                NULL::numeric                                 AS add_power,
                NULL::text                                    AS pa_id,
                COALESCE(po.supplier_id::text, poi.supplier_id::text) AS pa_sup_id,
                COALESCE(po.supplier_name, poi.supplier_name, '') AS pa_sup_name,
                CASE WHEN UPPER(COALESCE(po.document_type,'')) IN ('CHALLAN','BOTH') THEN COALESCE(po.document_no,'') ELSE '' END AS pa_challan,
                CASE WHEN UPPER(COALESCE(po.document_type,'')) IN ('INVOICE','BOTH') THEN COALESCE(po.document_no,'') ELSE '' END AS pa_invoice_no,
                COALESCE(poi.unit_price, 0)::numeric          AS pa_price,
                COALESCE(poi.qty_received, poi.qty_ordered, 0) AS pa_recv_qty,
                COALESCE(poi.batch_no, '')                    AS pa_batch_no,
                COALESCE(poi.expiry_date::text, '')           AS pa_expiry_date,
                COALESCE(po.document_date, poi.received_at::date, po.created_at::date)::text AS pa_document_date,
                TRUE                                          AS pa_locked,
                COALESCE(poi.status, po.status, 'PROCURED')   AS pa_status,
                COALESCE(po.invoice_file_path, pr.invoice_file_path, '') AS pa_invoice_file,
                COALESCE(po.supplier_name, poi.supplier_name, '') AS preferred_supplier_name,
                COALESCE(po.supplier_id::text, poi.supplier_id::text, '') AS preferred_supplier_id,
                0::numeric                                    AS last_purchase_price,
                TRUE                                          AS direct_invoice_upload,
                'DIRECT_PURCHASE'                             AS direct_audit_status,
                ''                                            AS supplier_product_name,
                ''                                            AS supplier_product_description,
                COALESCE(poi.batch_no, '')                    AS direct_batch_no,
                COALESCE(poi.expiry_date::text, '')           AS direct_expiry_date,
                TRUE                                          AS direct_purchase_entry
            FROM procurement_order_items poi
            JOIN procurement_orders po ON po.id = poi.procurement_order_id
            LEFT JOIN procurement_receipts pr ON pr.procurement_order_id = po.id
            WHERE {" AND ".join(_ledger_where)}
            ORDER BY COALESCE(po.document_date, poi.received_at::date, po.created_at::date) DESC,
                     COALESCE(po.document_no, po.procurement_no, '')
        """, _ledger_params) or []
        _rows.extend(_ledger_rows)

    if not _rows:
        st.success("✅ No stock lines pending purchase.")
        return

    try:
        from modules.procurement.procurement_ledger import ensure_queue_items
        ensure_queue_items(
            [str(r.get("line_id") or "") for r in _rows if not str(r.get("line_id") or "").startswith("direct_")],
            source="STOCK_PROCUREMENT",
        )
    except Exception as _pl_sync_err:
        st.caption(f"Procurement ledger sync pending: {_pl_sync_err}")

    # Enrich lens_params
    for _r in _rows:
        _lp = _r.get("lens_params") or {}
        if isinstance(_lp, str):
            try: _lp = _pj.loads(_lp)
            except: _lp = {}
        _r["_lp"] = _lp
        _r["_purchased"] = bool(_r.get("pa_id")) or float(_r.get("pa_price") or 0) > 0

    def _row_index_value(_row):
        _lp = _row.get("_lp") or {}
        return str(_lp.get("lens_index") or _lp.get("index") or _row.get("index_value") or "").strip()

    def _row_coating_value(_row):
        _lp = _row.get("_lp") or {}
        return str(_lp.get("coating") or _lp.get("coating_type") or _row.get("coating") or "").strip()

    def _batch_for_row(_row):
        _batch = str(_row.get("pa_batch_no") or _row.get("direct_batch_no") or "").strip()
        if not _batch:
            _alloc = (_row.get("_lp") or {}).get("batch_allocation") or []
            if isinstance(_alloc, list) and _alloc:
                _batch = str((_alloc[0] or {}).get("batch_no") or "").strip()
        return _batch

    def _expiry_for_row(_row):
        return str(_row.get("pa_expiry_date") or _row.get("direct_expiry_date") or "").strip()

    def _show_batch_for_row(_row):
        _group = f"{_row.get('main_group') or ''} {_row.get('category') or ''}".upper()
        _pname = str(_row.get("product_name") or "").upper()
        if bool(_row.get("is_batch_applicable")):
            return True
        if any(_kw in _group for _kw in ("CONTACT", "CL", "SOLUTION")):
            return True
        if any(_kw in _pname for _kw in ("CONTACT", "SOLUTION", "10PK", "6PK", "30PK", "90PK")):
            return True
        return False

    def _rate_text(_row):
        _price = float(_row.get("pa_price") or 0)
        try:
            _box_size = int(float(_row.get("box_size") or 1))
        except Exception:
            _box_size = 1
        if _box_size > 1 and _price > 0:
            return f"₹{_price:.2f}/box · ₹{(_price / _box_size):.2f}/pc"
        if _price > 0:
            return f"₹{_price:.2f}/pc"
        return "₹0.00"

    if procured_view:
        def _parse_pwr_filter(_raw, _is_axis=False):
            _txt = str(_raw or "").strip().replace(" ", "")
            if not _txt:
                return None
            try:
                return int(round(float(_txt))) if _is_axis else float(_txt)
            except Exception:
                return None

        _parsed_filters = {
            "sph": _parse_pwr_filter(_pwr_filters.get("sph")),
            "cyl": _parse_pwr_filter(_pwr_filters.get("cyl")),
            "axis": _parse_pwr_filter(_pwr_filters.get("axis"), _is_axis=True),
            "add_power": _parse_pwr_filter(_pwr_filters.get("add_power")),
        }
        _active_pwr_filters = {k: v for k, v in _parsed_filters.items() if v is not None}

        def _row_pwr_value(_row, _key):
            _val = _row.get(_key)
            if _val in (None, ""):
                _val = (_row.get("_lp") or {}).get(_key)
            if _val in (None, "") and _key == "add_power":
                _val = (_row.get("_lp") or {}).get("add")
            try:
                return float(_val)
            except Exception:
                return None

        def _direct_text_power_match(_row, _active_filters):
            if not (_row.get("direct_invoice_upload") or _row.get("direct_purchase_entry")):
                return False
            hay = " ".join(
                str(_row.get(k) or "")
                for k in (
                    "supplier_product_name", "product_name", "pa_invoice_no",
                    "pa_challan", "direct_audit_status"
                )
            )
            hay += " " + str((_row.get("_lp") or {}).get("supplier_product_description") or "")
            normalized = hay.replace("+", "").replace(" ", "")
            # Direct invoice rows often carry power only in supplier text
            # (for example "-00.00 HZ"). If SPH is the only meaningful filter
            # or the product is contact-lens-like, match on the text and ignore
            # unrelated CYL/AXIS/ADD fields left filled from a previous search.
            sph = _active_filters.get("sph")
            if sph is None:
                return False
            sph_tokens = {
                f"{sph:.2f}".replace("+", ""),
                f"{sph:+.2f}".replace("+", ""),
                f"{sph:.1f}".replace("+", ""),
            }
            if abs(float(sph)) < 0.01:
                sph_tokens.update({"0.00", "00.00", "-00.00", "+00.00", "0000", "-0000"})
            return any(tok.replace(" ", "") in normalized for tok in sph_tokens)

        if _active_pwr_filters:
            _filtered_rows = []
            for _row in _rows:
                if _direct_text_power_match(_row, _active_pwr_filters):
                    _filtered_rows.append(_row)
                    continue
                _match = True
                for _key, _want in _active_pwr_filters.items():
                    _got = _row_pwr_value(_row, _key)
                    if _got is None:
                        # CL/direct invoices often have SPH-only powers. Do not
                        # let stale CYL/AXIS/ADD boxes hide a valid SPH match.
                        if _key != "sph" and _row_pwr_value(_row, "sph") is not None:
                            continue
                        _match = False
                        break
                    if _key == "axis":
                        if int(round(_got)) != int(_want):
                            _match = False
                            break
                    elif abs(float(_got) - float(_want)) > 0.01:
                        _match = False
                        break
                if _match:
                    _filtered_rows.append(_row)
            _rows = _filtered_rows
            if not _rows:
                st.info("No procured lines match the selected SPH / CYL / AXIS / ADD filter.")
                return

        _pre_sales_rows = []
        if _movement_mode in ("Sales", "Both"):
            _pre_sales_params = {"dfrom": str(_sp_from), "dto": str(_sp_to)}
            _pre_sales_where = [
                "i.invoice_date BETWEEN %(dfrom)s AND %(dto)s",
                "COALESCE(i.is_deleted, FALSE) = FALSE",
                "COALESCE(il.is_deleted, FALSE) = FALSE",
                "UPPER(COALESCE(i.status,'')) NOT IN ('VOID','CANCELLED')",
            ]
            if _sp_search.strip():
                _pre_sales_where.append(
                    "(LOWER(COALESCE(il.product_name, p.product_name, '')) LIKE %(ss)s "
                    "OR LOWER(COALESCE(i.invoice_no,'')) LIKE %(ss)s "
                    "OR LOWER(COALESCE(c.challan_no,'')) LIKE %(ss)s "
                    "OR LOWER(COALESCE(o.order_no,'')) LIKE %(ss)s "
                    "OR regexp_replace(LOWER(COALESCE(o.order_no,'')), '[^a-z0-9]', '', 'g') LIKE %(ss_norm)s "
                    "OR LOWER(COALESCE(pt.party_name, o.party_name, o.patient_name, '')) LIKE %(ss)s "
                    "OR ol.sph::text = %(ss_exact)s "
                    "OR COALESCE(ol.lens_params->>'sph','') = %(ss_exact)s)"
                )
                _pre_sales_params["ss"] = f"%{_sp_search.strip().lower()}%"
                _pre_sales_params["ss_norm"] = f"%{_scan_norm(_sp_search)}%"
                _pre_sales_params["ss_exact"] = _sp_search.strip()
            if _party_filter_pick != "All":
                _pre_sales_where.append("COALESCE(pt.party_name, o.party_name, o.patient_name, '') = %(sparty)s")
                _pre_sales_params["sparty"] = _party_filter_pick
            _pre_sales_rows = _q2(f"""
                SELECT
                    i.invoice_date::text AS doc_date,
                    i.invoice_no,
                    COALESCE(c.challan_no, '') AS challan_no,
                    COALESCE(o.order_no, '') AS order_no,
                    COALESCE(pt.party_name, o.party_name, o.patient_name, '') AS party_name,
                    p.id::text AS product_id,
                    COALESCE(il.product_name, p.product_name, '—') AS product_name,
                    COALESCE(il.eye_side, ol.eye_side, '') AS eye_side,
                    COALESCE(il.quantity, 1) AS qty,
                    COALESCE(il.line_total, il.total_price, il.quantity * il.unit_price, 0)::numeric AS amount,
                    COALESCE(ol.lens_params->>'lens_index', ol.lens_params->>'index', p.index_value::text, '') AS lens_index,
                    COALESCE(ol.lens_params->>'coating', ol.lens_params->>'coating_type', p.coating, '') AS coating
                FROM invoice_lines il
                JOIN invoices i ON i.id = il.invoice_id
                LEFT JOIN challans c ON c.id = i.challan_id
                LEFT JOIN order_lines ol ON ol.id = il.order_line_id
                LEFT JOIN orders o ON o.id = ol.order_id
                LEFT JOIN parties pt ON pt.id = i.party_id
                LEFT JOIN products p ON p.id = ol.product_id
                WHERE {" AND ".join(_pre_sales_where)}
                ORDER BY i.invoice_date DESC, i.invoice_no DESC
                LIMIT 200
            """, _pre_sales_params) or []

        _prod_opts = sorted(
            {str(r.get("product_name") or "").split(" | ")[0] for r in _rows if r.get("product_name")} |
            {str(r.get("product_name") or "").split(" | ")[0] for r in _pre_sales_rows if r.get("product_name")}
        )
        _idx_opts = sorted(
            {_row_index_value(r) for r in _rows if _row_index_value(r)} |
            {str(r.get("lens_index") or "") for r in _pre_sales_rows if str(r.get("lens_index") or "")}
        )
        _coat_opts = sorted(
            {_row_coating_value(r) for r in _rows if _row_coating_value(r)} |
            {str(r.get("coating") or "") for r in _pre_sales_rows if str(r.get("coating") or "")}
        )
        _batch_opts = sorted({_batch_for_row(r) for r in _rows if _batch_for_row(r) and _show_batch_for_row(r)})
        _exp_opts = sorted({_expiry_for_row(r) for r in _rows if _expiry_for_row(r) and _show_batch_for_row(r)})

        with st.container(border=True):
            st.markdown(
                "<div style='font-size:0.74rem;color:#111827;font-weight:800;margin-bottom:4px'>"
                "Product movement filters</div>",
                unsafe_allow_html=True,
            )
            _mf1, _mf2, _mf3, _mf4, _mf5 = st.columns([2.5, 1.2, 1.6, 1.6, 1.4])
            _prod_pick = _mf1.selectbox("Product", ["All"] + _prod_opts, key="stkpr_prod_filter")
            _idx_pick = _mf2.selectbox("Index", ["All"] + _idx_opts, key="stkpr_idx_filter")
            _coat_pick = _mf3.selectbox("Coating", ["All"] + _coat_opts, key="stkpr_coat_filter")
            _batch_pick = _mf4.selectbox("Batch", ["All"] + _batch_opts, key="stkpr_batch_filter")
            _exp_pick = _mf5.selectbox("Expiry", ["All"] + _exp_opts, key="stkpr_exp_filter")

        _filtered_by_move = []
        for _row in _rows:
            if _prod_pick != "All" and str(_row.get("product_name") or "").split(" | ")[0] != _prod_pick:
                continue
            if _idx_pick != "All" and _row_index_value(_row) != _idx_pick:
                continue
            if _coat_pick != "All" and _row_coating_value(_row) != _coat_pick:
                continue
            if _batch_pick != "All" and _batch_for_row(_row) != _batch_pick:
                continue
            if _exp_pick != "All" and _expiry_for_row(_row) != _exp_pick:
                continue
            _filtered_by_move.append(_row)
        _rows = _filtered_by_move
        if not _rows and _movement_mode == "Purchase":
            st.info("No purchase lines match the selected product / index / coating / batch / expiry filters.")
            return

        _move_summary = {}
        for _row in _rows:
            _pn = str(_row.get("product_name") or "—").split(" | ")[0]
            _key = (_pn, _row_index_value(_row), _row_coating_value(_row))
            _rec = _move_summary.setdefault(
                _key,
                {"product": _pn, "index": _key[1], "coating": _key[2], "qty": 0.0, "value": 0.0, "docs": set()},
            )
            _qty_m = float(_row.get("pa_recv_qty") or _row.get("quantity") or 0)
            _rec["qty"] += _qty_m
            _rec["value"] += _qty_m * float(_row.get("pa_price") or 0)
            _doc_m = str(_row.get("pa_invoice_no") or _row.get("pa_challan") or _row.get("order_no") or "").strip()
            if _doc_m:
                _rec["docs"].add(_doc_m)

        with st.expander("🚀 Movement summary — fastest procured in current filter", expanded=False):
            _top_rows = sorted(_move_summary.values(), key=lambda x: (x["qty"], x["value"]), reverse=True)[:20]
            if _top_rows:
                for _mv in _top_rows:
                    st.markdown(
                        f"<div style='display:flex;gap:10px;flex-wrap:wrap;color:#111827;font-size:0.78rem;"
                        f"border-bottom:1px solid #e5e7eb;padding:4px 2px'>"
                        f"<b>{_mv['product']}</b>"
                        + (f"<span>Idx {_mv['index']}</span>" if _mv["index"] else "")
                        + (f"<span>{_mv['coating']}</span>" if _mv["coating"] else "")
                        + f"<span>Qty {_mv['qty']:g}</span>"
                        + f"<span>₹{_mv['value']:,.2f}</span>"
                        + f"<span>{len(_mv['docs'])} document(s)</span>"
                        + "</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("No movement in this filter.")

        _sales_params = {"dfrom": str(_sp_from), "dto": str(_sp_to)}
        _sales_where = [
            "i.invoice_date BETWEEN %(dfrom)s AND %(dto)s",
            "COALESCE(i.is_deleted, FALSE) = FALSE",
            "COALESCE(il.is_deleted, FALSE) = FALSE",
            "UPPER(COALESCE(i.status,'')) NOT IN ('VOID','CANCELLED')",
        ]
        if _sp_search.strip():
            _sales_where.append(
                "(LOWER(COALESCE(il.product_name, p.product_name, '')) LIKE %(ss)s "
                "OR LOWER(COALESCE(i.invoice_no,'')) LIKE %(ss)s "
                "OR LOWER(COALESCE(o.order_no,'')) LIKE %(ss)s "
                "OR regexp_replace(LOWER(COALESCE(o.order_no,'')), '[^a-z0-9]', '', 'g') LIKE %(ss_norm)s "
                "OR LOWER(COALESCE(pt.party_name, o.party_name, o.patient_name, '')) LIKE %(ss)s)"
            )
            _sales_params["ss"] = f"%{_sp_search.strip().lower()}%"
            _sales_params["ss_norm"] = f"%{_scan_norm(_sp_search)}%"
            _sales_params["ss_exact"] = _sp_search.strip()
            _sales_where[-1] = _sales_where[-1][:-1] + (
                " OR ol.sph::text = %(ss_exact)s "
                "OR COALESCE(ol.lens_params->>'sph','') = %(ss_exact)s)"
            )
        if _party_filter_pick != "All":
            _sales_where.append("COALESCE(pt.party_name, o.party_name, o.patient_name, '') = %(sparty)s")
            _sales_params["sparty"] = _party_filter_pick
        if _prod_pick != "All":
            _sales_where.append("COALESCE(il.product_name, p.product_name, '') = %(spn)s")
            _sales_params["spn"] = _prod_pick
        if _idx_pick != "All":
            _sales_where.append(
                "COALESCE(ol.lens_params->>'lens_index', ol.lens_params->>'index', p.index_value::text, '') = %(sidx)s"
            )
            _sales_params["sidx"] = _idx_pick
        if _coat_pick != "All":
            _sales_where.append(
                "COALESCE(ol.lens_params->>'coating', ol.lens_params->>'coating_type', p.coating, '') = %(scoat)s"
            )
            _sales_params["scoat"] = _coat_pick

        if _movement_mode == "Purchase":
            _sales_rows = []
        else:
            _sales_rows = _q2(f"""
            SELECT
                i.invoice_date::text AS doc_date,
                i.invoice_no,
                COALESCE(c.challan_no, '') AS challan_no,
                COALESCE(o.order_no, '') AS order_no,
                COALESCE(pt.party_name, o.party_name, o.patient_name, '') AS party_name,
                p.id::text AS product_id,
                COALESCE(il.product_name, p.product_name, '—') AS product_name,
                COALESCE(il.eye_side, ol.eye_side, '') AS eye_side,
                COALESCE(il.quantity, 1) AS qty,
                COALESCE(il.line_total, il.total_price, il.quantity * il.unit_price, 0)::numeric AS amount,
                COALESCE(ol.lens_params->>'lens_index', ol.lens_params->>'index', p.index_value::text, '') AS lens_index,
                COALESCE(ol.lens_params->>'coating', ol.lens_params->>'coating_type', p.coating, '') AS coating
            FROM invoice_lines il
            JOIN invoices i ON i.id = il.invoice_id
            LEFT JOIN challans c ON c.id = i.challan_id
            LEFT JOIN order_lines ol ON ol.id = il.order_line_id
            LEFT JOIN orders o ON o.id = ol.order_id
            LEFT JOIN parties pt ON pt.id = i.party_id
            LEFT JOIN products p ON p.id = ol.product_id
            WHERE {" AND ".join(_sales_where)}
            ORDER BY i.invoice_date DESC, i.invoice_no DESC
            LIMIT 200
            """, _sales_params) or []

        with st.expander(f"🟦 Sales movement — {_sales_rows and len(_sales_rows) or 0} line(s)", expanded=False):
            if not _sales_rows:
                st.caption("No sales lines in this date/search filter.")
            else:
                _sale_summary = {}
                for _sr in _sales_rows:
                    _pn = str(_sr.get("product_name") or "—").split(" | ")[0]
                    _key = (_pn, str(_sr.get("lens_index") or ""), str(_sr.get("coating") or ""))
                    _rec = _sale_summary.setdefault(
                        _key,
                        {"product": _pn, "index": _key[1], "coating": _key[2], "qty": 0.0, "value": 0.0},
                    )
                    _rec["qty"] += float(_sr.get("qty") or 0)
                    _rec["value"] += float(_sr.get("amount") or 0)
                st.markdown("<b style='color:#1d4ed8'>Fastest sold in this filter</b>", unsafe_allow_html=True)
                for _sv in sorted(_sale_summary.values(), key=lambda x: (x["qty"], x["value"]), reverse=True)[:15]:
                    st.markdown(
                        f"<div style='display:flex;gap:10px;flex-wrap:wrap;color:#111827;font-size:0.78rem;"
                        f"border-bottom:1px solid #e5e7eb;padding:4px 2px'>"
                        f"<span style='background:#dbeafe;color:#1d4ed8;padding:1px 6px;border-radius:3px;font-weight:800'>SALE</span>"
                        f"<b>{_sv['product']}</b>"
                        + (f"<span>Idx {_sv['index']}</span>" if _sv["index"] else "")
                        + (f"<span>{_sv['coating']}</span>" if _sv["coating"] else "")
                        + f"<span>Qty {_sv['qty']:g}</span>"
                        + f"<span>₹{_sv['value']:,.2f}</span>"
                        + "</div>",
                        unsafe_allow_html=True,
                    )

        if _sales_rows:
            _sales_groups = {}
            for _sr in _sales_rows:
                _sales_groups.setdefault(str(_sr.get("invoice_no") or "NO-INVOICE"), []).append(_sr)
            with st.expander(f"🟦 Sales invoice-wise view — {len(_sales_groups)} invoice(s)", expanded=False):
                for _inv_no, _lines in sorted(_sales_groups.items(), key=lambda kv: kv[0]):
                    _party = next((str(x.get("party_name") or "").strip() for x in _lines if x), "")
                    _total_qty = sum(float(x.get("qty") or 0) for x in _lines)
                    _total_value = sum(float(x.get("amount") or 0) for x in _lines)
                    _date = next((str(x.get("doc_date") or "")[:10] for x in _lines if x.get("doc_date")), "")
                    st.markdown(
                        f"<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;"
                        f"padding:8px 10px;margin:8px 0 4px'>"
                        f"<span style='color:#1d4ed8;font-weight:900'>{_inv_no}</span>"
                        f"<span style='color:#111827;font-size:0.75rem;margin-left:8px'>{_party}</span>"
                        f"<span style='color:#1d4ed8;font-size:0.75rem;margin-left:8px'>{len(_lines)} line(s)</span>"
                        f"<span style='color:#111827;font-size:0.75rem;margin-left:8px'>{_date}</span>"
                        f"<span style='color:#111827;font-size:0.75rem;margin-left:8px'>Qty {_total_qty:g}</span>"
                        f"<span style='color:#166534;font-size:0.75rem;margin-left:8px'>₹{_total_value:,.2f}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                    with st.popover("🖨️ View / print sales invoice", use_container_width=False):
                        try:
                            from modules.billing.smart_print import render_smart_invoice
                            _sales_html_doc = render_smart_invoice(_inv_no, return_html=True)
                            if _sales_html_doc:
                                st.download_button(
                                    "⬇️ Download invoice HTML",
                                    _sales_html_doc,
                                    file_name=f"sales_invoice_{_inv_no.replace('/', '-')}.html",
                                    mime="text/html",
                                    key=f"stkpr_sales_html_{_inv_no.replace('/', '_')}",
                                    use_container_width=True,
                                )
                                if st.button(
                                    "🖨️ Open print view",
                                    key=f"stkpr_sales_print_{_inv_no.replace('/', '_')}",
                                    use_container_width=True,
                                ):
                                    from modules.printing.print_opener import open_html_print
                                    open_html_print(_sales_html_doc, f"sales_invoice_{_inv_no.replace('/', '-')}.html")
                            else:
                                st.caption("Invoice HTML not generated.")
                        except Exception as _sales_print_err:
                            st.caption(f"Invoice print unavailable: {_sales_print_err}")
                    for _line in _lines:
                        _pname_s = str(_line.get("product_name") or "—").split(" | ")[0]
                        st.markdown(
                            f"<div style='display:flex;gap:8px;flex-wrap:wrap;font-size:0.74rem;"
                            f"padding:2px 8px;color:#111827'>"
                            f"<b>{_pname_s}</b>"
                            f"<span>{str(_line.get('eye_side') or '').upper()}</span>"
                            + (f"<span>Idx {str(_line.get('lens_index') or '')}</span>" if _line.get("lens_index") else "")
                            + (f"<span>{str(_line.get('coating') or '')}</span>" if _line.get("coating") else "")
                            + f"<span>Qty {float(_line.get('qty') or 0):g}</span>"
                            + f"<span>₹{float(_line.get('amount') or 0):,.2f}</span>"
                            + "</div>",
                            unsafe_allow_html=True,
                        )

        _product_ids = sorted({
            str(x.get("product_id") or "")
            for x in list(_rows) + list(_sales_rows or [])
            if str(x.get("product_id") or "").strip()
        })
        _stock_rows = []
        if _product_ids:
            _stock_rows = _q2("""
                SELECT product_id::text AS product_id,
                       COALESCE(SUM(quantity), 0)::numeric AS closing_qty,
                       COALESCE(SUM(allocated_qty), 0)::numeric AS allocated_qty
                FROM inventory_stock
                WHERE product_id = ANY(%(pids)s::uuid[])
                  AND COALESCE(is_active, TRUE) = TRUE
                GROUP BY product_id
            """, {"pids": _product_ids}) or []
        _closing_qty = sum(float(r.get("closing_qty") or 0) for r in _stock_rows)
        _allocated_qty = sum(float(r.get("allocated_qty") or 0) for r in _stock_rows)
        _proc_qty = sum(float(r.get("pa_recv_qty") or r.get("quantity") or 0) for r in _rows)
        _sold_qty = sum(float(r.get("qty") or 0) for r in (_sales_rows or []))
        _opening_qty = _closing_qty - _proc_qty + _sold_qty
        with st.expander("📦 Opening / Closing stock — current filter", expanded=False):
            _oc1, _oc2, _oc3, _oc4, _oc5 = st.columns(5)
            _oc1.metric("Opening", f"{_opening_qty:g}")
            _oc2.metric("Purchased", f"{_proc_qty:g}")
            _oc3.metric("Sold", f"{_sold_qty:g}")
            _oc4.metric("Closing", f"{_closing_qty:g}")
            _oc5.metric("Allocated", f"{_allocated_qty:g}")
            st.caption("Opening is calculated from current closing stock, less purchases in this filter, plus sales in this filter.")

        try:
            import html as _html_mov
            _purch_html = "".join(
                "<tr><td>{}</td><td>{}</td><td>{}</td><td>{:g}</td><td>{}</td></tr>".format(
                    _html_mov.escape(str(r.get("pa_document_date") or "")[:10]),
                    _html_mov.escape(str(r.get("pa_invoice_no") or r.get("pa_challan") or r.get("order_no") or "")),
                    _html_mov.escape(str(r.get("product_name") or "")),
                    float(r.get("pa_recv_qty") or r.get("quantity") or 0),
                    _html_mov.escape(_rate_text(r)),
                )
                for r in _rows[:500]
            )
            _sales_html = "".join(
                "<tr><td>{}</td><td>{}</td><td>{}</td><td>{:g}</td><td>₹{:,.2f}</td></tr>".format(
                    _html_mov.escape(str(r.get("doc_date") or "")[:10]),
                    _html_mov.escape(str(r.get("invoice_no") or "")),
                    _html_mov.escape(str(r.get("product_name") or "")),
                    float(r.get("qty") or 0),
                    float(r.get("amount") or 0),
                )
                for r in (_sales_rows or [])[:500]
            )
            _report_html = f"""<!doctype html><html><head><meta charset='utf-8'>
<style>
body{{font-family:Arial,sans-serif;color:#111;margin:20px}}
h1{{font-size:20px}} h2{{font-size:15px;margin-top:18px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#111827;color:white;text-align:left;padding:6px}}
td{{border-bottom:1px solid #ddd;padding:6px}}
.kpis span{{display:inline-block;border:1px solid #ddd;padding:8px 12px;margin:4px;border-radius:5px}}
.print{{position:fixed;right:10px;top:10px}} @media print{{.print{{display:none}}}}
</style></head><body>
<button class='print' onclick='window.print()'>Print</button>
<h1>Product Movement Report</h1>
<div>Period: {_html_mov.escape(str(_sp_from))} to {_html_mov.escape(str(_sp_to))}</div>
<div class='kpis'>
<span>Opening: {_opening_qty:g}</span><span>Purchased: {_proc_qty:g}</span>
<span>Sold: {_sold_qty:g}</span><span>Closing: {_closing_qty:g}</span>
<span>Allocated: {_allocated_qty:g}</span>
</div>
<h2>Purchase / Procurement Lines</h2>
<table><thead><tr><th>Date</th><th>Purchase Ref</th><th>Product</th><th>Qty</th><th>Rate</th></tr></thead><tbody>{_purch_html}</tbody></table>
<h2>Sales Lines</h2>
<table><thead><tr><th>Date</th><th>Sales Invoice</th><th>Product</th><th>Qty</th><th>Amount</th></tr></thead><tbody>{_sales_html}</tbody></table>
</body></html>"""
            st.download_button(
                "🖨️ Download / Print Product Movement HTML",
                _report_html,
                file_name="product_movement_report.html",
                mime="text/html",
                key="stkpr_movement_print",
                use_container_width=True,
            )
        except Exception as _print_e:
            st.caption(f"Movement print pending: {_print_e}")

    # ── Load all suppliers once ───────────────────────────────────────────
    _all_sups = _q2(
        "SELECT id::text AS id, party_name, COALESCE(mobile,'') AS mobile, "
        "COALESCE(email,'') AS email "
        "FROM parties WHERE UPPER(COALESCE(party_type,'')) IN "
        "('SUPPLIER','VENDOR','CONTACT_LENS_SUPPLIER','FRAME_SUPPLIER') "
        "AND COALESCE(is_active,TRUE)=TRUE ORDER BY party_name"
    )
    _sup_ids   = [s["id"]         for s in _all_sups]
    _sup_names = {s["id"]: s["party_name"] for s in _all_sups}
    _sup_mob   = {s["id"]: s.get("mobile","") for s in _all_sups}

    # ── Selection state ───────────────────────────────────────────────────
    if "sp_selected" not in st.session_state:
        st.session_state.sp_selected = set()

    _n_total     = len(_rows)
    _n_purchased = sum(1 for r in _rows if r["_purchased"])
    _n_pending   = _n_total - _n_purchased
    _bulk_mode_key = "sp_bulk_attach_mode" if procured_view else "sp_purchase_select_mode"
    if _bulk_mode_key not in st.session_state:
        st.session_state[_bulk_mode_key] = False
    _selection_mode = bool(st.session_state.get(_bulk_mode_key))
    if not _selection_mode and st.session_state.sp_selected:
        st.session_state.sp_selected.clear()

    # ── Summary bar ──────────────────────────────────────────────────────
    _sb1, _sb2, _sb3, _sb4 = st.columns([2, 2, 2, 2])
    _sb1.metric("Total Lines",      _n_total)
    _sb2.metric("Pending Purchase", _n_pending,   delta_color="inverse")
    _sb3.metric("Purchased",        _n_purchased)
    _sb4.metric("Selected",         len(st.session_state.sp_selected) if _selection_mode else 0)

    if procured_view:
        _mode_label = (
            "✅ Bulk attach / correction mode ON"
            if _selection_mode else
            "📎 Bulk attach / correction mode"
        )
        if st.button(_mode_label, key="sp_bulk_attach_toggle", use_container_width=True):
            st.session_state[_bulk_mode_key] = not _selection_mode
            if _selection_mode:
                st.session_state.sp_selected.clear()
            st.rerun()
        if _selection_mode:
            st.caption("Select procured lines only when you want to attach one invoice file or correct common supplier/document details together.")

    if procured_view and _rows:
        _groups = {}
        for _row in _rows:
            _ref = str(_row.get("pa_invoice_no") or _row.get("pa_challan") or _row.get("order_no") or "NO-REF").strip()
            _groups.setdefault(_ref, []).append(_row)
        with st.expander(f"📑 Purchase invoice-wise view — {len(_groups)} document(s)", expanded=False):
            for _ref, _lines in sorted(_groups.items(), key=lambda kv: kv[0]):
                _supplier = next((str(x.get("pa_sup_name") or x.get("patient_name") or "").strip() for x in _lines if x), "")
                _total_qty = sum(float(x.get("pa_recv_qty") or x.get("quantity") or 0) for x in _lines)
                _total_value = sum(float(x.get("pa_price") or 0) * float(x.get("pa_recv_qty") or x.get("quantity") or 1) for x in _lines)
                _direct_count = sum(1 for x in _lines if x.get("direct_invoice_upload"))
                _attached_path = next(
                    (
                        _invoice_path_from_notes(x.get("pa_invoice_file"))
                        for x in _lines
                        if _invoice_path_from_notes(x.get("pa_invoice_file"))
                    ),
                    "",
                )
                st.markdown(
                    f"<div style='background:#0f172a;border:1px solid #334155;border-radius:6px;"
                    f"padding:8px 10px;margin:8px 0 4px'>"
                    f"<span style='color:#f8fafc;font-weight:900'>{_ref}</span>"
                    f"<span style='color:#cbd5e1;font-size:0.75rem;margin-left:8px'>{_supplier}</span>"
                    f"<span style='color:#93c5fd;font-size:0.75rem;margin-left:8px'>{len(_lines)} line(s)</span>"
                    f"<span style='color:#e2e8f0;font-size:0.75rem;margin-left:8px'>Qty {_total_qty:g}</span>"
                    f"<span style='color:#22c55e;font-size:0.75rem;margin-left:8px'>₹{_total_value:,.2f}</span>"
                    + ("<span style='background:#f59e0b22;color:#fbbf24;font-size:0.65rem;"
                       "padding:1px 6px;border-radius:3px;margin-left:8px;font-weight:800'>"
                       "DIRECT PURCHASE / UPLOAD</span>" if _direct_count else "")
                    + "</div>",
                    unsafe_allow_html=True,
                )
                if _attached_path:
                    _render_attachment_popover(
                        "📎 View / download attached purchase invoice",
                        _attached_path,
                        f"stkpr_purch_attach_{_ref.replace('/', '_').replace(' ', '_')}",
                    )
                for _line in _lines:
                    _pname_sum = str(_line.get("product_name") or "—").split(" | ")[0]
                    _batch_sum = _batch_for_row(_line)
                    _expiry_sum = _expiry_for_row(_line)
                    _date_sum = str(_line.get("pa_document_date") or "")[:10]
                    _show_batch = _show_batch_for_row(_line)
                    st.markdown(
                        f"<div style='display:flex;gap:8px;flex-wrap:wrap;font-size:0.74rem;"
                        f"padding:2px 8px;color:#111827'>"
                        f"<b>{_pname_sum}</b>"
                        f"<span>{str(_line.get('eye_side') or '').upper()}</span>"
                        + (f"<span>{_date_sum}</span>" if _date_sum else "")
                        + (f"<span>Idx {_row_index_value(_line)}</span>" if _row_index_value(_line) else "")
                        + (f"<span>{_row_coating_value(_line)}</span>" if _row_coating_value(_line) else "")
                        + f"<span>Qty {float(_line.get('pa_recv_qty') or _line.get('quantity') or 0):g}</span>"
                        + f"<span>{_rate_text(_line)}</span>"
                        + (f"<span>Batch {_batch_sum}</span>" if _show_batch and _batch_sum else "")
                        + (f"<span>Exp {_expiry_sum}</span>" if _show_batch and _expiry_sum else "")
                        + ("<span style='color:#92400e;font-weight:800'>From Direct Purchase tab</span>"
                           if _line.get("direct_purchase_entry") else "")
                        + ("<span style='color:#92400e;font-weight:800'>Not from Procurement Queue</span>"
                           if _line.get("direct_invoice_upload") and not _line.get("direct_purchase_entry") else "")
                        + "</div>",
                        unsafe_allow_html=True,
                    )

    # ── Select-all toggle ────────────────────────────────────────────────
    if _selection_mode or not procured_view:
        _sel_col1, _sel_col2, _sel_col3 = st.columns([2, 2, 4])
        with _sel_col1:
            _sel_all_label = "☑ Select Visible Procured" if procured_view else "☑ Select All Pending"
            if st.button(_sel_all_label, key="sp_sel_all",
                         use_container_width=True):
                for _r in _rows:
                    if procured_view or not _r["_purchased"]:
                        st.session_state.sp_selected.add(_r["line_id"])
                st.rerun()
        with _sel_col2:
            if st.button("☐ Clear Selection", key="sp_clear_all",
                         use_container_width=True):
                st.session_state.sp_selected.clear()
                st.rerun()

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Item list ─────────────────────────────────────────────────────────
    for _ri, _row in enumerate(_rows):
        _lid      = _row["line_id"]
        _ono      = _row["order_no"]
        _pname    = str(_row.get("product_name","")).split(" | ")[0]
        _eye      = str(_row.get("eye_side","")).upper()
        _qty      = int(_row.get("quantity") or 1)
        _cat      = str(_row.get("category") or _row.get("main_group",""))
        _lp       = _row["_lp"]
        _purch    = _row["_purchased"]
        _pa_id    = str(_row.get("pa_id") or "")

        # Power string
        _pwr_parts = []
        for _pkey, _plbl in [("sph","SPH"),("cyl","CYL"),("axis","AX"),("add_power","ADD")]:
            _pv = _row.get(_pkey) or _lp.get(_pkey)
            if _pv is not None and str(_pv) not in ("","None","0","0.0"):
                try:
                    _pf = float(_pv)
                    if _pkey == "cyl" and abs(_pf) < 0.01: continue
                    if _pkey == "axis": _pwr_parts.append(f"AX {int(_pf)}°")
                    elif _pkey == "add_power" and _pf > 0: _pwr_parts.append(f"ADD +{_pf:.2f}")
                    else: _pwr_parts.append(f"{_plbl} {_pf:+.2f}")
                except: pass
        _pwr_str = "  ".join(_pwr_parts) or ""

        # Coating / index
        _coat  = str(_lp.get("coating") or _lp.get("coating_type") or "").strip()
        _idx   = str(_lp.get("lens_index") or _lp.get("index") or "").strip()
        _route = str(_lp.get("manufacturing_route","")).upper()

        # Route badge color
        _route_clr = {"STOCK":"#0d9488","VENDOR":"#3b82f6","EXTERNAL_LAB":"#a855f7","DIRECT_INVOICE":"#f59e0b"}.get(_route,"#64748b")
        _direct_tag = bool(_row.get("direct_invoice_upload"))
        _direct_purchase_tag = bool(_row.get("direct_purchase_entry"))
        _batch_no = _batch_for_row(_row)
        _expiry = _expiry_for_row(_row)
        _show_batch = _show_batch_for_row(_row)
        _rate_label = _rate_text(_row)
        _doc_date = str(_row.get("pa_document_date") or "")[:10]

        _is_selected = _selection_mode and (_lid in st.session_state.sp_selected)
        _card_bg     = "#052e16" if _purch else ("#0f172a" if not _is_selected else "#0f1e3a")
        _card_border = "#22c55e" if _purch else ("#3b82f6" if _is_selected else "#1e293b")

        with st.container():
            _lc1, _lc2 = st.columns([1, 11]) if (_selection_mode or not procured_view) else (None, st.container())

            # Checkbox: pending items can be selected for purchase entry.
            # In Procured view, purchased items can be selected for shared
            # invoice-file attachment / audit correction.
            if _selection_mode or not procured_view:
                with _lc1:
                    if (not _purch) or procured_view:
                        _chk = st.checkbox(
                            "sel",
                            value=_is_selected,
                            key=f"sp_chk_{_lid}",
                            label_visibility="collapsed",
                        )
                        if _chk != _is_selected:
                            if _chk:
                                st.session_state.sp_selected.add(_lid)
                            else:
                                st.session_state.sp_selected.discard(_lid)
                            st.rerun()
                    else:
                        st.markdown(
                            "<div style='text-align:center;color:#22c55e;"
                            "font-size:1.1rem;padding-top:6px'>✅</div>",
                            unsafe_allow_html=True
                        )
            else:
                _chk = False

            with _lc2:
                st.markdown(
                    f"<div style='background:{_card_bg};border:1px solid {_card_border};"
                    f"border-radius:6px;padding:7px 12px;margin-bottom:4px'>"
                    f"<div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap'>"
                    f"<span style='color:#cbd5e1;font-size:0.7rem;font-family:monospace;font-weight:700'>{_ono}</span>"
                    f"<span style='color:#f8fafc;font-weight:900;font-size:0.9rem'>{_pname}</span>"
                    + (f"<span style='color:#e2e8f0;font-size:0.72rem;font-weight:800'>{_eye}</span>" if _eye else "")
                    + (f"<span style='background:{_route_clr}22;color:{_route_clr};"
                       f"font-size:0.65rem;padding:1px 6px;border-radius:3px'>{_route}</span>")
                    + ("<span style='background:#f59e0b22;color:#fbbf24;"
                       "font-size:0.65rem;padding:1px 6px;border-radius:3px;font-weight:800'>"
                       "FROM DIRECT PURCHASE TAB</span>" if _direct_purchase_tag else "")
                    + ("<span style='background:#f59e0b22;color:#fbbf24;"
                       "font-size:0.65rem;padding:1px 6px;border-radius:3px;font-weight:800'>"
                       "NOT FROM PROCUREMENT QUEUE</span>" if _direct_tag and not _direct_purchase_tag else "")
                    + (f"<span style='color:#fde68a;font-size:0.7rem;font-weight:700'>"
                       f"Batch {_batch_no}</span>" if _show_batch and _batch_no else "")
                    + (f"<span style='color:#fde68a;font-size:0.7rem;font-weight:700'>"
                       f"Exp {_expiry}</span>" if _show_batch and _expiry else "")
                    + (f"<span style='color:#cbd5e1;font-size:0.72rem;font-weight:700'>{_doc_date}</span>" if _doc_date else "")
                    + (f"<code style='color:#60a5fa;font-size:0.75rem'>{_pwr_str}</code>" if _pwr_str else "")
                    + (f"<span style='color:#cbd5e1;font-size:0.72rem;font-weight:700'>{_coat}</span>" if _coat else "")
                    + (f"<span style='color:#cbd5e1;font-size:0.72rem;font-weight:700'>Idx {_idx}</span>" if _idx else "")
                    + f"<span style='color:#e2e8f0;font-size:0.72rem;font-weight:800'>Qty {_qty}</span>"
                    + ("  <span style='color:#22c55e;font-size:0.72rem;font-weight:700'>"
                       f"✅ {_rate_label} · {_row.get('pa_challan') or _row.get('pa_invoice_no','—')}"
                       "</span>" if _purch else "")
                    + f"</div></div>",
                    unsafe_allow_html=True,
                )
    # ── Bulk Purchase Panel (shown when items selected) ───────────────────
    _sel_lines = [r for r in _rows if r["line_id"] in st.session_state.sp_selected]

    if _sel_lines:
        st.markdown("---")
        _bulk_title = (
            f"📎 Attach / update {len(_sel_lines)} procured line(s)"
            if procured_view else
            f"💳 Purchase {len(_sel_lines)} selected line(s)"
        )
        st.markdown(
            f"<div style='color:#93c5fd;font-weight:700;font-size:0.9rem;margin-bottom:8px'>"
            f"{_bulk_title}</div>",
            unsafe_allow_html=True,
        )

        with st.container(border=True):
            # ── Supplier selector (preferred at top) ──────────────────────
            # Find the most common preferred supplier among selected lines
            _pref_sid = ""
            _pref_cnts = {}
            for _sl in _sel_lines:
                _ps = str(
                    _sl.get("pa_sup_id") if procured_view else _sl.get("preferred_supplier_id", "")
                ).strip()
                if _ps:
                    _pref_cnts[_ps] = _pref_cnts.get(_ps, 0) + 1
            if _pref_cnts:
                _pref_sid = max(_pref_cnts, key=_pref_cnts.get)

            # Build ordered supplier list — preferred first
            _ordered_sids = []
            if _pref_sid and _pref_sid in _sup_ids:
                _ordered_sids.append(_pref_sid)
            _ordered_sids += [s for s in _sup_ids if s != _pref_sid]
            _sup_labels = {}
            for _sid in _ordered_sids:
                _nm = _sup_names.get(_sid, _sid)
                _sup_labels[_sid] = ("⭐ " + _nm) if _sid == _pref_sid else _nm

            _pa1, _pa2 = st.columns([3, 2])
            with _pa1:
                _sel_sup_idx = 0
                if _pref_sid in _ordered_sids:
                    _sel_sup_idx = _ordered_sids.index(_pref_sid)
                _chosen_sup = st.selectbox(
                    "Supplier",
                    options=_ordered_sids,
                    index=_sel_sup_idx,
                    format_func=lambda x: _sup_labels.get(x, x),
                    key="sp_bulk_sup",
                )
                _chosen_sup_name = _sup_names.get(_chosen_sup, "")
                _chosen_sup_mob  = _sup_mob.get(_chosen_sup, "")

            with _pa2:
                st.markdown(
                    f"<div style='background:#0f172a;border:1px solid #1e293b;"
                    f"border-radius:6px;padding:8px 12px;margin-top:4px'>"
                    f"<div style='color:#64748b;font-size:0.68rem'>SUPPLIER</div>"
                    f"<div style='color:#e2e8f0;font-weight:700;font-size:0.85rem'>"
                    f"{_chosen_sup_name or '—'}</div>"
                    + (f"<div style='color:#94a3b8;font-size:0.72rem'>📞 {_chosen_sup_mob}</div>"
                       if _chosen_sup_mob else "")
                    + "</div>",
                    unsafe_allow_html=True,
                )

            # ── Invoice / Challan fields ──────────────────────────────────
            _pb1, _pb2, _pb3 = st.columns([3, 2, 2])
            _existing_refs = [
                str(x.get("pa_invoice_no") or x.get("pa_challan") or "").strip()
                for x in _sel_lines
                if str(x.get("pa_invoice_no") or x.get("pa_challan") or "").strip()
            ]
            _default_ref = _existing_refs[0] if _existing_refs and len(set(_existing_refs)) == 1 else ""
            _inv_no    = _pb1.text_input("Invoice / Challan No", key="sp_inv_no",
                                          value=_default_ref,
                                          placeholder="Supplier invoice or challan number")
            _inv_date  = _pb2.date_input("Invoice Date", value=_pd.date.today(),
                                          key="sp_inv_date", format="DD/MM/YYYY")
            _doc_type  = _pb3.selectbox("Document Type",
                                         ["Challan", "Invoice", "Both"],
                                         key="sp_doc_type")
            # Batch/expiry are entered per-line below.

            _line_prices = {}
            if procured_view:
                for _sl in _sel_lines:
                    _line_prices[_sl["line_id"]] = float(_sl.get("pa_price") or 0)
                st.caption(
                    "Attachment-only mode: existing purchase prices, quantities, batch and expiry are preserved."
                )
            else:
                # Per-line price entry
                st.markdown(
                    "<div style='color:#94a3b8;font-size:0.75rem;margin:8px 0 4px'>Unit prices:</div>",
                    unsafe_allow_html=True,
                )
                _pc_cols = st.columns(min(len(_sel_lines), 4))
                for _pi, _sl in enumerate(_sel_lines):
                    _def_price = float(_sl.get("pa_price") or
                                        _sl.get("last_purchase_price") or 0)
                    _pl = str(_sl.get("product_name","")).split(" | ")[0][:20]
                    _pe = str(_sl.get("eye_side","")).upper()
                    _col_i = _pi % min(len(_sel_lines), 4)
                    _line_prices[_sl["line_id"]] = _pc_cols[_col_i].number_input(
                        f"{_pe} {_pl}",
                        value=_def_price,
                        min_value=0.0, step=0.5, format="%.2f",
                        key=f"sp_price_{_sl['line_id'][:8]}",
                    )

                _total_val = sum(_line_prices.get(r["line_id"],0) * int(r.get("quantity",1))
                                 for r in _sel_lines)
                st.markdown(
                    f"<div style='text-align:right;color:#10b981;font-weight:700;"
                    f"font-size:0.9rem;margin:4px 0'>Total: ₹{_total_val:,.2f}</div>",
                    unsafe_allow_html=True,
                )

            # ── Invoice file upload ───────────────────────────────────────
            st.markdown(
                "<div style='color:#94a3b8;font-size:0.75rem;margin:8px 0 4px'>"
                "📎 Attach invoice PDF / image (optional — stored for audit):</div>",
                unsafe_allow_html=True,
            )
            _inv_file = st.file_uploader(
                "Invoice document",
                type=["pdf","png","jpg","jpeg","webp"],
                key="sp_inv_file",
                label_visibility="collapsed",
            )
            _inv_file_path = ""
            _inv_preview_path = ""
            if _inv_file is not None:
                try:
                    _inv_file_path, _inv_preview_path = _save_invoice_upload(
                        _inv_file,
                        _inv_no.strip() or "draft",
                    )
                except Exception as _scan_save_err:
                    st.warning(f"Invoice scan preview save warning: {_scan_save_err}")
                    _inv_file_path = ""
                # Preview
                if _inv_file.type.startswith("image"):
                    st.image(
                        _inv_preview_path if _inv_preview_path else _inv_file,
                        caption=_inv_file.name,
                        use_container_width=True,
                    )
                else:
                    st.markdown(
                        f"<div style='background:#1e293b;border-radius:6px;"
                        f"padding:6px 12px;font-size:0.8rem;color:#94a3b8'>"
                        f"📄 {_inv_file.name} ({len(_inv_file.getvalue())//1024} KB)</div>",
                        unsafe_allow_html=True,
                    )
                if _inv_file_path:
                    with st.expander("🔍 Scan / OCR Review", expanded=True):
                        try:
                            from modules.procurement.invoice_review_panel import render_invoice_scan_review_panel
                            render_invoice_scan_review_panel(
                                _inv_file_path,
                                supplier_id=str(_chosen_sup or ""),
                                supplier_name=str(_chosen_sup_name or ""),
                                key_prefix="sp_scan_review",
                            )
                        except Exception as _scan_err:
                            st.warning(f"Scan review unavailable: {_scan_err}")

            # ── Save buttons ──────────────────────────────────────────────
            # ── Per-line batch/expiry before save ─────────────────────
            if not procured_view:
                st.markdown('<div style="color:#a5b4fc;font-size:0.75rem">Batch / Expiry per line:</div>', unsafe_allow_html=True)
                for _sl_be in _sel_lines:
                    _lid_be = _sl_be['line_id']
                    _ey_be = str(_sl_be.get('eye_side','')).upper()
                    _pn_be = str(_sl_be.get('product_name','')).split(' | ')[0][:22]
                    st.markdown(f'<div style="font-size:0.72rem;color:#94a3b8">{_ey_be} {_pn_be}</div>', unsafe_allow_html=True)
                    _be1, _be2 = st.columns([2,2])
                    _be1.text_input('📦 Batch No.', key=f'sp_ln_batch_{_lid_be}', placeholder='CL required', label_visibility='collapsed')
                    _be2.date_input('📅 Expiry', value=None, key=f'sp_ln_expiry_{_lid_be}', format='DD/MM/YYYY', label_visibility='collapsed')
            _sv1, _sv2, _sv3 = st.columns([2, 2, 1])
            _do_save = _sv1.button("💾 Save Attachment" if procured_view else "💾 Save Purchase", key="sp_save",
                                    type="primary", use_container_width=True)
            _do_save_print = (
                False if procured_view else
                _sv2.button("💾 Save + Print", key="sp_save_print", use_container_width=True)
            )
            _sv3.button("✕ Cancel", key="sp_cancel", use_container_width=True,
                        on_click=lambda: st.session_state.sp_selected.clear())

            if _do_save or _do_save_print:
                if not _inv_no.strip():
                    st.error("❌ Invoice / Challan number is required.")
                elif not _chosen_sup:
                    st.error("❌ Please select a supplier.")
                elif (not procured_view) and any(float(_line_prices.get(r["line_id"], 0) or 0) <= 0 for r in _sel_lines):
                    _missing_price = [
                        str(r.get("product_name") or "Line").split(" | ")[0]
                        for r in _sel_lines
                        if float(_line_prices.get(r["line_id"], 0) or 0) <= 0
                    ]
                    st.error("❌ Purchase price is required for: " + ", ".join(_missing_price[:4]))
                else:
                    try:
                        from modules.core.date_guard import validate_not_future
                        _ok_dt, _msg_dt = validate_not_future(_inv_date, "Purchase document date")
                    except Exception as _dg_e:
                        _ok_dt, _msg_dt = False, f"Date validation failed: {_dg_e}"
                    if not _ok_dt:
                        st.error(_msg_dt)
                        return

                    _save_ok  = True
                    _saved_n  = 0

                    # ── Save invoice file if uploaded ─────────────────────
                    if _inv_file is not None:
                        try:
                            if not _inv_file_path:
                                _inv_file_path, _inv_preview_path = _save_invoice_upload(
                                    _inv_file,
                                    _inv_no.strip() or "draft",
                                )
                            # Also store base64 in DB (for portability)
                            _inv_b64 = _b64.b64encode(_inv_file.getvalue()).decode()
                        except Exception as _fe:
                            st.warning(f"File save warning: {_fe}")
                            _inv_b64 = ""
                    else:
                        _inv_b64 = ""

                    # ── Upsert purchase_acknowledgements per line ─────────
                    for _sl in _sel_lines:
                        _lid_s  = _sl["line_id"]
                        _qty_s  = int(_sl.get("quantity") or 1)
                        _price_s = _line_prices.get(_lid_s, 0)
                        _total_s = _price_s * _qty_s
                        _pa_id_s = _sl.get("pa_id") or ""

                        try:
                            if procured_view and _pa_id_s:
                                # Attachment/audit update only. Do not rewrite
                                # price, quantity, batch or expiry on already
                                # procured rows; old records may carry legacy
                                # unit math and must not be recalculated here.
                                _w2("""
                                    UPDATE purchase_acknowledgements SET
                                        supplier_id    = COALESCE(%(sid)s::uuid, purchase_acknowledgements.supplier_id),
                                        supplier_name  = COALESCE(NULLIF(%(sname)s,''), purchase_acknowledgements.supplier_name),
                                        challan_no     = CASE
                                            WHEN %(doc_type)s IN ('Challan','Both') THEN %(ref)s
                                            ELSE purchase_acknowledgements.challan_no
                                        END,
                                        invoice_no     = CASE
                                            WHEN %(doc_type)s IN ('Invoice','Both') THEN %(ref)s
                                            ELSE purchase_acknowledgements.invoice_no
                                        END,
                                        document_date  = COALESCE(%(idate)s::date, purchase_acknowledgements.document_date),
                                        notes          = COALESCE(NULLIF(%(fpath)s,''), purchase_acknowledgements.notes),
                                        updated_at     = NOW()
                                    WHERE id = %(paid)s::uuid
                                """, {
                                    "sid": _chosen_sup,
                                    "sname": _chosen_sup_name,
                                    "doc_type": _doc_type,
                                    "ref": _inv_no.strip(),
                                    "idate": str(_inv_date),
                                    "fpath": _inv_file_path or None,
                                    "paid": _pa_id_s,
                                })
                            elif _pa_id_s:
                                # Update existing
                                _w2("""
                                    UPDATE purchase_acknowledgements SET
                                        supplier_id       = %(sid)s::uuid,
                                        supplier_name     = %(sname)s,
                                        challan_no        = %(chal)s,
                                        invoice_no        = %(inv)s,
                                        document_date     = COALESCE(%(idate)s::date, purchase_acknowledgements.document_date),
                                        purchase_price    = %(price)s,
                                        total_value       = %(total)s,
                                        received_qty      = %(qty)s,
                                        billing_status    = 'PURCHASE_ACKED',
                                        notes             = COALESCE(NULLIF(%(fpath)s,''), purchase_acknowledgements.notes),
                                        batch_no          = COALESCE(NULLIF(%(batch)s,''), purchase_acknowledgements.batch_no),
                                        expiry_date       = COALESCE(NULLIF(%(expiry)s,'')::date, purchase_acknowledgements.expiry_date),
                                        updated_at        = NOW()
                                    WHERE id = %(paid)s::uuid
                                """, {
                                    "sid":   _chosen_sup,
                                    "sname": _chosen_sup_name,
                                    "chal":  _inv_no.strip() if _doc_type in ("Challan","Both") else "",
                                    "inv":   _inv_no.strip() if _doc_type in ("Invoice","Both") else "",
                                    "idate": str(_inv_date),
                                    "price": _price_s,
                                    "total": _total_s,
                                    "qty":   _qty_s,
                                    "fpath": _inv_file_path or None,
                                    "batch": st.session_state.get(f"sp_ln_batch_{_lid_s}", "").strip(),
                                    "expiry": str(st.session_state.get(f"sp_ln_expiry_{_lid_s}", "") or ""),
                                    "paid":  _pa_id_s,
                                })
                            else:
                                # Insert new
                                _w2("""
                                    INSERT INTO purchase_acknowledgements (
                                        order_line_id, supplier_id, supplier_name,
                                        challan_no, invoice_no, document_date,
                                        purchase_price, total_value, received_qty,
                                        billing_status, notes, batch_no, expiry_date, acknowledged_at
                                    ) VALUES (
                                        %(lid)s::uuid, %(sid)s::uuid, %(sname)s,
                                        %(chal)s, %(inv)s, %(idate)s::date,
                                        %(price)s, %(total)s, %(qty)s,
                                        'PURCHASE_ACKED', %(fpath)s, %(batch)s, NULLIF(%(expiry)s,'')::date,
                                        NOW()
                                    )
                                """, {
                                    "lid":   _lid_s,
                                    "sid":   _chosen_sup,
                                    "sname": _chosen_sup_name,
                                    "chal":  _inv_no.strip() if _doc_type in ("Challan","Both") else "",
                                    "inv":   _inv_no.strip() if _doc_type in ("Invoice","Both") else "",
                                    "idate": str(_inv_date),
                                    "price": _price_s,
                                    "total": _total_s,
                                    "qty":   _qty_s,
                                    "fpath": _inv_file_path or None,
                                    "batch": st.session_state.get(f"sp_ln_batch_{_lid_s}", "").strip(),
                                    "expiry": str(st.session_state.get(f"sp_ln_expiry_{_lid_s}", "") or ""),
                                })
                            _saved_n += 1

                            if procured_view:
                                continue

                            # Auto-advance the line's route-specific lifecycle field
                            # so it leaves the procurement queue and goes straight to
                            # billing-ready. This matches procurement_queue.py — once
                            # the bill is recorded, advancement is implicit.
                            try:
                                _lp_done = _sl.get("lens_params") or {}
                                if isinstance(_lp_done, str):
                                    import json as _jpd
                                    try: _lp_done = _jpd.loads(_lp_done)
                                    except Exception: _lp_done = {}
                                _lp_done = dict(_lp_done or {})
                                _route_done = str(
                                    _lp_done.get("manufacturing_route") or
                                    _sl.get("route") or ""
                                ).upper()
                                _flip = False
                                if _route_done in ("VENDOR", "EXTERNAL_LAB"):
                                    _lp_done["supplier_stage"] = "READY_FOR_BILLING"
                                    _flip = True
                                elif _route_done == "STOCK":
                                    _lp_done["replenishment_status"] = "PROCURED"
                                    _flip = True
                                if _flip:
                                    import json as _jpd2
                                    _w2(
                                        "UPDATE order_lines SET lens_params=%(lp)s::jsonb "
                                        "WHERE id=%(lid)s::uuid",
                                        {"lp": _jpd2.dumps(_lp_done), "lid": _lid_s},
                                    )
                            except Exception as _adv_e:
                                st.caption(f"Stage advance pending: {_adv_e}")
                        except Exception as _se2:
                            st.error(f"Save error for {_sl.get('product_name','?')}: {_se2}")
                            _save_ok = False

                    if _save_ok and _saved_n > 0:
                        try:
                            if not procured_view:
                                from modules.procurement.procurement_ledger import record_procurement_receipt
                                record_procurement_receipt(
                                    line_items=[
                                        {
                                            "line_id": r["line_id"],
                                            "qty": int(r.get("quantity") or 1),
                                            "price": _line_prices.get(r["line_id"], 0),
                                            "batch_no": st.session_state.get(f"sp_ln_batch_{r.get('line_id','')}", "").strip(),
                                            "expiry_date": str(st.session_state.get(f"sp_ln_expiry_{r.get('line_id','')}", "") or ""),
                                        }
                                        for r in _sel_lines
                                    ],
                                    supplier_id=str(_chosen_sup or ""),
                                    supplier_name=str(_chosen_sup_name or ""),
                                    document_no=_inv_no.strip(),
                                    document_type=str(_doc_type or "Invoice").upper(),
                                    document_date=str(_inv_date),
                                    invoice_file_path=_inv_file_path or "",
                                    source="STOCK_PROCUREMENT",
                                )
                        except Exception as _pl_save_err:
                            st.warning(f"Procurement ledger mirror pending: {_pl_save_err}")
                        if procured_view:
                            st.success(
                                f"✅ Attachment/details updated for {_saved_n} procured line(s) · "
                                f"Supplier: {_chosen_sup_name} · Ref: {_inv_no.strip()}"
                            )
                        else:
                            st.success(
                                f"✅ Purchase recorded for {_saved_n} line(s) · "
                                f"Supplier: {_chosen_sup_name} · "
                                f"Ref: {_inv_no.strip()} · Total: ₹{_total_val:,.2f}"
                            )
                        if _inv_file_path:
                            st.caption(f"📄 Invoice stored: {_inv_file_path}")
                        st.session_state.sp_selected.clear()

                        if _do_save_print:
                            # Inline print summary
                            _print_lines = "\n".join(
                                f"  {str(r.get('eye_side','')).upper()} {str(r.get('product_name','')).split(' | ')[0][:30]:<30} "
                                f"Qty:{int(r.get('quantity',1))}  ₹{_line_prices.get(r['line_id'],0):.2f}"
                                for r in _sel_lines
                            )
                            st.code(
                                f"PURCHASE RECORD — Parakh Optical\n"
                                f"Date: {str(_inv_date)}\n"
                                f"Supplier: {_chosen_sup_name}\n"
                                f"Ref No: {_inv_no.strip()}\n"
                                f"{'─'*50}\n"
                                f"{_print_lines}\n"
                                f"{'─'*50}\n"
                                f"TOTAL: ₹{_total_val:,.2f}\n",
                                language=None,
                            )
                        import time as _pt; _pt.sleep(0.6)
                        st.rerun()

    # ── Purchased items — view stored invoices ────────────────────────────
    _purchased_rows = [r for r in _rows if r["_purchased"]]
    if _purchased_rows:
        with st.expander(
            f"📋 Purchased — {len(_purchased_rows)} line(s) (tap to view / audit)",
            expanded=False
        ):
            for _pr in _purchased_rows:
                _pr_name  = str(_pr.get("product_name","")).split(" | ")[0]
                _pr_eye   = str(_pr.get("eye_side","")).upper()
                _pr_price = float(_pr.get("pa_price") or 0)
                _pr_status = str(_pr.get("pa_status") or "").upper()
                _pr_inv_ref = str(_pr.get("pa_invoice_no") or "").strip()
                _pr_chal_ref = str(_pr.get("pa_challan") or "").strip()
                _pr_ref = _pr_inv_ref if _pr_status == "INVOICED" and _pr_inv_ref else (_pr_chal_ref or _pr_inv_ref or "—")
                _pr_stage = "Moved to Register" if _pr_status == "INVOICED" else "Challan / Purchase Ack"
                _pr_sup   = _pr.get("pa_sup_name","—")

                with st.container(border=True):
                    _vc1, _vc2 = st.columns([5, 2])
                    with _vc1:
                        st.markdown(
                            f"<div style='font-size:0.82rem'>"
                            f"<span style='color:#e2e8f0;font-weight:700'>{_pr_name}</span>"
                            + (f" <span style='color:#64748b'>{_pr_eye}</span>" if _pr_eye else "")
                            + f"<br><span style='color:#94a3b8;font-size:0.72rem'>"
                            f"🏭 {_pr_sup} · {_pr_stage}: {_pr_ref} · ₹{_pr_price:.2f}</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                    with _vc2:
                        # ── Smart file resolution ────────────────────────
                        # notes may be: NULL | "/path/file.pdf" | "batch note | /path/file.pdf" | "b64:..."
                        _raw_file = _pr.get("pa_invoice_file") or ""

                        # Extract actual file path from notes (handle "note | path" format)
                        _resolved_path = ""
                        if _raw_file:
                            # Split on " | " and find the part that looks like a file path
                            for _part in reversed(_raw_file.split(" | ")):
                                _part = _part.strip()
                                if "/" in _part or _part.lower().endswith((".pdf",".png",".jpg",".jpeg",".webp")):
                                    _resolved_path = _part
                                    break
                            # If nothing looks like a path but the whole value is a path
                            if not _resolved_path and ("/" in _raw_file or _raw_file.lower().endswith((".pdf",".png",".jpg",".jpeg",".webp"))):
                                _resolved_path = _raw_file.strip()

                        _file_shown = False

                        # 1. Try disk path
                        if _resolved_path and _pos.path.exists(_resolved_path):
                            if _resolved_path.lower().endswith(".pdf"):
                                with open(_resolved_path,"rb") as _pvf:
                                    st.download_button(
                                        "📄 Invoice PDF",
                                        _pvf.read(),
                                        file_name=_pos.path.basename(_resolved_path),
                                        mime="application/pdf",
                                        key=f"sp_dl_{_pr['line_id'][:8]}",
                                        use_container_width=True,
                                    )
                            else:
                                st.image(_resolved_path, use_container_width=True)
                            _file_shown = True

                        # 2. Try b64 embedded (legacy or procurement_queue saves)
                        elif _raw_file.startswith("b64:") or _pr.get("_lp",{}).get("invoice_b64"):
                            _b64_data = (
                                _raw_file[4:] if _raw_file.startswith("b64:")
                                else _pr["_lp"].get("invoice_b64","")
                            )
                            if _b64_data:
                                try:
                                    st.download_button(
                                        "📄 Invoice (DB)",
                                        _b64.b64decode(_b64_data),
                                        file_name=f"invoice_{_pr['line_id'][:8]}.pdf",
                                        mime="application/pdf",
                                        key=f"sp_db_{_pr['line_id'][:8]}",
                                        use_container_width=True,
                                    )
                                    _file_shown = True
                                except Exception:
                                    pass

                        # 3. Path stored but file missing from disk (moved/deleted)
                        elif _resolved_path:
                            st.caption(f"⚠️ File moved: {_pos.path.basename(_resolved_path)}")

                        # 4. No file at all — show inline upload to attach one now
                        if not _file_shown:
                            _attach_key = f"sp_attach_{_pr['line_id'][:8]}"
                            _uploaded_now = st.file_uploader(
                                "📎 Attach invoice",
                                type=["pdf","png","jpg","jpeg","webp"],
                                key=_attach_key,
                                label_visibility="collapsed",
                            )
                            if _uploaded_now:
                                try:
                                    _save_path2, _preview2 = _save_invoice_upload(
                                        _uploaded_now,
                                        str(_pr.get("pa_challan") or _pr.get("pa_invoice_no") or "ref"),
                                    )
                                    # Save path back to notes
                                    _w2(
                                        "UPDATE purchase_acknowledgements SET notes=%(p)s WHERE id=%(id)s::uuid",
                                        {"p": _save_path2, "id": _pr.get("pa_id","")},
                                    )
                                    st.success("✅ Invoice attached")
                                    st.rerun()
                                except Exception as _ae:
                                    st.error(f"Attach failed: {_ae}")
                            else:
                                st.caption("No invoice file — upload above to attach")


# ══════════════════════════════════════════════════════════════════════════
# PROCUREMENT RX — unified central purchase manager
# ══════════════════════════════════════════════════════════════════════════


# ── Public entry points ──────────────────────────────────────────────────

def render_stock_pipeline(*args, **kwargs):
    return _render_stock_pipeline(*args, **kwargs)

def render_stock_procurement(*args, **kwargs):
    return _render_stock_procurement(*args, **kwargs)

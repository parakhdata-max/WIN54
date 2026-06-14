"""
modules/ui/frame_batch_loader.py
=================================
Frame Stock Manager — two tabs:

  ✏️  EDIT  — Download current frame stock → edit prices/qty → upload back
              Scan Code / Item Code is the scanning key — duplicate codes in upload = error

  ➕  ADD   — Upload your BatchData Excel (multi-sheet, StartCode, colour normalisation)
              New scan codes only — existing scan code in ADD tab = error (use EDIT tab to change price)
"""

import io
import uuid
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime


# ── Column map: Excel header → internal key ──────────────────────────────────
COL_MAP = {
    # ── Identity ──────────────────────────────────────────────────────────────
    "Product":              "product_name",
    "Product Name":         "product_name",
    "🔒 Product Name":      "product_name",
    "Brand":                "brand",
    "🔒 Brand":             "brand",
    "Barcode":              "barcode",
    "Scan Code":            "barcode",
    "Item Code":            "barcode",
    "Scan Code / Item Code": "barcode",
    "Scan Code / Item Code (for scanning)": "barcode",
    "SKU Code":             "barcode",
    "SKU CODE":             "barcode",
    "SKU":                  "barcode",
    "🔒 SKU Code":          "barcode",
    "🔒 Scan Code / Item Code": "barcode",
    "🔒 Scan Code / Item Code (for scanning)": "barcode",
    # ── Dimensions ────────────────────────────────────────────────────────────
    "ASize":                "size_a",
    "A Size":               "size_a",
    "A Size (mm)":          "size_a",
    "BSize":                "size_b",
    "B size":               "size_b",
    "B Size":               "size_b",
    "B Size (mm)":          "size_b",
    "DBL":                  "dbl",
    "DBL (mm)":             "dbl",
    "TempleLength":         "temple_length",
    "Temple Length":        "temple_length",
    "Temple Length (mm)":   "temple_length",
    # ── Material / Look ───────────────────────────────────────────────────────
    "BaseMaterial":         "base_material",
    "Base Material":        "base_material",
    "Material":             "base_material",
    "Finish":               "finish",
    "Colour":               "colour",
    "Color":                "colour",
    "ColourMix":            "colour_mix",
    "Colour Mix":           "colour_mix",
    "colour MIX":           "colour_mix",
    "Color Mix":            "colour_mix",
    "TempleColour":         "temple_colour",
    "Temple Colour":        "temple_colour",
    "Temple Color":         "temple_colour",
    "shape":                "shape",
    "Shape":                "shape",
    # ── Business ──────────────────────────────────────────────────────────────
    "StartCode":            "location",
    "Location":             "location",
    "Location / Box":       "location",
    "FrameGroup":           "frame_group",
    "Frame Group":          "frame_group",
    "FrameSeq":             "frame_seq",
    "Frame Seq":            "frame_seq",
    "Gender":               "gender",
    "FrameType":            "frame_type",
    "Frame Type":           "frame_type",
    "FRAME TYPE":           "frame_type",
    "Model":                "model",
    # ── Pricing ───────────────────────────────────────────────────────────────
    "Qty":                  "qty",
    "Quantity":             "qty",
    "CostPrice":            "purchase_rate",
    "Purchase price":       "purchase_rate",
    "Purchase Price":       "purchase_rate",
    "Purchase Price ₹":     "purchase_rate",
    "PurchasePrice":        "purchase_rate",
    "purchase_rate":        "purchase_rate",
    "selling_price":        "selling_price",
    "SellingPrice":         "selling_price",
    "Selling Price":        "selling_price",
    "Selling Price ₹":      "selling_price",
    "MRP":                  "mrp",
    "MRP ₹":                "mrp",
    # ── System ────────────────────────────────────────────────────────────────
    "ImagePath":            "image_path",
    "Image Path":           "image_path",
    "IsActive":             "is_active",
    "Is Active":            "is_active",
    "Active (Y/N)":         "is_active",
    "Active":               "is_active",
    # GSTPercent excluded — set via Product Master (auto-created with correct GST)
}

FINISH_NORM = {
    "matt": "Matt", "matte": "Matt", "mattt": "Matt", "matty": "Matt",
    "m att": "Matt", "m,att": "Matt",
    "glossy": "Glossy", "gloosy": "Glossy", "glosy": "Glossy",
}

SHAPE_NORM = {
    "square": "Square", "rectangle": "Rectangle", "rectancle": "Rectangle",
    "round": "Round", "oval": "Oval", "aviator": "Aviator",
    "cat eye": "Cat Eye", "butterfly": "Butterfly", "hexagon": "Hexagon",
    "hexa": "Hexagon", "supra": "Supra",
    "new": "Other", "new design": "Other", "new shape": "Other",
    "newdesign": "Other", "bear shaped": "Other",
}


# ── Helpers ───────────────────────────────────────────────────────────────────
def _norm(val, lookup):
    if not val or (isinstance(val, float) and np.isnan(val)):
        return ""
    v = str(val).strip()
    return lookup.get(v.lower(), v.title())

def _safe_str(val) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return str(val).strip()

def _safe_float(val) -> float:
    try:
        v = float(val)
        return 0.0 if np.isnan(v) else v
    except (TypeError, ValueError):
        return 0.0

def _safe_int(val, default=0) -> int:
    try:
        v = float(val)
        return default if np.isnan(v) else int(v)
    except (TypeError, ValueError):
        return default

def _parse_is_active(raw) -> bool:
    if raw is None or (isinstance(raw, float) and raw != raw):
        return True
    return str(raw).strip().upper() in ("1", "Y", "YES", "TRUE", "T", "ACTIVE")


# ══════════════════════════════════════════════════════════════════════════════
# DB FETCH — current frame stock
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_frame_stock() -> pd.DataFrame:
    """Pull all frame stock from inventory_stock JOIN products."""
    try:
        from modules.sql_adapter import run_query
        rows = run_query("""
            SELECT
                COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,'')) AS barcode,
                p.product_name,
                COALESCE(p.brand,'')                               AS brand,
                COALESCE(p.model,'')                               AS model,
                COALESCE(NULLIF(s.colour,''), '')                  AS colour,
                COALESCE(NULLIF(s.colour_mix,''), '')              AS colour_mix,
                COALESCE(NULLIF(s.temple_colour,''), '')           AS temple_colour,
                COALESCE(NULLIF(s.base_material,''), '')           AS base_material,
                COALESCE(NULLIF(s.shape,''), '')                   AS shape,
                COALESCE(NULLIF(s.finish,''), '')                  AS finish,
                s.size_a                                           AS size_a,
                s.size_b                                           AS size_b,
                s.dbl                                              AS dbl,
                s.temple_length                                    AS temple_length,
                COALESCE(s.location,'')                            AS location,
                COALESCE(s.frame_group,'')                         AS frame_group,
                COALESCE(s.quantity,0)                             AS qty,
                COALESCE(s.purchase_rate,0)                        AS purchase_rate,
                COALESCE(s.selling_price,0)                        AS selling_price,
                COALESCE(s.mrp,0)                                  AS mrp,
                COALESCE(s.is_active,true)                         AS is_active
            FROM inventory_stock s
            JOIN products p ON p.id = s.product_id
            WHERE LOWER(COALESCE(p.main_group,'')) IN ('frames','frame','sunglasses')
              AND COALESCE(s.is_active, true) = true
            ORDER BY p.product_name, COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,''))
        """)
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)
    except Exception as ex:
        st.error(f"DB fetch error: {ex}")
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER — edit download
# ══════════════════════════════════════════════════════════════════════════════

def _build_edit_excel(df: pd.DataFrame) -> bytes:
    """
    Build a formatted Excel for editing.
    Scan Code / Item Code column is locked (grey). Price + qty columns are white (editable).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    LOCKED_COLS = {"barcode", "product_name", "brand"}
    EDITABLE_COLS = [
        # Identity (locked)
        "barcode", "product_name", "brand",
        # Frame specs (all editable — stored in inventory_stock per SKU)
        "colour", "colour_mix", "temple_colour",
        "base_material", "shape", "finish",
        "size_a", "size_b", "dbl", "temple_length",
        # Stock & pricing
        "location", "frame_group",
        "qty", "purchase_rate", "selling_price", "mrp", "is_active",
    ]

    # Only keep columns that exist
    cols = [c for c in EDITABLE_COLS if c in df.columns]
    export_df = df[cols].copy()

    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Frame Stock"

    HDR_LOCKED  = "6B6B6B"
    HDR_EDIT    = "1A3C5E"
    DATA_LOCKED = "F2F2F2"
    DATA_EDIT   = "FFFFFF"
    HDR_TEXT    = "FFFFFF"
    ALT_ROW     = "E8F4FD"

    FRIENDLY = {
        "barcode":      "🔒 Scan Code / Item Code (for scanning)",
        "product_name":  "🔒 Product Name",
        "brand":         "🔒 Brand",
        "colour":        "Colour",
        "colour_mix":    "Colour Mix",
        "temple_colour": "Temple Colour",
        "base_material": "Material",
        "shape":         "Shape",
        "finish":        "Finish",
        "size_a":        "A Size (mm)",
        "size_b":        "B Size (mm)",
        "dbl":           "DBL (mm)",
        "temple_length": "Temple Length (mm)",
        "location":      "Location / Box",
        "frame_group":   "Frame Group",
        "qty":           "Qty",
        "purchase_rate": "Purchase Price ₹",
        "selling_price": "Selling Price ₹",
        "mrp":           "MRP ₹",
        "is_active":     "Active (Y/N)",
    }

    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=FRIENDLY.get(col, col))
        is_locked = col in LOCKED_COLS
        cell.font      = Font(bold=True, color=HDR_TEXT, size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color=HDR_LOCKED if is_locked else HDR_EDIT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(16, len(FRIENDLY.get(col, col)) + 4)
    ws.row_dimensions[1].height = 28

    # Data rows
    for ri, (_, row) in enumerate(export_df.iterrows(), 2):
        alt = (ri % 2 == 0)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            # Convert bool is_active to Y/N for easier editing
            if col == "is_active":
                val = "Y" if val else "N"
            cell = ws.cell(row=ri, column=ci, value=val)
            is_locked = col in LOCKED_COLS
            if is_locked:
                cell.fill = PatternFill("solid", start_color=DATA_LOCKED)
                cell.font = Font(color="888888", size=9, name="Arial")
            else:
                cell.fill = PatternFill("solid", start_color=ALT_ROW if alt else DATA_EDIT)
                cell.font = Font(size=9, name="Arial")
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Instructions sheet
    ws2 = wb.create_sheet("📖 Instructions", 0)
    ws2.column_dimensions["A"].width = 70
    instructions = [
        ("🕶️ Frame Stock — EDIT FILE", "title"),
        ("", None),
        ("🔒 LOCKED columns (grey) — DO NOT change:", "head"),
        ("   • Scan Code / Item Code, Product Name, Brand, Colour, Material", "body"),
        ("   These are identity fields — changing them breaks the match to DB.", "body"),
        ("   Scan Code / Item Code is the code used by scanners in punching and stock search.", "body"),
        ("   Batch No is not the frame scan field; keep batch numbers only for true batch/lot stock.", "body"),
        ("", None),
        ("✅ EDITABLE columns (white/blue):", "head"),
        ("   • Location / Box  — update box/tray location", "body"),
        ("   • Qty             — update stock quantity", "body"),
        ("   • Purchase Price  — cost price (optional)", "body"),
        ("   • Selling Price   — wholesale price (optional)", "body"),
        ("   • MRP             — retail price (required)", "body"),
        ("   • Active          — Y = active, N = inactive", "body"),
        ("", None),
        ("⚠️ RULES:", "head"),
        ("   • Do NOT add or delete rows", "body"),
        ("   • Do NOT rename or move columns", "body"),
        ("   • Scan Code / Item Code is the scanning key — it must match exactly", "body"),
        ("   • To add new frames, use the ➕ Add New tab (separate upload)", "body"),
    ]
    TITLE_FILL  = PatternFill("solid", start_color="1A3C5E")
    for text, style in instructions:
        cell = ws2.cell(row=ws2.max_row + 1, column=1, value=text)
        if style == "title":
            cell.fill = TITLE_FILL
            cell.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
            ws2.row_dimensions[cell.row].height = 28
        elif style == "head":
            cell.font = Font(bold=True, size=11, name="Arial", color="1A3C5E")
        else:
            cell.font = Font(size=10, name="Arial")
        cell.alignment = Alignment(vertical="center", indent=0)

    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL READER — add flow (BatchData format)
# ══════════════════════════════════════════════════════════════════════════════

# Known non-data sheet names to skip
_SKIP_SHEETS = {"📖 guide", "guide", "_meta", "instructions", "📋 guide", "readme"}

def read_frame_excel(file_bytes: bytes) -> pd.DataFrame:
    """
    Read BatchData Excel — multi-sheet, first DATA sheet has header.
    Skips guide/instruction sheets (📖 Guide, _meta etc.).
    """
    xl = pd.ExcelFile(io.BytesIO(file_bytes))

    # Separate data sheets from guide/meta sheets
    all_sheets  = [s for s in xl.sheet_names if s.strip()]
    data_sheets = [s for s in all_sheets
                   if s.strip().lower() not in _SKIP_SHEETS
                   and not s.strip().startswith("_")]

    # If no data sheets found, fall back to all sheets
    if not data_sheets:
        data_sheets = all_sheets

    frames = []
    canonical_cols = None

    for i, sheet in enumerate(data_sheets):
        try:
            if i == 0:
                # First data sheet always has a header row
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet)
                canonical_cols = list(df.columns)
            else:
                # Continuation sheets — no header
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None)
                if df.empty:
                    continue
                first_row = str(df.iloc[0, 0]).strip().lower()
                if first_row in ("product", "product name", "sku", "skucode"):
                    df.columns = canonical_cols[:len(df.columns)]
                    df = df.iloc[1:].reset_index(drop=True)
                else:
                    if canonical_cols:
                        df.columns = canonical_cols[:len(df.columns)]
            df = df.dropna(how="all")
            frames.append(df)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)

    # ── Case-insensitive column mapping ───────────────────────────────────────
    # Build a lookup: lowercase(excel_header) → db_column_name
    # This means "Scan Code", "SKU Code", "SKU CODE", "sku code" all work equally.
    _col_map_lower = {k.lower(): v for k, v in COL_MAP.items()}
    _rename = {
        col: _col_map_lower[col.lower()]
        for col in combined.columns
        if col.lower() in _col_map_lower
    }
    combined = combined.rename(columns=_rename)

    # ── Strip template special rows ───────────────────────────────────────────
    # Row 0 of template = orange example row ("⚠ EXAMPLE ROW...")
    # Row 1 of template = grey notes row ("Required · Unique · e.g. D10001...")
    # Both must be removed before import regardless of whether user deleted them.
    if "barcode" in combined.columns:
        # Remove example row
        combined = combined[
            ~combined["barcode"].astype(str).str.upper().str.startswith("⚠")
        ]
        # Remove notes row — notes contain "·" (middle dot) as separator, not a valid SKU
        combined = combined[
            ~combined["barcode"].astype(str).str.contains("·|Required|example|delete",
                                                             case=False, na=False)
        ]
    # Also strip by MRP — notes rows have text in MRP, not a number
    if "mrp" in combined.columns:
        combined = combined[
            pd.to_numeric(combined["mrp"], errors="coerce").notna()
        ]

    if "finish" in combined.columns:
        combined["finish"] = combined["finish"].apply(lambda x: _norm(x, FINISH_NORM))
    if "shape" in combined.columns:
        combined["shape"] = combined["shape"].apply(lambda x: _norm(x, SHAPE_NORM))
    if "colour" in combined.columns:
        combined["colour"] = combined["colour"].apply(
            lambda x: "" if (isinstance(x, float) and np.isnan(x)) else str(x).strip().title()
        )
    if "colour_mix" in combined.columns:
        combined["colour_mix"] = combined["colour_mix"].apply(
            lambda x: "" if (isinstance(x, float) and np.isnan(x)) else str(x).strip().title()
        )
    return combined


# ══════════════════════════════════════════════════════════════════════════════
# MAIN RENDER
# ══════════════════════════════════════════════════════════════════════════════

def render_frame_batch_loader():
    # ── Run migration once per session to ensure new columns exist ────────────
    if not st.session_state.get("_frame_migration_done"):
        try:
            from modules.loaders.frame_migration import run_frame_migration
            run_frame_migration(silent=True)
            st.session_state["_frame_migration_done"] = True
        except Exception:
            pass

    st.markdown(
        "<div style='background:#0f172a;border-left:4px solid #7dd3fc;"
        "padding:10px 16px;border-radius:6px;margin-bottom:12px'>"
        "<b style='color:#7dd3fc;font-size:1rem'>🕶️ Frame Stock Manager</b>"
        "<span style='color:#94a3b8;font-size:0.78rem;margin-left:10px'>"
        "Edit prices · Add new stock · Scan Code / Item Code is the scanner key</span>"
        "</div>",
        unsafe_allow_html=True
    )

    tab_edit, tab_add = st.tabs(["✏️  Edit Existing Stock", "➕  Add New Frames"])

    with tab_edit:
        _render_edit_tab()

    with tab_add:
        _render_add_tab()


# ══════════════════════════════════════════════════════════════════════════════
# EDIT TAB
# ══════════════════════════════════════════════════════════════════════════════

def _render_edit_tab():
    st.markdown("#### ✏️ Edit Existing Frame Stock")
    st.caption(
        "Download current stock → edit prices, qty, location → upload back. "
        "**Scan Code / Item Code is locked** — this is the field used by scanners in punching and stock search."
    )

    # ── Step 1: Download ──────────────────────────────────────────────────────
    with st.expander("📥 Step 1 — Download Current Frame Stock", expanded=True):
        c1, c2 = st.columns([2, 3])
        with c1:
            # Step 1a: Prepare — fetches DB, builds Excel, stores in session_state
            if st.button("⬇️ Prepare Edit File", key="frame_dl_edit",
                         type="primary", use_container_width=True):
                with st.spinner("Fetching frame stock from DB..."):
                    _edf = _fetch_frame_stock()
                if _edf is None or _edf.empty:
                    st.warning(
                        "⚠️ No frame stock in DB yet. "
                        "Use the ➕ Add New Frames tab to import your BatchData Excel first."
                    )
                    st.session_state.pop("_frame_edit_bytes", None)
                else:
                    try:
                        st.session_state["_frame_edit_bytes"] = _build_edit_excel(_edf)
                        st.session_state["_frame_edit_count"] = len(_edf)
                        st.session_state["_frame_edit_ts"]    = datetime.now().strftime("%Y%m%d_%H%M%S")
                        st.success(f"✅ {len(_edf)} frames ready — click Save below")
                    except Exception as _be:
                        st.error(f"Build error: {_be}")

            # Step 1b: Save — always visible once bytes are ready, survives reruns
            if st.session_state.get("_frame_edit_bytes"):
                _cnt = st.session_state["_frame_edit_count"]
                _ts  = st.session_state["_frame_edit_ts"]
                st.download_button(
                    label=f"💾 Save Edit File  ({_cnt} frames)",
                    data=st.session_state["_frame_edit_bytes"],
                    file_name=f"FRAME_EDIT_{_ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="frame_dl_save_btn",
                    use_container_width=True,
                )
                st.caption("Edit the white/blue columns → upload in Step 2 below.")
        with c2:
            st.info(
                "🔒 **Locked** (grey): Scan Code / Item Code, Product Name, Brand\n\n"
                "📷 **Scanning field**: Scan Code / Item Code. Do not use Batch No for frame scan codes.\n\n"
                "✏️ **Editable** (white): Colour, ColourMix, TempleColour, Material, "
                "BSize, Location, FrameGroup, Qty, Purchase Price, Selling Price, MRP, Active"
            )

    # ── Step 2: Upload ────────────────────────────────────────────────────────
    with st.expander("📤 Step 2 — Upload Edited File", expanded=True):
        uploaded = st.file_uploader(
            "Upload your edited Frame Stock file",
            type=["xlsx"],
            key="frame_edit_upload",
            help="Only upload files downloaded from Step 1 above."
        )
        if uploaded:
            _handle_edit_upload(uploaded)


def _handle_edit_upload(uploaded):
    """Process edit upload — Scan Code / Item Code is key, validate no duplicates, preview changes."""
    from modules.sql_adapter import run_query, run_write

    file_bytes = uploaded.read()

    # Read the uploaded file
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Frame Stock")
    except Exception:
        st.error("❌ Could not read 'Frame Stock' sheet — make sure you uploaded a file downloaded from Step 1.")
        return

    # Detect file type — EDIT file has a locked scan-code column; ADD template has Barcode.
    cols = list(df.columns)
    is_edit_file   = any("🔒" in str(c) for c in cols)
    is_add_template = "Barcode" in cols or "skucode" in [str(c).lower() for c in cols]

    if is_add_template and not is_edit_file:
        st.error(
            "❌ This looks like the **Add New template** (FRAME_ADD_TEMPLATE_...) "
            "not an Edit file. \n\n"
            "For editing prices, use **Step 1** above to download the current stock, "
            "then upload that file here."
        )
        return

    # Rename friendly headers back to internal keys
    reverse_friendly = {
        # Edit file headers (from _build_edit_excel FRIENDLY map)
        "🔒 SKU Code":       "barcode",
        "🔒 Scan Code / Item Code": "barcode",
        "🔒 Scan Code / Item Code (for scanning)": "barcode",
        "🔒 Product Name":   "product_name",
        "🔒 Brand":          "brand",
        "Colour":            "colour",
        "Colour Mix":        "colour_mix",
        "Temple Colour":     "temple_colour",
        "Material":          "base_material",
        "B Size (mm)":       "size_b",
        "Location / Box":    "location",
        "Frame Group":       "frame_group",
        "Qty":               "qty",
        "Purchase Price ₹":  "purchase_rate",
        "Selling Price ₹":   "selling_price",
        "MRP ₹":             "mrp",
        "Active (Y/N)":      "is_active",
    }
    df = df.rename(columns=reverse_friendly)

    if "barcode" not in df.columns:
        st.error(
            "❌ Scan Code / Item Code column not found. \n\n"
            "Make sure you are uploading a file downloaded from **Step 1 — Download Current Frame Stock** above, "
            "not the Add New template or any other file."
        )
        return

    # ── Validation ────────────────────────────────────────────────────────────
    issues = []

    # 1. Duplicate scan codes in uploaded file
    dup_skus = df[df.duplicated("barcode", keep=False) & df["barcode"].notna()]
    if not dup_skus.empty:
        issues.append(f"❌ **{len(dup_skus)} duplicate scan code(s)** in uploaded file — each code must appear once only: "
                      f"`{'`, `'.join(dup_skus['barcode'].dropna().unique()[:5])}`")

    # 2. Missing MRP
    no_mrp = df[df["mrp"].isna() | (pd.to_numeric(df["mrp"], errors="coerce") <= 0)]
    if not no_mrp.empty:
        issues.append(f"⚠️ {len(no_mrp)} row(s) have missing or zero MRP — they will be skipped.")

    if any(i.startswith("❌") for i in issues):
        for issue in issues:
            if issue.startswith("❌"):
                st.error(issue)
            else:
                st.warning(issue)
        st.stop()
        return
    for issue in issues:
        st.warning(issue)

    # ── Match to DB and show change preview ───────────────────────────────────
    all_skus = [_safe_str(s).upper() for s in df["barcode"].dropna()]
    if not all_skus:
        st.error("No valid Scan Code / Item Code values found.")
        return

    # Fetch current DB state for these SKUs
    placeholders = ",".join(["%s"] * len(all_skus))
    current_rows = run_query(f"""
        SELECT COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,'')) AS barcode,
               COALESCE(s.quantity,0)      AS qty,
               COALESCE(s.purchase_rate,0) AS purchase_rate,
               COALESCE(s.selling_price,0) AS selling_price,
               COALESCE(s.mrp,0)           AS mrp,
               COALESCE(s.location,'')     AS location,
               s.is_active
        FROM inventory_stock s
        JOIN products p ON p.id = s.product_id
        WHERE LOWER(COALESCE(p.main_group,'')) IN ('frames','frame','sunglasses')
          AND UPPER(COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,''))) IN ({placeholders})
    """, tuple(all_skus)) or []

    db_map = {str(r["barcode"]).upper(): r for r in current_rows}

    changes = []
    not_found = []
    # qty for frames: 0 = sold/out, 1 = in stock — no other value makes sense per-SKU
    PRICE_COLS = ["qty", "purchase_rate", "selling_price", "mrp",
                  "location", "frame_group",
                  "colour", "colour_mix", "temple_colour",
                  "base_material", "shape", "finish",
                  "size_a", "size_b", "dbl", "temple_length",
                  "is_active"]

    for _, row in df.iterrows():
        sku = _safe_str(row.get("barcode")).upper()
        if not sku:
            continue

        if sku not in db_map:
            not_found.append(sku)
            continue

        db_row = db_map[sku]
        row_changes = []

        for col in PRICE_COLS:
            excel_val = row.get(col)
            db_val    = db_row.get(col)

            if col == "is_active":
                new_val = _parse_is_active(excel_val)
                old_val = bool(db_val)
                if new_val != old_val:
                    row_changes.append({"SKU": sku, "Field": col,
                                        "Old": "Y" if old_val else "N",
                                        "New": "Y" if new_val else "N"})
            elif col == "location":
                new_val = _safe_str(excel_val)
                old_val = _safe_str(db_val)
                if new_val and new_val != old_val:
                    row_changes.append({"SKU": sku, "Field": col, "Old": old_val, "New": new_val})
            else:
                try:
                    new_val = float(excel_val) if pd.notna(excel_val) else None
                    old_val = float(db_val) if db_val is not None else 0.0
                    if new_val is not None and abs(new_val - old_val) > 0.001:
                        row_changes.append({"SKU": sku, "Field": col,
                                            "Old": f"{old_val:.2f}", "New": f"{new_val:.2f}"})
                except (TypeError, ValueError):
                    pass

        changes.extend(row_changes)

    # ── Show results ──────────────────────────────────────────────────────────
    if not_found:
        st.warning(
            f"⚠️ {len(not_found)} SKU(s) not found in DB — they will be skipped "
            f"(may have been deleted or wrong file): `{'`, `'.join(not_found[:5])}`"
        )

    if not changes:
        st.info("✅ No changes detected — uploaded file matches current database.")
        return

    st.markdown(f"### 🔍 Change Preview — {len(changes)} field change(s)")

    changes_df = pd.DataFrame(changes)
    # Colour-code by field type
    def _highlight(row):
        if row["Field"] in ("mrp", "selling_price", "purchase_rate"):
            return ["background-color:#fff7d6"] * len(row)
        if row["Field"] == "qty":
            return ["background-color:#e0f2fe"] * len(row)
        return [""] * len(row)

    st.dataframe(
        changes_df.style.apply(_highlight, axis=1),
        use_container_width=True, hide_index=True,
        column_config={
            "SKU":   st.column_config.TextColumn("Scan Code / Item Code"),
            "Field": st.column_config.TextColumn("Field"),
            "Old":   st.column_config.TextColumn("Current Value"),
            "New":   st.column_config.TextColumn("New Value"),
        }
    )
    st.caption("🟡 Yellow = price change  ·  🔵 Blue = qty change")

    # ── Confirm and apply ─────────────────────────────────────────────────────
    st.markdown("---")
    ac1, ac2 = st.columns(2)
    with ac1:
        if st.button("✅ Apply Changes", type="primary", key="frame_edit_apply",
                     use_container_width=True):
            _apply_edit(df, db_map)
    with ac2:
        if st.button("❌ Cancel", key="frame_edit_cancel", use_container_width=True):
            st.warning("Cancelled — no changes made.")


def _apply_edit(df: pd.DataFrame, db_map: dict):
    """Write approved changes to DB."""
    from modules.sql_adapter import run_query, run_write

    applied = skipped = 0
    errors = []

    for _, row in df.iterrows():
        sku = _safe_str(row.get("barcode")).upper()
        if not sku or sku not in db_map:
            skipped += 1
            continue
        try:
            mrp      = _safe_float(row.get("mrp"))
            purchase = row.get("purchase_rate")
            selling  = row.get("selling_price")
            # Qty for frames: only 0 (sold) or 1 (in stock) per SKU
            _raw_qty = row.get("qty")
            if pd.notna(_raw_qty):
                _qty_int = int(float(_raw_qty))
                qty_val = max(0, min(1, _qty_int))  # clamp to 0 or 1
            else:
                qty_val = None
            loc      = _safe_str(row.get("location"))
            is_act   = _parse_is_active(row.get("is_active"))

            size_b_val     = row.get("size_b")
            colour_mix_val = _safe_str(row.get("colour_mix"))
            temple_col_val = _safe_str(row.get("temple_colour"))
            grp_val        = _safe_str(row.get("frame_group"))

            # Extract all editable fields
            colour_val     = _safe_str(row.get("colour"))
            shape_val      = _safe_str(row.get("shape"))
            finish_val     = _safe_str(row.get("finish"))
            size_a_val     = row.get("size_a")
            dbl_val        = row.get("dbl")
            temple_val     = row.get("temple_length")

            run_write("""
                UPDATE inventory_stock SET
                    mrp           = %s,
                    purchase_rate = COALESCE(%s, purchase_rate),
                    selling_price = COALESCE(%s, selling_price),
                    quantity      = COALESCE(%s, quantity),
                    location      = COALESCE(NULLIF(%s,\'\'), location),
                    is_active     = %s,
                    colour        = COALESCE(NULLIF(%s,\'\'), colour),
                    colour_mix    = COALESCE(NULLIF(%s,\'\'), colour_mix),
                    temple_colour = COALESCE(NULLIF(%s,\'\'), temple_colour),
                    base_material = COALESCE(NULLIF(%s,\'\'), base_material),
                    shape         = COALESCE(NULLIF(%s,\'\'), shape),
                    finish        = COALESCE(NULLIF(%s,\'\'), finish),
                    size_a        = COALESCE(%s, size_a),
                    size_b        = COALESCE(%s, size_b),
                    dbl           = COALESCE(%s, dbl),
                    temple_length = COALESCE(%s, temple_length),
                    frame_group   = NULLIF(%s,\'\'),
                    updated_at    = NOW()
                WHERE UPPER(COALESCE(NULLIF(item_code,''), NULLIF(batch_no,''))) = %s
                  AND product_id IN (
                      SELECT id FROM products WHERE LOWER(COALESCE(main_group,'')) IN (\'frames\',\'frame\',\'sunglasses\')
                  )
            """, (
                mrp if mrp > 0 else None,
                float(purchase) if pd.notna(purchase) and float(purchase) > 0 else None,
                float(selling)  if pd.notna(selling)  and float(selling)  > 0 else None,
                int(float(qty_val)) if pd.notna(qty_val) else None,
                loc or None,
                is_act,
                colour_val     or None,
                colour_mix_val or None,
                temple_col_val or None,
                _safe_str(row.get("base_material")) or None,
                shape_val      or None,
                finish_val     or None,
                float(size_a_val) if pd.notna(size_a_val) else None,
                float(size_b_val) if pd.notna(size_b_val) else None,
                float(dbl_val)    if pd.notna(dbl_val)    else None,
                float(temple_val) if pd.notna(temple_val) else None,
                grp_val or None,
                sku
            ))
            applied += 1
        except Exception as ex:
            errors.append(f"{sku}: {ex}")
            skipped += 1

    if errors:
        st.warning(f"⚠️ {len(errors)} error(s)")
        with st.expander("Errors"):
            for e in errors:
                st.write(f"• {e}")
    st.success(f"✅ Applied changes to **{applied}** frame(s). {skipped} skipped.")


# ══════════════════════════════════════════════════════════════════════════════
# ADD TAB
# ══════════════════════════════════════════════════════════════════════════════

def _build_blank_template() -> bytes:
    """Build a blank Excel template with all frame columns and one example row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    COLUMNS = [
        # (Excel header,      example value,        notes)
        ("Barcode",           "D10001",             "Required · Unique · e.g. D10001"),
        ("Product",           "Butler 8305 Black",  "Required · Full product name"),
        ("Brand",             "Parakh",             "Brand name"),
        ("StartCode",         "D1",                 "Box/tray location code"),
        ("Colour",            "Black",              "Primary frame colour"),
        ("ColourMix",         "Gold",               "Secondary / accent colour"),
        ("TempleColour",      "Black",              "Temple arm colour (if different)"),
        ("shape",             "Square",             "Square / Round / Aviator / Rectangle / Cat Eye / Oval / Hexagon"),
        ("BaseMaterial",      "Plastic",            "Plastic / Metal / TR90 / Titanium / Acetate"),
        ("Finish",            "Matt",               "Matt / Glossy"),
        ("ASize",             "52",                 "A measurement — lens width (mm)"),
        ("BSize",             "38",                 "B measurement — lens height (mm)"),
        ("DBL",               "18",                 "Distance between lenses (mm)"),
        ("TempleLength",      "135",                "Temple arm length (mm)"),
        ("Qty",               "1",                  "Required · Stock quantity"),
        ("MRP",               "790",                "Required · Retail price ₹"),
        ("Purchase price",    "",                   "Cost price ₹ (optional)"),
        ("selling_price",     "",                   "Wholesale price ₹ (optional)"),
        ("FrameGroup",        "",                   "Dynamic tag e.g. Near Dead · Sale · Premium · Kids · New Arrival"),
        ("IsActive",          "Y",                  "Y = Active (default) · N = Inactive"),
    ]

    buf = __import__("io").BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Frame Stock"

    HDR_FILL   = PatternFill("solid", start_color="1A3C5E")
    REQ_FILL   = PatternFill("solid", start_color="1A5C3E")   # darker green = required
    EX_FILL    = PatternFill("solid", start_color="FF8C00")    # orange = example row
    NOTE_FILL  = PatternFill("solid", start_color="F0F0F0")    # grey = notes row
    WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    EX_FONT    = Font(italic=True, color="FFFFFF", size=9, name="Arial")
    NOTE_FONT  = Font(italic=True, color="666666", size=8, name="Arial")
    DATA_FONT  = Font(size=9, name="Arial")

    REQUIRED = {"Barcode", "Product", "Qty", "MRP"}

    # Row 1: Headers
    for ci, (hdr, _, _notes) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        is_req = hdr in REQUIRED
        cell.fill = REQ_FILL if is_req else HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(hdr) + 4)
    ws.row_dimensions[1].height = 28

    # Row 2: Example row (orange — delete before uploading)
    ws.cell(row=2, column=1).value = "⚠ EXAMPLE ROW — DELETE BEFORE UPLOADING"
    for ci, (_, example, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=ci)
        if ci > 1:
            cell.value = example
        cell.fill = EX_FILL
        cell.font = EX_FONT
        cell.alignment = Alignment(vertical="center")

    # Row 3: Notes row (grey — explains each column)
    for ci, (_, _, notes) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=ci, value=notes)
        cell.fill = NOTE_FILL
        cell.font = NOTE_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 40

    # Row 4 onwards: blank data rows with white fill + light border
    for ri in range(4, 54):   # 50 blank rows
        for ci in range(1, len(COLUMNS)+1):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = WHITE_FILL
            cell.font = DATA_FONT

    ws.freeze_panes = "A4"

    # Guide sheet
    ws2 = wb.create_sheet("📖 Guide", 0)
    ws2.column_dimensions["A"].width = 75
    guide_rows = [
        ("🕶️  Frame Stock — Blank Template", "title"),
        ("", None),
        ("✅ HOW TO USE", "head"),
        ("1. Delete the orange EXAMPLE row (Row 2) before uploading", "body"),
        ("2. The grey Notes row (Row 3) explains each column — delete it too, or keep for reference", "body"),
        ("3. Fill in your frames starting from Row 4 (or Row 2 after deleting example)", "body"),
        ("4. Upload via Frame Stock Manager → ➕ Add New Frames tab", "body"),
        ("", None),
        ("🟢 REQUIRED columns (dark green header):", "head"),
        ("   Scan Code / Item Code · Product · Qty · MRP", "body"),
        ("", None),
        ("📋 NEW COLUMNS:", "head"),
        ("   BSize        — B measurement (vertical lens height in mm)", "body"),
        ("   ColourMix    — secondary / accent colour on frame front", "body"),
        ("   TempleColour — temple arm colour (leave blank if same as Colour)", "body"),
        ("   FrameGroup   — dynamic tag for grouping / pricing rules", "body"),
        ("                  Examples: Near Dead · Sale · Premium · Kids · New Arrival", "body"),
        ("                  Leave blank for standard stock", "body"),
        ("", None),
        ("⚠️ RULES:", "head"),
        ("   • Scan Code / Item Code must be unique — this is what scanner reads in punching", "body"),
        ("   • For frames, this code is saved in inventory_stock.item_code, not Batch No", "body"),
        ("   • To update prices on existing frames, use the ✏️ Edit Existing tab", "body"),
        ("   • IsActive: Y or 1 = active (default), N or 0 = inactive", "body"),
        ("   • Prices in ₹: MRP = retail, selling_price = wholesale, Purchase price = cost", "body"),
        ("   • GST% is set on Product Master — not in this loader", "body"),
    ]
    TITLE_FILL = PatternFill("solid", start_color="1A3C5E")
    for text, style in guide_rows:
        cell = ws2.cell(row=ws2.max_row+1, column=1, value=text)
        if style == "title":
            cell.fill = TITLE_FILL
            cell.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
            ws2.row_dimensions[cell.row].height = 28
        elif style == "head":
            cell.font = Font(bold=True, size=11, name="Arial", color="1A3C5E")
            ws2.row_dimensions[cell.row].height = 20
        else:
            cell.font = Font(size=10, name="Arial")
        cell.alignment = Alignment(vertical="center", indent=0)

    wb.save(buf)
    return buf.getvalue()


def _render_add_tab():
    st.markdown("#### ➕ Add New Frames")
    st.caption(
        "Upload your BatchData Excel to add new frames. "
        "**Scan Code / Item Code values that already exist in the DB will be rejected** — "
        "use the ✏️ Edit tab to update prices/qty on existing frames."
    )

    # ── Blank template download ───────────────────────────────────────────────
    with st.expander("📥 Step 0 — Download Blank Template (recommended)", expanded=True):
        c1, c2 = st.columns([2, 3])
        with c1:
            # Build template — use version key so cache auto-refreshes when columns change
            _TPL_VERSION = "v6"   # ← bump this whenever COLUMNS list changes
            if st.session_state.get("_frame_tpl_ver") != _TPL_VERSION:
                try:
                    st.session_state["_frame_tpl_bytes"] = _build_blank_template()
                    st.session_state["_frame_tpl_ver"]   = _TPL_VERSION
                except Exception as _te:
                    st.error(f"Template build failed: {_te}")
                    st.session_state["_frame_tpl_bytes"] = None

            if st.session_state.get("_frame_tpl_bytes"):
                st.download_button(
                    label="⬇️ Download Blank Template",
                    data=st.session_state["_frame_tpl_bytes"],
                    file_name="FRAME_ADD_TEMPLATE.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="frame_tpl_dl",
                    use_container_width=True,
                    type="primary",
                )
        with c2:
            st.info(
                "Includes all columns: BSize, ColourMix, TempleColour, FrameGroup. "
                "Dark green header = required. Scan Code / Item Code is for scanning. "
                "Delete the orange example row before uploading."
            )

    with st.expander("📋 Expected Excel Format", expanded=False):
        st.markdown("""
| Column | Required | Notes |
|---|---|---|
| Product | ✅ | Full product name |
| Barcode / Scan Code / Item Code | ✅ | Scanner code — unique, rejected if already in DB |
| Brand | | e.g. Parakh |
| StartCode | | Box/tray location (D1, T3, BOX1) |
| Colour | | Primary colour |
| colour MIX | | Secondary colour |
| shape | | Square, Round, Aviator etc. |
| MRP | ✅ | Retail price |
| Qty | ✅ | Stock quantity |
| Purchase price | | Cost price (optional) |
| selling_price | | Wholesale price (optional) |
| BaseMaterial, Finish | | Frame specs |
| IsActive | | 1/Y = active (default), 0/N = inactive |
""")

    uploaded = st.file_uploader(
        "Upload BatchData Excel (.xlsx)",
        type=["xlsx"],
        key="frame_add_upload",
        help="Multi-sheet supported. First sheet must have column headers."
    )

    # ── Persist df across rerenders (button clicks reset file_uploader to None) ──
    if uploaded is not None:
        # New file uploaded — read and cache it
        file_bytes = uploaded.read()
        with st.spinner("Reading Excel..."):
            df = read_frame_excel(file_bytes)
        st.session_state["_frame_add_df"]   = df
        st.session_state["_frame_add_name"] = uploaded.name
    elif "_frame_add_df" in st.session_state:
        # Button was clicked — file_uploader reset but df is in cache
        df = st.session_state["_frame_add_df"]
    else:
        # Nothing uploaded yet
        return

    if df is None or df.empty:
        st.error("❌ No data found in file.")
        st.session_state.pop("_frame_add_df", None)
        st.session_state.pop("_frame_add_name", None)
        return

    fname = st.session_state.get("_frame_add_name", "file")
    data_rows = df["barcode"].notna().sum() if "barcode" in df.columns else len(df)
    st.success(f"✅ {data_rows} frame(s) loaded from **{fname}**")

    # ── Preview tabs — df persists via session_state ──────────────────────────
    tab_prev, tab_issues, tab_import = st.tabs(["📋 Preview", "⚠️ Issues", "📥 Import"])

    with tab_prev:
        _render_preview(df)
    with tab_issues:
        _render_issues(df)
    with tab_import:
        _render_import(df)


def _render_preview(df: pd.DataFrame):
    has_purchase = "purchase_rate" in df.columns and pd.to_numeric(df["purchase_rate"], errors="coerce").notna().any()
    has_selling  = "selling_price" in df.columns and pd.to_numeric(df["selling_price"], errors="coerce").notna().any()

    # Price status — inline, compact
    mrp_col = pd.to_numeric(df["mrp"], errors="coerce").dropna() if "mrp" in df.columns else pd.Series([])
    price_parts = [f"MRP ₹{mrp_col.min():.0f}–₹{mrp_col.max():.0f}"] if not mrp_col.empty else []
    if has_selling:  price_parts.append("Selling ✅")
    else:            price_parts.append("Selling —")
    if has_purchase: price_parts.append("Cost ✅")
    else:            price_parts.append("Cost —")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Frames",      len(df))
    c2.metric("Unique Scan Codes", df["barcode"].nunique() if "barcode" in df.columns else "—")
    c3.metric("Locations",   df["location"].dropna().astype(str).replace("","_").replace("_", pd.NA).dropna().nunique() if "location" in df.columns else "—")
    c4.metric("Colours",     df["colour"].dropna().astype(str).replace("","_").replace("_", pd.NA).dropna().nunique()   if "colour"   in df.columns else "—")
    c5.metric("MRP Range",   f"₹{mrp_col.min():.0f}–₹{mrp_col.max():.0f}" if not mrp_col.empty else "—")

    if not has_purchase and not has_selling:
        st.info("ℹ️ MRP only — purchase price and selling price are blank. Fine for now. Add later via ✏️ Edit tab.")

    st.markdown("---")
    fc1, fc2, fc3, fc4 = st.columns(4)
    colours = ["All"] + sorted(df["colour"].dropna().astype(str).unique()) if "colour" in df.columns else ["All"]
    sel_colour = fc1.selectbox("Colour", colours, key="fp_colour")
    shapes = ["All"] + sorted(df["shape"].dropna().astype(str).unique()) if "shape" in df.columns else ["All"]
    sel_shape = fc2.selectbox("Shape", shapes, key="fp_shape")
    locs = ["All"] + sorted(df["location"].dropna().astype(str).unique()) if "location" in df.columns else ["All"]
    sel_loc = fc3.selectbox("Location", locs, key="fp_loc")
    kw = fc4.text_input("Search name", placeholder="Butler", key="fp_kw")

    filtered = df.copy()
    if sel_colour != "All": filtered = filtered[filtered["colour"].astype(str).str.strip() == sel_colour]
    if sel_shape  != "All": filtered = filtered[filtered["shape"].astype(str).str.strip() == sel_shape]
    if sel_loc    != "All": filtered = filtered[filtered["location"].astype(str).str.strip() == sel_loc]
    if kw: filtered = filtered[filtered["product_name"].astype(str).str.contains(kw, case=False, na=False)]

    st.caption(f"Showing {len(filtered)} of {len(df)} rows")
    disp = [c for c in ["product_name","brand","barcode","location","colour","colour_mix",
                         "shape","finish","qty","purchase_rate","selling_price","mrp"] if c in filtered.columns]
    st.dataframe(filtered[disp].head(500), use_container_width=True, hide_index=True,
                 column_config={
                     "mrp":           st.column_config.NumberColumn("MRP ₹",      format="₹%.0f"),
                     "purchase_rate": st.column_config.NumberColumn("Cost ₹",     format="₹%.0f"),
                     "selling_price": st.column_config.NumberColumn("Selling ₹",  format="₹%.0f"),
                     "qty":           st.column_config.NumberColumn("Qty"),
                 })


def _render_issues(df: pd.DataFrame):
    try:
        from modules.sql_adapter import run_query
        _db_available = True
    except Exception:
        _db_available = False
        run_query = None

    issues = []

    no_mrp = df[df["mrp"].isna() | (df["mrp"] == 0)] if "mrp" in df.columns else pd.DataFrame()
    if not no_mrp.empty:
        issues.append(("❌ Missing MRP", no_mrp))

    no_sku = df[df["barcode"].isna() | (df["barcode"].astype(str).str.strip() == "")] if "barcode" in df.columns else pd.DataFrame()
    if not no_sku.empty:
        issues.append(("❌ Missing Scan Code / Item Code", no_sku))

    if "barcode" in df.columns:
        dup = df[df.duplicated("barcode", keep=False) & df["barcode"].notna()]
        if not dup.empty:
            issues.append((
                f"❌ Duplicate scan codes in Excel — {dup['barcode'].nunique()} code(s) appear more than once. "
                f"Each frame must have a unique Scan Code / Item Code. Check if code was mistyped — "
                f"only the FIRST occurrence will be kept during import.", dup
            ))

        # Check which scan codes already exist in DB — these will be REJECTED in ADD mode.
        all_skus = [_safe_str(s).upper() for s in df["barcode"].dropna() if _safe_str(s)]
        if all_skus and _db_available and run_query:
            try:
                ph = ",".join(["%s"] * len(all_skus))
                existing = run_query(
                    f"SELECT UPPER(COALESCE(NULLIF(item_code,''), NULLIF(batch_no,''))) AS sku FROM inventory_stock s "
                    f"JOIN products p ON p.id=s.product_id "
                    f"WHERE LOWER(COALESCE(p.main_group,'')) IN ('frames','frame','sunglasses') "
                    f"AND UPPER(COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,''))) IN ({ph})",
                    tuple(all_skus)
                ) or []
                existing_skus = {r["sku"] for r in existing}
                if existing_skus:
                    already_df = df[df["barcode"].astype(str).str.upper().isin(existing_skus)]
                    issues.append((
                        f"🚫 {len(existing_skus)} scan code(s) already in DB — will be SKIPPED in Add mode "
                        f"(use ✏️ Edit tab to update prices)", already_df
                    ))
            except Exception:
                pass

    no_name = df[df["product_name"].isna() | (df["product_name"].astype(str).str.strip() == "")] if "product_name" in df.columns else pd.DataFrame()
    if not no_name.empty:
        issues.append(("❌ Missing Product Name", no_name))

    # Check which product names are NOT in Product Master
    if "product_name" in df.columns and _db_available and run_query:
        try:
            all_names = df["product_name"].dropna().astype(str).str.strip().unique().tolist()
            if all_names:
                ph = ",".join(["%s"] * len(all_names))
                found = run_query(
                    f"""SELECT LOWER(TRIM(product_name)) AS name FROM products
                        WHERE LOWER(TRIM(product_name)) IN ({ph})
                          AND LOWER(COALESCE(main_group,'')) IN ('frames','frame','sunglasses')""",
                    tuple(n.lower() for n in all_names)
                ) or []
                found_names = {r["name"] for r in found}
                not_in_master = df[
                    ~df["product_name"].astype(str).str.strip().str.lower().isin(found_names)
                    & df["product_name"].notna()
                ]
                if not not_in_master.empty:
                    issues.append((
                        f"❌ {len(not_in_master)} product(s) not in Product Master — "
                        "add via Data Loader → Product Master first",
                        not_in_master
                    ))
        except Exception:
            pass

    if not issues:
        st.success("✅ No issues — all products found in master. Ready to import.")
        return

    for label, issue_df in issues:
        with st.expander(f"{label} — {len(issue_df)} row(s)", expanded="❌" in label or "🚫" in label):
            show = [c for c in ["product_name","barcode","location","colour","qty","mrp"] if c in issue_df.columns]
            st.dataframe(issue_df[show], use_container_width=True, hide_index=True)


def _render_import(df: pd.DataFrame):
    # NOTE: do NOT import sql_adapter here — it crashes the tab if DB unavailable
    # sql_adapter is imported inside the button callbacks only

    st.markdown("### 📥 Import New Frames")

    st.caption("Each Scan Code / Item Code = 1 unique physical frame · Qty always 1 · No stacking · duplicate codes kept first only")

    # Filter valid rows — use pd.to_numeric for safe conversion
    valid = df.copy()
    if "barcode" in valid.columns:
        valid = valid[valid["barcode"].notna() & (valid["barcode"].astype(str).str.strip() != "")]
    if "product_name" in valid.columns:
        valid = valid[valid["product_name"].notna() & (valid["product_name"].astype(str).str.strip() != "")]
    if "mrp" in valid.columns:
        valid = valid[pd.to_numeric(valid["mrp"], errors="coerce").fillna(0) > 0]

    # Remove in-Excel duplicates — keep first, warn user to fix source data
    if "barcode" in valid.columns:
        dupes_in_excel = valid.duplicated("barcode", keep="first").sum()
        if dupes_in_excel:
            _disp_cols = [c for c in ["barcode","product_name","colour","mrp","brand"] if c in valid.columns]
            _all_dupe_rows = valid[valid.duplicated("barcode", keep=False)].copy()

            st.warning(
                f"⚠️ **{dupes_in_excel} duplicate SKU row(s) dropped** — "
                f"rule: **first row in Excel is kept, all later rows are dropped.**"
            )

            # Show kept vs dropped side-by-side for each duplicate SKU
            for _sku, _grp in _all_dupe_rows.groupby("barcode"):
                _grp = _grp[_disp_cols].reset_index(drop=True)
                _grp.insert(0, "Action", ["✅ KEPT"] + ["❌ DROPPED"] * (len(_grp) - 1))
                st.dataframe(_grp, use_container_width=True, hide_index=True)

            st.caption(
                "Fix duplicates in your source Excel before uploading. "
                "Each Scan Code / Item Code must appear once only."
            )
        valid = valid.drop_duplicates("barcode", keep="first")

    # Force qty = 1 for all frames (each SKU = 1 physical frame)
    if "qty" in valid.columns:
        non_one = (pd.to_numeric(valid["qty"], errors="coerce").fillna(1) != 1).sum()
        if non_one:
            st.info(f"ℹ️ {non_one} row(s) had Qty ≠ 1 — reset to 1 (each scan code is one frame).")
        valid["qty"] = 1

    skipped_pre = len(df) - len(valid)
    st.info(f"**{len(valid)} valid frames** ready to add. {skipped_pre} rows skipped (missing SKU/name/MRP or duplicates).")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔍 Dry Run — Preview", key="frame_dry_run", use_container_width=True):
            _do_add_import(valid, dry_run=True)
    with c2:
        if st.button("✅ Add Frames to DB", key="frame_import_live", type="primary", use_container_width=True):
            if not st.session_state.get("frame_add_confirmed"):
                st.session_state["frame_add_confirmed"] = True
                st.warning("⚠️ Click **Add Frames to DB** again to confirm.")
            else:
                st.session_state["frame_add_confirmed"] = False
                _do_add_import(valid, dry_run=False)


# ─────────────────────────────────────────────────────────────────────────────
# PRODUCT MASTER HELPERS — check + auto-create
# ─────────────────────────────────────────────────────────────────────────────

def _is_sunglass_sku(sku: str) -> bool:
    """
    SKU prefix BS / TS (any case) = Sunglasses.
    Everything else = Frames.

    Examples:
      BS1001  → Sunglasses  (BS = Boys Sunglass / Brand Sunglass)
      TS2005  → Sunglasses  (TS = Trending Sunglass)
      D10001  → Frames
      t70694  → Frames
    """
    prefix = str(sku).strip().upper()[:2]
    return prefix in ("BS", "TS")


def _frame_product_meta(sku: str) -> dict:
    """
    Return correct main_group, hsn_code, gst_percent based on SKU prefix.

    HSN codes:
      90041000 — Sunglasses
      90030000 — Spectacle frames (using 9003)
    GST:
      18% — Sunglasses
       5% — Frames
    """
    if _is_sunglass_sku(sku):
        return {
            "main_group":  "Sunglasses",
            "hsn_code":    "90041000",
            "gst_percent": 18.0,
        }
    return {
        "main_group":  "Frames",
        "hsn_code":    "90030000",
        "gst_percent": 5.0,
    }


def _check_missing_products(product_names: list) -> list:
    """
    Return product names NOT in the products table under Frames/Sunglasses.
    Single batch query — no per-row DB calls.
    """
    if not product_names:
        return []
    try:
        from modules.sql_adapter import run_query
        rows = run_query(
            """SELECT LOWER(TRIM(product_name)) AS name
               FROM products
               WHERE LOWER(COALESCE(main_group,''))
                     IN ('frames','frame','sunglasses')"""
        ) or []
        existing = {r["name"] for r in rows}
        return [p for p in product_names if p.lower().strip() not in existing]
    except Exception:
        return []


def _auto_create_frame_products(df: pd.DataFrame, missing: list) -> tuple:
    """
    Create minimal-but-correct product master rows for missing frame products.

    For each missing name:
      - Looks up first matching row in df for brand, gender, model
      - Detects SKU prefix (BS/TS) to set Sunglasses vs Frames
      - Sets correct HSN + GST automatically
      - Does NOT put scan code in products table
        (frame scan code lives in inventory_stock.item_code)

    Returns (created_count, error_list).
    """
    try:
        from modules.sql_adapter import run_write
    except ImportError as ie:
        return 0, [f"DB unavailable: {ie}"]

    # Build name → first-row data map from df
    _row_map: dict = {}
    _cols = df.columns.tolist()
    for _, row in df.iterrows():
        pn = _safe_str(row.get("product_name")).strip()
        if pn and pn not in _row_map:
            _row_map[pn] = row

    created = 0
    errors  = []

    for prod_name in missing:
        row   = _row_map.get(prod_name)
        brand = _safe_str(row.get("brand")) if row is not None else ""
        model = _safe_str(row.get("model")) if row is not None else ""
        gender= _safe_str(row.get("gender")) if row is not None else ""
        sku   = _safe_str(row.get("barcode")) if row is not None else ""

        meta  = _frame_product_meta(sku)   # Sunglasses vs Frames from SKU prefix

        try:
            run_write(
                """
                INSERT INTO products (
                    id,
                    product_name,
                    brand,
                    main_group,
                    model,
                    gender,
                    hsn_code,
                    gst_percent,
                    is_batch_applicable,
                    is_active,
                    created_at,
                    updated_at,
                    created_source
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE,TRUE,NOW(),NOW(),'frame_auto_create')
                ON CONFLICT (product_name) DO NOTHING
                """,
                (
                    str(uuid.uuid4()),
                    prod_name,
                    brand  or None,
                    meta["main_group"],
                    model  or None,
                    gender or None,
                    meta["hsn_code"],
                    meta["gst_percent"],
                ),
            )
            created += 1
        except Exception as ex:
            errors.append(f"'{prod_name}': {ex}")

    return created, errors


def _do_add_import(df: pd.DataFrame, dry_run: bool):
    """
    Add new frames — one row per SKU, qty always 1.
    SKUs already in DB are skipped (use Edit tab to update prices).
    No qty stacking — frames are individual items identified by SKU.
    """
    try:
        from modules.sql_adapter import run_query, run_write
    except ImportError as _ie:
        st.error(f"❌ Database connection unavailable: {_ie}")
        return

    # ── Normalise column name: accept barcode/sku/item-code aliases ──────────
    # universal_loader_core maps Barcode → sku_code; this file maps it → barcode.
    # Whichever arrived, unify to 'barcode' so the rest of this function works.
    if "sku_code" in df.columns and "barcode" not in df.columns:
        df = df.rename(columns={"sku_code": "barcode"})

    # Hard stop if no SKU column at all — give a clear message
    if "barcode" not in df.columns:
        st.error(
            "❌ Scan Code / Item Code column not found in the uploaded file. "
            "Make sure the file has a column named **Barcode**, **Scan Code**, **Item Code** or **SKU Code**."
        )
        return

    inserted = skipped_existing = skipped_error = 0
    errors = []
    _seen_missing: set = set()          # deduplicate product-not-found errors — show once per name
    auto_created_products: list = []    # track which products were auto-created this run

    progress = st.progress(0, text="Processing...")
    total = max(len(df), 1)

    # Pre-fetch all existing scan codes for this batch to avoid per-row DB calls
    all_skus = [_safe_str(r).upper() for r in df["barcode"].dropna()]
    try:
        ph = ",".join(["%s"] * len(all_skus))
        existing_in_db = {
            r["sku"] for r in (run_query(
                f"SELECT UPPER(COALESCE(NULLIF(item_code,''), NULLIF(batch_no,''))) AS sku FROM inventory_stock s "
                f"JOIN products p ON p.id=s.product_id "
                f"WHERE LOWER(COALESCE(p.main_group,'')) IN ('frames','frame','sunglasses') "
                f"AND UPPER(COALESCE(NULLIF(s.item_code,''), NULLIF(s.batch_no,''))) IN ({ph})",
                tuple(all_skus)
            ) or [])
        }
    except Exception:
        existing_in_db = set()

    for i, row in df.iterrows():
        progress.progress(min((inserted+skipped_existing+skipped_error)/total, 1.0),
                          text=f"{inserted+skipped_existing+skipped_error}/{total}")

        sku = _safe_str(row.get("barcode")).upper()
        prod_name = _safe_str(row.get("product_name"))
        mrp = _safe_float(row.get("mrp"))

        if not sku or not prod_name or mrp <= 0:
            skipped_error += 1
            continue

        # STRICT: reject existing scan codes in ADD mode
        if sku in existing_in_db:
            skipped_existing += 1
            continue

        if dry_run:
            inserted += 1
            continue

        brand    = _safe_str(row.get("brand"))
        colour   = _safe_str(row.get("colour"))
        material = _safe_str(row.get("base_material"))
        location = _safe_str(row.get("location"))
        purchase = _safe_float(row.get("purchase_rate"))
        selling  = _safe_float(row.get("selling_price"))
        excel_qty = _safe_int(row.get("qty"), 1)
        is_active = _parse_is_active(row.get("is_active"))

        try:
            # ── Get or auto-create product in Product Master ──────────────────
            # If product already exists → use it.
            # If not → create it now from Excel data (brand, model, gender,
            #   HSN + GST + main_group decided by SKU prefix BS/TS).
            # This means one upload handles everything — no separate steps.
            prod_rows = run_query(
                """SELECT id FROM products
                   WHERE LOWER(TRIM(product_name)) = LOWER(TRIM(%s))
                     AND LOWER(COALESCE(main_group,'')) IN ('frames','frame','sunglasses')
                   LIMIT 1""",
                (prod_name,)
            )

            if not prod_rows:
                # Auto-create product from this row's data
                _meta   = _frame_product_meta(sku)
                _new_id = str(uuid.uuid4())
                run_write(
                    """
                    INSERT INTO products (
                        id, product_name, brand, main_group,
                        model, gender, hsn_code, gst_percent,
                        is_batch_applicable, is_active,
                        created_at, updated_at, created_source
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE,TRUE,NOW(),NOW(),'frame_auto')
                    ON CONFLICT (product_name) DO NOTHING
                    """,
                    (
                        _new_id,
                        prod_name,
                        _safe_str(row.get("brand"))  or None,
                        _meta["main_group"],
                        _safe_str(row.get("model"))  or None,
                        _safe_str(row.get("gender")) or None,
                        _meta["hsn_code"],
                        _meta["gst_percent"],
                    ),
                )
                # Re-fetch in case ON CONFLICT hit an existing row
                prod_rows = run_query(
                    """SELECT id FROM products
                       WHERE LOWER(TRIM(product_name)) = LOWER(TRIM(%s))
                       LIMIT 1""",
                    (prod_name,)
                )
                if prod_name not in _seen_missing:
                    _seen_missing.add(prod_name)
                    auto_created_products.append(prod_name)

            if not prod_rows:
                errors.append(f"'{prod_name}': could not create product — check DB permissions.")
                skipped_error += 1
                continue

            prod_id = str(prod_rows[0]["id"])

            # Insert inventory_stock (new SKU only) — all columns
            colour_mix_v  = _safe_str(row.get("colour_mix"))
            temple_col_v  = _safe_str(row.get("temple_colour"))
            grp_v         = _safe_str(row.get("frame_group"))
            size_a_v      = row.get("size_a")
            size_b_v      = row.get("size_b")
            dbl_v         = row.get("dbl")
            temple_len_v  = row.get("temple_length")
            base_mat_v    = _safe_str(row.get("base_material"))
            shape_v       = _safe_str(row.get("shape"))
            finish_v      = _safe_str(row.get("finish"))
            model_v       = _safe_str(row.get("model"))
            gender_v      = _safe_str(row.get("gender"))
            frame_type_v  = _safe_str(row.get("frame_type"))
            frame_seq_v   = row.get("frame_seq")
            image_path_v  = _safe_str(row.get("image_path"))

            def _safe_num(v):
                try:
                    return float(v) if v is not None and pd.notna(v) else None
                except Exception:
                    return None

            run_write("""
                INSERT INTO inventory_stock
                (id, product_id, item_code, batch_no, quantity,
                 purchase_rate, selling_price, mrp,
                 location, stock_type, is_active,
                 size_a, size_b, dbl, temple_length,
                 base_material, shape, finish,
                 colour, colour_mix, temple_colour,
                 frame_group, frame_seq,
                 model, gender, frame_type,
                 image_path,
                 created_at, updated_at)
                VALUES (
                    %s, %s, %s, NULL, %s,
                    %s, %s, %s,
                    %s, 'BATCH', %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s,
                    NOW(), NOW()
                )
            """, (
                str(uuid.uuid4()),   # id
                prod_id,             # product_id
                sku,                 # item_code: universal scan code used by scanners
                excel_qty,           # quantity
                purchase or None,    # purchase_rate
                selling  or None,    # selling_price
                mrp,                 # mrp
                location or None,    # location
                                     # stock_type = 'BATCH' (hardcoded)
                is_active,           # is_active
                _safe_num(size_a_v),       # size_a
                _safe_num(size_b_v),       # size_b
                _safe_num(dbl_v),          # dbl
                _safe_num(temple_len_v),   # temple_length
                base_mat_v   or None,      # base_material
                shape_v      or None,      # shape
                finish_v     or None,      # finish
                colour       or None,      # colour
                colour_mix_v or None,      # colour_mix
                temple_col_v or None,      # temple_colour
                grp_v        or None,      # frame_group
                _safe_num(frame_seq_v),    # frame_seq
                model_v      or None,      # model
                gender_v     or None,      # gender
                frame_type_v or None,      # frame_type
                image_path_v or None,      # image_path
            ))
            inserted += 1
            existing_in_db.add(sku)  # prevent double-insert within same batch

        except Exception as ex:
            err_msg = f"Row {i+2} ({sku}): {ex}"
            errors.append(err_msg)
            skipped_error += 1
            # Show first DB error immediately so silent failures are caught early
            if skipped_error == 1:
                st.warning(f"⚠️ First error (showing early): {err_msg}")

    progress.progress(1.0, text="Done!")

    if dry_run:
        st.info(
            f"🔍 **Dry run** — {inserted} would be inserted, "
            f"{skipped_existing} skipped (already in DB), "
            f"{skipped_error} skipped (invalid data). No changes made."
        )
        if skipped_existing:
            st.warning(f"ℹ️ {skipped_existing} SKU(s) already exist — use ✏️ Edit tab to update their prices.")
    else:
        # ── Auto-created products summary ────────────────────────────────────
        if auto_created_products:
            with st.expander(
                f"🆕 {len(auto_created_products)} product(s) auto-created in Product Master",
                expanded=True,
            ):
                for _p in auto_created_products:
                    # find a sku for this product to determine Sunglass vs Frame
                    _sku_for_p = ""
                    if "product_name" in df.columns and "barcode" in df.columns:
                        _prows = df[df["product_name"] == _p]["barcode"].dropna()
                        _sku_for_p = _safe_str(_prows.iloc[0]) if len(_prows) else ""
                    _meta = _frame_product_meta(_sku_for_p)
                    st.write(
                        f"• **{_p}** — {_meta['main_group']} · "
                        f"HSN {_meta['hsn_code']} · GST {_meta['gst_percent']}%"
                    )
                st.caption(
                    "These are now in Product Master. "
                    "Edit brand, gender, coating etc. there if needed."
                )

        if errors:
            with st.expander(f"⚠️ {len(errors)} error(s)"):
                for e in errors:
                    st.write(f"• {e}")

        msg = f"✅ **{inserted} frames added**"
        if auto_created_products: msg += f" · {len(auto_created_products)} product(s) auto-created"
        if skipped_existing:      msg += f" · {skipped_existing} skipped (already in DB — use Edit tab)"
        if skipped_error:         msg += f" · {skipped_error} skipped (invalid data)"
        st.success(msg)

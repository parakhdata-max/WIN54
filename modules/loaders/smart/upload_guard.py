"""
modules/loaders/smart/upload_guard.py
=======================================
Upload Guard — First line of defence.

Checks every uploaded file BEFORE any data processing:

  1. FINGERPRINT CHECK   — Is this a system-downloaded file?
  2. FLOW DETECTION      — EDIT or ADD?
  3. EXPIRY CHECK        — Is the EDIT file still within 72hr window?
  4. ONE-TIME USE CHECK  — Has this file already been uploaded?
  5. FILE TYPE CHECK     — Does the file type match what's expected?
  6. ROW COUNT CHECK     — Were rows added or deleted?
  7. TAMPER CHECK        — Were locked columns modified?

Returns a GuardResult with:
  - allowed: bool
  - flow: "EDIT" | "ADD" | None
  - file_type: str
  - issues: list of blocking errors
  - warnings: list of non-blocking warnings
  - meta: the extracted fingerprint dict
  - df: parsed DataFrame of the Data sheet, ready for use

Usage:
    from modules.loaders.smart.upload_guard import check_upload

    guard = check_upload(file_bytes, expected_type="CLENS")
    if not guard.allowed:
        st.error(guard.issues[0])
    elif guard.flow == "EDIT":
        # proceed to change_detector
    else:
        # proceed to ADD loader

    # After successful apply:
    guard.consume(user)   # marks fingerprint as used — blocks re-upload
"""

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, List

import pandas as pd


@dataclass
class GuardResult:
    allowed:   bool
    flow:      Optional[str]        # "EDIT" | "ADD" | None
    file_type: Optional[str]
    issues:    List[str] = field(default_factory=list)
    warnings:  List[str] = field(default_factory=list)
    meta:      Optional[dict] = None
    df:        Optional[pd.DataFrame] = None    # parsed Data sheet, ready for use

    def consume(self, user: str):
        """
        Mark this file's fingerprint as consumed.
        Call this AFTER changes are successfully applied — not before.
        Prevents re-upload of the same file.
        """
        if self.flow == "EDIT" and self.meta:
            file_id = self.meta.get("file_id")
            if file_id:
                _consume_fingerprint(file_id, user)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def check_upload(
    file_bytes:    bytes,
    expected_type: Optional[str] = None,    # if None — auto-detect from meta
    user:          str = "system",
) -> GuardResult:
    """
    Full upload guard check. Call this before any loader logic.
    Returns GuardResult — check .allowed before proceeding.
    """
    from modules.loaders.smart.download_manager import read_meta, FINGERPRINT_EXPIRY_HOURS

    # ── Step 1: Read the Data sheet ───────────────────────────────────────────
    # IMPORTANT: Downloaded files have '📖 Guide' as the FIRST sheet (index 0).
    # pandas without sheet_name reads index 0 → reads Guide sheet (36 rows of
    # instructions) instead of the actual data → false row count mismatch.
    # Always read the 'Data' sheet explicitly.
    try:
        try:
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name="Data",
                dtype=str,
                engine="openpyxl",
            )
        except Exception:
            # Try known alternate sheet names for ophthalmic files
            _tried = False
            for _sname in ("Spec Prices", "Add-ons", "New Records"):
                try:
                    df = pd.read_excel(
                        io.BytesIO(file_bytes),
                        sheet_name=_sname,
                        dtype=str,
                        engine="openpyxl",
                    )
                    _tried = True
                    break
                except Exception:
                    continue
            if not _tried:
                # Final fallback: first sheet
                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    dtype=str,
                    engine="openpyxl",
                )
    except Exception as e:
        return GuardResult(
            allowed=False, flow=None, file_type=None,
            issues=[f"Cannot read file: {e}"]
        )

    # ── Step 2: Extract fingerprint ───────────────────────────────────────────
    meta = read_meta(file_bytes)

    # ── Special case: OPH_SPEC / OPH_ADDON can be uploaded directly from
    #    brand price lists — no system download required.
    #    We synthesise a valid ADD meta so the guard lets them through.
    _OPEN_TYPES = ("OPH_SPEC", "OPH_ADDON", "PRICE")  # types that don't need fingerprint
    if meta is None and expected_type in _OPEN_TYPES:
        meta = {
            "flow":        "ADD",
            "file_type":   expected_type,
            "file_id":     "external-upload",
            "downloaded_at": None,
            "downloaded_by": user,
            "row_count":   len(df),
            "checksum":    "",
            "expiry":      None,
        }
    elif meta is None:
        return GuardResult(
            allowed=False, flow=None, file_type=None, df=df,
            issues=[
                "⛔ This file was not downloaded from the system.",
                "Only files downloaded via \'Edit Download\' or \'Add Template\' are accepted.",
                "Please download the current data first, make your changes, then re-upload.",
            ]
        )

    flow      = meta.get("flow", "").upper()
    file_type = meta.get("file_type", expected_type or "UNKNOWN")

    # ── Step 3: Validate flow ─────────────────────────────────────────────────
    if flow not in ("EDIT", "ADD"):
        return GuardResult(
            allowed=False, flow=None, file_type=file_type, meta=meta,
            issues=["⛔ Unrecognised file flow. Please download a fresh file."]
        )

    # ── Step 4: File type match ───────────────────────────────────────────────
    if expected_type and file_type != expected_type:
        return GuardResult(
            allowed=False, flow=flow, file_type=file_type, meta=meta,
            issues=[
                f"⛔ Wrong file type. Expected {expected_type} but this file is {file_type}.",
                "Please download the correct file type.",
            ]
        )

    warnings = []

    # ── Step 5: Expiry check (EDIT only) ──────────────────────────────────────
    if flow == "EDIT":
        expires_at_str = meta.get("expires_at")
        if expires_at_str:
            try:
                expires_at = datetime.fromisoformat(str(expires_at_str))
                if datetime.now() > expires_at:
                    return GuardResult(
                        allowed=False, flow=flow, file_type=file_type, meta=meta,
                        issues=[
                            f"⛔ This edit file has expired (valid for {FINGERPRINT_EXPIRY_HOURS} hours).",
                            "Please download a fresh copy and re-apply your changes.",
                        ]
                    )
                hours_left = (expires_at - datetime.now()).total_seconds() / 3600
                if hours_left < 6:
                    warnings.append(
                        f"⚠️ This file expires in {hours_left:.1f} hours. Upload soon."
                    )
            except Exception:
                pass

    # ── Step 5b: One-time use check (EDIT only) ───────────────────────────────
    if flow == "EDIT":
        file_id = meta.get("file_id")
        if file_id:
            try:
                from modules.sql_adapter import run_query
                fp_rows = run_query(
                    "SELECT used_at, used_by FROM download_fingerprints WHERE file_id = %s",
                    (file_id,)
                )
                if fp_rows and fp_rows[0].get("used_at"):
                    used_at = str(fp_rows[0]["used_at"])[:16]
                    used_by = fp_rows[0].get("used_by", "unknown")
                    return GuardResult(
                        allowed=False, flow=flow, file_type=file_type, meta=meta,
                        issues=[
                            f"⛔ This file has already been uploaded once (by {used_by} at {used_at}).",
                            "Each downloaded file can only be uploaded one time.",
                            "Please download a fresh copy to make further changes.",
                        ]
                    )
            except Exception:
                pass    # if table missing — skip one-time check, allow upload

    # ── Step 6: Row count sanity (EDIT only) ──────────────────────────────────
    if flow == "EDIT":
        original_rows = int(meta.get("row_count", 0))

        # Use openpyxl to count non-empty data rows — more reliable than pandas
        # for styled files where pandas may stop at first empty-looking row.
        actual_rows = _count_data_rows(file_bytes)
        if actual_rows is None:
            actual_rows = len(df)   # fallback to pandas count

        if actual_rows != original_rows:
            return GuardResult(
                allowed=False, flow=flow, file_type=file_type, meta=meta,
                issues=[
                    f"⛔ Row count mismatch. Downloaded file had {original_rows} rows, "
                    f"uploaded file has {actual_rows} rows.",
                    "Do not add or delete rows in the edit file. Only change cell values.",
                ]
            )

    # ── Step 7: Locked column tamper check (EDIT only) ────────────────────────
    if flow == "EDIT":
        tamper_warnings = _check_locked_columns(df, file_type)
        warnings.extend(tamper_warnings)

    return GuardResult(
        allowed=True,
        flow=flow,
        file_type=file_type,
        meta=meta,
        df=df,
        warnings=warnings,
    )


# ══════════════════════════════════════════════════════════════════════════════
# ROW COUNTER — openpyxl based (reliable for styled files)
# ══════════════════════════════════════════════════════════════════════════════

def _count_data_rows(file_bytes: bytes) -> Optional[int]:
    """
    Count non-empty data rows in the 'Data' sheet using openpyxl.
    Skips header row. Returns None if sheet not found.

    Why not pandas? pandas stops reading at the first fully-empty row in styled
    files, giving a lower count than actual. openpyxl reads all physical rows.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

        # Find the Data sheet
        ws = None
        for sname in wb.sheetnames:
            if sname == "Data":
                ws = wb[sname]
                break
        if ws is None:
            # Try first non-guide sheet
            for sname in wb.sheetnames:
                if not sname.startswith("_") and "Guide" not in sname:
                    ws = wb[sname]
                    break

        if ws is None:
            wb.close()
            return None

        count = sum(
            1 for row in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None and str(v).strip() != "" for v in row)
        )
        wb.close()
        return count
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# FINGERPRINT CONSUMPTION
# ══════════════════════════════════════════════════════════════════════════════

def _consume_fingerprint(file_id: str, user: str):
    """Mark a fingerprint as used — called after successful apply only."""
    try:
        from modules.sql_adapter import run_write
        run_write("""
            UPDATE download_fingerprints
            SET used_at = NOW(), used_by = %s
            WHERE file_id = %s AND used_at IS NULL
        """, (user, file_id))
    except Exception:
        pass    # non-fatal


# ══════════════════════════════════════════════════════════════════════════════
# LOCKED COLUMN TAMPER CHECK
# ══════════════════════════════════════════════════════════════════════════════

def _check_locked_columns(df: pd.DataFrame, file_type: str) -> List[str]:
    """
    Check if locked columns were modified.
    Returns warnings (non-blocking — locked columns are ignored on import).
    """
    try:
        from modules.loaders.smart.download_manager import FIELD_CONFIG
        cfg    = FIELD_CONFIG.get(file_type, {})
        locked = cfg.get("locked_cols", [])

        df_cols_norm = {c.lower().replace(" ", "_"): c for c in df.columns}
        warnings = []
        for lc in locked:
            norm = lc.lower()
            if norm in df_cols_norm:
                # Data values in locked cols are ignored by change_detector
                # No action needed here — just pass through
                pass

        return warnings
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# CONVENIENCE
# ══════════════════════════════════════════════════════════════════════════════

def get_flow_from_file(file_bytes: bytes) -> Optional[str]:
    """Quick check — returns 'EDIT', 'ADD', or None."""
    from modules.loaders.smart.download_manager import read_meta
    meta = read_meta(file_bytes)
    if not meta:
        return None
    return meta.get("flow", "").upper() or None

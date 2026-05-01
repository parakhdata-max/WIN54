"""
modules/loaders/smart/download_manager.py
==========================================
Download Manager — Two flows only:

  EDIT flow  → Downloads current DB data with fingerprint embedded.
               Only this file can be re-uploaded for editing.

  ADD flow   → Downloads a blank template for adding new records.
               No fingerprint. Additive only — cannot overwrite.

Fingerprint = hidden Excel sheet '_meta' containing:
  file_id, file_type, downloaded_at, downloaded_by,
  row_count, checksum, flow (EDIT/ADD), expiry

Usage:
    from modules.loaders.smart.download_manager import build_edit_download, build_add_template

    # EDIT — current DB data, fingerprinted
    excel_bytes, file_id = build_edit_download("CLENS", user="admin")

    # ADD — blank template
    excel_bytes = build_add_template("CLENS")
"""

import hashlib
import io
import json
import uuid
from datetime import datetime, timedelta
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Field definitions per file type ──────────────────────────────────────────

# ── Registry-driven config — NO hardcoded column lists ────────────────────────
# All column definitions live in db_schema_registry.py.
# Adding a Col() there + a migration in migrations.py is all that's needed.

def _get_cfg(file_type: str) -> dict:
    """
    Build runtime config from db_schema_registry.
    Replaces the old hardcoded FIELD_CONFIG dict.
    Called on every use — always reflects latest registry state.
    """
    from modules.loaders.db_schema_registry import DB_SCHEMA, get_download_cols

    schema = DB_SCHEMA.get(file_type, [])
    if not schema:
        return {}

    # Writable cols with excel_header = user-visible
    writable = [c for c in schema if c.writable and c.excel_header]
    download_cols = [c for c in writable if c.download]

    # db_column list for SELECT
    columns    = [c.db_column for c in download_cols]
    # locked = required cols (identity fields — cannot be changed)
    locked     = [c.db_column for c in writable if c.required]
    # product_cols = columns that live in products table (read-only in stock files)
    prod_cols  = [c.db_column for c in writable
                  if getattr(c, "notes", "").startswith("Stored on products") or
                     c.db_column in ("gst_percent",) and file_type not in ("PRODUCT", "FRAME")]

    # Table mapping
    _TABLE_MAP = {
        "PRODUCT":  "products",
        "FRAME":    "inventory_stock",
        "CLENS":    "inventory_stock",
        "OPHLENS":  "inventory_stock",
        "PRICE":    "inventory_stock",
        "SOL":      "batches",
        "BLANK":    "blank_inventory",
        "PARTY":    "parties",
        "PATIENT":  "patients",
    }
    _KEY_MAP = {
        "PRODUCT": "product_name",
        "FRAME":   "batch_no",
        "CLENS":   "batch_no",
        "OPHLENS": "product_name",
        "SOL":     "batch_no",
        "BLANK":   "brand",
        "PARTY":   "party_name",
        "PATIENT": "mobile",
        "OPH_SPEC":  "product",
        "PRICE":     "product_name",
        "OPH_ADDON": "addon_name",
    }
    _DISPLAY = {
        "PRODUCT": ("Product Master",       "📦"),
        "FRAME":   ("Frame Stock",          "🕶️"),
        "CLENS":   ("Contact Lens Stock",   "👁️"),
        "OPHLENS": ("Ophthalmic Lens Stock","🔍"),
        "SOL":     ("Solution / Batch",     "💊"),
        "BLANK":   ("Blank Inventory",      "⬜"),
        "PARTY":   ("Party Master",         "🏢"),
        "PATIENT": ("Patient Records",      "🏥"),
        "OPH_SPEC":  ("Ophthalmic Specs",    "🔬"),
        "PRICE":     ("Price Master",        "💰"),
        "OPH_ADDON": ("Ophthalmic Add-ons",  "➕"),
    }

    display_name, icon = _DISPLAY.get(file_type, (file_type, "📄"))

    return {
        "table":        _TABLE_MAP.get(file_type, ""),
        "key_col":      _KEY_MAP.get(file_type, columns[0] if columns else "id"),
        "columns":      columns,
        "locked_cols":  locked,
        "product_cols": prod_cols,
        "display_name": display_name,
        "icon":         icon,
        "_schema":      download_cols,   # full Col() objects for Excel builder
    }


# Keep FIELD_CONFIG as a thin compatibility shim — reads from registry
class _FieldConfigProxy:
    """Proxy that builds config from registry on access — zero hardcoding."""
    def get(self, file_type, default=None):
        cfg = _get_cfg(file_type)
        return cfg if cfg else default

    def __contains__(self, item):
        from modules.loaders.db_schema_registry import DB_SCHEMA
        return item in DB_SCHEMA

    def keys(self):
        from modules.loaders.db_schema_registry import DB_SCHEMA
        return DB_SCHEMA.keys()

    def __getitem__(self, key):
        cfg = _get_cfg(key)
        if not cfg:
            raise KeyError(key)
        return cfg

    def items(self):
        from modules.loaders.db_schema_registry import DB_SCHEMA
        return [(k, _get_cfg(k)) for k in DB_SCHEMA.keys()]


FIELD_CONFIG = _FieldConfigProxy()

# Fingerprint expiry window — files older than this are rejected on upload
FINGERPRINT_EXPIRY_HOURS = 72

# Module-level cache for information_schema column lookups (legacy PARTY path).
# Avoids hitting information_schema on every single PARTY download.
# Cleared on process restart (acceptable — schema rarely changes).
_legacy_col_cache: dict = {}

# ── Colour palette ────────────────────────────────────────────────────────────
COL_HEADER_EDIT   = "1A3C5E"   # dark blue  — editable column headers
COL_HEADER_LOCKED = "6B6B6B"   # dark grey  — locked column headers
COL_DATA_LOCKED   = "F2F2F2"   # light grey — locked data cells
COL_DATA_EDIT     = "FFFFFF"   # white      — editable data cells
COL_HEADER_TEXT   = "FFFFFF"   # white text on headers
COL_ACCENT        = "E8F4FD"   # light blue — alternate rows
COL_TEMPLATE_HDR  = "2E7D32"   # green      — blank template headers


# ══════════════════════════════════════════════════════════════════════════════
# FINGERPRINT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _compute_checksum(df: pd.DataFrame) -> str:
    """MD5 of the DataFrame contents — detects tampering of key columns."""
    raw = df.to_csv(index=False)
    return hashlib.md5(raw.encode()).hexdigest()


def _build_meta(file_type: str, df: pd.DataFrame, user: str, flow: str) -> dict:
    return {
        "file_id":       str(uuid.uuid4()),
        "file_type":     file_type,
        "flow":          flow,
        "downloaded_at": datetime.now().isoformat(),
        "downloaded_by": user,
        "row_count":     len(df),
        "checksum":      _compute_checksum(df),
        "expires_at":    (datetime.now() + timedelta(hours=FINGERPRINT_EXPIRY_HOURS)).isoformat(),
        "version":       "1.0",
    }


def _embed_meta(wb, meta: dict):
    """Write meta dict to hidden _meta sheet."""
    if "_meta" in wb.sheetnames:
        del wb["_meta"]
    ws = wb.create_sheet("_meta")
    ws.sheet_state = "hidden"
    ws["A1"] = "key"
    ws["B1"] = "value"
    for i, (k, v) in enumerate(meta.items(), start=2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = str(v)


def read_meta(file_bytes: bytes) -> Optional[dict]:
    """Extract _meta sheet from uploaded file. Returns None if not found."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        if "_meta" not in wb.sheetnames:
            return None
        ws = wb["_meta"]
        meta = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                meta[row[0]] = row[1]
        return meta if meta else None
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# DB FETCH
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_data(file_type: str, filters: dict = None) -> pd.DataFrame:
    """
    Fetch current DB data.
    Tries the live_schema_bridge first (auto-detects new DB columns).
    Falls through to proven per-type legacy SQL if bridge fails for any reason.
    This means: adding live_schema_bridge.py is an ENHANCEMENT — removing it
    or any failure in it never breaks the core download.
    """
    try:
        from modules.sql_adapter import run_query
    except ImportError:
        return pd.DataFrame()

    # ── Bridge: try first — handles all types + new DB columns automatically ─
    # PRICE and PATIENT use custom blocks only — skip bridge for these types
    if file_type not in ("PRICE", "PATIENT", "CLENS", "OPHLENS"):
        try:
            from modules.loaders.live_schema_bridge import build_download_sql, get_live_schema
            sql       = build_download_sql(file_type)
            live_cols = get_live_schema(file_type)
            rows = run_query(sql)
            if rows is not None:   # None = query error; [] = genuinely empty table
                fallback = [c.excel_header for c in live_cols]
                return pd.DataFrame(rows) if rows else pd.DataFrame(columns=fallback)
        except Exception as ex:
            import logging
            logging.warning(
                f"[download_manager] bridge failed for {file_type} ({ex}), "
                f"falling back to per-type SQL"
            )
            # DO NOT return here — fall through to proven legacy code below

    # ── OPH_SPEC / OPH_ADDON — custom JOIN query (no single-table download) ──
    if file_type in ("OPH_SPEC", "OPH_ADDON"):
        try:
            from modules.sql_adapter import run_query
            if file_type == "OPH_SPEC":
                rows = run_query("""
                    SELECT
                        p.brand              AS "Brand",
                        p.product_name       AS "Product",
                        s.lens_category      AS "LensCategory",
                        s.index_value::text  AS "Index",
                        s.coating            AS "Coating",
                        COALESCE(s.treatment,'Clear') AS "Treatment",
                        s.wlp_per_pair       AS "WLP_per_pair",
                        s.srp_per_pair       AS "SRP_per_pair",
                        s.purchase_rate      AS "PurchaseRate"
                    FROM ophthalmic_lens_specs s
                    JOIN products p ON p.id = s.product_id
                    WHERE s.is_active = TRUE
                    ORDER BY p.brand, p.product_name, s.index_value::numeric, s.coating
                """) or []
            else:  # OPH_ADDON
                rows = run_query("""
                    SELECT
                        a.brand              AS "Brand",
                        COALESCE(p.product_name,'') AS "Product",
                        a.addon_name         AS "AddonName",
                        a.addon_category     AS "AddonCategory",
                        a.applies_to         AS "AppliesTo",
                        a.wlp_addon          AS "WLP_Addon",
                        a.srp_addon          AS "SRP_Addon",
                        CASE WHEN a.is_percentage THEN 'YES' ELSE 'NO' END AS "IsPercentage",
                        a.sort_order         AS "SortOrder",
                        a.notes              AS "Notes"
                    FROM ophthalmic_addons a
                    LEFT JOIN products p ON p.id = a.product_id
                    WHERE a.is_active = TRUE
                    ORDER BY a.brand, a.sort_order, a.addon_name
                """) or []
            return pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=(["Brand","Product","LensCategory","Index","Coating","Treatment",
                          "WLP_per_pair","SRP_per_pair","PurchaseRate"]
                         if file_type == "OPH_SPEC"
                         else ["Brand","Product","AddonName","AddonCategory","AppliesTo",
                               "WLP_Addon","SRP_Addon","IsPercentage","SortOrder","Notes"])
            )
        except Exception as ex:
            import logging
            logging.warning(f"[download_manager] {file_type} query failed: {ex}")
            return pd.DataFrame()


    # ── PRICE: current prices for Contact Lenses (CLENS) + Solutions (SOL) ─────
    # Covers: Contact Lenses (inventory_stock stock_type=BATCH) + Solutions/Cleaners (batches).
    # NOT for: Ophthalmic lenses (→ OPH_SPEC) or Frames (→ FRAME loader).
    if file_type == "PRICE":
        # ── Price History View (MRP-grouped) ─────────────────────────────────
        # One row per unique MRP per product. qty in boxes + loose pcs + pairs-aware display.
        # Includes stock_type=BATCH (physical stock) + PRICE (price-only rows)
        _cols = ["brand","main_group","product_name","category","unit","mrp","selling_price",
                 "purchase_rate","qty_boxes","loose_pcs","qty_display","qty_unit",
                 "effective_from","company_product_name","notes"]
        try:
            rows = run_query("""
                SELECT
                    p.brand                                              AS brand,
                    COALESCE(p.main_group, '')                           AS main_group,
                    p.product_name                                       AS product_name,
                    COALESCE(p.category, p.main_group, '')               AS category,
                    COALESCE(p.unit, 'PCS')                              AS unit,
                    COALESCE(s.mrp, 0)                                   AS mrp,
                    COALESCE(s.selling_price, 0)                         AS selling_price,
                    COALESCE(s.purchase_rate, 0)                         AS purchase_rate,
                    SUM(COALESCE(s.quantity, 0))
                        / GREATEST(COALESCE(p.box_size::integer, 1), 1) AS qty_boxes,
                    SUM(COALESCE(s.quantity, 0))
                        % GREATEST(COALESCE(p.box_size::integer, 1), 1) AS loose_pcs,
                    -- Unit-aware display qty: PAIR products show in pairs, others in pcs
                    CASE WHEN UPPER(COALESCE(p.unit,'PCS')) = 'PAIR'
                         THEN ROUND(SUM(COALESCE(s.quantity,0))::numeric / 2, 2)
                         ELSE SUM(COALESCE(s.quantity, 0))::numeric
                    END                                                  AS qty_display,
                    CASE WHEN UPPER(COALESCE(p.unit,'PCS')) = 'PAIR'
                         THEN 'PAIRS' ELSE 'PCS'
                    END                                                  AS qty_unit,
                    MIN(COALESCE(
                        TO_CHAR(s.effective_from, 'DD-MM-YYYY'),
                        TO_CHAR(s.created_at,     'DD-MM-YYYY')
                    ))                                                   AS effective_from,
                    COALESCE(MAX(s.company_product_name),
                             MAX(p.company_product_name), '')            AS company_product_name,
                    '' AS notes
                FROM inventory_stock s
                JOIN products p ON p.id = s.product_id
                WHERE s.stock_type IN ('BATCH', 'PRICE')
                  AND LOWER(COALESCE(p.main_group, '')) NOT IN
                      ('ophthalmic lenses','ophthalmic lens','frame','frames')
                  AND COALESCE(s.is_active, TRUE) = TRUE
                GROUP BY p.brand, p.main_group, p.product_name,
                         p.category, p.unit, p.box_size, s.mrp, s.selling_price, s.purchase_rate

                UNION ALL

                -- Solutions / Cleaners (batches table)
                SELECT
                    p.brand,
                    COALESCE(p.main_group,''),
                    p.product_name,
                    COALESCE(p.category, p.main_group,''),
                    COALESCE(p.unit, 'PCS'),
                    COALESCE(b.mrp,0), COALESCE(b.selling_price,0), COALESCE(b.cost_price,0),
                    SUM(COALESCE(b.qty_available,0))
                        / GREATEST(COALESCE(p.box_size::integer,1),1),
                    SUM(COALESCE(b.qty_available,0))
                        % GREATEST(COALESCE(p.box_size::integer,1),1),
                    -- Solutions are always PCS
                    SUM(COALESCE(b.qty_available,0))::numeric,
                    'PCS',
                    TO_CHAR(MAX(b.created_at),'DD-MM-YYYY'),
                    COALESCE(MAX(b.company_product_name), MAX(p.company_product_name),''),
                    ''
                FROM batches b
                JOIN products p ON p.id = b.product_id
                WHERE COALESCE(b.is_active, TRUE) = TRUE
                GROUP BY p.brand, p.main_group, p.product_name, p.category, p.unit,
                         p.box_size, b.mrp, b.selling_price, b.cost_price

                ORDER BY brand ASC, main_group ASC, product_name ASC, mrp DESC
            """) or []
            return pd.DataFrame(rows) if rows else pd.DataFrame(columns=_cols)
        except Exception as _ex:
            import logging, traceback
            logging.error(f"[PRICE download] {_ex}\n{traceback.format_exc()}")
            return pd.DataFrame(columns=_cols)



    # ── Legacy fallback — proven per-type SQL, always reachable ──────────────
    cfg = _get_cfg(file_type)
    if not cfg:
        return pd.DataFrame()

    cols        = cfg["columns"]
    table       = cfg["table"]
    product_cols = set(cfg.get("product_cols", []))

    if not cols or not table:
        return pd.DataFrame()

    # ── FRAME: JOIN inventory_stock + products ────────────────────────────────
    if file_type == "FRAME":
        # Direct, explicit SQL — no schema loop to avoid column-source confusion.
        # product columns: product_name, brand  (only these live in products)
        # ALL other frame columns live in inventory_stock
        sql = """
            SELECT
                s.batch_no                                                AS "🔒 SKU Code",
                p.product_name                                            AS "🔒 Product Name",
                COALESCE(p.brand, '')                                     AS "🔒 Brand",
                COALESCE(s.colour,        '')                             AS "Colour",
                COALESCE(s.colour_mix,    '')                             AS "Colour Mix",
                COALESCE(s.temple_colour, '')                             AS "Temple Colour",
                COALESCE(s.base_material, '')                             AS "Material",
                COALESCE(s.shape,         '')                             AS "Shape",
                COALESCE(s.finish,        '')                             AS "Finish",
                s.size_a                                                  AS "A Size (mm)",
                s.size_b                                                  AS "B Size (mm)",
                s.dbl                                                     AS "DBL (mm)",
                s.temple_length                                           AS "Temple Length (mm)",
                COALESCE(s.location,      '')                             AS "Location / Box",
                COALESCE(s.frame_group,   '')                             AS "Frame Group",
                COALESCE(s.quantity,      1)                              AS "Qty",
                COALESCE(s.purchase_rate, 0)                              AS "Purchase Price \u20b9",
                COALESCE(s.selling_price, 0)                              AS "Selling Price \u20b9",
                COALESCE(s.mrp,           0)                              AS "MRP \u20b9",
                CASE WHEN COALESCE(s.is_active, true)
                     THEN 'Y' ELSE 'N' END                                AS "Active (Y/N)"
            FROM inventory_stock s
            JOIN products p ON p.id = s.product_id
            WHERE LOWER(COALESCE(p.main_group, ''))
                  IN ('frames', 'frame', 'sunglasses')
              AND COALESCE(s.is_active, true) = true
              AND COALESCE(p.is_active, true) = true
            ORDER BY p.product_name, s.batch_no
        """
        try:
            rows = run_query(sql)
            if not rows:
                return pd.DataFrame(columns=[
                    "🔒 SKU Code", "🔒 Product Name", "🔒 Brand",
                    "Colour", "Colour Mix", "Temple Colour", "Material", "Shape", "Finish",
                    "A Size (mm)", "B Size (mm)", "DBL (mm)", "Temple Length (mm)",
                    "Location / Box", "Frame Group", "Qty",
                    "Purchase Price \u20b9", "Selling Price \u20b9", "MRP \u20b9", "Active (Y/N)",
                ])
            return pd.DataFrame(rows)
        except Exception as ex:
            import logging
            logging.warning(f"[download_manager] FRAME fetch error: {ex}")
            return pd.DataFrame()
    # ── CLENS / OPHLENS: inventory_stock JOIN products ───────────────────────
    if file_type in ("CLENS", "OPHLENS"):
        stock_type = "BATCH" if file_type == "CLENS" else "POWER"
        order_by   = ("p.product_name, s.batch_no, s.sph"
                      if file_type == "CLENS"
                      else "p.product_name, s.sph, s.cyl, s.axis")
        select_parts = []
        for c in cols:
            if c == "product_name" or c in product_cols:
                select_parts.append(f'p."{c}"')
            else:
                select_parts.append(f's."{c}"')
        for c in product_cols:
            if c not in cols:
                select_parts.append(f'p."{c}"')

        # ── Always include Brand from products (first visual column after Product) ──
        # Gives full "Brand | Product" context in Excel without needing a lookup
        if 'p."brand"' not in select_parts:
            select_parts.insert(0, 'p.brand AS "Brand"')

        if file_type == "CLENS":
            # ── Box conversion: DB stores PCS, download shows BOXES ──────────
            box_expr = ("ROUND(s.quantity::numeric / "
                        "GREATEST(COALESCE(p.box_size::integer,1)::numeric, 1), 2) AS qty_boxes")
            pack_expr = "COALESCE(p.box_size, 1) AS box_size_info"
            # Validation: flag misaligned stock
            mismatch_expr = ("CASE WHEN COALESCE(p.box_size::integer,1) > 1 "
                             "AND (s.quantity % GREATEST(COALESCE(p.box_size::integer,1),1)) != 0 "
                             "THEN 'PACK MISMATCH' ELSE '' END AS pack_check")

            # ── PCS / PAIR unit-aware qty display ─────────────────────────────
            # Products with unit='PAIR' sell in pairs (2 pcs per pair).
            # Download shows qty in the natural unit (pairs or pcs).
            # Upload: loader reads this qty and converts back to pcs for DB.
            # Example: 10 pairs in DB → shows 5 in Excel → upload 5 → stores 10 pcs
            pair_qty_expr = (
                "CASE WHEN UPPER(COALESCE(p.unit, 'PCS')) = 'PAIR' "
                "THEN ROUND(s.quantity::numeric / 2, 2) "
                "ELSE s.quantity::numeric END AS qty_display"
            )
            pair_unit_expr = (
                "CASE WHEN UPPER(COALESCE(p.unit, 'PCS')) = 'PAIR' "
                "THEN 'PAIRS' ELSE 'PCS' END AS qty_unit"
            )
            # Also expose raw unit for upload roundtrip
            unit_expr = "COALESCE(p.unit, 'PCS') AS unit"

            select_parts_cl = select_parts + [
                box_expr, pack_expr, mismatch_expr,
                pair_qty_expr, pair_unit_expr, unit_expr,
            ]
        else:
            # OPHLENS: no box/pair conversion needed — just brand added above
            select_parts_cl = select_parts

        if file_type == "CLENS":
            where_extra = ("AND UPPER(COALESCE(p.main_group,'')) NOT IN "
                           "('OPHTHALMIC LENSES','OPHTHALMIC LENS','FRAME','FRAMES') ")
        else:  # OPHLENS
            where_extra = ("AND UPPER(COALESCE(p.main_group,'')) = 'OPHTHALMIC LENSES' "
                           "AND COALESCE(s.quantity, 0) > 0 ")

        sql = f"""
            SELECT {", ".join(select_parts_cl if file_type == "CLENS" else select_parts_cl)}
            FROM inventory_stock s
            LEFT JOIN products p ON p.id = s.product_id
            WHERE s.is_active = true AND s.stock_type = '{stock_type}'
              {where_extra}
            ORDER BY {order_by}
        """

    # ── SOL: batches JOIN products ────────────────────────────────────────────
    elif file_type == "SOL":
        select_parts = []
        for c in cols:
            if c == "product_name" or c in product_cols:
                select_parts.append(f'p."{c}"')
            else:
                select_parts.append(f'b."{c}"')
        for c in product_cols:
            if c not in cols:
                select_parts.append(f'p."{c}"')
        # Always add Brand from products table for clarity
        if 'p."brand"' not in select_parts and 'p.brand' not in " ".join(select_parts):
            select_parts.insert(0, 'p.brand AS "Brand"')
        sql = f"""
            SELECT {", ".join(select_parts)}
            FROM batches b
            JOIN products p ON p.id = b.product_id
            WHERE b.is_active = true
            ORDER BY p.brand, p.product_name
        """

    # ── PATIENT: needs JOIN across patients + patient_visits ─────────────────
    elif file_type == "PATIENT":
        pv_cols = {"visit_date", "visit_name", "right_sph", "right_cyl", "right_axis", "right_add",
                   "left_sph",  "left_cyl",  "left_axis",  "left_add"}
        pt_cols = {"master_name", "mobile", "record_no", "barcode", "is_temporary", "created_at"}

        select_parts = []
        for c in cfg["_schema"]:
            db  = c.db_column
            hdr = c.excel_header.replace('"', "")
            if db in pv_cols:
                expr = (f"TO_CHAR(pv.{db}, 'YYYY-MM-DD')"
                        if c.db_type == "date" else f"pv.{db}")
            elif db in pt_cols:
                expr = f"pt.{db}"
            else:
                expr = f"pt.{db}"
            select_parts.append(f'{expr} AS "{hdr}"')

        sql = f"""
            SELECT {", ".join(select_parts)}
            FROM patients pt
            LEFT JOIN patient_visits pv ON pv.patient_id = pt.id
            ORDER BY pt.master_name, pv.visit_date DESC NULLS LAST
            LIMIT 50000
        """

    # ── All other tables (incl. PARTY): registry-driven SELECT ───────────────
    else:
        # Ask DB which columns actually exist in this table.
        # Uses a module-level cache (_legacy_col_cache) so information_schema
        # is only queried once per table per process — not on every download.
        try:
            _c = _legacy_col_cache.get(table)
            if _c is None:
                rows_ic = run_query(
                    "SELECT column_name, data_type "
                    "FROM information_schema.columns "
                    "WHERE table_schema = 'public' AND table_name = %s",
                    (table,)
                ) or []
                _c = {r["column_name"]: r["data_type"] for r in rows_ic}
                _legacy_col_cache[table] = _c
            existing_cols = _c
        except Exception:
            existing_cols = {}  # fallback — try all, let DB raise if wrong

        select_parts = []
        for c in cfg["_schema"]:
            db  = c.db_column
            hdr = c.excel_header.replace('"', "")

            # Skip columns that don't exist in DB yet (missing migration)
            if existing_cols and db not in existing_cols:
                continue

            db_type = existing_cols.get(db, c.db_type)

            if c.db_type == "boolean" or db_type == "boolean":
                # Raw True/False breaks re-import — convert to YES/NO
                select_parts.append(
                    f"CASE WHEN \"{db}\" THEN 'YES' ELSE 'NO' END AS \"{hdr}\""
                )
            elif c.db_type == "date" or db_type in ("date",):
                select_parts.append(
                    f"TO_CHAR(\"{db}\", 'YYYY-MM-DD') AS \"{hdr}\""
                )
            elif "[]" in db_type or db_type in ("ARRAY",):
                # Postgres arrays (e.g. text[]) — join to comma string
                select_parts.append(
                    f"array_to_string(\"{db}\", ', ') AS \"{hdr}\""
                )
            elif "time" in db_type and "timestamp" not in db_type:
                # time without time zone — cast to text for Excel compat
                select_parts.append(
                    f"CAST(\"{db}\" AS text) AS \"{hdr}\""
                )
            else:
                select_parts.append(f'"{db}" AS "{hdr}"')

        sql = (f'SELECT {", ".join(select_parts)} '
               f'FROM {table} ORDER BY {cfg.get("key_col", cols[0])}')

    # Fallback column list uses excel_header names (matches SELECT aliases above)
    all_cols = (
        [c.excel_header for c in cfg["_schema"]]
        if file_type in ("PATIENT", "PARTY") or not product_cols
        else list(cols) + [c for c in product_cols if c not in cols]
    )
    try:
        rows = run_query(sql)
        if not rows:
            return pd.DataFrame(columns=all_cols)
        return pd.DataFrame(rows)
    except Exception as ex:
        import logging
        logging.warning(f"[download_manager] {file_type} fetch error: {ex}")
        return pd.DataFrame(columns=all_cols)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER — EDIT FILE (fingerprinted)
# ══════════════════════════════════════════════════════════════════════════════

def check_download_health(file_type: str, df) -> dict:
    """
    Pre-download data health check.
    Called before serving the edit file so user sees warnings in the UI.

    Returns dict with:
        warnings   : List[str]  — non-blocking issues
        alerts     : List[str]  — serious issues (flag prominently)
        stats      : dict       — row count, null counts, stale info
        healthy    : bool       — False only if data cannot be safely downloaded
    """
    import pandas as pd

    warnings = []
    alerts   = []
    stats    = {"total_rows": len(df), "empty_required": 0, "null_fields": {}}

    if df.empty:
        alerts.append("⚠️ No records found for this file type / filter. "
                      "Download will be empty.")
        return {"warnings": warnings, "alerts": alerts,
                "stats": stats, "healthy": False}

    # ── Check required fields for nulls ──────────────────────────────────────
    _REQUIRED = {
        "PRODUCT": ["product_name"],
        "CLENS":   ["product_name", "batch_no"],
        "OPHLENS": ["product_name"],
        "SOL":     ["product_name"],
        "FRAME":   ["product_name"],
        "PARTY":   ["party_name"],
        "PRICE":   ["product_name"],
    }
    for req_col in _REQUIRED.get(file_type, []):
        # try both db_column and excel header style
        for c in [req_col, req_col.replace("_", " ").title()]:
            if c in df.columns:
                null_count = df[c].isna().sum() + (df[c].astype(str) == "").sum()
                if null_count > 0:
                    alerts.append(
                        f"🔴 {null_count} row(s) have empty '{c}' — "
                        f"these rows will be uneditable."
                    )
                    stats["empty_required"] += null_count
                break

    # ── Check for stale prices (PRICE / CLENS) ────────────────────────────────
    if file_type in ("PRICE", "CLENS"):
        eff_col = next((c for c in ["effective_from", "EffectiveFrom"]
                        if c in df.columns), None)
        if eff_col:
            try:
                dates  = pd.to_datetime(df[eff_col], errors="coerce").dropna()
                if len(dates) > 0:
                    oldest = (pd.Timestamp.now() - dates.min()).days
                    if oldest > 365:
                        warnings.append(
                            f"⚠️ Oldest price entry is {oldest} days old — "
                            f"verify rate card is current."
                        )
            except Exception:
                pass

    # ── Check is_active coverage ──────────────────────────────────────────────
    for active_col in ["is_active", "IsActive"]:
        if active_col in df.columns:
            inactive = df[df[active_col].astype(str).str.lower().isin(
                ["no", "false", "0", "n"]
            )]
            if len(inactive) > 0:
                warnings.append(
                    f"ℹ️ {len(inactive)} inactive record(s) included in download "
                    f"(is_active=NO). They are shown greyed-out."
                )
            break

    # ── Check for very large download ─────────────────────────────────────────
    if len(df) > 5000:
        warnings.append(
            f"⚠️ Large download: {len(df):,} rows. "
            f"Use Brand/Sub-brand filter to reduce size for faster editing."
        )

    # ── Check for duplicate key column ───────────────────────────────────────
    _KEY_COLS = {
        "PRODUCT": "product_name",
        "PARTY":   "party_name",
        "FRAME":   "batch_no",
    }
    key = _KEY_COLS.get(file_type)
    if key and key in df.columns:
        dupes = df[key].duplicated().sum()
        if dupes > 0:
            alerts.append(
                f"🔴 {dupes} duplicate key(s) in '{key}' — "
                f"may cause comparison issues on upload."
            )

    stats["null_fields"] = {
        col: int(df[col].isna().sum())
        for col in df.columns
        if df[col].isna().sum() > 0
    }

    return {
        "warnings": warnings,
        "alerts":   alerts,
        "stats":    stats,
        "healthy":  len(alerts) == 0,
    }


def build_edit_download(
    file_type: str,
    user: str = "system",
    filters: dict = None,
) -> Tuple[bytes, str]:
    """
    Build a fingerprinted Excel file for EDITING existing records.

    Returns: (excel_bytes, file_id)

    The file contains:
      - Data sheet: current DB values, locked cols greyed out
      - _meta sheet: hidden fingerprint (file_id, checksum, expiry)
      - _guide sheet: instructions for the user
    """
    cfg = FIELD_CONFIG.get(file_type)
    if not cfg:
        raise ValueError(f"Unknown file type: {file_type}")

    df = _fetch_data(file_type, filters)
    meta = _build_meta(file_type, df, user, "EDIT")
    locked = set(cfg["locked_cols"])

    # ── Build Excel ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb["Data"]

    # ── Style headers ─────────────────────────────────────────────────────────
    cols_list = list(df.columns)
    for col_idx, col_name in enumerate(cols_list, start=1):
        cell = ws.cell(row=1, column=col_idx)
        is_locked = col_name in locked

        cell.font      = Font(bold=True, color=COL_HEADER_TEXT, size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color=COL_HEADER_LOCKED if is_locked else COL_HEADER_EDIT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add 🔒 prefix to locked column headers
        if is_locked:
            cell.value = f"🔒 {col_name}"

        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(col_name)) + 4)

    ws.row_dimensions[1].height = 30

    # ── Style data rows ───────────────────────────────────────────────────────
    for row_idx in range(2, ws.max_row + 1):
        alt = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(cols_list, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_locked = col_name in locked

            if is_locked:
                cell.fill = PatternFill("solid", start_color=COL_DATA_LOCKED)
                cell.font = Font(color="888888", size=9, name="Arial")
                cell.protection = cell.protection.copy(locked=True)
            else:
                bg = COL_ACCENT if alt else COL_DATA_EDIT
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(size=9, name="Arial")

            cell.alignment = Alignment(vertical="center")

    # ── Freeze panes + filter ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Guide sheet ───────────────────────────────────────────────────────────
    _add_guide_sheet(wb, file_type, cfg, "EDIT")

    # ── Embed fingerprint ─────────────────────────────────────────────────────
    _embed_meta(wb, meta)

    out = io.BytesIO()
    wb.save(out)
    _register_fingerprint(meta, user)   # ← register in DB (one-time use enforcement)
    return out.getvalue(), meta["file_id"]


# ══════════════════════════════════════════════════════════════════════════════
# FINGERPRINT REGISTRATION
# ══════════════════════════════════════════════════════════════════════════════

def _register_fingerprint(meta: dict, user: str):
    """Save fingerprint to DB so one-time use can be enforced on upload."""
    try:
        from modules.sql_adapter import run_write
        run_write("""
            INSERT INTO download_fingerprints
                (file_id, file_type, downloaded_by, downloaded_at, expires_at)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (file_id) DO NOTHING
        """, (
            meta["file_id"],
            meta["file_type"],
            user,
            meta["downloaded_at"],
            meta["expires_at"],
        ))
    except Exception:
        pass   # non-fatal — fingerprint check will fallback to expiry-only if table missing


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER — ADD TEMPLATE (blank)
# ══════════════════════════════════════════════════════════════════════════════

def build_add_template(file_type: str) -> bytes:
    """
    Build a blank Excel template for ADDING new records.

    No fingerprint. Additive only — upload cannot overwrite existing.
    Contains one example row (orange, delete before uploading).
    Columns are live-driven from the DB (via bridge) so new DB columns
    appear in the template automatically.
    """
    cfg = FIELD_CONFIG.get(file_type)
    if not cfg:
        raise ValueError(f"Unknown file type: {file_type}")

    # Try bridge first to get live column list (excel_header names)
    try:
        from modules.loaders.live_schema_bridge import get_live_schema
        live = get_live_schema(file_type)
        if live:
            cols   = [c.excel_header for c in live if c.writable and c.excel_header]
            locked = {c.excel_header for c in live if c.required}
        else:
            raise ValueError("empty")
    except Exception:
        cols   = cfg["columns"]
        locked = set(cfg["locked_cols"])

    buf = io.BytesIO()
    wb = __import__("openpyxl").Workbook()
    ws = wb.active
    ws.title = "New Records"

    # ── Headers ───────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=COL_HEADER_TEXT, size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color=COL_TEMPLATE_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 4)

    ws.row_dimensions[1].height = 28

    # ── Example row (orange background, instructional) ────────────────────────
    example = _get_example_row(file_type, cols)
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = PatternFill("solid", start_color="FF8C00")
        cell.font = Font(italic=True, color="FFFFFF", size=9, name="Arial")
        cell.alignment = Alignment(vertical="center")

    # Row 2 label
    ws.cell(row=2, column=1).value = "⚠ EXAMPLE ROW — DELETE BEFORE UPLOADING"

    ws.freeze_panes = "A3"

    _add_guide_sheet(wb, file_type, cfg, "ADD")
    _add_meta_add_flow(wb, file_type)

    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _add_guide_sheet(wb, file_type: str, cfg: dict, flow: str):
    """Add a _guide sheet with clear instructions."""
    if "_guide" in wb.sheetnames:
        del wb["_guide"]

    ws = wb.create_sheet("📖 Guide", 0)
    ws.column_dimensions["A"].width = 80

    title_fill  = PatternFill("solid", start_color="1A3C5E" if flow == "EDIT" else "2E7D32")
    title_font  = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    head_font   = Font(bold=True, size=11, name="Arial", color="1A3C5E")
    body_font   = Font(size=10, name="Arial")
    warn_font   = Font(size=10, name="Arial", color="C62828", bold=True)

    cfg_name = cfg.get("display_name", file_type)
    icon     = cfg.get("icon", "📄")

    rows = []

    if flow == "EDIT":
        rows = [
            (f"{icon}  {cfg_name} — EDIT FILE", "title"),
            ("", None),
            ("🔒  THIS FILE IS FINGERPRINTED", "warn"),
            ("Only this exact downloaded file can be re-uploaded for editing.", "body"),
            ("Do not copy data to another sheet or save as a new file.", "body"),
            ("", None),
            ("✅  HOW TO USE", "head"),
            ("1. Edit values in the WHITE / BLUE columns only.", "body"),
            ("2. GREY columns (🔒) are locked — changes there will be ignored.", "body"),
            ("3. Do NOT add or delete rows.", "body"),
            ("4. Save this file and upload it back.", "body"),
            ("5. The system will show you EXACTLY what will change before applying.", "body"),
            ("", None),
            ("⚠️  IMPORTANT RULES", "head"),
            ("• Do not rename this file.", "body"),
            ("• File expires in 72 hours — download fresh if needed.", "body"),
            ("• Do not change product name, batch no, or key identity columns.", "body"),
            ("", None),
            ("🔵  EDITABLE COLUMNS", "head"),
        ] + [(f"   • {c}", "body") for c in cfg["columns"] if c not in cfg["locked_cols"]] + [
            ("", None),
            ("🔒  LOCKED COLUMNS (will not be updated)", "head"),
        ] + [(f"   • {c}", "body") for c in cfg["locked_cols"]]

    else:  # ADD
        rows = [
            (f"{icon}  {cfg_name} — ADD NEW RECORDS TEMPLATE", "title"),
            ("", None),
            ("✅  HOW TO USE", "head"),
            ("1. DELETE the orange example row (Row 2) before uploading.", "body"),
            ("2. Fill in your new records starting from Row 2.", "body"),
            ("3. All columns must be filled for each row.", "body"),
            ("4. Upload — this file will ONLY ADD new records.", "body"),
            ("5. It cannot overwrite or modify existing records.", "body"),
            ("", None),
            ("⚠️  THIS IS AN ADD-ONLY FILE", "warn"),
            ("Uploading this will never change existing data in the database.", "body"),
            ("To edit existing records, use the EDIT DOWNLOAD option.", "body"),
            ("", None),
            ("📋  REQUIRED COLUMNS", "head"),
        ] + [(f"   • {c}", "body") for c in cfg["columns"]]

    for text, style in rows:
        cell = ws.cell(row=ws.max_row + 1, column=1, value=text)
        if style == "title":
            cell.fill = title_fill
            cell.font = title_font
            ws.row_dimensions[cell.row].height = 30
        elif style == "head":
            cell.font = head_font
            ws.row_dimensions[cell.row].height = 20
        elif style == "warn":
            cell.font = warn_font
        elif style == "body":
            cell.font = body_font
        cell.alignment = Alignment(vertical="center", indent=1 if style == "body" else 0)


def _add_meta_add_flow(wb, file_type: str):
    """Embed minimal meta for ADD flow (no fingerprint, just flow marker)."""
    if "_meta" in wb.sheetnames:
        del wb["_meta"]
    ws = wb.create_sheet("_meta")
    ws.sheet_state = "hidden"
    ws["A1"], ws["B1"] = "key", "value"
    ws["A2"], ws["B2"] = "flow", "ADD"
    ws["A3"], ws["B3"] = "file_type", file_type
    ws["A4"], ws["B4"] = "created_at", datetime.now().isoformat()


def _get_example_row(file_type: str, cols: list) -> list:
    """Return a realistic example row for the ADD template."""
    examples = {
        "product_name":   "Example Product Name",
        "brand":          "Acuvue",
        "batch_no":       "BATCH001",
        "sph":            "-2.50",
        "cyl":            "0.00",
        "axis":           "0",
        "add_power":      "0.00",
        "eye_side":       "B",
        "quantity":       "6",
        "purchase_rate":  "500.00",
        "selling_price":  "650.00",
        "mrp":            "800.00",
        "expiry_date":    "2027-12-31",
        "is_active":      "YES",
        "lens_design":    "SPHERICAL",
        "item_type":      "STOCK",
        "party_name":     "Example Supplier",
        "party_type":     "Supplier",
        "mobile":         "9876543210",
        "sku_code":       "SKU001",
        "qty":            "10",
        "cost_price":     "500.00",
        "main_group":     "Contact Lens",
        "category":       "Soft",
        "material":       "Silicone Hydrogel",
        "box_size":       "6",
        "unit":           "PCS",
        "is_batch_applicable": "YES",
        "is_eye_specific": "NO",
        "allow_loose":    "YES",
        "gender":         "Unisex",
        "gst_percent":    "12",       # ✅ GST rate example — 0/5/12/18/28
        "selling_price":  "750.00",
        "mrp":            "800.00",
    }
    return [examples.get(c, "") for c in cols]


# ══════════════════════════════════════════════════════════════════════════════
# FILENAME HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def make_edit_filename(file_type: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{file_type}_EDIT_{ts}.xlsx"


def make_add_filename(file_type: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{file_type}_ADD_TEMPLATE_{ts}.xlsx"

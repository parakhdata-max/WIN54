"""
patches/excel_export_enhancer.py
==================================
PATCH 2 — EXCEL EXPORT ENHANCER + METADATA SHEET WRITER
=========================================================

Drop-in upgrade for build_download_excel() in data_downloader.py.

Adds:
  ✅ Versioned filenames:  products_export_2026-02-22_v3.xlsx
  ✅ _metadata hidden sheet with self-describing info
  ✅ Schema hints in header row (type annotations as Excel comments)
  ✅ Data validation dropdowns (GST%, IsActive, EyeSide, RoleType, etc.)
  ✅ Sample "example" row at top of data so users understand instantly

USAGE — call instead of build_download_excel() from data_downloader.py:

    from patches.excel_export_enhancer import build_enhanced_excel, make_versioned_filename

    # Generate versioned filename
    filename = make_versioned_filename("PRODUCT", import_count=3)
    # → "PRODUCT_export_2026-02-22_v3.xlsx"

    # Build the enhanced Excel bytes
    excel_bytes = build_enhanced_excel(
        df           = df,
        meta         = meta,
        dataset_key  = "PRODUCT",
        diff_df      = None,       # optional
        schema_version = "1.2",
        system_version = "DV ERP v1",
        exported_by  = "admin",
    )

    st.download_button(label="Download", data=excel_bytes, file_name=filename)
"""

import io
import logging
from datetime import datetime, date
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

# ── Schema version (bump when columns change) ─────────────────────────────────
DEFAULT_SCHEMA_VERSION = "1.2"
DEFAULT_SYSTEM_VERSION = "DV ERP v1"

# ══════════════════════════════════════════════════════════════════════════════
# SCHEMA HINTS — shown as Excel comments on header cells
# Maps dataset_key → {column_name: hint_text}
# ══════════════════════════════════════════════════════════════════════════════

COLUMN_HINTS = {
    "PRODUCT": {
        "Product":           "Required | text | Primary key — do NOT change",
        "MainGroup":         "Optional | text | e.g. Ophthalmic, Contact",
        "Type":              "Optional | text | Product category",
        "LensCategory":      "Optional | text | e.g. SV, Bifocal, Progressive",
        "Brand":             "Optional | text",
        "Index":             "Optional | decimal | e.g. 1.50, 1.56, 1.67",
        "IsBatchApplicable": "Required | YES/NO",
        "IsEyeSpecific":     "Required | YES/NO",
        "IsActive":          "Required | YES/NO | Set NO to deactivate",
        "HSNCode":           "Optional | text | 8-digit GST HSN",
    },
    "FRAME": {
        "SKUCode":      "Required | text | Primary key — do NOT change",
        "Qty":          "Required | integer | ADD mode: adds to existing",
        "CostPrice":    "Optional | decimal | Purchase price",
        "MRP":          "Optional | decimal | Maximum retail price",
        "IsActive":     "Required | YES/NO",
    },
    "PARTY": {
        "PARTYNAME":  "Required | text | Primary key when MOBILE is empty",
        "ROLETYPE":   "Required | enum | Retail/Doctor/Optician/Supplier/Fitter/Wholesale",
        "MOBILE":     "Optional | 10-digit number | Used as primary key when filled",
        "ISACTIVE":   "Required | YES/NO",
    },
    "OPHLENS": {
        "Product":      "Required | text | Must match product master exactly",
        "SPH":          "Optional | decimal | e.g. -2.25, +1.00",
        "CYL":          "Optional | decimal | Negative only for most lens types",
        "AXIS":         "Optional | integer | 0–180, required when CYL is set",
        "ADD":          "Optional | decimal | Bifocal/progressive add power",
        "EyeSide":      "Required | enum | R / L / B",
        "Quantity":     "Required | integer | ADD mode adds; OPENING mode replaces",
        "MRP":          "Optional | decimal",
        "SellingPrice": "Optional | decimal",
        "PurchaseRate": "Optional | decimal",
    },
    "CLENS": {
        "Product":   "Required | text | Must match product master",
        "BatchNo":   "Required | text | Batch number",
        "Quantity":  "Required | integer",
        "MRP":       "Optional | decimal",
    },
    "SOL": {
        "Product":      "Required | text | Must match product master",
        "BatchNo":      "Optional | text",
        "Qty":          "Required | integer",
        "CostPrice":    "Optional | decimal",
        "SellingPrice": "Optional | decimal",
        "MRP":          "Optional | decimal",
    },
    "BLANK": {
        "brand":           "Required | text | Primary key part",
        "Category":        "Required | text | Primary key part",
        "Material":        "Required | text | Primary key part",
        "Add":             "Required | decimal | ADD power, primary key part",
        "qty_Right":       "Optional | integer | ADD mode adds; OPENING replaces",
        "qty_left":        "Optional | integer",
        "qty_independent": "Optional | integer",
    },
    "PATIENT": {
        "Client Name":   "Optional | text | Patient full name",
        "Mobile Number": "Optional | 10-digit | Primary key when filled",
        "Record No":     "Optional | text | Primary key when mobile empty",
        "Date":          "Optional | YYYY-MM-DD | Visit date",
        "Right Sph":     "Optional | decimal",
        "Right CYL":     "Optional | decimal",
        "Right AXIS":    "Optional | 0–180 integer",
    },
}

# ══════════════════════════════════════════════════════════════════════════════
# DATA VALIDATIONS — Excel dropdown / range rules per dataset
# Maps dataset_key → list of {col, type, values/min/max}
# ══════════════════════════════════════════════════════════════════════════════

COLUMN_VALIDATIONS = {
    "PRODUCT": [
        {"col": "IsBatchApplicable", "type": "list",  "values": ["YES", "NO"]},
        {"col": "IsEyeSpecific",     "type": "list",  "values": ["YES", "NO"]},
        {"col": "IsActive",          "type": "list",  "values": ["YES", "NO"]},
        {"col": "Index",             "type": "decimal", "min": 1.4, "max": 1.9},
    ],
    "FRAME": [
        {"col": "IsActive",  "type": "list",    "values": ["YES", "NO"]},
        {"col": "Qty",       "type": "whole",   "min": 0,  "max": 99999},
        {"col": "MRP",       "type": "decimal", "min": 0,  "max": 999999},
        {"col": "CostPrice", "type": "decimal", "min": 0,  "max": 999999},
    ],
    "PARTY": [
        {"col": "ROLETYPE", "type": "list",
         "values": ["Retail", "Doctor", "Optician", "Supplier", "Fitter", "Wholesale"]},
        {"col": "ISACTIVE", "type": "list", "values": ["YES", "NO"]},
    ],
    "OPHLENS": [
        {"col": "EyeSide",  "type": "list",  "values": ["R", "L", "B"]},
        {"col": "IsActive", "type": "list",  "values": ["YES", "NO"]},
        {"col": "AXIS",     "type": "whole", "min": 0, "max": 180},
        {"col": "Quantity", "type": "whole", "min": 0, "max": 99999},
    ],
    "CLENS": [
        {"col": "IsActive", "type": "list",  "values": ["YES", "NO"]},
        {"col": "Quantity", "type": "whole", "min": 0, "max": 99999},
    ],
    "SOL": [
        {"col": "IsActive", "type": "list", "values": ["YES", "NO"]},
    ],
    "BLANK": [],
    "PATIENT": [
        {"col": "Right AXIS", "type": "whole", "min": 0, "max": 180},
        {"col": "Left AXIS",  "type": "whole", "min": 0, "max": 180},
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE ROWS — one example row per dataset so users understand the format
# ══════════════════════════════════════════════════════════════════════════════

SAMPLE_ROWS = {
    "PRODUCT": {
        "Product": "EXAMPLE — DELETE THIS ROW",
        "MainGroup": "Ophthalmic",
        "Type": "Single Vision",
        "LensCategory": "SV",
        "Brand": "Essilor",
        "BrandProductGroup": "Crizal",
        "Material": "Organic",
        "Index": 1.56,
        "Coating": "AR",
        "coating_type": "Premium",
        "Colour": "Clear",
        "Gender": "Unisex",
        "WearSchedule": "",
        "unit": "PCS",
        "IsBatchApplicable": "YES",
        "IsEyeSpecific": "YES",
        "IsActive": "YES",
        "HSNCode": "90013000",
        "Box Size": 1,
        "Allow Loose": "NO",
    },
    "FRAME": {
        "SKUCode": "EXAMPLE-SKU-001",
        "Product": "EXAMPLE — DELETE THIS ROW",
        "Model": "ModelX",
        "Brand": "Ray-Ban",
        "ASize": 52,
        "DBL": 18,
        "TempleLength": 140,
        "BaseMaterial": "Metal",
        "Finish": "Matte",
        "Colour": "Black",
        "shape": "Oval",
        "Qty": 10,
        "CostPrice": 850.00,
        "MRP": 1999.00,
        "IsActive": "YES",
    },
    "PARTY": {
        "PARTYNAME": "EXAMPLE — DELETE THIS ROW",
        "ROLETYPE": "Retail",
        "MOBILE": "9876543210",
        "ADDRESS": "123 Main Street",
        "CITY": "Mumbai",
        "AREA": "Bandra",
        "ISACTIVE": "YES",
    },
    "OPHLENS": {
        "Product": "EXAMPLE — DELETE THIS ROW",
        "SPH": -2.25,
        "CYL": -0.50,
        "AXIS": 90,
        "ADD": "",
        "EyeSide": "B",
        "BatchNo": "BATCH001",
        "ExpiryDate": "2027-12-31",
        "Quantity": 5,
        "PurchaseRate": 120.00,
        "SellingPrice": 200.00,
        "MRP": 250.00,
        "lens_design": "TORIC",
        "Location": "STORE-A",
        "IsActive": "YES",
    },
    "CLENS": {
        "Product": "EXAMPLE — DELETE THIS ROW",
        "SPH": -1.00,
        "CYL": "",
        "AXIS": "",
        "ADD": "",
        "EyeSide": "B",
        "BatchNo": "CL-BATCH01",
        "ExpiryDate": "2026-06-30",
        "Quantity": 20,
        "PurchaseRate": 80.00,
        "SellingPrice": 140.00,
        "MRP": 160.00,
        "IsActive": "YES",
    },
    "SOL": {
        "Product": "EXAMPLE — DELETE THIS ROW",
        "BatchNo": "SOL-001",
        "ExpiryDate": "2026-12-31",
        "Qty": 30,
        "CostPrice": 110.00,
        "SellingPrice": 180.00,
        "MRP": 200.00,
        "IsActive": "YES",
    },
    "BLANK": {
        "brand": "EXAMPLE — DELETE",
        "Category": "SV",
        "Material": "CR39",
        "COLOUR": "Clear",
        "Add": 0.0,
        "qty_Right": 10,
        "qty_left": 10,
        "qty_independent": 5,
        "Recomended Base": 6.0,
        "Base 1 P": 6.0,
        "Base 2 P": 8.0,
        "Base 3P": "",
    },
    "PATIENT": {
        "Client Name": "EXAMPLE — DELETE THIS ROW",
        "Mobile Number": "9876543210",
        "Record No": "REC-001",
        "Date": "2026-02-22",
        "Right Sph": -1.50,
        "Right CYL": -0.25,
        "Right AXIS": 90,
        "Right Add Power": "",
        "Left SPH": -1.75,
        "Left CYL": "",
        "Left AXIS": "",
        "Left Add Power": "",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# VERSIONED FILENAME GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def make_versioned_filename(dataset_key: str, import_count: int = 1) -> str:
    """
    Returns:  PRODUCT_export_2026-02-22_v3.xlsx
    Use import_count from loader_import_log for the version number.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    return f"{dataset_key}_export_{today}_v{import_count}.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# METADATA SHEET WRITER
# ══════════════════════════════════════════════════════════════════════════════

def _write_metadata_sheet(wb, meta: dict, schema_version: str, system_version: str, exported_by: str):
    """
    Writes a hidden _metadata sheet with self-describing export info.
    This makes every Excel file independently auditable.
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return

    ws = wb.create_sheet("_metadata")
    ws.sheet_state = "hidden"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 52

    LABEL_FONT  = Font(bold=True, name="Calibri", size=10, color="1E3A5F")
    VALUE_FONT  = Font(name="Calibri", size=10)
    HEADER_FONT = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
    HDR_FILL    = PatternFill("solid", fgColor="1E3A5F")

    rows = [
        ("DV ERP — Export Metadata", ""),
        ("Field",                    "Value"),
        ("Exported At",              meta.get("downloaded", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
        ("Dataset",                  meta.get("label", meta.get("dataset", ""))),
        ("Dataset Key",              meta.get("dataset", "")),
        ("Loader Type",              meta.get("loader_type", "")),
        ("Rows Exported",            str(meta.get("rows", 0))),
        ("Filters Applied",          meta.get("filters", "none")),
        ("Exported By",              exported_by),
        ("Schema Version",           schema_version),
        ("System Version",           system_version),
        ("Roundtrip Safe",           "YES — upload this file directly to Loader"),
        ("Key Column(s)",            meta.get("key_col", "")),
        ("Instructions",             "See 'Info & Guide' sheet for editing rules"),
    ]

    for ri, (k, v) in enumerate(rows, 1):
        ck = ws.cell(row=ri, column=1, value=k)
        cv = ws.cell(row=ri, column=2, value=v)
        if ri == 1:
            ck.font = HEADER_FONT
            ck.fill = HDR_FILL
            cv.fill = HDR_FILL
            ws.row_dimensions[ri].height = 22
        elif ri == 2:
            ck.font = LABEL_FONT
            cv.font = LABEL_FONT
        else:
            ck.font = LABEL_FONT
            cv.font = VALUE_FONT
        ck.alignment = Alignment(vertical="center")
        cv.alignment = Alignment(vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# DATA VALIDATION APPLIER
# ══════════════════════════════════════════════════════════════════════════════

def _apply_data_validations(ws, df: pd.DataFrame, dataset_key: str):
    """
    Apply Excel data validation dropdowns and numeric range checks.
    Only applied to rows 3+ (row 1 = header, row 2 = sample).
    """
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return

    validations = COLUMN_VALIDATIONS.get(dataset_key, [])
    if not validations:
        return

    col_index = {col: i + 1 for i, col in enumerate(df.columns)}

    last_row = max(len(df) + 10, 200)  # extend validation to cover new rows users add

    for rule in validations:
        col_name = rule["col"]
        if col_name not in col_index:
            continue

        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_index[col_name])
        cell_range = f"{col_letter}3:{col_letter}{last_row}"

        vtype = rule["type"]

        if vtype == "list":
            quoted = [f'"{v}"' for v in rule["values"]]
            formula = ",".join(quoted)
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(rule["values"])}"',
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error=f"Must be one of: {', '.join(rule['values'])}",
            )
        elif vtype == "whole":
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=str(int(rule.get("min", 0))),
                formula2=str(int(rule.get("max", 99999))),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid number",
                error=f"Must be a whole number between {rule.get('min',0)} and {rule.get('max',99999)}",
            )
        elif vtype == "decimal":
            dv = DataValidation(
                type="decimal",
                operator="between",
                formula1=str(rule.get("min", 0)),
                formula2=str(rule.get("max", 999999)),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error=f"Must be a number between {rule.get('min',0)} and {rule.get('max',999999)}",
            )
        else:
            continue

        dv.sqref = cell_range
        ws.add_data_validation(dv)


# ══════════════════════════════════════════════════════════════════════════════
# SCHEMA HINT COMMENTS ON HEADER ROW
# ══════════════════════════════════════════════════════════════════════════════

def _add_header_comments(ws, df: pd.DataFrame, dataset_key: str):
    """Add Excel comments to header cells explaining each column."""
    try:
        from openpyxl.comments import Comment
        from openpyxl.styles import Font
    except ImportError:
        return

    hints = COLUMN_HINTS.get(dataset_key, {})
    if not hints:
        return

    for ci, col in enumerate(df.columns, 1):
        hint = hints.get(col)
        if hint:
            cell = ws.cell(row=1, column=ci)
            comment = Comment(hint, "DV ERP Schema")
            comment.width  = 300
            comment.height = 60
            cell.comment   = comment


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE ROW INSERTER
# ══════════════════════════════════════════════════════════════════════════════

def _prepend_sample_row(df: pd.DataFrame, dataset_key: str) -> pd.DataFrame:
    """
    Prepend a labelled example row at position 0 so users see the expected
    data format immediately. The row is clearly labelled for deletion.
    """
    sample = SAMPLE_ROWS.get(dataset_key)
    if not sample:
        return df

    # Build a sample row aligned to df columns
    sample_row = {col: sample.get(col, "") for col in df.columns}
    sample_df  = pd.DataFrame([sample_row])

    return pd.concat([sample_df, df], ignore_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENHANCED EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_enhanced_excel(
    df:             pd.DataFrame,
    meta:           dict,
    dataset_key:    str,
    diff_df=None,
    schema_version: str = DEFAULT_SCHEMA_VERSION,
    system_version: str = DEFAULT_SYSTEM_VERSION,
    exported_by:    str = "system",
    include_sample: bool = True,
) -> bytes:
    """
    Enhanced drop-in replacement for build_download_excel().

    Adds on top of the original:
      ✅ _metadata hidden sheet (self-describing export)
      ✅ Schema hints as Excel comments on every header cell
      ✅ Data validation dropdowns (enums, numeric ranges)
      ✅ Sample row at row 2 so users understand format immediately
      ✅ Versioned-filename-aware (caller uses make_versioned_filename())

    Returns raw bytes for st.download_button.
    """
    buf = io.BytesIO()

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        df.to_excel(buf, index=False)
        buf.seek(0)
        return buf.getvalue()

    wb = openpyxl.Workbook()

    # ── Style palette (matching original data_downloader.py) ─────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
    SAMP_FILL = PatternFill("solid", fgColor="FFF7ED")   # warm orange tint for sample row
    NEW_FILL  = PatternFill("solid", fgColor="DCFCE7")
    UPD_FILL  = PatternFill("solid", fgColor="FEF9C3")
    DEL_FILL  = PatternFill("solid", fgColor="FEE2E2")
    UNCH_FILL = PatternFill("solid", fgColor="F8FAFC")

    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    KEY_FONT  = Font(bold=True, name="Calibri", size=10)
    NORM_FONT = Font(name="Calibri", size=10)
    SAMP_FONT = Font(italic=True, color="B45309", name="Calibri", size=10)
    THIN      = Border(
        bottom=Side(style="thin", color="D1D5DB"),
        right =Side(style="thin", color="D1D5DB"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(vertical="center")
    WRAP   = Alignment(vertical="center", wrap_text=True)

    def _write_sheet(ws, dataframe: pd.DataFrame, title: str, row_status=None,
                     is_main_data: bool = False):
        ws.title = title

        # Header row (row 1)
        for ci, col_name in enumerate(dataframe.columns, 1):
            c = ws.cell(row=1, column=ci, value=col_name)
            c.fill      = HDR_FILL
            c.font      = HDR_FONT
            c.alignment = CENTER
            c.border    = THIN
        ws.row_dimensions[1].height = 22

        # Data rows start at row 2
        for ri, (_, row_data) in enumerate(dataframe.iterrows(), 2):
            is_sample = is_main_data and (ri == 2) and include_sample
            status    = row_status.get(ri - 2, None) if row_status else None

            fill = (SAMP_FILL if is_sample else
                    NEW_FILL  if status == "NEW"       else
                    UPD_FILL  if status == "UPDATED"   else
                    DEL_FILL  if status == "DELETED"   else
                    UNCH_FILL if status == "UNCHANGED" else
                    ALT_FILL  if ri % 2 == 0 else None)

            for ci, (col, val) in enumerate(row_data.items(), 1):
                if col == "_status":
                    continue
                c = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    c.fill = fill
                c.font      = SAMP_FONT if is_sample else (KEY_FONT if ci == 1 else NORM_FONT)
                c.border    = THIN
                c.alignment = LEFT

        # Column widths
        for ci, col in enumerate(dataframe.columns, 1):
            if col == "_status":
                continue
            col_letter = get_column_letter(ci)
            try:
                max_len = max(
                    len(str(col)),
                    dataframe[col].astype(str).str.len().max() if len(dataframe) > 0 else 0
                )
            except Exception:
                max_len = len(str(col))
            ws.column_dimensions[col_letter].width = min(max(int(max_len) + 3, 12), 42)

        # Freeze header, autofilter
        ws.freeze_panes = "A2"
        if len(dataframe) > 0:
            last_col = get_column_letter(len(dataframe.columns))
            ws.auto_filter.ref = f"A1:{last_col}1"

    # ── Prepare data with sample row ──────────────────────────────────────────
    display_df = _prepend_sample_row(df, dataset_key) if include_sample else df.copy()

    # ── Sheet 1: Data ─────────────────────────────────────────────────────────
    ws1 = wb.active
    _write_sheet(ws1, display_df, title="Data", is_main_data=True)

    # Add schema hints (comments on header)
    _add_header_comments(ws1, display_df, dataset_key)

    # Add data validation dropdowns (from row 3 onward — row 2 is sample)
    _apply_data_validations(ws1, display_df, dataset_key)

    # ── Sheet 2: Info & Guide (keep identical to original) ───────────────────
    ws2 = wb.create_sheet("Info & Guide")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 72

    BOLD_FONT = Font(bold=True, name="Calibri", size=10)

    guide_rows = [
        ("📊 DOWNLOAD INFORMATION", ""),
        ("Dataset",         meta.get("label", "")),
        ("Downloaded",      meta.get("downloaded", "")),
        ("Rows",            meta.get("rows", 0)),
        ("Loader Type",     meta.get("loader_type", "")),
        ("Schema Version",  schema_version),
        ("Key Column(s)",   meta.get("key_col", "")),
        ("Filters Used",    meta.get("filters", "none")),
        ("Exported By",     exported_by),
        ("System Version",  system_version),
        ("", ""),
        ("✏️ HOW TO EDIT & RE-IMPORT", ""),
        ("Step 1", "Delete the orange sample row (row 2) before importing"),
        ("Step 2", "Edit the Data sheet — change values, add rows at bottom"),
        ("Step 3", "Do NOT rename or delete any column header"),
        ("Step 4", "Save as .xlsx"),
        ("Step 5", "Open Data Loader → Upload & Import tab"),
        ("Step 6", "Drop this file → DRY RUN → review → SHADOW → LIVE"),
        ("", ""),
        ("📌 EDIT RULES", ""),
        ("Key column",  meta.get("key_col", "") + " — do NOT change these values"),
        ("New rows",    "Add at bottom → loader will INSERT"),
        ("Updates",     "Change values → loader will UPDATE"),
        ("Deactivate",  "Set IsActive=NO (do not delete the row)"),
        ("Stock qty",   "ADD mode = adds qty | OPENING mode = replaces qty"),
        ("", ""),
        ("⚠️ IMPORTANT RULES", ""),
        ("Rule 1", "Column names must remain EXACTLY as downloaded"),
        ("Rule 2", "Data sheet tab must stay named 'Data'"),
        ("Rule 3", "Save as .xlsx — not .xls or .csv"),
        ("Rule 4", "Run DRY RUN first every time before LIVE"),
        ("Rule 5", "Delete the orange example row before importing"),
        ("", ""),
        ("💡 EDIT GUIDE", meta.get("edit_guide", "")),
    ]

    SECTION_ROWS = {0, 11, 19, 26}
    for ri, (k, v) in enumerate(guide_rows, 1):
        ck = ws2.cell(row=ri, column=1, value=k)
        cv = ws2.cell(row=ri, column=2, value=v)
        if ri - 1 in SECTION_ROWS:
            ck.font = Font(bold=True, size=11, color="1E3A5F", name="Calibri")
        else:
            ck.font = BOLD_FONT
            cv.font = NORM_FONT
        ck.alignment = LEFT
        cv.alignment = WRAP
        ws2.row_dimensions[ri].height = 18

    ws2.freeze_panes = "A1"

    # ── Sheet 3: Changes (diff) ───────────────────────────────────────────────
    if diff_df is not None and not diff_df.empty:
        ws3 = wb.create_sheet("Changes")
        status_map = {i: row.get("_status") for i, (_, row) in enumerate(diff_df.iterrows())}
        _write_sheet(ws3, diff_df, title="Changes", row_status=status_map)

        legend_row = len(diff_df) + 4
        ws3.cell(row=legend_row, column=1, value="Legend").font = BOLD_FONT
        for i, (status, fill, label) in enumerate([
            ("NEW",       NEW_FILL,  "New row → will be INSERTED"),
            ("UPDATED",   UPD_FILL,  "Changed → will be UPDATED"),
            ("DELETED",   DEL_FILL,  "Missing in edited file → NOT auto-deleted (set IsActive=NO)"),
            ("UNCHANGED", UNCH_FILL, "No changes"),
        ], 1):
            c1 = ws3.cell(row=legend_row + i, column=1, value=status)
            c1.fill = fill
            c1.font = BOLD_FONT
            c2 = ws3.cell(row=legend_row + i, column=2, value=label)
            c2.font = NORM_FONT

    # ── _metadata hidden sheet ────────────────────────────────────────────────
    _write_metadata_sheet(wb, meta, schema_version, system_version, exported_by)

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

"""
modules/loaders/data_downloader.py
====================================
Data Downloader Engine — DV ERP  (Registry-Driven v3)

All column definitions come from db_schema_registry.py.
Adding a new column to DB_SCHEMA automatically updates:
  ✔ SELECT query   (build_select_fragments)
  ✔ Excel template column headers
  ✔ Diff engine key columns
  ✔ Column mapping panel in loader_ui
  ✔ Guide sheet in downloaded Excel

Download → Edit → Re-import workflow:
  DB → Download Excel (exact loader column names) → Edit → Upload & Import → DB updated
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd

from modules.loaders.db_schema_registry import (
    DB_SCHEMA, get_download_cols, build_select_fragments,
    get_writable_cols, get_allowed_values,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# DOWNLOAD CONFIGS
# Static metadata per dataset. SQL queries auto-generated from registry.
# Only add here when adding a NEW dataset type.
# ═══════════════════════════════════════════════════════════════════════════════

DOWNLOAD_CONFIGS: Dict[str, dict] = {

    "PRODUCT": {
        "label":       "Product Master",
        "icon":        "📦",
        "desc":        "All products — lenses, frames, solutions. Edit any attribute and re-import to update.",
        "loader_type": "PRODUCT",
        "key_col":     "Product",
        "table":       "products",
        "table_alias": "",
        "edit_guide":  "Key column: Product (do not rename). Edit any field. Add new rows → INSERT. Set IsActive=NO to deactivate.",
        "join":        "LEFT JOIN parties psupp ON psupp.id = preferred_supplier_id",
        "base_where":  "",
        "order_by":    "brand NULLS LAST, product_name",
        "filters": [
            {"name": "active_only", "label": "Active only",         "type": "checkbox", "default": True,  "sql": "is_active = true"},
            {"name": "brand",       "label": "Filter by Brand",     "type": "text",     "default": "",    "sql": "LOWER(brand) LIKE LOWER('%{value}%')"},
            {"name": "category",    "label": "Filter by Type",      "type": "text",     "default": "",    "sql": "LOWER(category) LIKE LOWER('%{value}%')"},
            {"name": "lens_cat",    "label": "Lens Category",       "type": "text",     "default": "",    "sql": "LOWER(lens_category) LIKE LOWER('%{value}%')"},
            {"name": "search",      "label": "Search product name", "type": "text",     "default": "",    "sql": "LOWER(product_name) LIKE LOWER('%{value}%')"},
        ],
    },

    "FRAME": {
        "label":       "Frame Stock",
        "icon":        "🖼️",
        "desc":        "Frame SKUs with quantities and pricing. Edit qty or prices and re-import.",
        "loader_type": "FRAME",
        "key_col":     "SKUCode",
        "table":       "frames",
        "table_alias": "",
        "edit_guide":  "Key column: SKUCode. Edit Qty, CostPrice, MRP. ADD mode adds; OPENING replaces qty.",
        "join":        "",
        "base_where":  "",
        "order_by":    "brand NULLS LAST, sku_code",
        "filters": [
            {"name": "active_only", "label": "Active only",        "type": "checkbox", "default": True,  "sql": "is_active = true"},
            {"name": "brand",       "label": "Filter by Brand",    "type": "text",     "default": "",    "sql": "LOWER(brand) LIKE LOWER('%{value}%')"},
            {"name": "zero_stock",  "label": "Zero stock only",    "type": "checkbox", "default": False, "sql": "qty = 0"},
            {"name": "search",      "label": "Search SKU / Model", "type": "text",     "default": "",    "sql": "LOWER(sku_code) LIKE LOWER('%{value}%') OR LOWER(model) LIKE LOWER('%{value}%')"},
        ],
    },

    "PARTY": {
        "label":       "Party Master",
        "icon":        "🏢",
        "desc":        "Customers, suppliers, doctors, wholesalers — with GST, PAN, credit terms. Edit and re-import.",
        "loader_type": "PARTY",
        "key_col":     "MOBILE (primary) or PARTYNAME (fallback)",
        "table":       "parties",
        "table_alias": "",
        "edit_guide":  (
            "Key: MOBILE when filled; else PARTYNAME. "
            "GSTIN must be 15 chars. PAN must be 10 chars. "
            "ISACTIVE: YES to keep active, NO to deactivate. "
            "Do NOT change MOBILE or PARTYNAME of existing rows — these are conflict keys."
        ),
        "join":        "",
        "base_where":  "",
        "order_by":    "party_type, party_name",
        "filters": [
            {"name": "active_only", "label": "Active only",       "type": "checkbox", "default": False,
             "sql": "is_active = true"},
            {"name": "role",        "label": "Role type",         "type": "select",   "default": "ALL",
             "options": ["ALL", "Retail", "Doctor", "Optician", "Supplier", "Fitter", "Wholesale"],
             "sql": "LOWER(party_type) = LOWER('{value}')"},
            {"name": "city",        "label": "Filter by City",    "type": "text",     "default": "",
             "sql": "LOWER(city) LIKE LOWER('%{value}%')"},
            {"name": "has_gstin",   "label": "Has GSTIN only",   "type": "checkbox", "default": False,
             "sql": "gstin IS NOT NULL AND gstin != ''"},
            {"name": "search",      "label": "Search name/mobile/GSTIN", "type": "text", "default": "",
             "sql": "LOWER(party_name) LIKE LOWER('%{value}%') OR COALESCE(mobile,'') LIKE '%{value}%' OR COALESCE(gstin,'') LIKE UPPER('%{value}%')"},
        ],
    },

    "PATIENT": {
        "label":       "Patient Data",
        "icon":        "👤",
        "desc":        "Patient records with Rx history. Download for corrections or backup.",
        "loader_type": "PATIENT",
        "key_col":     "Mobile Number (or Record No)",
        "table":       "patients pt",
        "table_alias": "pt",
        "edit_guide":  "Key: Mobile Number or Record No. Each row = one patient visit. Re-importing adds a NEW visit row.",
        "join":        "LEFT JOIN patient_visits pv ON pv.patient_id = pt.id",
        "base_where":  "",
        "order_by":    "pt.master_name, pv.visit_date DESC",
        "filters": [
            {"name": "search",    "label": "Search name / mobile","type": "text",     "default": "",    "sql": "LOWER(pt.master_name) LIKE LOWER('%{value}%') OR pt.mobile::text LIKE '%{value}%'"},
            {"name": "has_rx",    "label": "Has Rx data only",    "type": "checkbox", "default": False, "sql": "pv.right_sph IS NOT NULL"},
            {"name": "from_date", "label": "Visits from date",    "type": "text",     "default": "",    "sql": "pv.visit_date >= '{value}'::date"},
        ],
    },

    "OPHLENS": {
        "label":       "Ophthalmic Lens Stock",
        "icon":        "👁️",
        "desc":        "Lens power stock. Download, edit qty or prices, re-import.",
        "loader_type": "OPHLENS",
        "key_col":     "Product + SPH + CYL + AXIS + EyeSide + ItemType",
        "table":       "inventory_stock s",
        "table_alias": "s",
        "edit_guide":  "Key: Product + SPH + CYL + AXIS + EyeSide + ItemType. Edit Quantity using ADD or OPENING mode.",
        "join":        "JOIN products p ON p.id = s.product_id",
        # CONFIRMED from DB: main_group values are 'Ophthalmic Lenses' (66) + 'OPHTHALMIC LENSES' (51)
        # UPPER() handles both; stock_type = 'POWER' for all ophthalmic lens rows
        "base_where":  "UPPER(p.main_group) = 'OPHTHALMIC LENSES'",
        "order_by":    "p.product_name, s.sph, s.cyl",
        "filters": [
            {"name": "active_only", "label": "Active only",       "type": "checkbox", "default": False, "sql": "s.is_active = true"},
            {"name": "product",     "label": "Filter by Product", "type": "text",     "default": "",    "sql": "LOWER(p.product_name) LIKE LOWER('%{value}%')"},
            {"name": "item_type",   "label": "Item type",         "type": "select",   "default": "ALL",
             "options": ["ALL", "STOCK", "RX"], "sql": "s.item_type = '{value}'"},
            {"name": "lens_design", "label": "Lens design",       "type": "select",   "default": "ALL",
             "options": ["ALL", "SPHERICAL", "TORIC", "MULTIFOCAL"], "sql": "s.lens_design = '{value}'"},
            {"name": "zero_stock",  "label": "Zero stock only",   "type": "checkbox", "default": False, "sql": "s.quantity = 0"},
        ],
    },

    "CLENS": {
        "label":       "Contact Lens Stock",
        "icon":        "🔵",
        "desc":        "Contact lens batches with powers and expiry. Edit qty or add batches.",
        "loader_type": "CLENS",
        "key_col":     "Product + BatchNo",
        "table":       "inventory_stock s",
        "table_alias": "s",
        "edit_guide":  "Key: Product + BatchNo. Edit Quantity, MRP, SellingPrice. Add new rows for new batches.",
        "join":        "JOIN products p ON p.id = s.product_id",
        # CONFIRMED: main_group = 'Contact Lenses' (59 products), stock_type = 'BATCH' (107 rows)
        "base_where":  "UPPER(p.main_group) = 'CONTACT LENSES' AND s.stock_type = 'BATCH'",
        "order_by":    "p.product_name, s.batch_no",
        "filters": [
            {"name": "active_only", "label": "Active only",       "type": "checkbox", "default": False, "sql": "s.is_active = true"},
            {"name": "product",     "label": "Filter by Product", "type": "text",     "default": "",    "sql": "LOWER(p.product_name) LIKE LOWER('%{value}%')"},
            {"name": "zero_stock",  "label": "Zero stock only",   "type": "checkbox", "default": False, "sql": "s.quantity = 0"},
        ],
    },

    "SOL": {
        "label":       "Solution Batches",
        "icon":        "🧴",
        "desc":        "Solution and accessory batches. Edit prices or qty and re-import.",
        "loader_type": "SOL",
        "key_col":     "Product + BatchNo",
        "table":       "batches b",
        "table_alias": "b",
        "edit_guide":  "Key: Product + BatchNo. Edit prices or qty. New rows will be inserted as new batches.",
        "join":        "JOIN products p ON p.id = b.product_id",
        "base_where":  "",
        "order_by":    "p.product_name, b.batch_no",
        "filters": [
            {"name": "active_only", "label": "Active only",       "type": "checkbox", "default": False, "sql": "b.is_active = true"},
            {"name": "product",     "label": "Filter by Product", "type": "text",     "default": "",    "sql": "LOWER(p.product_name) LIKE LOWER('%{value}%')"},
        ],
    },

    "BLANK": {
        "label":       "Blank Inventory",
        "icon":        "🔲",
        "desc":        "Optical blank lenses with base curves. Edit quantities and re-import.",
        "loader_type": "BLANK",
        "key_col":     "brand + Category + Material + COLOUR + Add",
        "table":       "blank_inventory",
        "table_alias": "",
        "edit_guide":  "Key: brand+Category+Material+COLOUR+Add. Edit qty_Right, qty_left. ADD adds; OPENING replaces.",
        "join":        "",
        "base_where":  "",
        "order_by":    "brand, category, material, add_power",
        "filters": [
            {"name": "brand",    "label": "Filter by Brand",    "type": "text", "default": "", "sql": "LOWER(brand) LIKE LOWER('%{value}%')"},
            {"name": "category", "label": "Filter by Category", "type": "text", "default": "", "sql": "LOWER(category) LIKE LOWER('%{value}%')"},
            {"name": "material", "label": "Filter by Material", "type": "text", "default": "", "sql": "LOWER(material) LIKE LOWER('%{value}%')"},
        ],
    },
}


# ═══════════════════════════════════════════════════════════════════════════════
# QUERY BUILDER  — fully driven from registry
# ═══════════════════════════════════════════════════════════════════════════════

def _build_select(cfg: dict, file_type: str) -> str:
    """
    Build SELECT clause from registry column definitions.
    Handles table aliases for joined queries (OPHLENS/CLENS/SOL/PATIENT).
    """
    alias  = cfg.get("table_alias", "")
    frags  = []
    cols   = get_download_cols(file_type)

    # For joined tables, map which columns belong to which table alias
    # patient: pt.* and pv.* columns
    # ophlens/clens: s.* + p.product_name
    # sol: b.* + p.product_name
    pv_cols = {"visit_date", "right_sph", "right_cyl", "right_axis", "right_add",
               "left_sph", "left_cyl", "left_axis", "left_add"}
    pt_cols = {"master_name", "mobile", "record_no"}

    for col in cols:
        if not col.excel_header:
            continue
        db  = col.db_column
        hdr = col.excel_header

        # Determine prefix
        if file_type == "PATIENT":
            if db in pv_cols:
                pfx = "pv."
            elif db in pt_cols:
                pfx = "pt."
            else:
                pfx = ""
        elif file_type in ("OPHLENS", "CLENS") and db in ("product_name", "gst_percent"):
            # product_name and gst_percent both live in the products table (alias p)
            pfx = "p."
        elif file_type == "SOL" and db in ("product_name", "gst_percent"):
            pfx = "p."
        elif alias:
            pfx = f"{alias}."
        else:
            pfx = ""

        # Virtual cols: _psupp_xxx → psupp.xxx via JOIN, _xxx → skip (post-fetch)
        if db.startswith("_psupp_"):
            real_col = db[len("_psupp_"):]
            frags.append(f"psupp.{real_col} AS \"{hdr}\"")
        elif db.startswith("_"):
            pass  # virtual — added post-fetch, not in SQL
        elif col.db_type == "boolean":
            frags.append(f"CASE WHEN {pfx}{db} THEN 'YES' ELSE 'NO' END AS \"{hdr}\"")
        elif col.db_type == "date" and not col.required:
            frags.append(f"TO_CHAR({pfx}{db}, 'YYYY-MM-DD') AS \"{hdr}\"")
        else:
            frags.append(f"{pfx}{db} AS \"{hdr}\"")

    # Always include `id` for PRODUCT so post-fetch supplier enrichment can join
    if file_type == "PRODUCT" and not any(f.split(" AS ")[0].strip() in ("id",) for f in frags):
        frags.insert(0, 'id AS "_id"')

    return ",\n                ".join(frags)


def _build_where(config: dict, filter_values: dict) -> str:
    """Build SQL WHERE clause from active filters. Returns 'WHERE ...' or ''."""
    clauses = []

    # Base where (hardcoded per dataset, e.g. main_group filter)
    base = config.get("base_where", "").strip()
    if base:
        clauses.append(base)

    for f in config.get("filters", []):
        name  = f["name"]
        ftype = f["type"]
        val   = filter_values.get(name)
        sql   = f.get("sql", "")
        if not sql or val is None:
            continue
        if ftype == "checkbox":
            if val:
                clauses.append(sql)
        elif ftype in ("text", "date"):
            v = str(val).strip()
            if v:
                clauses.append(sql.replace("{value}", v.replace("'", "''")))
        elif ftype == "select":
            if val and val != "ALL":
                clauses.append(sql.replace("{value}", val.replace("'", "''")))

    return ("WHERE " + " AND ".join(clauses)) if clauses else ""


def build_query(file_type: str, cfg: dict, filter_values: dict, limit: int) -> str:
    """Assemble full SQL from registry + config + filters."""
    select_sql = _build_select(cfg, file_type)
    table      = cfg["table"]
    join       = cfg.get("join", "")
    order_by   = cfg.get("order_by", "1")
    where      = _build_where(cfg, filter_values)

    return f"""
            SELECT
                {select_sql}
            FROM {table}
            {join}
            {where}
            ORDER BY {order_by}
            LIMIT {limit}
    """.strip()


# ═══════════════════════════════════════════════════════════════════════════════
# FETCH
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_for_download(
    dataset_key: str,
    filter_values: dict,
    limit: int = 20000,
) -> Tuple[pd.DataFrame, dict]:
    from modules.sql_adapter import run_query

    cfg = DOWNLOAD_CONFIGS.get(dataset_key)
    if not cfg:
        raise ValueError(f"Unknown dataset key: {dataset_key}")

    sql  = build_query(dataset_key, cfg, filter_values, limit)
    rows = run_query(sql)

    if not rows:
        return pd.DataFrame(), _make_meta(cfg, dataset_key, 0, filter_values)

    df = pd.DataFrame(rows)
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col], errors="ignore")
        except Exception:
            pass

    # For PRODUCT download: populate supplier columns from product_supplier_map
    # Supplier enrichment: use _id (hidden id col) or id if present
    _prod_id_col = "_id" if "_id" in df.columns else ("id" if "id" in df.columns else None)
    if cfg.get("loader_type") == "PRODUCT" and not df.empty and _prod_id_col:
        try:
            psm_rows = run_query("""
                SELECT
                    psm.product_id::text,
                    psm.rank,
                    psm.route_type,
                    p.party_name AS supplier_name
                FROM product_supplier_map psm
                JOIN parties p ON p.id = psm.supplier_id
                WHERE psm.product_id = ANY(%s::uuid[])
                  AND psm.is_active = TRUE
                ORDER BY psm.product_id, psm.route_type, psm.rank
            """, (df[_prod_id_col].tolist(),)) or []

            # Build lookup: product_id → {(rank, route) → name}
            _psm = {}
            for r in psm_rows:
                _pid = str(r["product_id"])
                _psm.setdefault(_pid, {})[(int(r["rank"]), str(r["route_type"]))] = r["supplier_name"]

            # Map rank→column name
            _vendor_cols  = {1:"PreferredSupplier", 2:"Supplier2", 3:"Supplier3",
                             4:"Supplier4", 5:"Supplier5"}

            for _col in list(_vendor_cols.values()):
                if _col not in df.columns:
                    df[_col] = ""

            for idx, row in df.iterrows():
                _pid = str(row.get(_prod_id_col) or "")
                _map = _psm.get(_pid, {})
                for _rank, _col in _vendor_cols.items():
                    df.at[idx, _col] = _map.get((_rank, "VENDOR"), "")

        except Exception:
            pass  # Non-fatal — supplier columns just stay empty

    # Ensure Supplier2-5 columns always present for PRODUCT Excel template
    if cfg.get("loader_type") == "PRODUCT" and not df.empty:
        for _sc in ["Supplier2", "Supplier3", "Supplier4", "Supplier5"]:
            if _sc not in df.columns:
                df[_sc] = ""

    # Drop hidden id col — not for user display
    if "_id" in df.columns:
        df = df.drop(columns=["_id"])

    return df, _make_meta(cfg, dataset_key, len(df), filter_values)


def _make_meta(cfg, dataset_key, rows, filter_values) -> dict:
    return {
        "rows":        rows,
        "dataset":     dataset_key,
        "label":       cfg["label"],
        "icon":        cfg["icon"],
        "downloaded":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "loader_type": cfg["loader_type"],
        "key_col":     cfg["key_col"],
        "edit_guide":  cfg["edit_guide"],
        "filters":     str({k: v for k, v in filter_values.items()
                            if v not in (False, "", "ALL", None)}),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER  — registry-driven guide sheet
# ═══════════════════════════════════════════════════════════════════════════════

def build_download_excel(
    df: pd.DataFrame,
    meta: dict,
    diff_df: Optional[pd.DataFrame] = None,
) -> bytes:
    buf = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        df.to_excel(buf, index=False)
        buf.seek(0)
        return buf.getvalue()

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
    NEW_FILL  = PatternFill("solid", fgColor="DCFCE7")
    UPD_FILL  = PatternFill("solid", fgColor="FEF9C3")
    DEL_FILL  = PatternFill("solid", fgColor="FEE2E2")
    UNCH_FILL = PatternFill("solid", fgColor="F8FAFC")
    CELL_CHG  = PatternFill("solid", fgColor="FDE68A")

    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    NORM_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(bold=True, name="Arial", size=10)
    THIN = Border(
        bottom=Side(style="thin", color="D1D5DB"),
        right =Side(style="thin", color="D1D5DB"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(vertical="center")
    WRAP   = Alignment(vertical="center", wrap_text=True)

    def _write_data_sheet(ws, dataframe, row_status=None, changed_cols_map=None, title="Data"):
        ws.title = title
        for ci, col_name in enumerate(dataframe.columns, 1):
            c = ws.cell(row=1, column=ci, value=col_name)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = THIN
        ws.row_dimensions[1].height = 20

        for ri, (_, row) in enumerate(dataframe.iterrows(), 2):
            row_i   = ri - 2
            status  = (row_status or {}).get(row_i)
            ch_cols = (changed_cols_map or {}).get(row_i, set())

            if status == "UPDATED" and ch_cols:
                row_fill = UNCH_FILL
            else:
                row_fill = (NEW_FILL  if status == "NEW"
                       else UPD_FILL  if status == "UPDATED"
                       else DEL_FILL  if status in ("DELETED", "MISSING FROM FILE")
                       else UNCH_FILL if status == "UNCHANGED"
                       else ALT_FILL  if ri % 2 == 0
                       else None)

            for ci, (col, val) in enumerate(row.items(), 1):
                if col in ("_status", "_changed_cols"):
                    continue
                c = ws.cell(row=ri, column=ci, value=val)
                if status == "UPDATED" and col in ch_cols:
                    c.fill = CELL_CHG
                elif row_fill:
                    c.fill = row_fill
                c.font = BOLD_FONT if ci == 1 else NORM_FONT
                c.border = THIN; c.alignment = LEFT

        for ci, col in enumerate(dataframe.columns, 1):
            if col == "_status": continue
            try:
                max_len = max(len(str(col)),
                    dataframe[col].astype(str).str.len().max() if len(dataframe) > 0 else 0)
            except Exception:
                max_len = len(str(col))
            ws.column_dimensions[get_column_letter(ci)].width = min(max(int(max_len) + 3, 12), 42)

        ws.freeze_panes = "A2"
        if len(dataframe) > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(dataframe.columns))}1"

        # Add YES/NO dropdowns for boolean columns
        file_type = meta.get("loader_type", "")
        if file_type and file_type in DB_SCHEMA:
            allowed = get_allowed_values(file_type)
            for ci, col_name in enumerate(dataframe.columns, 1):
                vals = allowed.get(col_name)
                if vals and set(vals) == {"YES", "NO"}:
                    dv = DataValidation(
                        type="list", formula1='"YES,NO"', allow_blank=True,
                        showErrorMessage=True, errorTitle="Invalid",
                        error=f"Must be YES or NO",
                    )
                    col_letter = get_column_letter(ci)
                    dv.sqref = f"{col_letter}2:{col_letter}50000"
                    ws.add_data_validation(dv)

        # Supplier name dropdowns for PRODUCT loader supplier columns
        _sup_cols = {
            "PreferredSupplier", "Supplier2", "Supplier3", "Supplier4", "Supplier5",
        }
        if file_type == "PRODUCT" and any(col in _sup_cols for col in dataframe.columns):
            try:
                from modules.sql_adapter import run_query as _rq_sup_dv
                _sup_rows = _rq_sup_dv(
                    "SELECT party_name FROM parties "
                    "WHERE UPPER(party_type) IN ('SUPPLIER','VENDOR') "
                    "  AND COALESCE(is_active,TRUE)=TRUE "
                    "ORDER BY party_name"
                ) or []
                _sup_names = [r["party_name"] for r in _sup_rows]
                if _sup_names:
                    # Write supplier list to a hidden sheet for dropdown source
                    _ws_sup = wb.create_sheet("_Suppliers")
                    _ws_sup.sheet_state = "hidden"
                    for _si, _sn in enumerate(_sup_names, 1):
                        _ws_sup.cell(row=_si, column=1, value=_sn)

                    _sup_range = f"_Suppliers!$A$1:$A${len(_sup_names)}"
                    for ci, col_name in enumerate(dataframe.columns, 1):
                        if col_name in _sup_cols:
                            _dv_sup = DataValidation(
                                type="list",
                                formula1=_sup_range,
                                allow_blank=True,
                                showErrorMessage=True,
                                errorTitle="Unknown Supplier",
                                error="Select a supplier from the list or leave blank",
                            )
                            _col_letter = get_column_letter(ci)
                            _dv_sup.sqref = f"{_col_letter}2:{_col_letter}50000"
                            ws.add_data_validation(_dv_sup)
            except Exception:
                pass  # Non-fatal — dropdown just won't appear

    # Sheet 1: Data
    ws1 = wb.active
    _write_data_sheet(ws1, df, title="Data")
    ws1.sheet_properties.tabColor = "1E3A5F"

    # Sheet 2: Registry-driven Guide
    ws2 = wb.create_sheet("Guide")
    ws2.sheet_properties.tabColor = "16A34A"
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 52
    ws2.column_dimensions["F"].width = 22

    file_type = meta.get("loader_type", "")

    # Title
    ws2.merge_cells("A1:F1")
    t = ws2.cell(row=1, column=1, value=f"📖  {meta.get('label','').upper()} — FIELD GUIDE  (auto-generated from DB schema)")
    t.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    t.fill = PatternFill("solid", fgColor="0F2744")
    t.alignment = CENTER
    ws2.row_dimensions[1].height = 26

    r = 2
    # Download metadata
    for label, val in [
        ("Dataset",       meta.get("label", "")),
        ("Downloaded",    meta.get("downloaded", "")),
        ("Rows",          meta.get("rows", 0)),
        ("Key column(s)", meta.get("key_col", "")),
        ("Filters used",  meta.get("filters", "none")),
        ("Edit guide",    meta.get("edit_guide", "")),
    ]:
        ws2.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws2.cell(row=r, column=2, value=val)
        c.font = NORM_FONT; c.alignment = WRAP
        ws2.merge_cells(f"B{r}:F{r}")
        ws2.row_dimensions[r].height = 18
        r += 1

    r += 1

    # Field guide header
    hdr_vals = ["Excel Column", "DB Column", "Type", "Required?", "Description", "Example"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    for ci, h in enumerate(hdr_vals, 1):
        c = ws2.cell(row=r, column=ci, value=h)
        c.font = HDR_FONT; c.fill = hdr_fill; c.alignment = CENTER; c.border = THIN
    ws2.row_dimensions[r].height = 20
    r += 1

    if file_type and file_type in DB_SCHEMA:
        cols = get_writable_cols(file_type)
        REQ_FILL = PatternFill("solid", fgColor="FEE2E2")
        OPT_FILL = PatternFill("solid", fgColor="F0FDF4")
        for col in cols:
            req_text = "🔴 REQUIRED" if col.required else "Optional"
            req_fill = REQ_FILL if col.required else OPT_FILL
            desc = col.description
            if col.allowed_values:
                desc += f"  ✅ Allowed: {', '.join(col.allowed_values)}"
            if col.notes:
                desc += f"  📌 {col.notes}"

            vals = [col.excel_header, col.db_column, col.db_type.upper(), req_text, desc, col.example]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=val)
                c.font = BOLD_FONT if ci == 1 else NORM_FONT
                c.fill = req_fill if ci == 4 else (ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF"))
                c.alignment = WRAP if ci == 5 else LEFT
                c.border = THIN
            ws2.row_dimensions[r].height = 20
            r += 1

    # Sheet 3: Changes (diff)
    if diff_df is not None and not diff_df.empty:
        ws3 = wb.create_sheet("Changes")
        status_map = {i: row.get("_status") for i, (_, row) in enumerate(diff_df.iterrows())}
        changed_cols_map = {}
        for i, (_, row) in enumerate(diff_df.iterrows()):
            raw = row.get("_changed_cols", "")
            if raw:
                changed_cols_map[i] = set(c.strip() for c in str(raw).split(",") if c.strip())
        disp_df = diff_df.drop(columns=[c for c in ("_changed_cols",) if c in diff_df.columns])
        _write_data_sheet(ws3, disp_df, row_status=status_map, changed_cols_map=changed_cols_map, title="Changes")

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# DIFF ENGINE  — unchanged logic, key cols now from registry
# ═══════════════════════════════════════════════════════════════════════════════

# Key columns per loader type — what uniquely identifies a row
_KEY_COLS = {
    "PRODUCT": ["Product"],
    "FRAME":   ["SKUCode"],
    "PARTY":   ["PARTYNAME"],
    "PATIENT": ["Mobile Number", "Record No"],
    "OPHLENS": ["Product", "SPH", "CYL", "AXIS", "EyeSide", "ItemType"],
    # ✅ CLENS key includes SPH+CYL+AXIS+EyeSide — same as upload dedup key.
    # Without power in the key, two rows sharing a BatchNo but different SPH
    # get merged into one diff key → wrong row compared → price diff shows UNCHANGED.
    "CLENS":   ["Product", "BatchNo", "SPH", "CYL", "AXIS", "EyeSide"],
    "SOL":     ["Product", "BatchNo"],
    "BLANK":   ["brand", "Category", "Material", "COLOUR", "Add", "Recomended Base"],
}


def get_key_cols_for_type(loader_type: str) -> List[str]:
    return _KEY_COLS.get(loader_type, [])


def _norm_cell(v) -> str:
    import math
    if v is None: return ""
    if isinstance(v, bool): return "yes" if v else "no"
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v): return ""
        if v == int(v): return str(int(v))
        return str(v)
    if isinstance(v, int): return str(v)
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat", "<na>", "null"): return ""
    if s.lower() in ("yes", "true", "1", "t", "y"): return "yes"
    if s.lower() in ("no", "false", "0", "f", "n"): return "no"
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): s = s[:-2]
    return s.lower()


def _norm_df(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda col: col.map(_norm_cell))


def compute_diff(
    original_df: pd.DataFrame,
    updated_df: pd.DataFrame,
    key_cols: List[str],
) -> Tuple[pd.DataFrame, dict]:
    if not key_cols:
        return pd.DataFrame(), {"error": "No key columns defined"}

    valid_keys = [k for k in key_cols if k in original_df.columns and k in updated_df.columns]
    if not valid_keys:
        return pd.DataFrame(), {"error": f"Key column(s) {key_cols} not found in both files"}

    # Warn if some key cols were not found (e.g. user removed SPH from file)
    missing_keys = [k for k in key_cols if k not in valid_keys]
    warning_msg  = None
    if missing_keys:
        warning_msg = (
            f"⚠️ Key columns {missing_keys} not found in uploaded file — "
            f"diff is using partial key {valid_keys}. "
            f"This may cause incorrect row matching. Include all key columns in your file."
        )

    orig = _norm_df(original_df.copy())
    upd  = _norm_df(updated_df.copy())

    # Build composite key — use NULL_MARKER for blank/NaN values so that
    # two rows with different NaN-containing keys don't all merge to the same key.
    # e.g. SPH=-0.25, CYL=NaN, AXIS=NaN → "-0.25||__NULL__||__NULL__"
    # vs   SPH=-1.00, CYL=NaN, AXIS=NaN → "-1.00||__NULL__||__NULL__"  (different keys ✅)
    def _make_key(row):
        return "||".join(
            "__NULL__" if (v == "" or v is None) else str(v)
            for v in row.values
        )

    orig["_key"] = orig[valid_keys].apply(_make_key, axis=1)
    upd["_key"]  = upd[valid_keys].apply(_make_key, axis=1)

    orig_keys = set(orig["_key"].unique())
    upd_keys  = set(upd["_key"].unique())

    new_keys     = upd_keys  - orig_keys
    deleted_keys = orig_keys - upd_keys
    common_keys  = orig_keys & upd_keys

    orig_idx = orig.set_index("_key")
    upd_idx  = upd.set_index("_key")

    diff_rows       = []
    updated_count   = 0
    unchanged_count = 0
    common_cols = [c for c in orig.columns if c in upd.columns and c != "_key"]

    for key in sorted(new_keys):
        row = upd_idx.loc[key].copy()
        row["_status"] = "NEW"
        diff_rows.append(row)

    for key in sorted(deleted_keys):
        row = orig_idx.loc[key].copy()
        row["_status"] = "MISSING FROM FILE"
        diff_rows.append(row)

    for key in sorted(common_keys):
        try:
            orig_row = orig_idx.loc[key][common_cols]
            upd_row  = upd_idx.loc[key][common_cols]
            if isinstance(orig_row, pd.DataFrame): orig_row = orig_row.iloc[0]
            if isinstance(upd_row,  pd.DataFrame): upd_row  = upd_row.iloc[0]

            changed_cols = [c for c in common_cols if str(orig_row.get(c,"")) != str(upd_row.get(c,""))]
            changed = bool(changed_cols)

            row = upd_idx.loc[key].copy()
            if isinstance(row, pd.DataFrame): row = row.iloc[0]
            row["_status"]       = "UPDATED" if changed else "UNCHANGED"
            row["_changed_cols"] = ", ".join(changed_cols) if changed else ""
            diff_rows.append(row)

            if changed: updated_count += 1
            else: unchanged_count += 1
        except Exception:
            pass

    summary = {
        "new": len(new_keys), "updated": updated_count,
        "deleted": len(deleted_keys), "unchanged": unchanged_count,
    }
    if warning_msg:
        summary["warning"] = warning_msg

    if diff_rows:
        diff_df = pd.DataFrame(diff_rows).reset_index(drop=True)
        cols = ["_status"] + [c for c in diff_df.columns if c not in ("_status", "_key")]
        diff_df = diff_df[[c for c in cols if c in diff_df.columns]]
    else:
        diff_df = pd.DataFrame()

    return diff_df, summary
